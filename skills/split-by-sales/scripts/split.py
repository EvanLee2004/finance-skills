# -*- coding: utf-8 -*-
"""
应收 all 按销售人员拆分 · split.py （财务部 Agent 技能）
================================================================
输入：一张"应收all"（receivables-merge 的产物，含销售人员等 17 列）。
输出：按【销售人员】拆成一人一份带下拉框的 Excel（亮晶姐模板：前8列蓝、后9列黄、下拉、冻结首行）。
口径（项目五已验收，会变的放 config/拆分规则.md）：
  - 坏账桶（如 高美杰1）整行忽略、不分给任何人；
  - 名字形如「X-高美杰」的行归到前头的人 X；其中"GM 单独成 sheet"的接手人（如于占国）的高美杰行
    单独放进「GM订单」sheet，其余人（如梁玲玲）直接并入主表；
  - 销售人员为空 → 单独出「_销售人员为空_请人工处理.xlsx」；
  - 跑完做对账：分出去的 + 空名 + 忽略 == 输入行数。

用法：python3 split.py --input <应收all.xlsx> [--out-dir <目录>] [--date 0604]
缺文件/格式不对会清晰报错，不裸崩。
"""
import os
import re
import sys
import argparse
import datetime

HERE = os.path.dirname(os.path.abspath(__file__))
SKILL_DIR = os.path.dirname(HERE)
CONFIG_DIR = os.path.join(SKILL_DIR, "config")
WORK_INPUT = os.path.join(SKILL_DIR, "工作区", "input")
WORK_OUTPUT = os.path.join(SKILL_DIR, "工作区", "output")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("✗ 缺 openpyxl → pip install openpyxl")
    sys.exit(1)

# ===== 模板（亮晶姐模板，稳定；要改在这里）=====
GM_SUFFIX = "-高美杰"
GM_SHEET_NAME = "GM订单"
# 业务可变规则的内置默认（优先读 config/拆分规则.md 覆盖）
_DEFAULT_IGNORE = {"高美杰1"}
_DEFAULT_GM_OWNERS = {"于占国"}

HEADERS = [
    "年度", "销售人员", "客户名称", "新智云单号", "文件名", " 应收金额 ",
    "交付月份", "账龄(月份）", "结算阶段(请筛选分类)",
    "{d}销售预计回款日期\n（必须年月日，格式：20241210） ",
    "销售解释说明", "有无合同", " 合同分类\n（框架合同；具体金额） ",
    " 框架合同是否存在PO单或下单记录 ",
    " 应收金额是否有客户正式确认（盖章/对公邮件） ",
    "客户结算周期", "是否按月给客户发结算单",
]
DROPDOWNS = {
    9:  '"未对账,已对账，待开票,已对账，已开票,已开票，未回款,已开票，已回款,已回款，未核销,已回款，已核销"',
    12: '"有,无"',
    13: '"具体金额,1年框架,2年框架,3年框架,4年框架,5年框架,长期框架"',
    14: '"无,有单次报价PO单，通过邮件确认,有单次报价PO单，通过微信确认,有单次报价PO单，通过飞书确认,无单次报价PO单，通过微信/飞书聊天确认,其他"',
    15: '"有,无"',
    16: '"批次结,每月结,每季度,半年结,年结"',
    17: '"是,否"',
}
COL_WIDTHS = {"A": 7.8, "B": 12.2, "C": 20.1, "D": 14, "E": 30, "F": 15.8,
              "G": 15.8, "H": 10, "I": 17.6, "J": 20.1, "K": 24, "L": 9.6,
              "M": 21.0, "N": 37.1, "O": 21.0, "P": 13.1, "Q": 12}
HEADER_KEYS = ["年度", "销售人员", "客户名称", "新智云单号", "文件名", "应收金额",
               "交付月份", "账龄", "结算阶段", "回款日期", "销售解释", "有无合同",
               "合同分类", "PO单", "客户正式确认", "客户结算周期", "是否按月"]

THIN = Border(*[Side(style="thin", color="999999")] * 4)
HDR_FONT = Font(name="等线", size=10.5, bold=True)
HDR_FILL_BLUE = PatternFill("solid", fgColor="BDD7EE")
HDR_FILL_YELLOW = PatternFill("solid", fgColor="FFFF00")
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_FONT = Font(name="等线", size=10.5)


def log(m): print(m, flush=True)
def norm(s): return re.sub(r"\s+", "", str(s or ""))
def safe_name(s): return re.sub(r'[\\/:*?"<>|]', "_", str(s)).strip()


# ===== 读 config/拆分规则.md（活配置；缺/坏用默认）=====
def _parse_md_table(lines):
    rows, seen = [], False
    for ln in lines:
        s = ln.strip()
        if not s.startswith("|"):
            continue
        cells = [c.strip() for c in s.strip("|").split("|")]
        if cells and all(c and set(c) <= set("-: ") for c in cells):
            seen = True; continue
        if not seen:
            continue
        rows.append(cells)
    return rows


def load_split_rules():
    """读 config/拆分规则.md → (ignore_sales, gm_owners)。缺/坏 → 默认。"""
    p = os.path.join(CONFIG_DIR, "拆分规则.md")
    if not os.path.isfile(p):
        return set(_DEFAULT_IGNORE), set(_DEFAULT_GM_OWNERS)
    try:
        with open(p, encoding="utf-8") as f:
            text = f.read()
    except Exception as e:
        log(f"⚠ 读 拆分规则.md 失败({e})，用内置默认。")
        return set(_DEFAULT_IGNORE), set(_DEFAULT_GM_OWNERS)
    ignore, gm = set(), set()
    for part in re.split(r"(?m)^##\s+", text):
        ls = part.splitlines()
        head = ls[0] if ls else ""
        rows = _parse_md_table(ls)
        if "不分配" in head or "忽略" in head:
            for c in rows:
                if c and c[0]: ignore.add(c[0].strip())
        elif "GM" in head or "单独" in head:
            for c in rows:
                if c and c[0]: gm.add(c[0].strip())
    return (ignore or set(_DEFAULT_IGNORE)), (gm or set(_DEFAULT_GM_OWNERS))


def report_date(override=None):
    if override:
        return override
    return datetime.date.today().strftime("%m%d")


def _date_from_text(s):
    """从文件名/sheet名里抠日期 → 'MMDD'。如 '2026.6.4应收all' → '0604'；抠不到返回 None。"""
    m = re.search(r"(\d{4})[.\-/年]\s*(\d{1,2})[.\-/月]\s*(\d{1,2})", str(s or ""))
    return f"{int(m.group(2)):02d}{int(m.group(3)):02d}" if m else None


# ===== 输出样式 =====
def style_sheet(ws, n_rows, date_str):
    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(1, i, h.format(d=date_str) if "{d}" in h else h)
        c.font, c.alignment, c.border = HDR_FONT, HDR_ALIGN, THIN
        c.fill = HDR_FILL_BLUE if i <= 8 else HDR_FILL_YELLOW
    ws.row_dimensions[1].height = 40.5
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    last = max(n_rows + 300, 1000)
    for col, formula in DROPDOWNS.items():
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        letter = openpyxl.utils.get_column_letter(col)
        dv.add(f"{letter}2:{letter}{last + 1}")
        ws.add_data_validation(dv)


def write_rows(ws, rows):
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row[:17], 1):
            cell = ws.cell(r, c, v)
            cell.font = BODY_FONT
            cell.border = THIN
            if c == 6 and isinstance(v, (int, float)):
                cell.number_format = "#,##0.00"


# ===== 读输入 =====
def find_data_sheet(wb):
    dated, cands = [], []
    for name in wb.sheetnames:
        if "销售反馈" in name:
            continue
        ws = wb[name]
        hdr = [norm(ws.cell(1, i).value) for i in range(1, 18)]
        if sum(1 for k in HEADER_KEYS if any(k in h for h in hdr)) >= 12:
            cands.append((ws.max_row, name))
            if re.fullmatch(r"\d{4}[.年]\d{1,2}[.月]\d{1,2}日?", name.strip()):
                dated.append((ws.max_row, name))
    pool = dated or cands
    if not pool:
        return None
    pool.sort(reverse=True)
    if len(pool) > 1:
        log(f"  注意：多个候选 sheet {[n for _, n in pool]}，取「{pool[0][1]}」")
    return pool[0][1]


def map_columns(ws):
    actual = {c: norm(ws.cell(1, c).value) for c in range(1, ws.max_column + 1) if norm(ws.cell(1, c).value)}
    mapping = []
    for key in HEADER_KEYS:
        mapping.append(next((c for c, h in actual.items() if key in h and c not in mapping), None))
    return mapping


def read_rows(ws):
    mapping = map_columns(ws)
    missing = [HEADER_KEYS[i] for i, c in enumerate(mapping) if c is None]
    if missing:
        log(f"  ⚠ sheet「{ws.title}」缺列 {missing}，对应列留空")
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        row = [raw[c - 1] if (c and c <= len(raw)) else None for c in mapping]
        if any(v is not None and str(v).strip() != "" for v in row):
            rows.append(row)
    return rows, missing


# ===== 排序（让每人表整齐、好催账）=====
def _sort_rows(rows):
    """每人文件内排序：账龄(月份)【降序】——账龄最久的老账排最上面（最该催的在前）；
    账龄空的排最后；相同账龄按客户名归拢（同一客户的单挨在一起，整齐）。
    （要改排序键就改这里：如想按客户排，把 key 改成 (cust, ...)。）"""
    def key(r):
        try:
            a = int(r[7])              # 账龄(月份)
        except (TypeError, ValueError):
            a = None
        cust = str(r[2] or "")         # 客户名称
        return (0, -a, cust) if a is not None else (1, 0, cust)
    return sorted(rows, key=key)


# ===== 拆分主流程 =====
def do_split(input_path, out_dir, date_str, ignore_sales, gm_owners):
    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    sheet = find_data_sheet(wb)
    if not sheet:
        wb.close()
        raise RuntimeError("没找到数据 sheet（表头需含 年度/销售人员/客户名称… 这17列）。")
    log(f"· 数据 sheet：{sheet}")
    rows, missing = read_rows(wb[sheet])
    wb.close()
    log(f"· 读到 {len(rows)} 行")
    os.makedirs(out_dir, exist_ok=True)

    groups, gm_groups, unassigned, ignored_rows = {}, {}, [], []
    for row in rows:
        name = norm(row[1])
        if not name:
            unassigned.append(row); continue
        if name in ignore_sales:
            ignored_rows.append(row); continue
        if name.endswith(GM_SUFFIX) and len(name) > len(GM_SUFFIX):
            owner = name[:-len(GM_SUFFIX)]
            (gm_groups if owner in gm_owners else groups).setdefault(owner, []).append(row)
        else:
            groups.setdefault(name, []).append(row)

    all_names = sorted(set(groups) | set(gm_groups))
    per_person = []
    for name in all_names:
        main_rows = _sort_rows(groups.get(name, []))   # 账龄降序，整齐好催
        gm_rows = _sort_rows(gm_groups.get(name, []))
        out = openpyxl.Workbook(); ws = out.active; ws.title = safe_name(name)[:31]  # openpyxl sheet名≤31
        style_sheet(ws, len(main_rows), date_str); write_rows(ws, main_rows)
        if gm_rows:
            ws2 = out.create_sheet(GM_SHEET_NAME)
            style_sheet(ws2, len(gm_rows), date_str); write_rows(ws2, gm_rows)
        fname = f"{safe_name(name)}-应收{date_str}.xlsx"
        out.save(os.path.join(out_dir, fname))
        per_person.append((fname, len(main_rows), len(gm_rows)))
        log(f"  {fname}：{len(main_rows)}行" + (f"（含{GM_SHEET_NAME} {len(gm_rows)}行）" if gm_rows else ""))

    if unassigned:
        out = openpyxl.Workbook(); ws = out.active; ws.title = "未分配"
        style_sheet(ws, len(unassigned), date_str); write_rows(ws, _sort_rows(unassigned))
        out.save(os.path.join(out_dir, "_销售人员为空_请人工处理.xlsx"))
        log(f"  ⚠ {len(unassigned)} 行「销售人员」为空 → _销售人员为空_请人工处理.xlsx")

    total_out = sum(len(v) for v in groups.values()) + sum(len(v) for v in gm_groups.values())
    balanced = total_out + len(unassigned) + len(ignored_rows) == len(rows)
    log(f"· 对账：分出 {total_out} + 空名 {len(unassigned)} + 忽略 {len(ignored_rows)} = "
        f"{total_out + len(unassigned) + len(ignored_rows)} / 输入 {len(rows)} "
        f"{'✓ 对得上' if balanced else '✗ 对不上，请检查！'}")
    if ignored_rows:
        log(f"· 已按规则忽略（坏账桶等）{sorted(ignore_sales)} 共 {len(ignored_rows)} 行")
    return {
        "input_rows": len(rows), "people": len(all_names), "out_rows": total_out,
        "unassigned": len(unassigned), "ignored": len(ignored_rows),
        "balanced": balanced, "missing_cols": missing, "per_person": per_person,
        "out_dir": out_dir,
    }


def main():
    ap = argparse.ArgumentParser(description="应收 all 按销售人员拆分")
    ap.add_argument("--input", help="应收all.xlsx（绝对路径）")
    ap.add_argument("--out-dir", help="输出目录（绝对路径）")
    ap.add_argument("--date", help="文件名/表头日期，如 0604；默认运行当天")
    ap.add_argument("--input-dir", default=WORK_INPUT, help="没给 --input 时去这里找最新 xlsx")
    a = ap.parse_args()

    inp = a.input
    auto_picked = not a.input          # 没显式给 → 自动从 工作区/input 取
    if not inp:  # 没显式给 → 去 input-dir 找最新 xlsx
        cands = [os.path.join(a.input_dir, f) for f in sorted(os.listdir(a.input_dir))] \
            if os.path.isdir(a.input_dir) else []
        cands = [p for p in cands if p.lower().endswith(".xlsx") and not os.path.basename(p).startswith(("~$", "."))]
        if not cands:
            log(f"✗ 没给 --input，也没在 {a.input_dir}/ 找到 xlsx。"); sys.exit(1)
        inp = max(cands, key=os.path.getmtime)
    if not os.path.isfile(inp):
        log(f"✗ 输入文件不存在：{inp}"); sys.exit(1)
    if not inp.lower().endswith((".xlsx", ".xlsm")):
        log(f"✗ 输入要是 .xlsx（收到 {os.path.basename(inp)}）。"); sys.exit(1)

    ignore_sales, gm_owners = load_split_rules()
    date_str = a.date
    if not date_str:                   # 没显式给日期 → 优先从 all 文件名取（别用今天，免得文件名标错期）
        date_str = _date_from_text(os.path.basename(inp))
        if date_str:
            log(f"· 日期自动取自文件名：{date_str}（要改用 --date）")
        else:
            date_str = report_date()
            log(f"· ⚠ 文件名里没日期，用今天 {date_str}；当期不是今天的话请用 --date 指定！")
    if a.out_dir:
        out_dir = a.out_dir
    elif auto_picked:                  # 自动取的 → 结果落 工作区/output，不和输入混一起
        out_dir = os.path.join(WORK_OUTPUT, f"拆分_{date_str}")
    else:                              # 用户给了具体 all → 结果就放它旁边，好找
        out_dir = os.path.join(os.path.dirname(os.path.abspath(inp)), f"拆分_{date_str}")
    log(f"· 输入：{os.path.basename(inp)}")
    log(f"· 拆分规则：忽略 {sorted(ignore_sales)}；GM单独成sheet {sorted(gm_owners)}")
    try:
        rep = do_split(inp, out_dir, date_str, ignore_sales, gm_owners)
    except Exception as e:
        log(f"✗ 拆分出错：{e}"); sys.exit(1)
    log(f"\n✓ 完成：{rep['people']} 位销售，输出在 {rep['out_dir']}"
        + ("" if rep["balanced"] else "  ⚠ 对账没对上，先别发，检查！"))


if __name__ == "__main__":
    main()
