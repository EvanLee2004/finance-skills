# -*- coding: utf-8 -*-
"""
拆分 skill 稳定性回归测试（合成假数据，不依赖真实财务数据）。
跑：python3 tests/test_robustness.py   全过打印 ALL PASS
"""
import os
import sys
import subprocess
import tempfile
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SPLIT = os.path.join(HERE, "..", "scripts", "split.py")
PY = sys.executable
PASS = FAIL = 0


def check(d, c):
    global PASS, FAIL
    print(f"  {'✓' if c else '✗'} {d}")
    PASS += c == True
    FAIL += c != True


def run(args):
    r = subprocess.run([PY, SPLIT] + args, capture_output=True, text=True)
    return r.returncode, r.stdout + r.stderr


HEADERS17 = ["年度", "销售人员", "客户名称", "新智云单号", "文件名", "应收金额", "交付月份",
             "账龄(月份)", "结算阶段", "0604销售预计回款日期", "销售解释说明", "有无合同",
             "合同分类", "框架合同是否存在PO单", "应收金额是否有客户正式确认", "客户结算周期", "是否按月给客户发结算单"]


def make_all(path):
    """合成一张应收 all：含普通销售、于占国-高美杰(→GM)、梁玲玲-高美杰(→并入)、高美杰1(→忽略)、空销售(→未分配)。"""
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "2026.6.4"
    ws.append(HEADERS17)
    data = [
        ["2026", "张健", "甲公司", "SO1", "f1", 100, "202603", 2] + [""] * 9,
        ["2026", "张健", "乙公司", "SO2", "f2", 200, "202512", 8] + [""] * 9,   # 账龄不同，测排序
        ["2026", "张健", "丙公司", "SO2b", "f2b", 150, "202509", 5] + [""] * 9,
        ["2026", "于占国", "丙公司", "SO3", "f3", 300, "202603", 2] + [""] * 9,
        ["2018", "于占国-高美杰", "华电", "GM1", "f4", 400, "201803", 90] + [""] * 9,   # → 于占国 GM订单
        ["2018", "梁玲玲-高美杰", "二冶", "GM2", "f5", 500, "201803", 90] + [""] * 9,   # → 梁玲玲 主表
        ["2018", "高美杰1", "中国机械", "GM3", "f6", 600, "201803", 90] + [""] * 9,     # → 忽略
        ["2026", "", "丁公司", "SO7", "f7", 700, "202603", 2] + [""] * 9,               # → 未分配
    ]
    for r in data:
        ws.append(r)
    wb.save(path)


def main():
    tmp = tempfile.mkdtemp(prefix="split_test_")
    allp = os.path.join(tmp, "应收all.xlsx"); make_all(allp)
    out = os.path.join(tmp, "out")

    print("【1】正常拆分 → 对账对得上、GM/忽略/空名都对")
    rc, log = run(["--input", allp, "--out-dir", out, "--date", "0604"])
    check("退出码 0", rc == 0)
    check("对账对得上", "对得上" in log)
    check("于占国含 GM订单", "GM订单" in log)
    check("高美杰1 被忽略", "忽略" in log and "高美杰1" in log)
    check("空销售 → 待人工文件", "人工处理" in log)
    if os.path.isdir(out):
        files = os.listdir(out)
        check("于占国文件生成", any("于占国" in f for f in files))
        check("生成了'待人工'文件", any("人工" in f for f in files))
        check("高美杰1 没单独出文件", not any("高美杰1" in f for f in files))
        # 于占国文件应有 GM订单 sheet
        yf = next((f for f in files if "于占国" in f), None)
        if yf:
            wb = openpyxl.load_workbook(os.path.join(out, yf))
            check("于占国文件里有 GM订单 sheet", "GM订单" in wb.sheetnames)
        # 排序：张健文件账龄应【降序】（最该催的老账在前）
        zf = next((f for f in files if "张健" in f), None)
        if zf:
            ws = openpyxl.load_workbook(os.path.join(out, zf), data_only=True).active
            ages = [r[7] for r in ws.iter_rows(min_row=2, values_only=True) if isinstance(r[7], int)]
            check("张健文件账龄降序排好", len(ages) >= 2 and ages == sorted(ages, reverse=True))

    print("【2】缺输入文件 → 清晰报错(退出1)")
    rc, log = run(["--input", os.path.join(tmp, "nope.xlsx"), "--out-dir", out])
    check("退出码 1 且提示不存在", rc == 1 and "不存在" in log)

    print("【3】非 xlsx 当输入 → 清晰报错")
    txt = os.path.join(tmp, "x.txt"); open(txt, "w").write("x")
    rc, log = run(["--input", txt, "--out-dir", out])
    check("退出码 1 且提示要 .xlsx", rc == 1 and "xlsx" in log)

    print("【4】确定性：同输入两次 → 某人文件行数一致")
    o1, o2 = os.path.join(tmp, "o1"), os.path.join(tmp, "o2")
    run(["--input", allp, "--out-dir", o1, "--date", "0604"])
    run(["--input", allp, "--out-dir", o2, "--date", "0604"])

    def rowcount(d, who):
        f = next((x for x in os.listdir(d) if who in x), None)
        if not f:
            return -1
        ws = openpyxl.load_workbook(os.path.join(d, f)).active
        return ws.max_row
    check("两次张健文件行数一致", rowcount(o1, "张健") == rowcount(o2, "张健") > 1)

    print("【5】不给 --date → 自动从文件名取日期（不用今天）")
    dated = os.path.join(tmp, "2026.6.4应收all.xlsx"); make_all(dated)
    o5 = os.path.join(tmp, "o5")
    rc, log = run(["--input", dated, "--out-dir", o5])
    check("日志显示自动取自文件名", "自动取自文件名" in log and "0604" in log)
    if os.path.isdir(o5):
        check("产物文件名含 0604（非今天）", any("0604" in f for f in os.listdir(o5)))

    print(f"\n{'='*40}\n通过 {PASS} / 失败 {FAIL}  →  {'ALL PASS ✓' if FAIL == 0 else 'HAS FAILURES ✗'}")
    sys.exit(0 if FAIL == 0 else 1)


if __name__ == "__main__":
    main()
