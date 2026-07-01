#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""工资表清洗与信息匹配 · 回归测试
合成用例（金标）+ 真实数据冒烟（测试数据/ 在才跑，缺则跳过）。
跑：python3 tests/test_robustness.py
"""
import os
import sys
import tempfile
from collections import defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(os.path.dirname(HERE), "scripts"))
import clean_match as cm  # noqa
import openpyxl as opx

PASS = 0
FAIL = 0


def check_eq(label, got, want):
    global PASS, FAIL
    if got == want:
        PASS += 1
    else:
        FAIL += 1
        print(f"  ✗ {label}: got={got!r} want={want!r}")


def check_true(label, cond):
    check_eq(label, bool(cond), True)


PAYROLL_HEADER = ["序号", "姓名", "部门", "地区", "身份证号码", "基本工资", "养老", "失业",
                  "基本医疗", None, "住房公积金", "税前工资", "个人所得税", "实发工资"]


def _build_payroll_ws(wb, sheet_name, people, summary=None, trailing_junk=0):
    """people: list of dict(姓名,部门,地区,身份证号码,基本工资..实发工资等)。
    模拟真实结构：表头在第3行，person 行紧接着；可选一个『合计』行；trailing_junk 模拟合计前的杂散行数。"""
    ws = wb.create_sheet(sheet_name)
    ws.append(["工    资    表"])
    ws.append(["编制单位：xxx"])
    ws.append(PAYROLL_HEADER)
    for i, p in enumerate(people, 1):
        row = [i, p["姓名"], p.get("部门"), p.get("地区", "湖南"), p.get("身份证号码"),
               p.get("基本工资"), p.get("养老"), p.get("失业"), p.get("基本医疗"), None,
               p.get("住房公积金"), p.get("税前工资"), p.get("个人所得税"), p.get("实发工资")]
        ws.append(row)
    # 模拟真实数据里"最后一人和合计行之间"的杂散/空行
    for _ in range(trailing_junk):
        ws.append([None] * 14)
    if summary:
        row = ["合计", None, None, None, None] + [summary.get(k) for k in
            ("基本工资", "养老", "失业", "基本医疗", None, "住房公积金", "税前工资", "个人所得税", "实发工资")][:9]
        # 上面拼接容易出错，直接显式写
        row = ["合计", None, None, None, None,
               summary.get("基本工资"), summary.get("养老"), summary.get("失业"),
               summary.get("基本医疗"), None, summary.get("住房公积金"),
               summary.get("税前工资"), summary.get("个人所得税"), summary.get("实发工资")]
        ws.append(row)
    return ws


def test_sheet_detection_and_person_row_range():
    print("· sheet识别 + 人员行范围（含合计前杂散行、合计后VLOOKUP参考块不被读到）")
    cm.CONFIG.update({"SUMMARY_KEYWORDS": ["合计", "总计", "小计", "共计"], "BLANK_ROW_STOP": 10,
                       "TOLERANCE": 0.05, "DECIMALS": 2})
    wb = opx.Workbook()
    wb.remove(wb.active)
    people = [
        dict(姓名="谢柯", 部门="人力部", 身份证号码="430600000000000001",
             基本工资=8960, 养老=344.64, 失业=12.92, 基本医疗=101.16, 住房公积金=105,
             税前工资=8396.28, 个人所得税=101.89, 实发工资=8294.39),
        dict(姓名="戴文怡", 部门="资源部", 身份证号码="430600000000000002",
             基本工资=5015, 养老=344.64, 失业=12.92, 基本医疗=101.16, 住房公积金=105,
             税前工资=4451.28, 个人所得税=0, 实发工资=4451.28),
    ]
    summary = dict(基本工资=13975, 养老=689.28, 失业=25.84, 基本医疗=202.32, 住房公积金=210,
                   税前工资=12847.56, 个人所得税=101.89, 实发工资=12745.67)
    ws = _build_payroll_ws(wb, "6月湖南", people, summary=summary, trailing_junk=3)
    # 合计行之后再加一段VLOOKUP参考块（不同列位置放真实姓名），不应被当成人员
    r0 = ws.max_row
    for i, nm in enumerate(["谢柯", "戴文怡"]):
        ws.cell(r0 + 2 + i, 17, nm)  # 列Q，远超header宽度(14列)
        ws.cell(r0 + 2 + i, 18, 8960)

    payroll_alias, master_alias, company_map = cm.load_aliases()
    sheets = cm.detect_payroll_sheets(wb, payroll_alias)
    check_eq("识别到1个工资sheet", len(sheets), 1)
    cleaned, company, mismatches, warns = cm.clean_one_sheet(wb, sheets[0], payroll_alias, company_map)
    check_eq("正确读到2条人员行(不含合计/杂散/VLOOKUP块)", len(cleaned), 2)
    check_eq("合计校验通过(容差内)", mismatches, [])
    check_eq("所属公司按sheet名前缀映射", company, "湖南分公司")
    check_eq("J列(表头为空)未被当成有效字段读入", "None" not in [str(k) for k in cleaned[0]["numeric"].keys()], True)


def test_blank_row_stop_regression():
    """真实数据踩过的坑：阈值太小(2)会在扫到合计行之前就误停，导致合计校验被跳过。
       这里显式验证：3行杂散/空行 + 合计行 的结构下，阈值=10能读到合计，阈值=2读不到。"""
    print("· 空行停止阈值回归(BLANK_ROW_STOP 太小会跳过合计行)")
    wb = opx.Workbook()
    wb.remove(wb.active)
    people = [dict(姓名="张三", 身份证号码="1", 基本工资=1000, 养老=1, 失业=1, 基本医疗=1,
                   住房公积金=1, 税前工资=996, 个人所得税=0, 实发工资=996)]
    summary = dict(基本工资=1000, 养老=1, 失业=1, 基本医疗=1, 住房公积金=1, 税前工资=996,
                   个人所得税=0, 实发工资=996)
    ws = _build_payroll_ws(wb, "6月湖南", people, summary=summary, trailing_junk=3)
    payroll_alias, _, company_map = cm.load_aliases()
    sheets = cm.detect_payroll_sheets(wb, payroll_alias)

    cm.CONFIG["BLANK_ROW_STOP"] = 2
    _, _, mismatches_small, _ = cm.clean_one_sheet(wb, sheets[0], payroll_alias, company_map)
    check_eq("阈值=2 时误停，读不到合计行(mismatches=None)", mismatches_small, None)

    cm.CONFIG["BLANK_ROW_STOP"] = 10
    _, _, mismatches_ok, _ = cm.clean_one_sheet(wb, sheets[0], payroll_alias, company_map)
    check_eq("阈值=10 时能正确读到合计行并校验通过", mismatches_ok, [])


def test_blank_column_detection():
    print("· 无表头空列检测：正例(全空该删) + 反例(看似空但有值不该删) + 多候选不猜")
    header_1 = ["姓名", "地区", "", "部门", "工资"]  # 候选=idx2
    rows_all_empty = [(1, ["a", "湖南", None, "x", 100]), (2, ["b", "湖南", "", "y", 200])]
    confirmed, warn = cm.find_blank_columns(header_1, rows_all_empty)
    check_eq("单候选且全空→确认删除", confirmed, [2])
    check_eq("单候选无警告", warn, None)

    rows_one_has_value = [(1, ["a", "湖南", None, "x", 100]), (2, ["b", "湖南", "真有值", "y", 200])]
    confirmed2, warn2 = cm.find_blank_columns(header_1, rows_one_has_value)
    check_eq("候选列里有一行非空→不确认删除(反例)", confirmed2, [])

    header_2 = ["姓名", "地区", "", "部门", "", "工资"]  # 两个候选 idx2, idx4
    rows2 = [(1, ["a", "湖南", None, "x", None, 100])]
    confirmed3, warn3 = cm.find_blank_columns(header_2, rows2)
    check_eq(">1个确认候选→不猜、不删除", confirmed3, [])
    check_true(">1个候选→有警告", warn3)


def test_formula_missing_not_zero():
    print("· 公式未缓存(None)不当0、四舍五入两位小数")
    check_eq("正常数值四舍五入", cm.to_number(3936.2799999999997, 2), 3936.28)
    check_eq("None不当0", cm.to_number(None, 2), None)
    check_eq("字符串数字", cm.to_number("1,234.5", 2), 1234.5)


def test_match_three_way():
    print("· 匹配三分类：未匹配/匹配成功(含花名册重复登记去重)/重名-待核实")
    index = defaultdict(list)
    index["谢柯"].append(dict(部门="人力中心", 地区="湖南", 岗位="HRBP", 电话=111, 身份证号="A1"))
    # 花名册重复登记同一人(身份证号相同)——真实数据踩到的案例(宋霈森/刘笑容原型)
    index["刘笑容"].append(dict(部门="湖南分公司", 地区="湖南", 岗位="兼职客服", 电话=222, 身份证号="B1"))
    index["刘笑容"].append(dict(部门="湖南分公司", 地区="湖南", 岗位="兼职客服", 电话=222, 身份证号="B1"))
    # 真重名：同名不同身份证号
    index["张伟"].append(dict(部门="项目中心", 地区="湖南", 岗位="项目经理", 电话=333, 身份证号="C1"))
    index["张伟"].append(dict(部门="资源中心", 地区="湖南", 岗位="客服", 电话=444, 身份证号="C2"))

    s, rec, note = cm.match_one("谢柯", index)
    check_eq("谢柯→匹配成功", s, "匹配成功")
    check_eq("谢柯填部门正确", rec["部门"], "人力中心")

    s, rec, note = cm.match_one("王五", index)
    check_eq("花名册无此人→未匹配-待核实", s, "未匹配-待核实")
    check_eq("未匹配不填字段", rec, None)

    s, rec, note = cm.match_one("刘笑容", index)
    check_eq("花名册重复登记(身份证号相同)→仍算匹配成功", s, "匹配成功")
    check_true("提示里说明已自动去重", "去重" in note)

    s, rec, note = cm.match_one("张伟", index)
    check_eq("真重名(身份证号不同)→重名-待核实", s, "重名-待核实")
    check_eq("重名不自动填字段", rec, None)
    check_true("提示里列出候选区分信息", "项目中心" in note and "资源中心" in note)


def test_month_scoping_no_cross_month():
    print("· 月份限定：只读当月sheet，不跨月扫描(避免把跨月重复误判成重名)")
    wb = opx.Workbook()
    wb.remove(wb.active)
    master_alias = cm._MASTER_ALIAS_DEFAULT
    hdr = ["序号", "姓名", "地区", "部门", "岗位", "身份证号", "电话"]
    for month, people in {
        "202605": [("谢柯", "湖南", "人力部", "HRBP", "D1", 111)],
        "202606": [("谢柯", "湖南", "人力部", "HRBP", "D1", 111), ("戴文怡", "湖南", "资源部", "专员", "D2", 222)],
    }.items():
        ws = wb.create_sheet(month)
        ws.append(hdr)
        for i, p in enumerate(people, 1):
            ws.append([i, p[0], p[1], p[2], p[3], p[4], p[5]])
    path = os.path.join(tempfile.gettempdir(), "_t_master.xlsx")
    wb.save(path)

    idx606, avail = cm.load_employee_master(path, "202606", master_alias)
    check_eq("202606月能查到2人", sorted(idx606.keys()), ["戴文怡", "谢柯"])
    s, rec, note = cm.match_one("戴文怡", idx606)
    check_eq("戴文怡在当月sheet里→匹配成功", s, "匹配成功")

    idx605, _ = cm.load_employee_master(path, "202605", master_alias)
    s2, rec2, note2 = cm.match_one("戴文怡", idx605)
    check_eq("戴文怡只在下个月出现、当月(202605)sheet没有→未匹配-待核实(不跨月找)", s2, "未匹配-待核实")

    check_eq("谢柯两个月sheet各出现一次→各自月份内都只有1条，不被误判重名",
             (len(idx605["谢柯"]), len(idx606["谢柯"])), (1, 1))

    try:
        cm.load_employee_master(path, "209912", master_alias)
        check_true("月份不存在应抛异常", False)
    except cm.MonthNotFoundError as e:
        check_true("月份不存在→报错并列出现有月份", "202605" in e.available and "202606" in e.available)
    os.remove(path)


def test_real_smoke():
    """真实数据冒烟：测试数据不在则跳过。数字来自 2026-07-01 用真实数据跑通、人工核实过的结果。"""
    财务部skills = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(HERE))))
    D = os.path.join(财务部skills, "技能", "工资表清洗与信息匹配", "测试数据")
    payroll = os.path.join(D, "明妹工资原始表_未处理版.xlsx")
    employee = os.path.join(D, "湖南分公司员工信息（财务）.xlsx")
    if not (os.path.isfile(payroll) and os.path.isfile(employee)):
        print(f"· 真实冒烟：跳过（测试数据不在 {D}）")
        return
    print("· 真实数据冒烟")
    cm.load_rules()
    payroll_alias, master_alias, company_map = cm.load_aliases()
    wb = cm.openpyxl.load_workbook(payroll, data_only=True)
    sheets = cm.detect_payroll_sheets(wb, payroll_alias)
    check_eq("识别到2个工资sheet(6月湖南/6月湖南科技)", sorted(s["sheet"] for s in sheets),
              ["6月湖南", "6月湖南科技"])
    master_index, avail = cm.load_employee_master(employee, "202606", master_alias)
    check_eq("员工信息表有9个月份sheet", len(avail), 9)

    all_status = []
    for si in sheets:
        cleaned, company, mismatches, warns = cm.clean_one_sheet(wb, si, payroll_alias, company_map)
        check_eq(f"{si['sheet']} 合计校验通过", mismatches, [])
        for rec in cleaned:
            s, rec2, note = cm.match_one(rec["name"], master_index)
            all_status.append((rec["name"], s))
    wb.close()
    check_eq("总人数=82(77+5)", len(all_status), 82)
    from collections import Counter as _Counter
    cnt = _Counter(s for _, s in all_status)
    check_eq("匹配成功=73", cnt.get("匹配成功", 0), 73)
    check_eq("未匹配-待核实=9", cnt.get("未匹配-待核实", 0), 9)
    check_eq("重名-待核实=0(本批82人内部无重名)", cnt.get("重名-待核实", 0), 0)
    by_name = dict(all_status)
    check_eq("谢柯→匹配成功", by_name.get("谢柯"), "匹配成功")
    check_eq("罗欢(湖南科技,花名册未收录)→未匹配-待核实", by_name.get("罗欢"), "未匹配-待核实")


if __name__ == "__main__":
    test_sheet_detection_and_person_row_range()
    test_blank_row_stop_regression()
    test_blank_column_detection()
    test_formula_missing_not_zero()
    test_match_three_way()
    test_month_scoping_no_cross_month()
    test_real_smoke()
    print(f"\n{'='*40}\nPASS={PASS}  FAIL={FAIL}")
    sys.exit(1 if FAIL else 0)
