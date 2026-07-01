#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
工资表清洗与信息匹配
把出纳（明妹）给的原始工资表清洗干净（删无表头空列、公式转数值两位小数、剔除合计/参考数据），
再按姓名从权威员工信息表（按月）匹配电话/部门/岗位/地区，产出一张干净底表。

用法：
  python3 clean_match.py --inspect [--input-dir DIR]        # 只认文件/看建议月份，不跑
  python3 clean_match.py --payroll 工资原始表.xlsx --employee 员工信息表.xlsx --month 202606 [--out 结果.xlsx]
  缺 --payroll/--employee 时按内容从 --input-dir(默认 工作区/input) 自动认；两候选以上不猜、报出来讓人选。
  --month 必须显式传（正式跑不接受"猜的"月份，见 config/业务规则.md 二·月份怎么定）。

规则（容差/小数位/合计关键字/公司映射）在 config/业务规则.md、config/列名别名.json，改表不改码。
"""
import os
import re
import sys
import json
import datetime
import argparse
from collections import defaultdict, Counter

import openpyxl

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

HERE = os.path.dirname(os.path.abspath(__file__))
SKILL_DIR = os.path.dirname(HERE)
CONFIG_DIR = os.path.join(SKILL_DIR, "config")
WORK_INPUT = os.path.join(SKILL_DIR, "工作区", "input")
WORK_OUTPUT = os.path.join(SKILL_DIR, "工作区", "output")

NUMERIC_FIELDS = ["基本工资", "养老", "失业", "基本医疗", "住房公积金", "税前工资", "个人所得税", "实发工资"]

CONFIG = {
    "TOLERANCE": 0.05,
    "DECIMALS": 2,
    "SUMMARY_KEYWORDS": ["合计", "总计", "小计", "共计"],
    # 真实数据实测：明妹原表最后一个人和「合计」行之间常隔 2~3 行空/杂散行(如VLOOKUP渗入的碎片单元格)，
    # 阈值定太小(如2)会在到达合计行之前就误停、导致合计校验直接跳过——所以留够余量。
    "BLANK_ROW_STOP": 10,
}

_PAYROLL_ALIAS_DEFAULT = {
    "姓名": ["姓名"],
    "身份证号": ["身份证号码", "身份证号", "证件号", "身份证"],
    "部门": ["部门"],
    "地区": ["地区"],
    "基本工资": ["基本工资"],
    "养老": ["养老", "养老保险"],
    "失业": ["失业", "失业保险"],
    "基本医疗": ["基本医疗", "医疗保险"],
    "住房公积金": ["住房公积金", "公积金"],
    "税前工资": ["税前工资"],
    "个人所得税": ["个人所得税", "个税"],
    "实发工资": ["实发工资"],
}
_MASTER_ALIAS_DEFAULT = {
    "姓名": ["姓名"],
    "地区": ["地区"],
    "部门": ["部门"],
    "岗位": ["岗位"],
    "身份证号": ["身份证号", "身份证号码", "证件号", "身份证"],
    "电话": ["电话", "手机", "联系电话", "手机号"],
}
_COMPANY_MAP_DEFAULT = {"湖南": "湖南分公司", "湖南科技": "湖南子公司"}


def log(msg):
    print(msg, file=sys.stderr)


# ----------------- 配置加载 -----------------
def load_aliases():
    p = os.path.join(CONFIG_DIR, "列名别名.json")
    if os.path.isfile(p):
        try:
            with open(p, encoding="utf-8") as f:
                d = json.load(f)
            return (d.get("工资表_列别名", _PAYROLL_ALIAS_DEFAULT),
                    d.get("员工信息表_列别名", _MASTER_ALIAS_DEFAULT),
                    d.get("_所属公司映射", _COMPANY_MAP_DEFAULT))
        except Exception as e:
            log(f"⚠ 读 列名别名.json 失败({e})，用内置默认。")
    return dict(_PAYROLL_ALIAS_DEFAULT), dict(_MASTER_ALIAS_DEFAULT), dict(_COMPANY_MAP_DEFAULT)


def load_rules():
    """从 config/业务规则.md 的『可调参数』表读容差/小数位/关键字/空行阈值 → 覆盖默认。缺/解析失败不崩。"""
    p = os.path.join(CONFIG_DIR, "业务规则.md")
    if not os.path.isfile(p):
        return
    try:
        text = open(p, encoding="utf-8").read()
    except Exception as e:
        log(f"⚠ 读 业务规则.md 失败({e})，用内置默认。")
        return
    for line in text.splitlines():
        if not line.strip().startswith("|"):
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if len(cells) < 2:
            continue
        key, val = cells[0], cells[1]
        num = re.search(r"-?\d+(?:\.\d+)?", val)
        if key == "合计校验容差" and num:
            CONFIG["TOLERANCE"] = float(num.group())
        elif key == "数值保留小数位" and num:
            CONFIG["DECIMALS"] = int(float(num.group()))
        elif key == "合计行识别关键字" and val:
            CONFIG["SUMMARY_KEYWORDS"] = [w.strip() for w in re.split(r"[、,，/]", val) if w.strip()]
        elif key == "空行停止阈值" and num:
            CONFIG["BLANK_ROW_STOP"] = int(float(num.group()))


# ----------------- 工具 -----------------
def _strip_row(row):
    return [str(c).strip() if c is not None else "" for c in row]


def _find_col(header_stripped, aliases):
    for a in aliases:
        if a in header_stripped:
            return header_stripped.index(a)
    return None


def norm_id(v):
    if v is None:
        return ""
    s = str(v).strip().upper()
    return re.sub(r"\s+", "", s)


def to_number(v, decimals=None):
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return round(float(v), decimals) if decimals is not None else float(v)
    if isinstance(v, (datetime.datetime, datetime.date, datetime.time)):
        return None
    s = str(v).strip().replace(",", "").replace("，", "")
    if s in ("", "-", "#N/A", "NA", "nan", "None"):
        return None
    try:
        f = float(s)
        return round(f, decimals) if decimals is not None else f
    except ValueError:
        return None


def excel_serial_to_ym(v):
    """粗略把疑似 Excel 日期序列号的数字转成 (year, month)；不在合理范围内返回 None。仅用于 --inspect 的建议、不用于正式判断。"""
    if not isinstance(v, (int, float)) or isinstance(v, bool):
        return None
    if not (42000 <= v <= 55000):  # 约 2015~2050，避开工资/金额类数字
        return None
    try:
        d = datetime.date(1899, 12, 30) + datetime.timedelta(days=int(v))
        return d.year, d.month
    except Exception:
        return None


# ----------------- 识别工资表 sheet -----------------
def _header_row_for_payroll(ws, alias, max_scan=6):
    name_al = alias.get("姓名", _PAYROLL_ALIAS_DEFAULT["姓名"])
    id_al = alias.get("身份证号", _PAYROLL_ALIAS_DEFAULT["身份证号"])
    amt_groups = [alias.get(k, _PAYROLL_ALIAS_DEFAULT[k]) for k in ("基本工资", "实发工资", "税前工资")]
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        cells = _strip_row(row)
        has_name = any(a in cells for a in name_al)
        has_id = any(a in cells for a in id_al)
        hits = sum(1 for grp in amt_groups if any(a in cells for a in grp))
        if has_name and has_id and hits >= 2:
            w = len(cells)
            while w > 0 and cells[w - 1] == "":
                w -= 1
            return ri, list(row[:w])
    return None


def detect_payroll_sheets(wb, alias):
    out = []
    for name in wb.sheetnames:
        hdr = _header_row_for_payroll(wb[name], alias)
        if hdr is not None:
            header_row0, header = hdr
            out.append(dict(sheet=name, header_row0=header_row0, header=header))
    return out


def company_of(sheet_name, company_map):
    m = re.match(r"^\d{1,2}月(.+)$", sheet_name)
    key = m.group(1).strip() if m else sheet_name
    if key in company_map:
        return company_map[key], None
    return key, f"sheet「{sheet_name}」按前缀『{key}』没查到所属公司映射，暂用『{key}』当公司名，请去 config/列名别名.json 的 _所属公司映射 补一行"


# ----------------- 人员行范围 + 空列检测 -----------------
def scan_person_rows(ws, header_row0, width, name_idx, summary_keywords, blank_stop):
    rows_out = []
    summary_raw = None
    blank_streak = 0
    r = header_row0 + 2  # 1-based：header_row0 是 0-based 表头行号，数据从下一行开始
    max_row = ws.max_row or 0
    while r <= max_row:
        raw = [ws.cell(r, c + 1).value for c in range(width)]
        strs = _strip_row(raw)
        if any(s in summary_keywords for s in strs):
            summary_raw = raw
            break
        name_val = raw[name_idx] if name_idx is not None else None
        if name_val not in (None, ""):
            rows_out.append((r, raw))
            blank_streak = 0
        else:
            if all(v is None for v in raw):
                blank_streak += 1
                if blank_streak >= blank_stop:
                    break
            # 有名字空但行内还有杂散值(如下方VLOOKUP参考区渗进来的碎片)→跳过，不计入空行streak、不停止
        r += 1
    return rows_out, summary_raw


def find_blank_columns(header_stripped, person_rows):
    """候选=表头空+左右表头都非空；候选须在所有人员行里100%为空才确认。>1个确认候选→不猜，报警告。"""
    candidates = [i for i in range(1, len(header_stripped) - 1)
                  if header_stripped[i] == "" and header_stripped[i - 1] != "" and header_stripped[i + 1] != ""]
    confirmed = []
    for i in candidates:
        if all((raw[i] is None or str(raw[i]).strip() == "") for _, raw in person_rows):
            confirmed.append(i)
    if len(confirmed) > 1:
        return [], f"检测到{len(confirmed)}个候选无表头空列(位置{confirmed})，为避免误删不自动处理，请人工确认"
    return confirmed, None


# ----------------- 员工信息表（权威数据源） -----------------
def _header_row_for_master(ws, alias, max_scan=5):
    name_al = alias.get("姓名", _MASTER_ALIAS_DEFAULT["姓名"])
    region_al = alias.get("地区", _MASTER_ALIAS_DEFAULT["地区"])
    dept_al = alias.get("部门", _MASTER_ALIAS_DEFAULT["部门"])
    id_al = alias.get("身份证号", _MASTER_ALIAS_DEFAULT["身份证号"])
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        cells = _strip_row(row)
        if all(any(a in cells for a in grp) for grp in (name_al, region_al, dept_al, id_al)):
            w = len(cells)
            while w > 0 and cells[w - 1] == "":
                w -= 1
            return ri, list(row[:w])
    return None


class MonthNotFoundError(Exception):
    def __init__(self, month, available):
        self.month = month
        self.available = available
        super().__init__(f"员工信息表里没有月份 sheet「{month}」，现有：{available}")


def load_employee_master(path, month, alias):
    wb = openpyxl.load_workbook(path, data_only=True)
    available = [s for s in wb.sheetnames if re.match(r"^\d{6}$", s)]
    if month not in wb.sheetnames:
        raise MonthNotFoundError(month, available)
    ws = wb[month]
    hdr = _header_row_for_master(ws, alias)
    if hdr is None:
        raise ValueError(f"员工信息表 sheet「{month}」没识别到表头(需含姓名+地区+部门+身份证号)")
    header_row0, header = hdr
    header_stripped = _strip_row(header)
    col = {f: _find_col(header_stripped, alias.get(f, _MASTER_ALIAS_DEFAULT[f])) for f in _MASTER_ALIAS_DEFAULT}
    index = defaultdict(list)
    r = header_row0 + 2
    while r <= (ws.max_row or 0):
        raw = [ws.cell(r, c + 1).value for c in range(len(header_stripped))]
        name = raw[col["姓名"]] if col.get("姓名") is not None else None
        if name not in (None, ""):
            rec = dict(
                部门=raw[col["部门"]] if col.get("部门") is not None else None,
                地区=raw[col["地区"]] if col.get("地区") is not None else None,
                岗位=raw[col["岗位"]] if col.get("岗位") is not None else None,
                电话=raw[col["电话"]] if col.get("电话") is not None else None,
                身份证号=norm_id(raw[col["身份证号"]]) if col.get("身份证号") is not None else "",
            )
            index[str(name).strip()].append(rec)
        r += 1
    wb.close()
    return index, available


def match_one(name, index):
    """三分类决策树：未匹配 / 匹配成功(含花名册内重复登记去重) / 重名-待核实。"""
    candidates = index.get(name, [])
    if not candidates:
        return "未匹配-待核实", None, "员工信息表(当月)无此姓名，可能是新员工/未录入，请人工核实"
    ids = set(c["身份证号"] for c in candidates if c["身份证号"])
    if len(candidates) == 1 or len(ids) <= 1:
        note = "花名册重复登记同一人，已自动去重" if len(candidates) > 1 else ""
        return "匹配成功", candidates[0], note
    desc = []
    for c in candidates:
        tail = c["身份证号"][-4:] if c["身份证号"] else "无证号"
        desc.append(f"{c.get('部门') or '?'}/{c.get('岗位') or '?'}(证号尾{tail})")
    note = f"花名册有{len(candidates)}个同名，区分信息：" + "；".join(desc) + "——人工确认对应哪个"
    return "重名-待核实", None, note


# ----------------- 主清洗流程 -----------------
def clean_one_sheet(wb, sheet_info, payroll_alias, company_map):
    sheet = sheet_info["sheet"]
    header_row0 = sheet_info["header_row0"]
    header_stripped = _strip_row(sheet_info["header"])
    width = len(header_stripped)
    col = {f: _find_col(header_stripped, payroll_alias.get(f, _PAYROLL_ALIAS_DEFAULT[f])) for f in _PAYROLL_ALIAS_DEFAULT}
    warns = []
    if col.get("姓名") is None or col.get("身份证号") is None:
        return [], None, None, [f"{sheet}: 缺姓名或身份证号列，跳过整个sheet"]
    ws = wb[sheet]
    person_rows, summary_raw = scan_person_rows(
        ws, header_row0, width, col["姓名"], CONFIG["SUMMARY_KEYWORDS"], CONFIG["BLANK_ROW_STOP"])
    blank_cols, blank_warn = find_blank_columns(header_stripped, person_rows)
    if blank_warn:
        warns.append(f"{sheet}: {blank_warn}")
    company, cwarn = company_of(sheet, company_map)
    if cwarn:
        warns.append(cwarn)

    decimals = CONFIG["DECIMALS"]
    cleaned = []
    for r, raw in person_rows:
        name = str(raw[col["姓名"]]).strip()
        idno = raw[col["身份证号"]] if col.get("身份证号") is not None else None
        numeric = {}
        missing = []
        for field in NUMERIC_FIELDS:
            idx = col.get(field)
            if idx is None:
                continue
            v = to_number(raw[idx], decimals)
            numeric[field] = v
            if v is None:
                missing.append(field)
        cleaned.append(dict(row=r, name=name, idno=idno, numeric=numeric, missing=missing))

    mismatches = None
    if summary_raw is not None:
        mismatches = []
        for field in NUMERIC_FIELDS:
            idx = col.get(field)
            if idx is None:
                continue
            computed = round(sum((c["numeric"].get(field) or 0.0) for c in cleaned), decimals)
            reported = to_number(summary_raw[idx], decimals)
            if reported is not None and abs(computed - reported) > CONFIG["TOLERANCE"]:
                mismatches.append(dict(field=field, computed=computed, reported=reported,
                                        diff=round(computed - reported, decimals)))

    return cleaned, company, mismatches, warns


def run(payroll_path, employee_path, month, out_path):
    load_rules()
    payroll_alias, master_alias, company_map = load_aliases()
    wb = openpyxl.load_workbook(payroll_path, data_only=True)
    payroll_sheets = detect_payroll_sheets(wb, payroll_alias)
    if not payroll_sheets:
        log("✗ 工资表里没识别到任何符合特征(姓名+身份证号+≥2个金额列)的 sheet。")
        sys.exit(2)

    try:
        master_index, available_months = load_employee_master(employee_path, month, master_alias)
    except MonthNotFoundError as e:
        log(f"✗ {e}")
        sys.exit(2)

    all_records = []
    report_rows = []
    all_warns = []
    missing_counter = Counter()
    for sheet_info in payroll_sheets:
        cleaned, company, mismatches, warns = clean_one_sheet(wb, sheet_info, payroll_alias, company_map)
        all_warns.extend(warns)
        if not cleaned and not company:
            continue
        report_rows.append(dict(sheet=sheet_info["sheet"], company=company, 人数=len(cleaned),
                                 合计校验=mismatches))
        for rec in cleaned:
            status, mrec, note = match_one(rec["name"], master_index)
            notes = []
            if note:
                notes.append(note)
            if rec["missing"]:
                notes.append(f"{','.join(rec['missing'])} 数值缺失(公式未缓存计算结果)，人工核实")
                missing_counter.update(rec["missing"])
            out = dict(
                姓名=rec["name"],
                部门=(mrec or {}).get("部门"),
                岗位=(mrec or {}).get("岗位"),
                地区=(mrec or {}).get("地区"),
                身份证号码=rec["idno"],
                电话=(mrec or {}).get("电话"),
                所属公司=company,
                匹配状态=status,
                核对提示="；".join(notes),
            )
            for field in NUMERIC_FIELDS:
                out[field] = rec["numeric"].get(field)
            all_records.append(out)
    wb.close()

    for x in all_warns:
        log(f"⚠ {x}")
    write_output(out_path, all_records, report_rows, all_warns, month, missing_counter)
    log(f"· 共 {len(all_records)} 人；{Counter(r['匹配状态'] for r in all_records)}")
    log(f"· 已写出 → {out_path}（4个sheet：主表 / 待人工核实 / 匹配成功 / 运行报告，跟用户汇报时按这4个原名说，别改写）")
    return out_path


# ----------------- 输出 -----------------
COLUMNS_ORDER = (["姓名", "部门", "岗位", "地区", "身份证号码", "电话"] + NUMERIC_FIELDS
                  + ["所属公司", "匹配状态", "核对提示"])


def _style(ws, ncol, yellow_rows):
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    hfont = Font(name="等线", size=10.5, bold=True)
    hfill = PatternFill("solid", fgColor="BDD7EE")
    yfill = PatternFill("solid", fgColor="FFFF00")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(*[Side(style="thin", color="999999")] * 4)
    for i in range(1, ncol + 1):
        c = ws.cell(1, i)
        c.font = hfont; c.fill = hfill; c.alignment = align; c.border = border
    for ri in yellow_rows:
        for i in range(1, ncol + 1):
            ws.cell(ri, i).fill = yfill
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"


def write_output(out_path, records, report_rows, warnings, month, missing_counter=None):
    import pandas as pd
    df = pd.DataFrame(records, columns=["序号"] + COLUMNS_ORDER)
    df["序号"] = range(1, len(df) + 1)
    need_review = df[df["核对提示"] != ""].copy()
    clean_ok = df[(df["匹配状态"] == "匹配成功") & (df["核对提示"] == "")].copy()

    report_lines = []
    cnt = Counter(records and [r["匹配状态"] for r in records] or [])
    for k, v in sorted(cnt.items()):
        report_lines.append((k, v, ""))
    report_lines.append(("合计人数", len(records), ""))
    report_lines.append(("待人工核实合计", len(need_review), ""))
    if missing_counter:
        for field, n in missing_counter.most_common():
            report_lines.append((f"数值缺失-{field}", n,
                                  "常见于实习生的社保类列(养老/失业/医疗/公积金)是正常现象；"
                                  "若集中出现在正式员工的税前/实发/个税列，建议先回明妹打开原表触发一次公式重算再贴数"))
    for rr in report_rows:
        report_lines.append((f"sheet「{rr['sheet']}」({rr['company']})人数", rr["人数"], ""))
        if rr["合计校验"] is None:
            report_lines.append((f"sheet「{rr['sheet']}」合计校验", "未找到合计行，跳过", ""))
        elif not rr["合计校验"]:
            report_lines.append((f"sheet「{rr['sheet']}」合计校验", "通过(容差内)", ""))
        else:
            for m in rr["合计校验"]:
                report_lines.append((f"sheet「{rr['sheet']}」{m['field']}合计不符",
                                      f"清洗后求和={m['computed']} vs 原表合计={m['reported']}",
                                      f"差{m['diff']}"))
    for w in warnings:
        report_lines.append(("警告", w, ""))
    report_df = pd.DataFrame(report_lines, columns=["项目", "值", "备注"])

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="主表", index=False)
        need_review.to_excel(w, sheet_name="待人工核实", index=False)
        clean_ok.to_excel(w, sheet_name="匹配成功", index=False)
        report_df.to_excel(w, sheet_name="运行报告", index=False)
        yellow_rows = [i + 2 for i, v in enumerate(df["核对提示"]) if v]
        _style(w.sheets["主表"], len(df.columns), yellow_rows)
        _style(w.sheets["待人工核实"], len(need_review.columns), list(range(2, len(need_review) + 2)))
        _style(w.sheets["匹配成功"], len(clean_ok.columns), [])
        _style(w.sheets["运行报告"], len(report_df.columns), [])
    return out_path


# ----------------- 认文件 / inspect -----------------
def _classify_file(path, payroll_alias):
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return None
    is_master = bool(wb.sheetnames) and all(re.match(r"^\d{6}$", s) for s in wb.sheetnames)
    is_payroll = False
    if not is_master:
        try:
            wb2 = openpyxl.load_workbook(path, data_only=True)
            is_payroll = bool(detect_payroll_sheets(wb2, payroll_alias))
            wb2.close()
        except Exception:
            pass
    wb.close()
    if is_master:
        return "master"
    if is_payroll:
        return "payroll"
    return None


def find_inputs(input_dir, payroll_alias):
    payroll_cands, master_cands = [], []
    if os.path.isdir(input_dir):
        for fn in sorted(os.listdir(input_dir)):
            if not fn.lower().endswith((".xlsx", ".xls")) or fn.startswith("~$"):
                continue
            kind = _classify_file(os.path.join(input_dir, fn), payroll_alias)
            if kind == "payroll":
                payroll_cands.append(fn)
            elif kind == "master":
                master_cands.append(fn)
    return payroll_cands, master_cands


def inspect_mode(input_dir):
    payroll_alias, master_alias, company_map = load_aliases()
    payroll_cands, master_cands = find_inputs(input_dir, payroll_alias)
    print(f"识别输入目录：{input_dir}")
    print(f"  工资原始表候选：{payroll_cands or '（没认到）'}")
    print(f"  员工信息表候选：{master_cands or '（没认到）'}")
    if len(payroll_cands) != 1:
        print("  ⚠ 工资原始表候选不是恰好1个，需要用 --payroll 明确指定")
    if len(master_cands) != 1:
        print("  ⚠ 员工信息表候选不是恰好1个，需要用 --employee 明确指定")
    if len(payroll_cands) == 1:
        p = os.path.join(input_dir, payroll_cands[0])
        wb = openpyxl.load_workbook(p, data_only=True)
        sheets = detect_payroll_sheets(wb, payroll_alias)
        print(f"  工资表识别到 {len(sheets)} 个符合特征的 sheet：")
        month_digits = set()
        date_hints = set()
        for s in sheets:
            company, cwarn = company_of(s["sheet"], company_map)
            print(f"    - 「{s['sheet']}」→ 所属公司猜测：{company}" + (f"（{cwarn}）" if cwarn else ""))
            m = re.match(r"^(\d{1,2})月", s["sheet"])
            if m:
                month_digits.add(int(m.group(1)))
        for name in wb.sheetnames:
            for row in wb[name].iter_rows(min_row=1, max_row=3, values_only=True):
                for v in row:
                    ym = excel_serial_to_ym(v)
                    if ym:
                        date_hints.add(ym)
        wb.close()
        print(f"  月份线索：sheet名数字={month_digits or '无'}；表头区疑似日期={date_hints or '无'}")
        if len(month_digits) == 1 and date_hints:
            mdigit = next(iter(month_digits))
            hit = [ym for ym in date_hints if ym[1] == mdigit]
            if hit:
                y, m = hit[0]
                print(f"  → 建议 --month {y}{m:02d}（两个线索吻合），但正式跑请显式传入、不要让脚本替你决定")
    if len(master_cands) == 1:
        p = os.path.join(input_dir, master_cands[0])
        wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
        print(f"  员工信息表月份sheet：{sorted(wb.sheetnames)}")
        wb.close()


def main():
    load_rules()
    ap = argparse.ArgumentParser(description="工资表清洗与信息匹配")
    ap.add_argument("--payroll", dest="payroll_path")
    ap.add_argument("--employee", dest="employee_path")
    ap.add_argument("--month", dest="month")
    ap.add_argument("--out")
    ap.add_argument("--input-dir", default=WORK_INPUT)
    ap.add_argument("--inspect", action="store_true")
    a = ap.parse_args()
    if a.inspect:
        inspect_mode(a.input_dir)
        return
    payroll_alias, _, _ = load_aliases()
    pp, ep = a.payroll_path, a.employee_path
    if not pp or not ep:
        pcands, mcands = find_inputs(a.input_dir, payroll_alias)
        if not pp:
            if len(pcands) == 1:
                pp = os.path.join(a.input_dir, pcands[0])
            else:
                log(f"✗ 工资原始表候选{len(pcands)}个({pcands})，请用 --payroll 明确指定。")
                sys.exit(2)
        if not ep:
            if len(mcands) == 1:
                ep = os.path.join(a.input_dir, mcands[0])
            else:
                log(f"✗ 员工信息表候选{len(mcands)}个({mcands})，请用 --employee 明确指定。")
                sys.exit(2)
    if not a.month:
        log("✗ 必须显式传 --month YYYYMM（不接受自动猜测，见 config/业务规则.md）。先跑 --inspect 看建议值。")
        sys.exit(2)
    out = a.out
    if not out:
        base = os.path.dirname(os.path.abspath(pp))
        out = os.path.join(base, f"工资底表_{a.month}.xlsx")
    elif os.path.isdir(out):
        # 真实 opencode 模拟测试踩到的坑：agent 有时把 --out 传成目录本身，
        # 直接 open() 会是生硬的 IsADirectoryError；这里兜底补上默认文件名。
        out = os.path.join(out, f"工资底表_{a.month}.xlsx")
    run(pp, ep, a.month, out)


if __name__ == "__main__":
    main()
