#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
银行日记账 · 收入提取汇总（bank-income-extract）
===============================================
把一份"银行日记账.xlsx"里【每个 sheet（=每个银行/现金账户）】的"收入"记录
筛出来，汇总到一张表；顺带把"疑似漏标/缺客户名"的可疑行挑出来给人核对。

设计原则（该死的死、该活的活）：
  - 确定性逻辑（读表、筛选、去空行、算小计）写死在本脚本；
  - 会变的业务规则（哪些词算收入、列叫什么名、哪些 sheet 跳过、哪些是外币）
    放 config/识别规则.md，改表不改码。config 缺失时用脚本内兜底默认值。

收入判定：某行「类型」列含"收入"类关键词 且「金额(借方)」列有数字 → 收入。
  （日记账里常有"类型=收入但金额空"的预填行，靠"有金额"过滤掉——最容易踩的坑。）

用法：
    python3 extract_income.py 银行日记账.xlsx
    python3 extract_income.py 银行日记账.xlsx -o 收入汇总.xlsx
    python3 extract_income.py 银行日记账.xlsx --config /别处/识别规则.md
不指定 -o 时，输出「收入汇总_<输入名>_<日期>.xlsx」放输入文件同目录。
"""
import sys, os, argparse, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill

SKILL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_CONFIG = os.path.join(SKILL_DIR, "config", "识别规则.md")

# ── config 缺失时的兜底默认（和 config/识别规则.md 初始内容一致）──────────
FALLBACK = {
    "收入判定关键词": ["收入", "到账"],
    "支出类关键词": ["手续费", "利息", "支出", "转出", "服务费", "退款", "费用"],
    "日期列表头词": ["日期", "交易日期", "记账日期"],
    "客户名列表头词": ["摘要", "客户", "对方户名", "汇款人", "收款人", "对方"],
    "金额列表头词": ["借方", "增加", "收入金额", "收方"],
    "类型列表头词": ["类型", "类别"],
    "跳过的 sheet": ["说明", "目录", "封面", "模板", "汇总表"],
}
FALLBACK_CURRENCY = {  # 关键词 => 币种
    "美元": "美元", "美金": "美元", "USD": "美元",
    "欧元": "欧元", "EUR": "欧元", "英镑": "英镑", "GBP": "英镑",
    "港币": "港币", "港元": "港币", "HKD": "港币", "日元": "日元", "JPY": "日元",
}
HEADER_SCAN_ROWS = 15


# ── config 解析 ──────────────────────────────────────────────────────
def load_config(path):
    """读 config/识别规则.md：按二级标题分节，节下 '- ' 列表项收集。
    外币节支持 '关键词 => 币种'。读不到就用兜底。"""
    lists = {k: list(v) for k, v in FALLBACK.items()}
    currency = dict(FALLBACK_CURRENCY)
    if not path or not os.path.isfile(path):
        return lists, currency
    section = None
    cur_from_config = {}
    try:
        with open(path, encoding="utf-8") as f:
            for raw in f:
                line = raw.rstrip("\n")
                s = line.strip()
                if s.startswith("## "):
                    section = s[3:].strip()
                    if section == "外币 sheet":
                        cur_from_config = {}
                    elif section in lists:
                        lists[section] = []       # 命中已知节 → 清空兜底、用配置
                    continue
                if s.startswith("- "):
                    item = s[2:].strip()
                    if not item:
                        continue
                    if section == "外币 sheet":
                        if "=>" in item:
                            k, v = item.split("=>", 1)
                            cur_from_config[k.strip()] = v.strip()
                    elif section in lists:
                        lists[section].append(item)
        if cur_from_config:
            currency = cur_from_config
    except Exception as e:
        print(f"⚠ 读 config 失败（{e}），用内置默认规则继续。")
    # 任何节被配成空 → 回退兜底，防误清空
    for k, v in FALLBACK.items():
        if not lists.get(k):
            lists[k] = list(v)
    return lists, currency


def _norm(v):
    return "" if v is None else str(v).strip()


def to_number(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("，", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def fmt_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return _norm(v)


# ── 识别表头 / 列 / 币种 ──────────────────────────────────────────────
def find_header_row(ws, cfg):
    date_keys = cfg["日期列表头词"]
    type_keys = cfg["类型列表头词"]
    amt_keys = cfg["金额列表头词"]
    for r in range(1, min(HEADER_SCAN_ROWS, ws.max_row) + 1):
        cells = [_norm(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
        has_date = any(any(k in x for k in date_keys) for x in cells)
        has_type = any(any(k in x for k in type_keys) for x in cells)
        has_amt = any(any(k in x for k in amt_keys) for x in cells)
        if has_date and (has_type or has_amt):
            return r
    return None


def map_columns(ws, header_row, cfg):
    colmap = {}
    headers = [(c, _norm(ws.cell(row=header_row, column=c).value))
               for c in range(1, ws.max_column + 1)]
    logical_keys = {
        "日期": cfg["日期列表头词"], "摘要": cfg["客户名列表头词"],
        "借方": cfg["金额列表头词"], "类型": cfg["类型列表头词"],
    }
    for logical, keys in logical_keys.items():
        for c, text in headers:
            if text and any(k in text for k in keys):
                colmap[logical] = c
                break
    return colmap


def detect_currency(ws, sheet_name, currency_map):
    up = sheet_name.upper()
    for kw, cur in currency_map.items():
        if kw in sheet_name or kw.upper() in up:
            return cur
    for row in ws.iter_rows(values_only=True):
        for v in row:
            s = _norm(v)
            if s.startswith("币种") and len(s) > 2:
                return s[2:]
    return "人民币"


# ── 核心提取 ──────────────────────────────────────────────────────────
def is_income(type_text, cfg):
    return any(k in type_text for k in cfg["收入判定关键词"])


def is_known_expense(type_text, cfg):
    return any(k in type_text for k in cfg["支出类关键词"])


def extract(path, cfg, currency_map):
    wb = openpyxl.load_workbook(path, data_only=True)
    rows_out = []      # [渠道, 日期, 客户名称, 金额, 币种]
    flags = []         # 待人工：[渠道, 位置(行号), 问题, 相关值]
    report = []        # 每 sheet 小结
    for sheet_name in wb.sheetnames:          # 保持工作簿 sheet 先后顺序
        if any(sk in sheet_name for sk in cfg["跳过的 sheet"]):
            report.append((sheet_name, "跳过", "命中『跳过的 sheet』规则"))
            continue
        ws = wb[sheet_name]
        hr = find_header_row(ws, cfg)
        if hr is None:
            report.append((sheet_name, "跳过", "没找到表头（日期/类型/金额），可能不是日记账"))
            continue
        colmap = map_columns(ws, hr, cfg)
        missing = [k for k in ("类型", "借方") if k not in colmap]
        if missing:
            report.append((sheet_name, "跳过", f"缺关键列 {missing}（对不上『列表头词』），无法判定收入"))
            continue
        currency = detect_currency(ws, sheet_name, currency_map)

        n_income = 0
        for r in range(hr + 1, ws.max_row + 1):
            typ = _norm(ws.cell(row=r, column=colmap["类型"]).value)
            amount = to_number(ws.cell(row=r, column=colmap["借方"]).value)
            cust = _norm(ws.cell(row=r, column=colmap["摘要"]).value) if "摘要" in colmap else ""
            date_v = fmt_date(ws.cell(row=r, column=colmap["日期"]).value) if "日期" in colmap else ""

            if is_income(typ, cfg):
                if amount is None or amount == 0:
                    continue                      # 空预填收入行 → 跳过
                rows_out.append([sheet_name, date_v, cust, amount, currency])
                n_income += 1
                if not cust:                      # 收入但没客户名 → 下游创建回款要用，挑出来
                    flags.append([sheet_name, f"第{r}行", "收入行缺客户名(摘要空)", f"{amount:,.2f} {currency}"])
            else:
                # 有进账金额、却没标成收入、也不是已知支出 → 疑似漏标收入
                if amount and amount > 0 and not is_known_expense(typ, cfg):
                    flags.append([sheet_name, f"第{r}行",
                                  "有进账金额但未标收入(疑似漏标)",
                                  f"类型='{typ or '空'}' 金额={amount:,.2f} 摘要='{cust}'"])
        report.append((sheet_name, "完成", f"收入 {n_income} 笔（币种：{currency}）"))
    return rows_out, flags, report


# ── 写输出 ────────────────────────────────────────────────────────────
def write_output(rows_out, flags, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "收入汇总"
    headers = ["渠道（银行/账户）", "日期", "到账客户名称", "金额", "币种"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    for row in rows_out:
        ws.append(row)
        if not _norm(row[2]):                     # 缺客户名的行标黄
            for c in range(1, 6):
                ws.cell(row=ws.max_row, column=c).fill = warn_fill
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=4).number_format = "#,##0.00"
    for i, w in enumerate([18, 12, 34, 15, 8], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # 待人工核对（有才建）
    if flags:
        wf = wb.create_sheet("待人工核对")
        wf.append(["渠道", "位置", "问题", "相关值"])
        for cell in wf[1]:
            cell.font = Font(bold=True)
        for f in flags:
            wf.append(f)
        for i, w in enumerate([16, 10, 30, 40], start=1):
            wf.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        wf.freeze_panes = "A2"
    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser(description="银行日记账收入提取汇总")
    ap.add_argument("input", help="银行日记账 .xlsx")
    ap.add_argument("-o", "--output", help="输出 xlsx（默认放输入同目录）")
    ap.add_argument("--config", default=DEFAULT_CONFIG, help="识别规则.md 路径")
    args = ap.parse_args()

    if not os.path.isfile(args.input):
        sys.exit(f"找不到文件：{args.input}")
    cfg, currency_map = load_config(args.config)

    out_path = args.output
    if not out_path:
        base = os.path.splitext(os.path.basename(args.input))[0]
        today = datetime.date.today().strftime("%Y%m%d")
        out_path = os.path.join(os.path.dirname(os.path.abspath(args.input)),
                                f"收入汇总_{base}_{today}.xlsx")

    rows_out, flags, report = extract(args.input, cfg, currency_map)

    print("—— 各 sheet 处理小结 ——")
    for name, status, detail in report:
        print(f"  [{name}] {status}：{detail}")
    print(f"\n合计提取收入 {len(rows_out)} 笔")

    from collections import defaultdict
    by_cur = defaultdict(lambda: [0, 0.0])
    for _, _, _, amt, cur in rows_out:
        by_cur[cur][0] += 1
        by_cur[cur][1] += amt
    if len(by_cur) > 1:
        print("⚠ 多币种，未跨币种加总：")
    for cur, (cnt, total) in by_cur.items():
        print(f"    {cur}：{cnt} 笔，小计 {total:,.2f}")

    if flags:
        print(f"\n⚠ 有 {len(flags)} 处需人工核对（见输出的『待人工核对』sheet）：")
        for f in flags[:10]:
            print(f"    [{f[0]}] {f[1]}  {f[2]}  {f[3]}")
        if len(flags) > 10:
            print(f"    …… 还有 {len(flags) - 10} 处")

    write_output(rows_out, flags, out_path)
    print(f"\n✓ 已写出 → {out_path}")


if __name__ == "__main__":
    main()
