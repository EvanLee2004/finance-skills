#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""bank-income-extract 回归测试（合成数据，不含真实财务内容）。
跑：python3 tests/test_extract.py  →  期望 All passed。"""
import os, sys, datetime, tempfile
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SKILL = os.path.dirname(HERE)
sys.path.insert(0, os.path.join(SKILL, "scripts"))
import extract_income as E

CFG, CUR = E.load_config(os.path.join(SKILL, "config", "识别规则.md"))
passed = failed = 0


def check(name, cond, detail=""):
    global passed, failed
    if cond:
        passed += 1; print(f"  ✓ {name}")
    else:
        failed += 1; print(f"  ✗ {name}  {detail}")


def make_book(path):
    wb = openpyxl.Workbook()
    # 现金（排第一）
    ws = wb.active; ws.title = "现金"
    ws.append(["日期", "凭证号", "摘要", "支票根号", "借方（增加）", "贷方（减少）", "余额", "类型"])
    ws.append([None, None, None, None, None, None, 0, None])
    ws.append([datetime.datetime(2026, 6, 20), None, "现金收款-张三", None, 500, None, 500, "收入"])
    ws.append([datetime.datetime(2026, 6, 21), None, "零星支出", None, None, 50, 450, "支出"])
    # 工行：多笔收入 + 手续费 + 空预填收入 + 缺日期有金额 + 缺客户名收入 + 疑似漏标
    w = wb.create_sheet("工行")
    w.append(["日期", "凭证号", "摘要", "支票根号", "借方（增加）", "贷方（减少）", "余额", "类型"])
    w.append([None, None, None, None, None, None, 10000, None])
    w.append([datetime.datetime(2026, 6, 24), None, "徐州重型机械", None, 52550.41, None, 62550.41, "收入"])
    w.append([datetime.datetime(2026, 6, 24), None, "银行手续费", None, None, 25, 62525.41, "手续费"])
    w.append([datetime.datetime(2026, 6, 25), None, "利洁时中国", None, 679.99, None, 63205.40, "收入"])
    w.append([None, None, "雷沃重工(补登日期)", None, 3021.11, None, 66226.51, "收入"])   # 缺日期有金额→保留
    w.append([None, None, None, None, None, None, 66226.51, "收入"])                     # 空预填→跳过
    w.append([datetime.datetime(2026, 6, 26), None, "", None, 8000, None, 74226.51, "收入"])  # 收入缺客户名→保留+flag
    w.append([datetime.datetime(2026, 6, 27), None, "某到账没标类型", None, 999, None, 75225.51, None])  # 疑似漏标→flag不入表
    # 美元
    u = wb.create_sheet("美元")
    u.append(["日期", "凭证号", "摘要", "支票根号", "借方（增加）", "贷方（减少）", "余额", "类型"])
    u.append([datetime.datetime(2026, 6, 24), None, "SKYWORK AI", None, 12219.00, None, 12219, "收入"])
    # 非日记账页
    s = wb.create_sheet("说明")
    s.append(["这是说明页", "无表头"])
    wb.save(path)


def run():
    tmp = tempfile.mkdtemp()
    src = os.path.join(tmp, "日记账.xlsx")
    make_book(src)
    rows, flags, report = E.extract(src, CFG, CUR)

    # 汇总应含 6 笔：现金1 + 工行4(52550/679/3021/8000) + 美元1
    check("收入总笔数=6", len(rows) == 6, f"实际 {len(rows)}")
    # sheet 顺序：现金在工行前
    channels = [r[0] for r in rows]
    check("现金排在工行前", channels.index("现金") < channels.index("工行"))
    # 空预填收入行被跳过（工行只4笔非8笔）
    gh = [r for r in rows if r[0] == "工行"]
    check("工行=4笔(空预填被跳过)", len(gh) == 4, f"实际 {len(gh)}")
    # 手续费/支出不进表
    check("无手续费/支出行", all("手续费" not in r[2] and r[2] != "零星支出" for r in rows))
    # 缺日期有金额的保留
    check("缺日期有金额保留", any("雷沃" in r[2] for r in rows))
    # 美元币种正确、不与人民币混
    usd = [r for r in rows if r[0] == "美元"]
    check("美元行币种=美元", usd and usd[0][4] == "美元")
    rmb = [r for r in rows if r[4] == "人民币"]
    check("人民币5笔", len(rmb) == 5, f"实际 {len(rmb)}")

    # flags：缺客户名(8000那笔) + 疑似漏标(999那笔) = 至少2
    kinds = [f[2] for f in flags]
    check("捕获『收入缺客户名』", any("缺客户名" in k for k in kinds))
    check("捕获『疑似漏标』", any("漏标" in k for k in kinds))
    # 疑似漏标那笔(999)不进汇总
    check("疑似漏标不入汇总", all(r[3] != 999 for r in rows))
    # 说明页被跳过
    check("说明页被跳过", any(n == "说明" and st == "跳过" for n, st, _ in report))

    # 写文件不报错 + 待人工 sheet 存在
    out = os.path.join(tmp, "out.xlsx")
    E.write_output(rows, flags, out)
    wb = openpyxl.load_workbook(out)
    check("输出含收入汇总sheet", "收入汇总" in wb.sheetnames)
    check("输出含待人工核对sheet", "待人工核对" in wb.sheetnames)

    # config 生效：口头新增"到账"关键词已在默认里；验证支出词表能挡漏标
    check("config收入词含到账", "到账" in CFG["收入判定关键词"])

    print(f"\n{passed} passed, {failed} failed")
    return failed == 0


if __name__ == "__main__":
    ok = run()
    print("All passed ✅" if ok else "FAILED ❌")
    sys.exit(0 if ok else 1)
