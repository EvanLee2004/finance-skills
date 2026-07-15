# -*- coding: utf-8 -*-
"""
合规文件抽查 skill 稳定性回归（合成假数据，驱动真实 recommend.py，不重实现算法）。
跑：python3 tests/test_robustness.py   全过打印 ALL PASS
"""
import os
import sys
import csv
import subprocess
import tempfile
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SKILL = os.path.dirname(HERE)
RECOMMEND = os.path.join(SKILL, "scripts", "recommend.py")
# 本地资料侧测试数据（与 skill 源码平行）
FIXTURE_DIR = os.path.normpath(os.path.join(
    SKILL, "..", "..", "..", "技能", "合规文件抽查", "测试数据"
))
PY = sys.executable
PASS = FAIL = 0

# 与 make_fixtures 一致的虚构名
SALES_BIG = "测销甲"
SALES_SMALL = "测销乙"
SALES_UNFED = "测销丙"
SALES_NORMAL = "测销丁"
SALES_EMPTY = "测销戊"
SALES_RESIGNED = "已离职测"
SPECIAL_CUST = "方圆测科"


def check(d, c):
    global PASS, FAIL
    ok = bool(c)
    print(f"  {'✓' if ok else '✗'} {d}")
    PASS += ok
    FAIL += (not ok)


def run(args):
    r = subprocess.run([PY, RECOMMEND] + args, capture_output=True, text=True)
    return r.returncode, r.stdout + r.stderr


HEADERS17 = [
    "年度", "销售人员", "客户名称", "新智云单号", "文件名", "应收金额", "交付月份",
    "账龄(月份）", "结算阶段", "0604销售预计回款日期", "销售解释说明", "有无合同",
    "合同分类", "框架合同是否存在PO单", "应收金额是否有客户正式确认", "客户结算周期", "是否按月给客户发结算单",
]


def _row(year, sales, cust, so, amt, month, aging):
    return [year, sales, cust, so, f"f-{so}", amt, month, aging] + [""] * 9


def make_main_all(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2026.6.4"
    ws.append(HEADERS17)
    data = [
        _row(2026, SALES_BIG, "星河科技有限公司", "SO-A1", 280000, "202412", 8),
        _row(2026, SALES_BIG, "星河科技有限公司", "SO-A1b", 20000, "202412", 8),
        _row(2026, SALES_BIG, "云帆信息", "SO-A2", 50000, "202501", 7),
        _row(2026, SALES_BIG, "小额甲客", "SO-A3", 800, "202505", 3),
        _row(2026, SALES_SMALL, "青禾工作室", "SO-B1", 3200, "202411", 9),
        _row(2026, SALES_SMALL, "青禾工作室", "SO-B1b", 800, "202411", 9),
        _row(2026, SALES_SMALL, "蓝屿工作室", "SO-B2", 1500, "202412", 8),
        _row(2026, SALES_SMALL, "微尘工作室", "SO-B3", 500, "202503", 5),
        _row(2026, SALES_UNFED, "赤峰数科", "SO-C1", 18000, "202410", 10),
        _row(2026, SALES_UNFED, "赤峰数科", "SO-C2", 12000, "202411", 9),
        _row(2026, SALES_UNFED, "白杨数据", "SO-C3", 9000, "202501", 7),
        _row(2026, SALES_NORMAL, SPECIAL_CUST, "SO-D1", 25000, "202505", 2),
        _row(2026, SALES_NORMAL, SPECIAL_CUST, "SO-D2", 22000, "202504", 3),
        _row(2026, SALES_NORMAL, SPECIAL_CUST, "SO-D3", 21000, "202503", 1),
        _row(2026, SALES_NORMAL, "长账龄客户己", "SO-D4", 15000, "202409", 11),
        _row(2026, SALES_EMPTY, "瞬时客一", "SO-E1", 200, "202505", 1),
        _row(2026, SALES_EMPTY, "瞬时客二", "SO-E2", 300, "202506", 0),
        _row(2026, SALES_RESIGNED, "旧客户庚", "SO-R1", 99999, "202401", 20),
        _row(2018, "高美杰1", "挂账客户", "SO-G1", 600, "201803", 90),
    ]
    for r in data:
        ws.append(r)
    wb.save(path)


def make_hist(path, rows):
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["销售", "客户", "交付月份", "抽查日期", "反馈状态"])
        for r in rows:
            w.writerow(r)


def main():
    tmp = tempfile.mkdtemp(prefix="compliance_spot_")
    allp = os.path.join(tmp, "合成_应收all.xlsx")
    make_main_all(allp)
    hist = os.path.join(tmp, "hist.csv")
    make_hist(hist, [
        [SALES_UNFED, "赤峰数科", "202410", "2026-06-01", "未反馈"],
        [SALES_BIG, "星河科技有限公司", "202412", "2026-05-20", "已反馈"],
    ])
    out = os.path.join(tmp, "建议.txt")

    print("【0】全流程 happy path")
    rc, log = run(["--input", allp, "--history", hist, "--out", out])
    check("退出码 0", rc == 0)
    check("产出非空清单文件", os.path.isfile(out) and os.path.getsize(out) > 20)
    text = open(out, encoding="utf-8").read() if os.path.isfile(out) else ""
    check("含表头 营销人员/客户名称/交付月份", "营销人员" in text and "客户名称" in text and "交付月份" in text)
    check("默认带理由列", "理由" in text.splitlines()[0] if text else False)
    check("离职销售不在清单", SALES_RESIGNED not in text)
    check("坏账桶不在清单", "高美杰1" not in text)
    check("未反馈销售可出现", SALES_UNFED in text)
    check("小盘销售可出现（相对金额）", SALES_SMALL in text)
    # 已反馈的 星河+202412 应排除
    star_checked_out = any(
        SALES_BIG in ln and "星河" in ln and "202412" in ln
        for ln in text.splitlines()[1:]
    )
    check("已反馈单位排除", not star_checked_out)
    # 特殊客户短账龄仍可进
    special_in = any(SPECIAL_CUST in ln for ln in text.splitlines())
    check("特殊客户短账龄可进名单", special_in)
    # 确定性
    out2 = os.path.join(tmp, "建议2.txt")
    rc2, _ = run(["--input", allp, "--history", hist, "--out", out2])
    t2 = open(out2, encoding="utf-8").read() if os.path.isfile(out2) else ""
    check("同输入两次清单一致", rc2 == 0 and text == t2)

    # —— 边界 1：某销售无符合条件订单 ——
    print("【1】边界：某销售无符合条件订单（测销戊）")
    rc, log = run(["--input", allp, "--history", hist, "--out", os.path.join(tmp, "b1.txt")])
    check("退出码 0", rc == 0)
    check("日志提示无符合条件候选含测销戊", "无符合条件" in log and SALES_EMPTY in log)
    b1_text = open(os.path.join(tmp, "b1.txt"), encoding="utf-8").read()
    check("清单中无测销戊行", SALES_EMPTY not in b1_text)

    # —— 边界 2：全部客户未反馈 ——
    print("【2】边界：历史全部未反馈")
    hist_all = os.path.join(tmp, "hist_all_unfed.csv")
    make_hist(hist_all, [
        [SALES_BIG, "星河科技有限公司", "202412", "2026-06-01", "未反馈"],
        [SALES_SMALL, "青禾工作室", "202411", "2026-06-01", "未反馈"],
        [SALES_UNFED, "赤峰数科", "202410", "2026-06-01", "未反馈"],
        [SALES_NORMAL, "长账龄客户己", "202409", "2026-06-01", "未反馈"],
    ])
    rc, log = run(["--input", allp, "--history", hist_all, "--out", os.path.join(tmp, "b2.txt")])
    check("退出码 0", rc == 0)
    b2 = open(os.path.join(tmp, "b2.txt"), encoding="utf-8").read()
    check("理由或日志含未反馈", "未反馈" in b2 or "未反馈" in log)
    check("清单非空", len(b2.strip().splitlines()) > 1)

    # —— 边界 3：历史记录文件不存在 ——
    print("【3】边界：历史文件不存在仍可跑")
    missing_hist = os.path.join(tmp, "no_such_history.csv")
    rc, log = run(["--input", allp, "--history", missing_hist, "--out", os.path.join(tmp, "b3.txt")])
    check("退出码 0", rc == 0)
    check("提示历史不存在并继续", "不存在" in log and ("继续" in log or "无历史" in log or "按无历史" in log))
    check("仍产出清单", os.path.isfile(os.path.join(tmp, "b3.txt")) and os.path.getsize(os.path.join(tmp, "b3.txt")) > 20)

    # —— 边界 4a：表头完全不对 → 找不到 sheet ——
    print("【4a】边界：表头全错 → 没找到数据 sheet")
    bad = os.path.join(tmp, "bad.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "坏表"
    ws.append(["年度", "某某列", "另一列"])
    ws.append([2026, "x", "y"])
    wb.save(bad)
    rc, log = run(["--input", bad, "--out", os.path.join(tmp, "b4.txt")])
    check("退出码 1", rc == 1)
    check("报错含没找到数据 sheet", "没找到数据 sheet" in log)

    # —— 边界 4b：像应收但缺账龄列 → 精确缺列报错 ——
    print("【4b】边界：缺账龄列 → 应收表缺必要列")
    miss = os.path.join(tmp, "miss_aging.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2026.6.4"
    # 够认 sheet（≥5 个 HEADER_KEYS），但缺账龄
    ws.append(["年度", "销售人员", "客户名称", "新智云单号", "文件名", "应收金额", "交付月份", "结算阶段"])
    ws.append([2026, "测销甲", "甲公司", "SO1", "f1", 1000, "202501", ""])
    wb.save(miss)
    rc, log = run(["--input", miss, "--out", os.path.join(tmp, "b4b.txt")])
    check("退出码 1", rc == 1)
    check("报错含缺必要列与账龄", "缺必要列" in log and "账龄" in log)

    # —— 纯函数：聚合同月两单 ——
    print("【5】纯函数：同客户同月多单金额合计")
    sys.path.insert(0, os.path.join(SKILL, "scripts"))
    import recommend as R  # noqa
    units = R.group_units([
        {"销售": "A", "客户": "C", "金额": 100.0, "交付月份": "202401", "账龄": 3},
        {"销售": "A", "客户": "C", "金额": 50.0, "交付月份": "202401", "账龄": 5},
    ])
    check("聚合成 1 个单位", len(units) == 1)
    check("金额合计 150", abs(units[0]["金额"] - 150) < 1e-6)
    check("账龄取最大 5", units[0]["账龄"] == 5)
    check("销售名归一（去-高美杰）", units[0]["销售"] == "A")
    units_gm = R.group_units([
        {"销售": "张三-高美杰", "客户": "C", "金额": 10.0, "交付月份": "202401", "账龄": 1},
    ])
    check("X-高美杰 归张三", units_gm[0]["销售"] == "张三")

    # 资格：三条 OR，无魔法 5/2
    print("【5b】资格规则纯函数")
    cfg = R.load_rules()
    fake_units = [
        {"销售": "S", "客户": "老客", "交付月份": "202401", "金额": 100.0, "账龄": 8.0, "订单数": 1},
        {"销售": "S", "客户": "大客", "交付月份": "202402", "金额": 20000.0, "账龄": 1.0, "订单数": 1},
        {"销售": "S", "客户": "中客", "交付月份": "202403", "金额": 5000.0, "账龄": 4.0, "订单数": 1},
        {"销售": "S", "客户": "渣客", "交付月份": "202404", "金额": 100.0, "账龄": 0.0, "订单数": 1},
    ]
    scored = R.score_units(fake_units, [], cfg)
    by_cust = {u["客户"]: u for u in scored}
    check("长账龄够格", by_cust["老客"]["资格"] is True)
    check("金额兜底够格", by_cust["大客"]["资格"] is True)
    check("相对金额+够账龄够格", by_cust["中客"]["资格"] is True)
    check("短账龄小额不够格", by_cust["渣客"]["资格"] is False)

    # 配置五项可读
    print("【6】配置五项待确认默认存在")
    check("weekly_cap 有值", int(cfg["weekly_cap"]) > 0)
    check("include_reason 默认 True", cfg["include_reason"] is True)
    check("special 含默认客户", any("方圆" in str(x) for x in cfg["special_customers"]))
    check("history_mode 非空", bool(cfg.get("history_mode")))
    check("acceptor 默认可为空字符串", cfg.get("acceptor") == "" or cfg.get("acceptor") is not None)
    check("相对金额门槛在配置", 0 < float(cfg["relative_amount_pct"]) <= 1)

    # 若仓库内 测试数据 存在，额外跑一次（全路径）
    if os.path.isdir(FIXTURE_DIR):
        fix_all = os.path.join(FIXTURE_DIR, "合成_应收all_合规抽查.xlsx")
        fix_hist = os.path.join(FIXTURE_DIR, "合成_抽查历史.csv")
        if os.path.isfile(fix_all):
            print("【7】仓库测试数据全流程")
            fout = os.path.join(tmp, "fixture_out.txt")
            rc, log = run(["--input", fix_all, "--history", fix_hist, "--out", fout])
            check("测试数据退出 0", rc == 0)
            check("测试数据清单非空", os.path.isfile(fout) and os.path.getsize(fout) > 20)

    print(f"\n{'='*40}\n通过 {PASS} / 失败 {FAIL}  →  {'ALL PASS ✓' if FAIL == 0 else 'HAS FAILURES ✗'}")
    sys.exit(0 if FAIL == 0 else 1)


if __name__ == "__main__":
    main()
