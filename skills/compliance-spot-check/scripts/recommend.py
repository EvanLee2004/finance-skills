# -*- coding: utf-8 -*-
"""
合规文件抽查 · recommend.py（财务部 Agent 技能）
================================================================
输入：一张「应收all」（receivables-merge 产物，17 列含销售/客户/交付月份/账龄/金额）
     + 可选抽查历史 CSV（技能自建格式：销售|客户|交付月份|抽查日期|反馈状态）。
输出：本周抽查建议清单（文字版可粘进抽查表；可选 xlsx）。

核心规则（方案锁定，阈值在 config/抽查规则.md）：
  - 抽查单位 = 销售 + 客户 + 交付月份（多单合计）
  - 金额：销售内部相对比较 + ≥1 万绝对兜底
  - 账龄 ≥6 月门槛；特殊客户（如方圆）账龄相对化
  - 覆盖全部在职销售（有候选时每人≥1）再按分填满
  - 此前通知未反馈优先；离职销售跳过

用法：
  python3 recommend.py --input <应收all.xlsx> [--history <抽查历史.csv>] [--out <清单.txt|xlsx>]
缺文件/缺列会清晰报错，不裸崩。
"""
from __future__ import annotations

import os
import re
import sys
import csv
import json
import argparse
import datetime
from collections import defaultdict
from typing import Any, Dict, List, Optional, Sequence, Tuple

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

HERE = os.path.dirname(os.path.abspath(__file__))
SKILL_DIR = os.path.dirname(HERE)
CONFIG_DIR = os.path.join(SKILL_DIR, "config")
WORK_INPUT = os.path.join(SKILL_DIR, "工作区", "input")
WORK_OUTPUT = os.path.join(SKILL_DIR, "工作区", "output")

try:
    import openpyxl
except ImportError:
    print("✗ 缺 openpyxl → pip install openpyxl")
    sys.exit(1)

# 与 split-by-sales 对齐的 17 列认列键（只要求本技能用到的几列必有）
HEADER_KEYS = [
    "年度", "销售人员", "客户名称", "新智云单号", "文件名", "应收金额",
    "交付月份", "账龄", "结算阶段", "回款日期", "销售解释", "有无合同",
    "合同分类", "PO单", "客户正式确认", "客户结算周期", "是否按月",
]
# 本技能必填列（缺则清晰报错）
REQUIRED_KEYS = ["销售人员", "客户名称", "应收金额", "交付月份", "账龄"]

# 内置默认（config 覆盖）
_DEFAULTS = {
    "aging_threshold": 6,
    "amount_floor": 10000.0,
    "weekly_cap": 20,
    "min_per_sales": 1,
    "unfed_boost": 1.5,
    "include_reason": True,
    "history_mode": "技能自建新格式",
    "acceptor": "",
    "already_checked_policy": "排除",
    "ignore_sales": {"高美杰1"},
    "resigned_sales": {"已离职测"},
    "active_sales": [],  # 空 = 从数据推导
    "special_customers": ["方圆"],
}


def log(m: str) -> None:
    print(m, flush=True)


def norm(s: Any) -> str:
    return re.sub(r"\s+", "", str(s or ""))


# ===== config 解析（MD 表，与先例一致）=====
def _parse_md_table(lines: Sequence[str]) -> List[List[str]]:
    rows, seen = [], False
    for ln in lines:
        s = ln.strip()
        if not s.startswith("|"):
            continue
        cells = [c.strip() for c in s.strip("|").split("|")]
        if cells and all(c and set(c) <= set("-: ") for c in cells):
            seen = True
            continue
        if not seen:
            continue
        rows.append(cells)
    return rows


def _num_from(val: str, default: float) -> float:
    m = re.search(r"-?\d+(?:\.\d+)?", val or "")
    return float(m.group()) if m else default


def load_rules() -> Dict[str, Any]:
    """读 config/抽查规则.md → 规则 dict。缺/坏用内置默认。"""
    cfg = {k: (set(v) if isinstance(v, set) else (list(v) if isinstance(v, list) else v))
           for k, v in _DEFAULTS.items()}
    cfg["ignore_sales"] = set(_DEFAULTS["ignore_sales"])
    cfg["resigned_sales"] = set(_DEFAULTS["resigned_sales"])
    cfg["special_customers"] = list(_DEFAULTS["special_customers"])
    cfg["active_sales"] = list(_DEFAULTS["active_sales"])

    p = os.path.join(CONFIG_DIR, "抽查规则.md")
    if not os.path.isfile(p):
        return cfg
    try:
        text = open(p, encoding="utf-8").read()
    except Exception as e:
        log(f"⚠ 读 抽查规则.md 失败({e})，用内置默认。")
        return cfg

    for part in re.split(r"(?m)^##\s+", text):
        ls = part.splitlines()
        head = ls[0] if ls else ""
        rows = _parse_md_table(ls)

        if "待确认" in head or "5 个" in head or "五个" in head:
            for r in rows:
                if len(r) < 2:
                    continue
                k, v = r[0], r[1]
                if "历史表衔接" in k:
                    cfg["history_mode"] = v
                elif "特殊客户" in k:
                    names = [x.strip() for x in re.split(r"[、,，/]", v) if x.strip()]
                    if names:
                        cfg["special_customers"] = names
                elif "每周抽查条数" in k or "每周建议条数" in k:
                    cfg["weekly_cap"] = int(_num_from(v, cfg["weekly_cap"]))
                elif "理由" in k:
                    cfg["include_reason"] = v.strip() not in ("否", "不", "不带", "无", "false", "0")
                elif "验收人" in k:
                    cfg["acceptor"] = v.strip()

        elif "可调阈值" in head or "阈值" in head:
            for r in rows:
                if len(r) < 2:
                    continue
                k, v = r[0], r[1]
                if "账龄门槛" in k:
                    cfg["aging_threshold"] = int(_num_from(v, cfg["aging_threshold"]))
                elif "金额绝对兜底" in k or "绝对兜底" in k:
                    cfg["amount_floor"] = float(_num_from(v, cfg["amount_floor"]))
                elif "每周建议条数上限" in k or "每周抽查条数" in k:
                    cfg["weekly_cap"] = int(_num_from(v, cfg["weekly_cap"]))
                elif "每人最少" in k:
                    cfg["min_per_sales"] = int(_num_from(v, cfg["min_per_sales"]))
                elif "未反馈加成" in k:
                    cfg["unfed_boost"] = float(_num_from(v, cfg["unfed_boost"]))
                elif "已抽过" in k:
                    cfg["already_checked_policy"] = v.strip() or cfg["already_checked_policy"]
                elif "忽略销售" in k:
                    names = [x.strip() for x in re.split(r"[、,，/]", v) if x.strip()]
                    if names:
                        cfg["ignore_sales"] = set(names)

        elif "特殊客户" in head:
            specs = []
            for r in rows:
                if r and r[0] and r[0] not in ("客户关键词", "客户"):
                    specs.append(r[0].strip())
            if specs:
                # 合并待确认表里的名单与本表（去重保序）
                merged, seen = [], set()
                for n in list(cfg["special_customers"]) + specs:
                    if n and n not in seen:
                        seen.add(n)
                        merged.append(n)
                cfg["special_customers"] = merged

        elif "离职" in head:
            names = set()
            for r in rows:
                if r and r[0] and r[0] not in ("销售名", "销售"):
                    names.add(norm(r[0]))
            if names:
                cfg["resigned_sales"] = names

        elif "在职销售" in head or "覆盖口径" in head:
            names = []
            for r in rows:
                if r and r[0] and r[0] not in ("销售名", "销售") and r[0].strip():
                    names.append(r[0].strip())
            cfg["active_sales"] = names

    return cfg


def load_aliases() -> Dict[str, List[str]]:
    p = os.path.join(CONFIG_DIR, "列名别名.json")
    default = {
        "销售人员": ["销售人员", "销售", "营销人员"],
        "客户名称": ["客户名称", "客户"],
        "应收金额": ["应收金额", "订单折合本币", "金额"],
        "交付月份": ["交付月份", "项目交付", "销售确认"],
        "账龄": ["账龄(月份）", "账龄(月份)", "账龄", "账龄月份"],
    }
    if not os.path.isfile(p):
        return default
    try:
        with open(p, encoding="utf-8") as f:
            d = json.load(f)
        aliases = d.get("COLUMN_ALIASES") or d
        out = dict(default)
        for k, v in aliases.items():
            if k.startswith("_"):
                continue
            if isinstance(v, list) and v:
                out[k] = v
        return out
    except Exception as e:
        log(f"⚠ 读 列名别名.json 失败({e})，用内置默认。")
        return default


# ===== 读应收 all（对齐 split-by-sales）=====
def find_data_sheet(wb) -> Optional[str]:
    dated, cands = [], []
    for name in wb.sheetnames:
        if "销售反馈" in name or "建议" in name:
            continue
        ws = wb[name]
        hdr = [norm(ws.cell(1, i).value) for i in range(1, min(ws.max_column, 30) + 1)]
        if sum(1 for k in HEADER_KEYS if any(k in h for h in hdr)) >= 5:
            cands.append((ws.max_row, name))
            if re.fullmatch(r"\d{4}[.年]\d{1,2}[.月]\d{1,2}日?", name.strip()):
                dated.append((ws.max_row, name))
    pool = dated or cands
    if not pool:
        return None
    pool.sort(reverse=True)
    if len(pool) > 1:
        log(f"  注意：多个候选 sheet {[n for _, n in pool]}，取「{pool[0][1]}」")
    return pool[0][1]


def map_required_columns(ws, aliases: Dict[str, List[str]]) -> Dict[str, int]:
    """返回 规范键 → 1-based 列号。缺 REQUIRED_KEYS 则抛错。"""
    actual = {}
    for c in range(1, ws.max_column + 1):
        h = norm(ws.cell(1, c).value)
        if h:
            actual[c] = h
    mapping: Dict[str, int] = {}
    for key in REQUIRED_KEYS:
        alts = aliases.get(key, [key])
        found = None
        for c, h in actual.items():
            if c in mapping.values():
                continue
            if any(norm(a) in h or h in norm(a) for a in alts):
                # 账龄别名勿误匹配到「账龄」以外的长串：优先最短包含
                if key == "账龄" and "账龄" not in h and "账龄" not in "".join(alts):
                    continue
                found = c
                break
            # 子串匹配（与 split 的 key in h 一致）
            for a in alts:
                if norm(a) and norm(a) in h:
                    found = c
                    break
            if found:
                break
        if found is None:
            # 再试 HEADER_KEYS 风格：key 子串
            for c, h in actual.items():
                if c in mapping.values():
                    continue
                short = key.replace("人员", "").replace("名称", "")
                if key in h or short in h:
                    found = c
                    break
        if found is not None:
            mapping[key] = found

    missing = [k for k in REQUIRED_KEYS if k not in mapping]
    if missing:
        raise RuntimeError(
            f"应收表缺必要列：{missing}。"
            f"表头需含 销售人员/客户名称/应收金额/交付月份/账龄（可有别名，见 config/列名别名.json）。"
        )
    return mapping


def to_number(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("，", "").replace(" ", "")
    if s in ("", "-", "#N/A", "NA", "nan", "None"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def to_month(v: Any) -> str:
    """交付月份 → 'YYYYMM' 字符串。抠不到返回 ''。"""
    if v is None:
        return ""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        n = int(v)
        if 190001 <= n <= 210012:
            return f"{n:06d}"
        s = str(n)
        if len(s) == 6:
            return s
    s = str(v).strip()
    m = re.search(r"(20\d{2})\s*[.\-/年]?\s*(\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}{int(m.group(2)):02d}"
    digits = re.sub(r"\D", "", s)
    if len(digits) >= 6:
        return digits[:6]
    return s


def to_aging(v: Any) -> Optional[float]:
    n = to_number(v)
    return n


def read_receivable_rows(ws, aliases: Dict[str, List[str]]) -> List[Dict[str, Any]]:
    mapping = map_required_columns(ws, aliases)
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        def cell(key: str):
            c = mapping[key]
            return raw[c - 1] if c <= len(raw) else None

        sales = str(cell("销售人员") or "").strip()
        cust = str(cell("客户名称") or "").strip()
        amt = to_number(cell("应收金额"))
        month = to_month(cell("交付月份"))
        aging = to_aging(cell("账龄"))
        if not sales and not cust and amt is None:
            continue
        if not any([sales, cust, month, amt is not None]):
            continue
        rows.append({
            "销售": sales,
            "客户": cust,
            "金额": amt if amt is not None else 0.0,
            "交付月份": month,
            "账龄": aging if aging is not None else 0.0,
        })
    return rows


# ===== 历史 =====
HistoryRec = Dict[str, str]


def load_history(path: Optional[str]) -> List[HistoryRec]:
    """读技能自建 CSV。文件不存在 → 空列表（不报错）。"""
    if not path:
        return []
    if not os.path.isfile(path):
        log(f"· ⚠ 抽查历史文件不存在：{path}（按无历史继续跑）")
        return []
    recs: List[HistoryRec] = []
    with open(path, encoding="utf-8-sig", newline="") as f:
        # 尝试自动识别 delimiter
        sample = f.read(2048)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        except Exception:
            dialect = csv.excel
        reader = csv.DictReader(f, dialect=dialect)
        if not reader.fieldnames:
            return []
        # 归一列名
        def pick(row, *cands):
            for k, v in row.items():
                kn = norm(k)
                for c in cands:
                    if norm(c) in kn or kn in norm(c):
                        return str(v or "").strip()
            return ""

        for row in reader:
            sales = pick(row, "销售", "营销人员", "销售人员")
            cust = pick(row, "客户", "客户名称")
            month = to_month(pick(row, "交付月份", "月份"))
            status = pick(row, "反馈状态", "反馈", "状态")
            date = pick(row, "抽查日期", "日期")
            if not sales and not cust:
                continue
            recs.append({
                "销售": sales,
                "客户": cust,
                "交付月份": month,
                "反馈状态": status,
                "抽查日期": date,
            })
    return recs


# ===== 聚合 / 打分 / 覆盖（纯函数，测试直接调）=====
Unit = Dict[str, Any]


def group_units(rows: List[Dict[str, Any]]) -> List[Unit]:
    """按 销售+客户+交付月份 聚合：金额合计、账龄取最大。"""
    bag: Dict[Tuple[str, str, str], Unit] = {}
    for r in rows:
        key = (r["销售"], r["客户"], r["交付月份"])
        if key not in bag:
            bag[key] = {
                "销售": r["销售"],
                "客户": r["客户"],
                "交付月份": r["交付月份"],
                "金额": 0.0,
                "账龄": 0.0,
                "订单数": 0,
            }
        bag[key]["金额"] += float(r["金额"] or 0)
        bag[key]["账龄"] = max(float(bag[key]["账龄"]), float(r["账龄"] or 0))
        bag[key]["订单数"] += 1
    return list(bag.values())


def _is_special(customer: str, specials: Sequence[str]) -> bool:
    cn = norm(customer)
    for s in specials:
        sn = norm(s)
        if sn and (sn in cn or cn in sn):
            return True
    return False


def _customer_aging_median(units: List[Unit], customer: str) -> float:
    ages = [float(u["账龄"]) for u in units if norm(u["客户"]) == norm(customer)]
    if not ages:
        return 1.0
    ages = sorted(ages)
    mid = len(ages) // 2
    if len(ages) % 2:
        med = ages[mid]
    else:
        med = (ages[mid - 1] + ages[mid]) / 2.0
    return max(med, 1.0)


def _sales_amount_percentile(units: List[Unit], sales: str, amount: float) -> float:
    """该销售内部金额分位 0~1（越大金额越高）。同额并列取中位。"""
    amts = sorted(float(u["金额"]) for u in units if u["销售"] == sales)
    if not amts:
        return 0.0
    if len(amts) == 1:
        return 1.0
    # 严格小于的比例
    less = sum(1 for a in amts if a < amount)
    equal = sum(1 for a in amts if a == amount)
    return (less + 0.5 * equal) / len(amts)


def score_units(
    units: List[Unit],
    history: List[HistoryRec],
    cfg: Dict[str, Any],
) -> List[Unit]:
    """为每个单位打分并标注理由片段。过滤离职/忽略/已抽过（非未反馈）。"""
    resigned = {norm(x) for x in cfg["resigned_sales"]}
    ignore = {norm(x) for x in cfg["ignore_sales"]}
    specials = cfg["special_customers"]
    aging_th = int(cfg["aging_threshold"])
    amount_floor = float(cfg["amount_floor"])
    unfed_boost = float(cfg["unfed_boost"])
    exclude_checked = "排除" in str(cfg.get("already_checked_policy", "排除"))

    # 历史索引
    unfed_sales = set()
    checked_keys = set()  # (销售norm, 客户norm, 月份) 已抽且非未反馈
    unfed_keys = set()    # 未反馈的 key → 仍优先再抽
    for h in history:
        sn, cn, m = norm(h["销售"]), norm(h["客户"]), h["交付月份"]
        st = (h.get("反馈状态") or "").strip()
        if "未反馈" in st:
            unfed_sales.add(sn)
            unfed_keys.add((sn, cn, m))
        else:
            checked_keys.add((sn, cn, m))

    # 特殊客户中位数缓存
    med_cache: Dict[str, float] = {}

    scored: List[Unit] = []
    for u in units:
        sn = norm(u["销售"])
        if not sn or not u["客户"] or not u["交付月份"]:
            continue
        if sn in resigned or sn in ignore:
            continue
        # X-高美杰 → 归前头人（与 split 一致，覆盖口径用真实在职名）
        sales_display = u["销售"]
        if sn.endswith(norm("-高美杰")) and len(sn) > len(norm("-高美杰")):
            owner = u["销售"][: -len("-高美杰")] if u["销售"].endswith("-高美杰") else u["销售"]
            sales_display = owner
            sn = norm(owner)
            if sn in resigned or sn in ignore:
                continue

        key = (sn, norm(u["客户"]), u["交付月份"])
        if exclude_checked and key in checked_keys and key not in unfed_keys:
            continue

        amount = float(u["金额"] or 0)
        aging = float(u["账龄"] or 0)
        is_spec = _is_special(u["客户"], specials)

        reasons = []
        # 账龄分 0~10
        if is_spec:
            if u["客户"] not in med_cache:
                med_cache[u["客户"]] = _customer_aging_median(units, u["客户"])
            med = med_cache[u["客户"]]
            if aging >= med:
                aging_score = 8.0 + min(2.0, (aging - med))
                reasons.append(f"特殊客户相对账龄{aging:.0f}月≥中位{med:.0f}")
            else:
                aging_score = max(0.0, aging / med * 3.0)
                reasons.append(f"特殊客户相对账龄{aging:.0f}月(中位{med:.0f})")
            aging_ok = aging >= med
        else:
            if aging >= aging_th:
                aging_score = 8.0 + min(2.0, (aging - aging_th) / 6.0)
                reasons.append(f"账龄{aging:.0f}月≥{aging_th}")
                aging_ok = True
            else:
                aging_score = aging / max(aging_th, 1) * 3.0
                aging_ok = False

        # 金额分 0~10：销售内部分位 *10；绝对兜底
        sales_units = [x for x in units if norm(x["销售"]) == norm(u["销售"])]
        pct = _sales_amount_percentile(sales_units, u["销售"], amount)
        amount_score = pct * 10.0
        amount_ok = False
        if amount >= amount_floor:
            amount_score = max(amount_score, 10.0)
            reasons.append(f"金额{amount:.0f}≥兜底{amount_floor:.0f}")
            amount_ok = True
        else:
            reasons.append(f"销售内金额分位{pct:.0%}")
            amount_ok = pct >= 0.5  # 相对偏大也算有资格

        # 资格：账龄达标 OR 金额≥兜底 OR（特殊且相对达标）
        qualifies = aging_ok or amount >= amount_floor or (is_spec and aging_ok)
        # 小盘销售：即使金额 < 兜底，相对分位高仍可 qualifies（方案：陈霞类）
        if not qualifies and amount_score >= 5.0 and aging_score >= 2.0:
            qualifies = True
            reasons.append("销售内相对金额+账龄综合")

        base = aging_score + amount_score
        unfed = sn in unfed_sales or key in unfed_keys
        if unfed:
            base *= unfed_boost
            reasons.append("未反馈优先")
        if key in unfed_keys:
            reasons.append("本客户月上次未反馈")

        scored.append({
            "销售": sales_display,
            "客户": u["客户"],
            "交付月份": u["交付月份"],
            "金额": amount,
            "账龄": aging,
            "订单数": u.get("订单数", 1),
            "分数": base,
            "资格": qualifies,
            "未反馈": unfed,
            "理由": "；".join(reasons),
        })

    scored.sort(key=lambda x: (-x["分数"], x["销售"], x["客户"], x["交付月份"]))
    return scored


def select_recommendations(
    scored: List[Unit],
    cfg: Dict[str, Any],
    all_sales_in_data: Optional[Sequence[str]] = None,
) -> Tuple[List[Unit], Dict[str, Any]]:
    """覆盖约束：在职销售有候选时每人≥min_per；再按分填满 weekly_cap。"""
    weekly_cap = int(cfg["weekly_cap"])
    min_per = int(cfg["min_per_sales"])
    resigned = {norm(x) for x in cfg["resigned_sales"]}
    ignore = {norm(x) for x in cfg["ignore_sales"]}

    if cfg.get("active_sales"):
        active = [s for s in cfg["active_sales"] if s and norm(s) not in resigned]
    else:
        pool = all_sales_in_data or sorted({u["销售"] for u in scored})
        active = [s for s in pool if norm(s) not in resigned and norm(s) not in ignore]

    active_norm = {norm(s): s for s in active}
    # 候选：优先有资格的
    by_sales: Dict[str, List[Unit]] = defaultdict(list)
    for u in scored:
        by_sales[norm(u["销售"])].append(u)

    picked: List[Unit] = []
    picked_keys = set()
    no_candidate_sales = []

    def add(u: Unit) -> bool:
        k = (norm(u["销售"]), norm(u["客户"]), u["交付月份"])
        if k in picked_keys:
            return False
        if len(picked) >= weekly_cap:
            return False
        picked.append(u)
        picked_keys.add(k)
        return True

    # 第一轮：每人至少 min_per（优先有资格、分数高；无资格时用低分兜底覆盖）
    for sn, display in active_norm.items():
        cands = by_sales.get(sn, [])
        if not cands:
            no_candidate_sales.append(display)
            continue
        qual = [c for c in cands if c.get("资格")]
        pool = qual if qual else cands  # 无符合条件时不强塞，记入 no_candidate
        if not qual:
            no_candidate_sales.append(display)
            continue
        n = 0
        for c in pool:
            if n >= min_per:
                break
            if add(c):
                n += 1

    # 第二轮：全局按分填满
    for u in scored:
        if len(picked) >= weekly_cap:
            break
        if not u.get("资格"):
            continue
        add(u)

    meta = {
        "active_sales": list(active_norm.values()),
        "no_candidate_sales": no_candidate_sales,
        "weekly_cap": weekly_cap,
        "picked": len(picked),
    }
    # 输出稳定排序：按分数再销售
    picked.sort(key=lambda x: (-x["分数"], x["销售"], x["客户"], x["交付月份"]))
    return picked, meta


# ===== 输出 =====
def format_text_list(picked: List[Unit], include_reason: bool) -> str:
    headers = ["营销人员", "客户名称", "交付月份"]
    if include_reason:
        headers.append("理由")
    lines = ["\t".join(headers)]
    for u in picked:
        row = [str(u["销售"]), str(u["客户"]), str(u["交付月份"])]
        if include_reason:
            row.append(str(u.get("理由") or ""))
        lines.append("\t".join(row))
    return "\n".join(lines) + ("\n" if lines else "")


def write_xlsx(path: str, picked: List[Unit], include_reason: bool) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "本周抽查建议"
    headers = ["营销人员", "客户名称", "交付月份"]
    if include_reason:
        headers.append("理由")
    ws.append(headers)
    for u in picked:
        row = [u["销售"], u["客户"], u["交付月份"]]
        if include_reason:
            row.append(u.get("理由") or "")
        ws.append(row)
    wb.save(path)


def write_text(path: str, text: str) -> None:
    with open(path, "w", encoding="utf-8") as f:
        f.write(text)


# ===== 主流程 =====
def recommend(
    input_path: str,
    history_path: Optional[str] = None,
    out_path: Optional[str] = None,
    cfg: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """端到端：读 all → 聚合 → 打分 → 覆盖 → 写清单。返回报告 dict。"""
    cfg = cfg or load_rules()
    aliases = load_aliases()

    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    sheet = find_data_sheet(wb)
    if not sheet:
        wb.close()
        raise RuntimeError(
            "没找到数据 sheet（表头需含 销售人员/客户名称/应收金额/交付月份/账龄 等列）。"
        )
    log(f"· 数据 sheet：{sheet}")
    rows = read_receivable_rows(wb[sheet], aliases)
    wb.close()
    log(f"· 读到 {len(rows)} 行订单")

    units = group_units(rows)
    log(f"· 聚合为 {len(units)} 个抽查单位（销售+客户+交付月份）")

    history = load_history(history_path)
    if history_path:
        log(f"· 抽查历史：{len(history)} 条（模式：{cfg.get('history_mode')}）")
    else:
        log("· 未提供抽查历史（可选；按无历史跑）")

    scored = score_units(units, history, cfg)
    # 在职名单推导用：数据里出现的销售（复合名归并）
    sales_in_data = []
    seen = set()
    for u in units:
        name = u["销售"]
        if str(name).endswith("-高美杰"):
            name = name[: -len("-高美杰")]
        n = norm(name)
        if n and n not in seen and n not in {norm(x) for x in cfg["resigned_sales"]} \
                and n not in {norm(x) for x in cfg["ignore_sales"]}:
            seen.add(n)
            sales_in_data.append(name)

    picked, meta = select_recommendations(scored, cfg, sales_in_data)
    text = format_text_list(picked, bool(cfg.get("include_reason", True)))

    if not out_path:
        base = os.path.splitext(os.path.basename(input_path))[0]
        out_path = os.path.join(
            os.path.dirname(os.path.abspath(input_path)),
            f"本周抽查建议_{base}.txt",
        )
    os.makedirs(os.path.dirname(os.path.abspath(out_path)) or ".", exist_ok=True)
    ext = os.path.splitext(out_path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        write_xlsx(out_path, picked, bool(cfg.get("include_reason", True)))
        # 同步旁路 txt 便于粘贴
        txt_side = os.path.splitext(out_path)[0] + ".txt"
        write_text(txt_side, text)
        log(f"· 已写 Excel：{out_path}")
        log(f"· 已写文字版：{txt_side}")
    else:
        write_text(out_path, text)
        log(f"· 已写文字版：{out_path}")

    log(f"· 建议 {len(picked)} 条（上限 {meta['weekly_cap']}）；在职口径 {len(meta['active_sales'])} 人")
    if meta["no_candidate_sales"]:
        log(f"· ⚠ 无符合条件候选的销售：{', '.join(meta['no_candidate_sales'])}")
    if cfg.get("acceptor"):
        log(f"· 验收人（配置）：{cfg['acceptor']}")

    return {
        "input_rows": len(rows),
        "units": len(units),
        "scored": len(scored),
        "picked": picked,
        "text": text,
        "out_path": out_path,
        "meta": meta,
        "cfg": cfg,
    }


def main() -> None:
    ap = argparse.ArgumentParser(description="合规文件抽查 · 本周建议清单")
    ap.add_argument("--input", help="应收all.xlsx（绝对路径）")
    ap.add_argument("--history", help="抽查历史 CSV（可选；不存在则按无历史继续）")
    ap.add_argument("--out", help="输出路径 .txt 或 .xlsx（可省：源文件同目录）")
    ap.add_argument("--input-dir", default=WORK_INPUT, help="没给 --input 时去这里找最新 xlsx")
    a = ap.parse_args()

    inp = a.input
    auto_picked = not a.input
    if not inp:
        cands = []
        if os.path.isdir(a.input_dir):
            cands = [
                os.path.join(a.input_dir, f)
                for f in sorted(os.listdir(a.input_dir))
                if f.lower().endswith(".xlsx") and not f.startswith(("~$", "."))
            ]
        if not cands:
            log(f"✗ 没给 --input，也没在 {a.input_dir}/ 找到 xlsx。")
            sys.exit(1)
        inp = max(cands, key=os.path.getmtime)
    if not os.path.isfile(inp):
        log(f"✗ 输入文件不存在：{inp}")
        sys.exit(1)
    if not inp.lower().endswith((".xlsx", ".xlsm")):
        log(f"✗ 输入要是 .xlsx（收到 {os.path.basename(inp)}）。")
        sys.exit(1)

    cfg = load_rules()
    log(f"· 输入：{os.path.basename(inp)}")
    log(
        f"· 规则：账龄≥{cfg['aging_threshold']}月 / 金额兜底{cfg['amount_floor']:.0f} / "
        f"每周上限{cfg['weekly_cap']} / 理由列{'开' if cfg['include_reason'] else '关'} / "
        f"特殊客户{cfg['special_customers']} / 离职{sorted(cfg['resigned_sales'])}"
    )

    out = a.out
    if not out:
        if auto_picked:
            os.makedirs(WORK_OUTPUT, exist_ok=True)
            out = os.path.join(
                WORK_OUTPUT,
                f"本周抽查建议_{datetime.date.today().strftime('%Y%m%d')}.txt",
            )
        else:
            out = os.path.join(
                os.path.dirname(os.path.abspath(inp)),
                f"本周抽查建议_{os.path.splitext(os.path.basename(inp))[0]}.txt",
            )

    try:
        rep = recommend(inp, a.history, out, cfg)
    except Exception as e:
        log(f"✗ 推荐出错：{e}")
        sys.exit(1)

    # 清单正文打到 stdout，便于直接复制 / 测试捕获
    print("\n===== 本周抽查建议清单（可粘贴）=====", flush=True)
    print(rep["text"], end="" if rep["text"].endswith("\n") else "\n", flush=True)
    print("===== 清单结束 =====", flush=True)
    log(f"\n✓ 完成：{len(rep['picked'])} 条建议 → {rep['out_path']}")


if __name__ == "__main__":
    main()
