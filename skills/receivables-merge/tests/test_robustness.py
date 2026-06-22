# -*- coding: utf-8 -*-
"""
稳定性回归测试 · test_robustness.py
================================================================
用【合成假数据】（不依赖真实财务数据）验证 merge.py 的稳定性：
缺文件/坏参数要清晰报错、坏输入要优雅降级+告警、同参数两次结果一致。

跑：  python3 tests/test_robustness.py      （全过打印 ALL PASS）
"""
import os
import sys
import subprocess
import tempfile
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
MERGE = os.path.join(HERE, "..", "scripts", "merge.py")
PY = sys.executable
PASS = FAIL = 0


def check(desc, cond):
    global PASS, FAIL
    if cond:
        print(f"  ✓ {desc}"); PASS += 1
    else:
        print(f"  ✗ {desc}"); FAIL += 1


def run(args):
    """跑 merge.py，返回 (returncode, 合并的stdout+stderr)。"""
    r = subprocess.run([PY, MERGE] + args, capture_output=True, text=True)
    return r.returncode, r.stdout + r.stderr


def make_source(path, rename_amount=False, unknown_sheet=False):
    """造一份合成源台账：2026/2025 两个年份 sheet + 6月批量。"""
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    amt_hdr = "应收款" if rename_amount else "应收金额"
    for yr in ("2026", "2025"):
        ws = wb.create_sheet(yr)
        ws.append(["销售人员", "客户名称", "单号", "新智云单号", "文件名", amt_hdr, "项目交付"])
        ws.append(["张三", "甲公司", f"SO{yr}1", f"SO{yr}1", "f1", 1000, f"{yr}03"])
        ws.append(["李四", "乙公司", f"SO{yr}2", f"SO{yr}2", "f2", 0, f"{yr}04"])  # 应收0→应删
        ws.append(["钱七", "己公司", f"SO{yr}3", f"SO{yr}3", "f3", 0.5, f"{yr}05"])  # ≤1→应删(新口径)
    b = wb.create_sheet("6月批量")
    b.append(["销售", "客户", "订单号", "名称", "完成时间", "订单折合本币"])
    b.append(["王五", "丙公司", "GM1", "n1", "2024-06-09", 500])
    if unknown_sheet:
        u = wb.create_sheet("9月批量")
        u.append(["销售人员", "客户名称", "单号", "应收金额", "项目交付"])
        u.append(["赵六", "丁公司", "X9", 100, "202509"])
    wb.save(path)


def main():
    tmp = tempfile.mkdtemp(prefix="recv_test_")
    src = os.path.join(tmp, "源台账.xlsx"); make_source(src)
    out = os.path.join(tmp, "out.xlsx")

    print("【1】正常跑 → 成功、产物含主表")
    rc, log = run(["--source", src, "--base-month", "202601", "--out", out])
    check("退出码 0", rc == 0)
    check("产物存在", os.path.isfile(out))
    if os.path.isfile(out):
        wb = openpyxl.load_workbook(out)
        check("含 主表/认列告警/运行报告 sheet",
              all(s in wb.sheetnames for s in ("主表", "认列告警", "运行报告")))
        rows = len(list(wb["主表"].iter_rows())) - 1
        check("应收=0/≤1 的行都被删（7源行→3行）", rows == 3)
        names = {r[1] for r in wb["主表"].iter_rows(min_row=2, values_only=True)}
        check("≤1 元行(钱七0.5)被删（新口径）", "钱七" not in names)

    print("【2】缺 --source → 清晰报错(退出1)")
    rc, log = run(["--source", os.path.join(tmp, "nope.xlsx"), "--out", out])
    check("退出码 1 且提示文件不存在", rc == 1 and "不存在" in log)

    print("【3】坏 --base-month → 清晰报错(退出1)")
    rc, log = run(["--source", src, "--base-month", "abc", "--out", out])
    check("退出码 1 且提示 YYYYMM", rc == 1 and "YYYYMM" in log)

    print("【4】xlsx 当 --rules → 不崩、优雅跳过")
    rc, log = run(["--source", src, "--rules", src, "--base-month", "202601", "--out", out])
    check("退出码 0 且提示应为 .md", rc == 0 and ".md" in log)

    print("【4b】显式 --rules 缺文件 → 硬报错(退出1)，不静默跳过归属")
    rc, log = run(["--source", src, "--rules", os.path.join(tmp, "no_rules.md"), "--base-month", "202601", "--out", out])
    check("退出码 1 且提示维护表不存在", rc == 1 and "维护表文件不存在" in log)

    print("【5】关键列改名 → 大声告警(不静默丢)")
    bad = os.path.join(tmp, "bad.xlsx"); make_source(bad, rename_amount=True, unknown_sheet=True)
    out2 = os.path.join(tmp, "out2.xlsx")
    rc, log = run(["--source", bad, "--base-month", "202601", "--out", out2])
    check("退出码 0（不崩）", rc == 0)
    check("日志含『关键列没认出』", "关键列没认出" in log)
    check("日志含『未知 sheet』", "未知 sheet" in log)
    if os.path.isfile(out2):
        warn = list(openpyxl.load_workbook(out2)["认列告警"].iter_rows(values_only=True))
        check("认列告警 sheet 有内容", len(warn) >= 3)  # 表头 + 2 告警

    print("【6】确定性：同参数两次 → 主表完全一致")
    o1 = os.path.join(tmp, "d1.xlsx"); o2 = os.path.join(tmp, "d2.xlsx")
    run(["--source", src, "--base-month", "202601", "--out", o1])
    run(["--source", src, "--base-month", "202601", "--out", o2])

    def sig(p):
        ws = openpyxl.load_workbook(p, data_only=True)["主表"]
        return [tuple("" if c is None else str(c) for c in r) for r in ws.iter_rows(values_only=True)]
    check("两次主表逐格一致", sig(o1) == sig(o2))

    print(f"\n{'='*40}\n通过 {PASS} / 失败 {FAIL}  →  {'ALL PASS ✓' if FAIL == 0 else 'HAS FAILURES ✗'}")
    sys.exit(0 if FAIL == 0 else 1)


if __name__ == "__main__":
    main()
