# -*- coding: utf-8 -*-
"""
稳定性回归测试 · test_robustness.py（真 pytest）
用【合成假数据】验证 merge.py：缺文件/坏参数清晰报错、坏输入优雅降级、两次结果一致。
原自跑脚本 check() 断言语义全部保留，仅改为 assert。
"""
import os
import sys
import subprocess
import tempfile

import openpyxl
import pytest

HERE = os.path.dirname(os.path.abspath(__file__))
MERGE = os.path.join(HERE, "..", "scripts", "merge.py")
PY = sys.executable


def run(args):
    """跑 merge.py，返回 (returncode, 合并的stdout+stderr)。"""
    r = subprocess.run([PY, MERGE] + args, capture_output=True, text=True)
    return r.returncode, r.stdout + r.stderr


def make_source(path, rename_amount=False, unknown_sheet=False):
    """造一份合成源台账：2026/2025 两个年份 sheet + 6月批量。"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    amt_hdr = "应收款" if rename_amount else "应收金额"
    for yr in ("2026", "2025"):
        ws = wb.create_sheet(yr)
        ws.append(["销售人员", "客户名称", "单号", "新智云单号", "文件名", amt_hdr, "项目交付"])
        ws.append(["张三", "甲公司", f"SO{yr}1", f"SO{yr}1", "f1", 1000, f"{yr}03"])
        ws.append(["李四", "乙公司", f"SO{yr}2", f"SO{yr}2", "f2", 0, f"{yr}04"])  # 应收0→应删
        ws.append(["钱七", "己公司", f"SO{yr}3", f"SO{yr}3", "f3", 0.5, f"{yr}05"])  # ≤1→应删(新口径)
    b = wb.create_sheet("6月批量")
    b.append(["销售", "客户", "订单号", "名称", "完成时间", "订单折合本币"])
    b.append(["王五", "丙公司", "GM1", "n1", "2024-06-09", 500])
    if unknown_sheet:
        u = wb.create_sheet("9月批量")
        u.append(["销售人员", "客户名称", "单号", "应收金额", "项目交付"])
        u.append(["赵六", "丁公司", "X9", 100, "202509"])
    wb.save(path)


@pytest.fixture
def tmp():
    d = tempfile.mkdtemp(prefix="recv_test_")
    yield d


def test_happy_path_main_table(tmp):
    """【1】正常跑 → 成功、产物含主表；应收=0/≤1 删除口径。"""
    src = os.path.join(tmp, "源台账.xlsx")
    make_source(src)
    out = os.path.join(tmp, "out.xlsx")
    rc, log = run(["--source", src, "--base-month", "202601", "--out", out])
    assert rc == 0, log
    assert os.path.isfile(out)
    wb = openpyxl.load_workbook(out)
    assert all(s in wb.sheetnames for s in ("主表", "认列告警", "运行报告"))
    rows = len(list(wb["主表"].iter_rows())) - 1
    assert rows == 3, "应收=0/≤1 的行都被删（7源行→3行）"
    names = {r[1] for r in wb["主表"].iter_rows(min_row=2, values_only=True)}
    assert "钱七" not in names, "≤1 元行(钱七0.5)被删（新口径）"


def test_missing_source(tmp):
    """【2】缺 --source → 清晰报错(退出1)。"""
    out = os.path.join(tmp, "out.xlsx")
    rc, log = run(["--source", os.path.join(tmp, "nope.xlsx"), "--out", out])
    assert rc == 1 and "不存在" in log


def test_bad_base_month(tmp):
    """【3】坏 --base-month → 清晰报错(退出1)。"""
    src = os.path.join(tmp, "源台账.xlsx")
    make_source(src)
    out = os.path.join(tmp, "out.xlsx")
    rc, log = run(["--source", src, "--base-month", "abc", "--out", out])
    assert rc == 1 and "YYYYMM" in log


def test_rules_xlsx_skipped_gracefully(tmp):
    """【4】xlsx 当 --rules → 不崩、优雅跳过。"""
    src = os.path.join(tmp, "源台账.xlsx")
    make_source(src)
    out = os.path.join(tmp, "out.xlsx")
    rc, log = run(["--source", src, "--rules", src, "--base-month", "202601", "--out", out])
    assert rc == 0 and ".md" in log


def test_missing_explicit_rules_hard_fail(tmp):
    """【4b】显式 --rules 缺文件 → 硬报错(退出1)，不静默跳过归属。"""
    src = os.path.join(tmp, "源台账.xlsx")
    make_source(src)
    out = os.path.join(tmp, "out.xlsx")
    rc, log = run([
        "--source", src,
        "--rules", os.path.join(tmp, "no_rules.md"),
        "--base-month", "202601",
        "--out", out,
    ])
    assert rc == 1 and "维护表文件不存在" in log


def test_renamed_columns_and_unknown_sheet_warn(tmp):
    """【5】关键列改名 → 大声告警(不静默丢)；未知 sheet 告警。"""
    bad = os.path.join(tmp, "bad.xlsx")
    make_source(bad, rename_amount=True, unknown_sheet=True)
    out2 = os.path.join(tmp, "out2.xlsx")
    rc, log = run(["--source", bad, "--base-month", "202601", "--out", out2])
    assert rc == 0, log
    assert "关键列没认出" in log
    assert "未知 sheet" in log
    assert os.path.isfile(out2)
    warn = list(openpyxl.load_workbook(out2)["认列告警"].iter_rows(values_only=True))
    assert len(warn) >= 3  # 表头 + 2 告警


def test_deterministic_main_table(tmp):
    """【6】确定性：同参数两次 → 主表完全一致。"""
    src = os.path.join(tmp, "源台账.xlsx")
    make_source(src)
    o1 = os.path.join(tmp, "d1.xlsx")
    o2 = os.path.join(tmp, "d2.xlsx")
    run(["--source", src, "--base-month", "202601", "--out", o1])
    run(["--source", src, "--base-month", "202601", "--out", o2])

    def sig(p):
        ws = openpyxl.load_workbook(p, data_only=True)["主表"]
        return [
            tuple("" if c is None else str(c) for c in r)
            for r in ws.iter_rows(values_only=True)
        ]

    assert sig(o1) == sig(o2)


def test_ownership_scope_rules(tmp):
    """【7】归属范围：『仅高美杰』客户键只对高美杰名下生效。"""
    src2 = os.path.join(tmp, "源2.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2026"
    ws.append(["销售人员", "客户名称", "单号", "新智云单号", "文件名", "应收金额", "项目交付"])
    ws.append(["高美杰", "特变电工新疆", "A1", "A1", "f", 100, "202601"])
    ws.append(["骆利飞", "特变电工新疆新能源股份有限公司", "A2", "A2", "f", 100, "202601"])
    ws.append(["骆利飞", "中国国际电视总公司", "A3", "A3", "f", 100, "202601"])
    wb.save(src2)
    o3 = os.path.join(tmp, "scope.xlsx")
    rc, log = run(["--source", src2, "--base-month", "202601", "--out", o3])
    assert rc == 0 and os.path.isfile(o3), log
    m = {
        r[3]: r[1]
        for r in openpyxl.load_workbook(o3, data_only=True)["主表"].iter_rows(
            min_row=2, values_only=True
        )
    }
    assert m.get("A1") == "于占国-高美杰", "高美杰的特变→于占国-高美杰（仅高美杰生效）"
    assert m.get("A2") == "王雄", "骆利飞的特变→王雄（仅高美杰不及骆利飞，落段一）"
    assert m.get("A3") == "于占国", "骆利飞的中国国际电视→于占国（通用规则生效）"
