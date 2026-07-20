#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
追觅应收进度对比（dreame-ar-progress-diff）

把多份格式相近的「应收进度」Excel 按人名列智能对齐后做跨版本 diff：
  期间并集 · 值变化 · 底色变化 · 预计付款忽略公式只比备注 · 列结构变化

用法：
  python3 compare.py --inspect [--input-dir DIR]
  python3 compare.py --files a.xlsx b.xlsx [c.xlsx ...] [--out 报告.xlsx]
  python3 compare.py --input-dir DIR [--out 报告.xlsx]

规则在 config/：业务规则.md · 颜色图例.json · 子列识别.json（改表不改码）
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

HERE = os.path.dirname(os.path.abspath(__file__))
SKILL_DIR = os.path.dirname(HERE)
CONFIG_DIR = os.path.join(SKILL_DIR, "config")
WORK_INPUT = os.path.join(SKILL_DIR, "工作区", "input")
WORK_OUTPUT = os.path.join(SKILL_DIR, "工作区", "output")

# ----------------- 默认配置（config 可覆盖） -----------------
CONFIG: Dict[str, Any] = {
    "SHEET_NAME": "应收进度",
    "HEADER_ROW": 2,
    "DATA_START_ROW": 3,
    "PERIOD_COL": 1,
    "PERSON_START_COL": 2,
    "MAX_SUBCOLS": 3,
    "IGNORE_FORMULA": True,
    "DETAIL_ONLY_CHANGED_OR_TEXT": True,
    "INCLUDE_SUMMARY_SHEET": True,
}

STOP_PREFIXES_DEFAULT = [
    "为方便", "待付款", "流程中", "未收到PO", "总计", "合计", "图例", "说明",
]

SUBCOL_MAP_DEFAULT: Dict[str, List[str]] = {
    "金额": ["金额", "应收金额", "本期金额"],
    "PO时间": ["PO时间", "PO 时间", "收到PO时间", "PO"],
    "发票上传时间": ["发票上传时间", "上传发票时间", "发票时间"],
    "预计付款时间": ["预计付款时间", "预计付款", "付款时间"],
}
FIELD_ORDER_DEFAULT = ["金额", "PO时间", "发票上传时间", "预计付款时间"]

COLOR_NAMES_DEFAULT = {
    "FF00B050": "绿色(待付款)",
    "FFFED4A4": "橙色(流程中/已收PO)",
    "FFFFF258": "黄色(未收到PO)",
    "FFFFFF00": "黄色(未收到PO)",
    "FF2EA121": "深绿(进度更新)",
    "FFDEE0E3": "浅灰",
    "FF8F959E": "灰色(已回款)",
    "FFF9CBAA": "浅橙",
    "FFF8F9FA": "极浅灰",
}

STOP_PREFIXES: List[str] = list(STOP_PREFIXES_DEFAULT)
SUBCOL_MAP: Dict[str, List[str]] = {k: list(v) for k, v in SUBCOL_MAP_DEFAULT.items()}
FIELD_ORDER: List[str] = list(FIELD_ORDER_DEFAULT)
COLOR_NAMES: Dict[str, str] = dict(COLOR_NAMES_DEFAULT)

# styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="微软雅黑")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
TITLE_FONT = Font(bold=True, size=14, name="微软雅黑")
BOLD = Font(bold=True, size=11, name="微软雅黑")
NORMAL = Font(size=10, name="微软雅黑")
ITALIC_GRAY = Font(size=10, name="微软雅黑", italic=True, color="808080")
CHANGED_FILL = PatternFill("solid", fgColor="FFFF00")
MISS_FILL = PatternFill("solid", fgColor="D9D9D9")
RENAME_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="center")


def log(msg: str) -> None:
    print(msg, file=sys.stderr)


def load_config() -> None:
    """读 config 覆盖默认；缺文件/解析失败不崩。"""
    global STOP_PREFIXES, SUBCOL_MAP, FIELD_ORDER, COLOR_NAMES

    # 子列识别.json
    p_sub = os.path.join(CONFIG_DIR, "子列识别.json")
    if os.path.isfile(p_sub):
        try:
            with open(p_sub, encoding="utf-8") as f:
                d = json.load(f)
            for k in ("金额", "PO时间", "发票上传时间", "预计付款时间"):
                if k in d and isinstance(d[k], list):
                    SUBCOL_MAP[k] = [str(x) for x in d[k]]
            if isinstance(d.get("字段显示顺序"), list) and d["字段显示顺序"]:
                FIELD_ORDER = [str(x) for x in d["字段显示顺序"]]
        except Exception as e:
            log(f"⚠ 读 子列识别.json 失败({e})，用内置默认。")

    # 颜色图例.json
    p_color = os.path.join(CONFIG_DIR, "颜色图例.json")
    if os.path.isfile(p_color):
        try:
            with open(p_color, encoding="utf-8") as f:
                d = json.load(f)
            colors = d.get("颜色") or {}
            if isinstance(colors, dict) and colors:
                COLOR_NAMES = {str(k).upper(): str(v) for k, v in colors.items()}
        except Exception as e:
            log(f"⚠ 读 颜色图例.json 失败({e})，用内置默认。")

    # 业务规则.md
    p_rules = os.path.join(CONFIG_DIR, "业务规则.md")
    if not os.path.isfile(p_rules):
        return
    try:
        text = open(p_rules, encoding="utf-8").read()
    except Exception as e:
        log(f"⚠ 读 业务规则.md 失败({e})，用内置默认。")
        return

    for line in text.splitlines():
        if not line.strip().startswith("|"):
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if len(cells) < 2:
            continue
        key, val = cells[0], cells[1]
        if key in ("参数", "---", "标准字段名") or set(key) <= set("-: "):
            continue
        if key == "目标 sheet 名" and val:
            CONFIG["SHEET_NAME"] = val
        elif key == "表头所在行":
            m = re.search(r"\d+", val)
            if m:
                CONFIG["HEADER_ROW"] = int(m.group())
        elif key == "数据起始行":
            m = re.search(r"\d+", val)
            if m:
                CONFIG["DATA_START_ROW"] = int(m.group())
        elif key == "期间列":
            m = re.search(r"\d+", val)
            if m:
                CONFIG["PERIOD_COL"] = int(m.group())
        elif key == "人名起始列":
            m = re.search(r"\d+", val)
            if m:
                CONFIG["PERSON_START_COL"] = int(m.group())
        elif key == "子列最大个数":
            m = re.search(r"\d+", val)
            if m:
                CONFIG["MAX_SUBCOLS"] = int(m.group())
        elif key == "忽略公式":
            CONFIG["IGNORE_FORMULA"] = val not in ("否", "false", "False", "0", "不")
        elif key == "明细只含有变化或有文本":
            CONFIG["DETAIL_ONLY_CHANGED_OR_TEXT"] = val not in ("否", "false", "False", "0")
        elif key == "输出含文字结论 sheet":
            CONFIG["INCLUDE_SUMMARY_SHEET"] = val not in ("否", "false", "False", "0")

    # 停止词：从「三、期间停止词」段落后的列表行读取
    in_stop = False
    stops: List[str] = []
    for line in text.splitlines():
        if "期间停止词" in line:
            in_stop = True
            continue
        if in_stop:
            if line.startswith("## "):
                break
            m = re.match(r"^[-*]\s+(.+)$", line.strip())
            if m:
                stops.append(m.group(1).strip())
    if stops:
        STOP_PREFIXES = stops


# ----------------- 工具 -----------------
def color_name(rgb: str) -> str:
    if not rgb or rgb in ("00000000", "None"):
        return "无"
    return COLOR_NAMES.get(rgb.upper(), rgb)


def get_bg(cell) -> str:
    fill = cell.fill
    if not fill or not fill.start_color:
        return ""
    rgb = fill.start_color.rgb
    if rgb is None:
        return ""
    s = str(rgb)
    if s in ("00000000", "None"):
        return ""
    # theme colors without rgb → skip (can't reliably name)
    if not re.match(r"^[0-9A-Fa-f]{6,8}$", s):
        return ""
    if len(s) == 6:
        s = "FF" + s
    return s.upper()


def is_formula(v: Any) -> bool:
    return isinstance(v, str) and v.startswith("=")


def fmt_val(v: Any) -> str:
    if v is None:
        return ""
    if CONFIG["IGNORE_FORMULA"] and is_formula(v):
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        return f"{v:.2f}"
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def norm_compare(v: Any) -> str:
    """文本对比用：公式当空。"""
    if v is None:
        return ""
    if CONFIG["IGNORE_FORMULA"] and is_formula(v):
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, float):
        # 稳定数值比较：去掉多余小数噪声
        if abs(v - round(v)) < 1e-9:
            return str(int(round(v)))
        return f"{v:.6g}"
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def normalize_name(n: str) -> str:
    n = n.lower().replace(" ", "").replace("\n", "").replace("\r", "")
    n = n.replace("（", "(").replace("）", ")")
    n = n.replace("，", ",").replace("、", ",")
    return n


def extract_label(path: str, index: int) -> str:
    """从文件名抽版本标签：优先 MMDD / 月日 / 完整日期。"""
    base = os.path.splitext(os.path.basename(path))[0]
    # 20260703 / 2026.7.3 / 2026-07-03
    m = re.search(r"(20\d{2})[.\-_]?(\d{1,2})[.\-_]?(\d{1,2})", base)
    if m:
        return f"{int(m.group(2)):02d}{int(m.group(3)):02d}"
    # Martin0709 / -0703 / 0709 / 7.13
    m = re.search(r"(?<!\d)(0?\d{1,2})[.\-_]?(\d{2})(?!\d)", base)
    if m:
        return f"{int(m.group(1)):02d}{m.group(2)}"
    m = re.search(r"(?<!\d)(\d{4})(?!\d)", base)
    if m:
        return m.group(1)
    # 下载 等后缀清理后仍无
    clean = re.sub(r"[^\w\u4e00-\u9fff]+", "", base)[-8:]
    return clean or f"v{index + 1}"


def unique_labels(paths: List[str]) -> List[str]:
    raw = [extract_label(p, i) for i, p in enumerate(paths)]
    seen: Dict[str, int] = {}
    out = []
    for lab in raw:
        if lab not in seen:
            seen[lab] = 0
            out.append(lab)
        else:
            seen[lab] += 1
            out.append(f"{lab}_{seen[lab] + 1}")
    return out


def find_sheet(wb) -> Optional[str]:
    target = CONFIG["SHEET_NAME"]
    if target in wb.sheetnames:
        return target
    for sn in wb.sheetnames:
        if target in sn or "应收进度" in sn:
            return sn
    return None


def score_as_progress_sheet(ws) -> int:
    """粗评一张 sheet 是不是「应收进度」宽表。"""
    hr = CONFIG["HEADER_ROW"]
    score = 0
    names = 0
    subs = 0
    for c in range(CONFIG["PERSON_START_COL"], min(ws.max_column or 1, 60) + 1):
        v = ws.cell(hr, c).value
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        if s in _all_sub_titles():
            subs += 1
        else:
            names += 1
    if names >= 3:
        score += 30
    if names >= 8:
        score += 20
    if subs >= 3:
        score += 30
    # 期间列有月份字样
    months = 0
    for r in range(CONFIG["DATA_START_ROW"], min(CONFIG["DATA_START_ROW"] + 20, (ws.max_row or 1) + 1)):
        p = ws.cell(r, CONFIG["PERIOD_COL"]).value
        if p and re.search(r"月|年", str(p)):
            months += 1
    if months >= 3:
        score += 20
    return score


def _all_sub_titles() -> set:
    s = set()
    for titles in SUBCOL_MAP.values():
        s.update(titles)
    return s


def classify_header(text: str) -> Optional[str]:
    """表头文本 → 标准字段名；人名列返回 None。"""
    t = str(text).replace("\n", " ").strip()
    for field, aliases in SUBCOL_MAP.items():
        if field == "金额":
            continue
        if t in aliases:
            return field
    if t in SUBCOL_MAP.get("金额", []):
        return "金额"
    return None


# ----------------- 解析 -----------------
def parse_groups(ws) -> List[Dict[str, Any]]:
    """解析人名列组。每组: name, cols[field]=col_index, col_list 顺序。"""
    hr = CONFIG["HEADER_ROW"]
    start = CONFIG["PERSON_START_COL"]
    max_sub = CONFIG["MAX_SUBCOLS"]
    groups = []
    col = start
    max_col = ws.max_column or start

    while col <= max_col:
        v = ws.cell(hr, col).value
        if v is None or str(v).strip() == "":
            col += 1
            continue
        title = str(v).replace("\n", " ").strip()
        # 若本格本身是子列标题（异常结构），跳过
        if classify_header(title) and classify_header(title) != "金额":
            col += 1
            continue

        name = title
        cols_map: Dict[str, int] = {"金额": col}
        col_list = [col]
        # 向右吃子列
        for offset in range(1, max_sub + 1):
            cc = col + offset
            if cc > max_col:
                break
            sv = ws.cell(hr, cc).value
            if sv is None or str(sv).strip() == "":
                break
            st = str(sv).replace("\n", " ").strip()
            field = classify_header(st)
            if field and field != "金额":
                cols_map[field] = cc
                col_list.append(cc)
            else:
                # 撞上下一个人名列
                break
        groups.append({
            "name": name,
            "cols_map": cols_map,
            "col_list": col_list,
            "start_col": col,
        })
        col = col + len(col_list)
    return groups


def get_periods(ws) -> List[str]:
    periods = []
    pc = CONFIG["PERIOD_COL"]
    for r in range(CONFIG["DATA_START_ROW"], (ws.max_row or 0) + 1):
        p = ws.cell(r, pc).value
        if p is None:
            continue
        p_str = str(p).strip()
        if not p_str:
            continue
        if any(p_str.startswith(x) for x in STOP_PREFIXES):
            break
        # 空白行跳过但不终止（防止中间空行截断）
        if p_str not in periods:
            periods.append(p_str)
    return periods


def period_row_map(ws) -> Dict[str, int]:
    """period -> row index"""
    m: Dict[str, int] = {}
    pc = CONFIG["PERIOD_COL"]
    for r in range(CONFIG["DATA_START_ROW"], (ws.max_row or 0) + 1):
        p = ws.cell(r, pc).value
        if p is None:
            continue
        p_str = str(p).strip()
        if not p_str:
            continue
        if any(p_str.startswith(x) for x in STOP_PREFIXES):
            break
        if p_str not in m:
            m[p_str] = r
    return m


# ----------------- 对齐 -----------------
def align_score(n0: str, ng: str) -> int:
    if n0 == ng:
        return 100
    core0 = re.sub(r"\(.*?\)", "", n0).strip()
    coreg = re.sub(r"\(.*?\)", "", ng).strip()
    score = 0
    if core0 and core0 == coreg:
        score = 80
    if n0 in ng or ng in n0:
        score = max(score, 50)
    # 核心名互相包含
    if core0 and coreg and (core0 in coreg or coreg in core0):
        score = max(score, 60)
    return score


def align_groups(groups_by_label: Dict[str, List[Dict]]) -> Tuple[List[Dict], List[str]]:
    """
    以第一个 label 为锚，对齐其余。
    返回 aligned 列表 + 存疑说明。
    """
    labels = list(groups_by_label.keys())
    if not labels:
        return [], []
    anchor = labels[0]
    used = {lab: set() for lab in labels[1:]}
    aligned = []
    doubts = []

    for g0 in groups_by_label[anchor]:
        entry = {
            "name": g0["name"],
            "groups": {anchor: g0},
            "status": {anchor: "存在"},
        }
        n0 = normalize_name(g0["name"])
        for lab in labels[1:]:
            candidates = []
            for gi, g in enumerate(groups_by_label[lab]):
                if gi in used[lab]:
                    continue
                sc = align_score(n0, normalize_name(g["name"]))
                if sc >= 50:
                    candidates.append((sc, gi, g))
            candidates.sort(key=lambda x: (-x[0], x[1]))
            if not candidates:
                entry["status"][lab] = "缺失"
                continue
            best_sc, best_gi, best_g = candidates[0]
            # 存疑：第二名同分
            if len(candidates) > 1 and candidates[1][0] == best_sc:
                doubts.append(
                    f"{g0['name']} 在 {lab} 有并列候选: "
                    f"{best_g['name']} / {candidates[1][2]['name']}（分={best_sc}）"
                )
            used[lab].add(best_gi)
            entry["groups"][lab] = best_g
            if n0 != normalize_name(best_g["name"]):
                entry["status"][lab] = f"名称变化: {best_g['name']}"
            else:
                entry["status"][lab] = "存在"
        aligned.append(entry)

    # 后续文件独有列（锚文件没有的）
    label_index = {lab: i for i, lab in enumerate(labels)}
    for lab in labels[1:]:
        for gi, g in enumerate(groups_by_label[lab]):
            if gi in used[lab]:
                continue
            entry = {
                "name": g["name"],
                "groups": {lab: g},
                "status": {l: ("存在" if l == lab else "缺失") for l in labels},
            }
            n = normalize_name(g["name"])
            # 与更后版本对齐
            for lab2 in labels:
                if label_index[lab2] <= label_index[lab]:
                    continue
                best, best_sc, best_gi = None, 0, None
                for gi2, g2 in enumerate(groups_by_label[lab2]):
                    if gi2 in used[lab2]:
                        continue
                    sc = align_score(n, normalize_name(g2["name"]))
                    if sc > best_sc:
                        best_sc, best, best_gi = sc, g2, gi2
                if best is not None and best_sc >= 50:
                    used[lab2].add(best_gi)
                    entry["groups"][lab2] = best
                    if n != normalize_name(best["name"]):
                        entry["status"][lab2] = f"名称变化: {best['name']}"
                    else:
                        entry["status"][lab2] = "存在"
            aligned.append(entry)

    return aligned, doubts


# ----------------- 对比核心 -----------------
def find_cell(
    sheets, row_maps, groups, lab: str, period: str, field: str
) -> Tuple[Any, str, bool, bool]:
    """
    return (value, bg, is_formula, exists)
    exists=False → 该版本无此人列 或 无该期间
    """
    if lab not in groups:
        return None, "", False, False
    g = groups[lab]
    cmap = g["cols_map"]
    if field not in cmap:
        # 字段不存在：仍算 exists=True（人在、期间可能在），值为空
        if period not in row_maps[lab]:
            return None, "", False, False
        return None, "", False, True
    if period not in row_maps[lab]:
        return None, "", False, False
    r = row_maps[lab][period]
    c = cmap[field]
    cell = sheets[lab].cell(r, c)
    v = cell.value
    return v, get_bg(cell), is_formula(v), True


def compare_all(
    labels: List[str],
    sheets: Dict[str, Any],
    row_maps: Dict[str, Dict[str, int]],
    aligned: List[Dict],
    period_union: List[str],
) -> Dict[str, Any]:
    rows_detail = []
    val_changes = []
    color_changes = []
    pay_remark_changes = []
    pay_color_only = []
    formula_notes = []

    for entry in aligned:
        name = entry["name"]
        groups = entry["groups"]
        for period in period_union:
            for field in FIELD_ORDER:
                # 至少有一个版本有这个 field 列
                has_field = any(field in g["cols_map"] for g in groups.values())
                if not has_field:
                    # 金额永远在 cols_map；若单列组只有金额，其它 field 跳过
                    if field != "金额":
                        continue
                    # 金额：有 col_list 就有金额
                    if not any(g.get("cols_map") for g in groups.values()):
                        continue

                vals, bgs, formulas, exists = {}, {}, {}, {}
                for lab in labels:
                    v, bg, fml, ex = find_cell(sheets, row_maps, groups, lab, period, field)
                    vals[lab] = v
                    bgs[lab] = bg
                    formulas[lab] = fml
                    exists[lab] = ex

                if not any(exists[lab] for lab in labels if lab in groups):
                    continue

                # --- 值变化 ---
                text_set = set()
                for lab in labels:
                    if lab not in groups or not exists[lab]:
                        text_set.add("(无该期间)")
                    else:
                        t = norm_compare(vals[lab])
                        text_set.add(t if t else "(空)")
                non_empty = [t for t in text_set if t not in ("(空)", "(无该期间)")]
                val_changed = len(text_set) > 1 and bool(non_empty)

                # --- 底色变化：只在同时存在期间的版本间比 ---
                present_colors = []
                for lab in labels:
                    if lab in groups and exists[lab]:
                        present_colors.append(bgs[lab] if bgs[lab] else "(无)")
                color_changed = len(set(present_colors)) > 1 if len(present_colors) >= 2 else False

                any_text = any(
                    norm_compare(vals[lab])
                    for lab in labels
                    if lab in groups and exists[lab]
                )
                any_fml = any(
                    formulas[lab] for lab in labels if lab in groups and exists[lab]
                )
                any_bg = any(
                    bgs[lab] for lab in labels if lab in groups and exists[lab]
                )

                # 公式-only note（预计付款）
                if field == "预计付款时间":
                    fml_flags = [
                        bool(formulas[lab])
                        for lab in labels
                        if lab in groups and exists[lab]
                    ]
                    texts = [
                        norm_compare(vals[lab])
                        for lab in labels
                        if lab in groups and exists[lab]
                    ]
                    if (
                        fml_flags
                        and len(set(fml_flags)) > 1
                        and not any(texts)
                        and not color_changed
                    ):
                        formula_notes.append({
                            "name": name,
                            "period": period,
                            "vals": {lab: vals[lab] for lab in labels},
                        })

                    remark_set = set()
                    for lab in labels:
                        if lab not in groups or not exists[lab]:
                            remark_set.add("(无该期间)")
                        else:
                            t = norm_compare(vals[lab])
                            remark_set.add(t if t else "(空)")
                    remark_non_empty = [
                        x for x in remark_set if x not in ("(空)", "(无该期间)")
                    ]
                    remark_changed = len(remark_set) > 1 and bool(remark_non_empty)
                    if remark_changed:
                        pay_remark_changes.append({
                            "name": name, "period": period, "vals": vals, "bgs": bgs
                        })
                    elif color_changed:
                        pay_color_only.append({
                            "name": name, "period": period, "bgs": bgs
                        })

                if not any_text and not any_fml and not any_bg and not val_changed and not color_changed:
                    continue

                change_parts = []
                if val_changed:
                    change_parts.append("值变化")
                if color_changed:
                    seq = "→".join(
                        color_name(bgs[lab])
                        if lab in groups and exists[lab] and bgs[lab]
                        else ("无该期间" if lab not in groups or not exists[lab] else "无")
                        for lab in labels
                    )
                    change_parts.append(f"底色变化: {seq}")

                row = {
                    "name": name,
                    "period": period,
                    "field": field,
                    "vals": vals,
                    "bgs": bgs,
                    "formulas": formulas,
                    "exists": exists,
                    "in_groups": {lab: lab in groups for lab in labels},
                    "val_changed": val_changed,
                    "color_changed": color_changed,
                    "change_desc": "; ".join(change_parts),
                }
                rows_detail.append(row)
                if val_changed:
                    val_changes.append(row)
                if color_changed:
                    color_changes.append(row)

    return {
        "rows_detail": rows_detail,
        "val_changes": val_changes,
        "color_changes": color_changes,
        "pay_remark_changes": pay_remark_changes,
        "pay_color_only": pay_color_only,
        "formula_notes": formula_notes,
    }


# ----------------- 写报告 -----------------
def style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_bg(cell, bg: str) -> None:
    if not bg:
        return
    fc = bg[2:] if bg.startswith("FF") and len(bg) == 8 else bg
    try:
        cell.fill = PatternFill("solid", fgColor=fc)
    except Exception:
        pass


def write_val_cell(cell, v, bg, is_fml=False) -> None:
    cell.border = THIN
    cell.alignment = WRAP
    if is_fml and CONFIG["IGNORE_FORMULA"]:
        cell.value = "(公式，已忽略)"
        cell.font = ITALIC_GRAY
    else:
        cell.value = fmt_val(v) if not (is_fml and not CONFIG["IGNORE_FORMULA"]) else str(v)
        cell.font = NORMAL
    apply_bg(cell, bg)


def build_conclusion_lines(
    labels, paths, groups_by_label, period_union, result, doubts, aligned
) -> List[str]:
    lines = []
    lines.append(
        f"对比版本：{'  →  '.join(labels)}（共 {len(labels)} 份，sheet={CONFIG['SHEET_NAME']}）"
    )
    lines.append(
        "文件：" + "；".join(f"{lab}={os.path.basename(p)}" for lab, p in zip(labels, paths))
    )
    gcounts = "，".join(f"{lab}={len(groups_by_label[lab])}组" for lab in labels)
    lines.append(f"人名列组：{gcounts}")
    lines.append(f"期间并集（{len(period_union)}）：{', '.join(period_union)}")
    lines.append("")
    lines.append(
        f"值变化 {len(result['val_changes'])} 处；"
        f"底色变化 {len(result['color_changes'])} 处；"
        f"预计付款备注变化 {len(result['pay_remark_changes'])} 处；"
        f"预计付款仅底色 {len(result['pay_color_only'])} 处；"
        f"公式差异已忽略 {len(result['formula_notes'])} 处"
    )
    # 列结构要点
    rename = []
    missing = []
    for e in aligned:
        for lab in labels[1:]:
            st = e["status"].get(lab, "")
            if st.startswith("名称变化"):
                rename.append(f"{e['name']} → {st.replace('名称变化: ', '')}（{lab}）")
            if st == "缺失" and labels[0] in e["groups"]:
                missing.append(f"{e['name']} 在 {lab} 缺失")
            if st == "存在" and labels[0] not in e["groups"]:
                pass
    new_only = [
        e["name"]
        for e in aligned
        if labels[0] not in e["groups"]
    ]
    if rename:
        lines.append("改名对齐：" + "；".join(rename[:8]) + ("…" if len(rename) > 8 else ""))
    if missing:
        lines.append("列缺失：" + "；".join(missing[:8]) + ("…" if len(missing) > 8 else ""))
    if new_only:
        lines.append("仅后期出现的列：" + "；".join(new_only[:8]))
    if doubts:
        lines.append("对齐存疑：" + "；".join(doubts[:5]))

    # 业务向摘要（从变化里抽）
    lines.append("")
    lines.append("【关键变化摘要】")
    # 回款
    settled = [
        r for r in result["val_changes"]
        if r["field"] == "金额" and any(
            "回款" in norm_compare(r["vals"].get(lab)) for lab in labels
        )
    ]
    if settled:
        lines.append(
            "· 回款/核销：" + "；".join(
                f"{r['name']}|{r['period']}" for r in settled[:6]
            )
        )
    # 催促
    urge = [
        r for r in result["val_changes"]
        if any("催促" in norm_compare(r["vals"].get(lab)) for lab in labels)
    ]
    if urge:
        lines.append(f"· 含「催促」字样的值变化：{len(urge)} 处（详见「值变化」sheet）")
    # 收到po
    po = [
        r for r in result["val_changes"]
        if any(
            re.search(r"收到\s*po|收到PO|收到Po", norm_compare(r["vals"].get(lab)), re.I)
            for lab in labels
        )
    ]
    if po:
        lines.append(f"· 含「收到PO」类表述的值变化：{len(po)} 处")
    # 新增期间有内容
    early_periods = set()
    # first file periods
    # approximate: periods only in later labels
    lines.append(
        "· 建议优先看：「值变化」「颜色变化」两 sheet；预计付款公式差异不计入业务备注。"
    )
    return lines


def write_report(
    out_path: str,
    labels: List[str],
    paths: List[str],
    groups_by_label: Dict[str, List],
    period_union: List[str],
    aligned: List[Dict],
    result: Dict[str, Any],
    doubts: List[str],
) -> None:
    out = openpyxl.Workbook()

    # --- 结论摘要 ---
    if CONFIG["INCLUDE_SUMMARY_SHEET"]:
        ws0 = out.active
        ws0.title = "结论摘要"
    else:
        ws0 = out.active
        ws0.title = "汇总说明"

    ws0["A1"] = "追觅应收进度对比报告"
    ws0["A1"].font = TITLE_FONT
    lines = build_conclusion_lines(
        labels, paths, groups_by_label, period_union, result, doubts, aligned
    )
    row = 3
    ws0.cell(row, 1, "对比范围").font = BOLD
    ws0.cell(row, 2, " → ".join(labels) + f"  （{CONFIG['SHEET_NAME']}）")
    row = 4
    ws0.cell(row, 1, "生成时间").font = BOLD
    ws0.cell(row, 2, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    row = 6
    ws0.cell(row, 1, "统计").font = BOLD
    ws0.cell(row + 1, 2, f"值变化：{len(result['val_changes'])} 处")
    ws0.cell(row + 2, 2, f"底色变化：{len(result['color_changes'])} 处")
    ws0.cell(row + 3, 2, f"预计付款·备注文本变化：{len(result['pay_remark_changes'])} 处")
    ws0.cell(row + 4, 2, f"预计付款·仅底色变化：{len(result['pay_color_only'])} 处")
    ws0.cell(row + 5, 2, f"预计付款·仅公式差异（已忽略）：{len(result['formula_notes'])} 处")
    row = 13
    ws0.cell(row, 1, "结构").font = BOLD
    ws0.cell(
        row + 1, 2,
        "；".join(f"{lab}={len(groups_by_label[lab])}组" for lab in labels),
    )
    ws0.cell(row + 2, 2, f"期间并集（{len(period_union)}）：{', '.join(period_union)}")
    row = 17
    ws0.cell(row, 1, "结论摘要").font = BOLD
    for i, line in enumerate(lines):
        ws0.cell(row + 1 + i, 1, line).font = NORMAL
    row = 17 + len(lines) + 3
    ws0.cell(row, 1, "Sheet 说明").font = BOLD
    for j, t in enumerate([
        "① 列结构对比 — 人名列存在/改名/缺失",
        "② 值变化 — 仅文本/数字变化（推荐先看）",
        "③ 颜色变化 — 仅底色变化",
        "④ 预计付款时间 — 备注或底色（公式已忽略）",
        "⑤ 明细对比 — 有文本或有变化的并排明细",
        "⑥ 运行报告 — 文件/对齐/存疑（给 agent 看）",
    ]):
        ws0.cell(row + 1 + j, 1, t)
    ws0.column_dimensions["A"].width = 22
    ws0.column_dimensions["B"].width = 100

    # --- 列结构 ---
    ws1 = out.create_sheet("列结构对比")
    headers1 = ["人名列"] + labels + [f"子列数({labels[0]})"]
    for i, h in enumerate(headers1, 1):
        ws1.cell(1, i, h)
    style_header(ws1, len(headers1))
    for ri, entry in enumerate(aligned, 2):
        c = ws1.cell(ri, 1, entry["name"])
        c.font = NORMAL
        c.border = THIN
        for fi, lab in enumerate(labels):
            st = entry["status"].get(lab, "缺失")
            cell = ws1.cell(ri, 2 + fi, st)
            cell.font = NORMAL
            cell.border = THIN
            if st == "缺失":
                cell.fill = MISS_FILL
            elif st.startswith("名称变化"):
                cell.fill = RENAME_FILL
        g0 = entry["groups"].get(labels[0])
        nsub = len(g0["col_list"]) if g0 else ""
        cell = ws1.cell(ri, 2 + len(labels), nsub)
        cell.border = THIN
        cell.font = NORMAL
    ws1.column_dimensions["A"].width = 36
    for col in range(2, 3 + len(labels)):
        ws1.column_dimensions[get_column_letter(col)].width = 28

    # --- 值变化 ---
    wsV = out.create_sheet("值变化")
    headersV = (
        ["人名列", "期间", "字段"]
        + [f"{lab}值" for lab in labels]
        + [f"{lab}底色" for lab in labels]
    )
    for i, h in enumerate(headersV, 1):
        wsV.cell(1, i, h)
    style_header(wsV, len(headersV))
    for ri, row in enumerate(result["val_changes"], 2):
        for col, key in enumerate(["name", "period", "field"], 1):
            cell = wsV.cell(ri, col, row[key])
            cell.font = NORMAL
            cell.border = THIN
        for fi, lab in enumerate(labels):
            write_val_cell(
                wsV.cell(ri, 4 + fi),
                row["vals"].get(lab),
                row["bgs"].get(lab, ""),
                is_fml=row["formulas"].get(lab, False),
            )
        base = 4 + len(labels)
        for fi, lab in enumerate(labels):
            bg = row["bgs"].get(lab, "")
            cell = wsV.cell(ri, base + fi, color_name(bg) if bg else "无")
            cell.font = NORMAL
            cell.border = THIN
            apply_bg(cell, bg)
    wsV.column_dimensions["A"].width = 28
    wsV.column_dimensions["B"].width = 12
    wsV.column_dimensions["C"].width = 14
    for col in range(4, 4 + 2 * len(labels)):
        wsV.column_dimensions[get_column_letter(col)].width = 26

    # --- 颜色变化 ---
    ws4 = out.create_sheet("颜色变化")
    headers4 = (
        ["人名列", "期间", "字段"]
        + [f"{lab}色" for lab in labels]
        + ["变化路径", f"{labels[-1]}当前值"]
    )
    for i, h in enumerate(headers4, 1):
        ws4.cell(1, i, h)
    style_header(ws4, len(headers4))
    for ri, row in enumerate(result["color_changes"], 2):
        for col, key in enumerate(["name", "period", "field"], 1):
            cell = ws4.cell(ri, col, row[key])
            cell.font = NORMAL
            cell.border = THIN
        for fi, lab in enumerate(labels):
            bg = row["bgs"].get(lab, "")
            if not row["in_groups"].get(lab) or not row["exists"].get(lab):
                label = "无该期间"
            elif bg:
                label = color_name(bg)
            else:
                label = "无"
            cell = ws4.cell(ri, 4 + fi, label)
            cell.font = NORMAL
            cell.border = THIN
            apply_bg(cell, bg)
        seq = " → ".join(
            color_name(row["bgs"][lab])
            if row["in_groups"].get(lab) and row["exists"].get(lab) and row["bgs"].get(lab)
            else ("无该期间" if not row["in_groups"].get(lab) or not row["exists"].get(lab) else "无")
            for lab in labels
        )
        cell = ws4.cell(ri, 4 + len(labels), seq)
        cell.font = NORMAL
        cell.border = THIN
        last = labels[-1]
        write_val_cell(
            ws4.cell(ri, 5 + len(labels)),
            row["vals"].get(last),
            row["bgs"].get(last, ""),
            is_fml=row["formulas"].get(last, False),
        )
    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 12
    ws4.column_dimensions["C"].width = 14
    for col in range(4, 6 + len(labels)):
        ws4.column_dimensions[get_column_letter(col)].width = 22

    # --- 预计付款 ---
    ws3 = out.create_sheet("预计付款时间")
    headers3 = ["人名列", "期间"]
    for lab in labels:
        headers3 += [f"{lab}备注", f"{lab}底色"]
    headers3.append("变化说明")
    for i, h in enumerate(headers3, 1):
        ws3.cell(1, i, h)
    style_header(ws3, len(headers3))
    ri = 2
    for row in result["rows_detail"]:
        if row["field"] != "预计付款时间":
            continue
        remark_set = set()
        for lab in labels:
            if not row["in_groups"].get(lab) or not row["exists"].get(lab):
                remark_set.add("(无该期间)")
            else:
                t = norm_compare(row["vals"].get(lab))
                remark_set.add(t if t else "(空)")
        remark_non_empty = [x for x in remark_set if x not in ("(空)", "(无该期间)")]
        remark_changed = len(remark_set) > 1 and bool(remark_non_empty)
        if not remark_changed and not row["color_changed"]:
            continue
        parts = []
        if remark_changed:
            parts.append("备注文本变化")
        if row["color_changed"]:
            parts.append(
                "底色变化: "
                + "→".join(
                    color_name(row["bgs"][lab])
                    if row["in_groups"].get(lab) and row["exists"].get(lab) and row["bgs"].get(lab)
                    else "无"
                    for lab in labels
                )
            )
        ws3.cell(ri, 1, row["name"]).border = THIN
        ws3.cell(ri, 1).font = NORMAL
        ws3.cell(ri, 2, row["period"]).border = THIN
        ws3.cell(ri, 2).font = NORMAL
        for fi, lab in enumerate(labels):
            write_val_cell(
                ws3.cell(ri, 3 + fi * 2),
                row["vals"].get(lab),
                row["bgs"].get(lab, ""),
                is_fml=row["formulas"].get(lab, False),
            )
            bg = row["bgs"].get(lab, "")
            cell = ws3.cell(ri, 4 + fi * 2, color_name(bg) if bg else "无")
            cell.border = THIN
            cell.font = NORMAL
            apply_bg(cell, bg)
        cell = ws3.cell(ri, 3 + len(labels) * 2, "; ".join(parts))
        cell.border = THIN
        cell.font = NORMAL
        cell.fill = CHANGED_FILL
        ri += 1
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 12
    for col in range(3, 4 + len(labels) * 2):
        ws3.column_dimensions[get_column_letter(col)].width = 22

    # --- 明细 ---
    ws2 = out.create_sheet("明细对比")
    headers2 = ["人名列", "期间", "字段"]
    for lab in labels:
        headers2 += [f"{lab}值", f"{lab}底色"]
    headers2.append("变化说明")
    for i, h in enumerate(headers2, 1):
        ws2.cell(1, i, h)
    style_header(ws2, len(headers2))
    ri = 2
    detail_n = 0
    for row in result["rows_detail"]:
        has_text = any(
            norm_compare(row["vals"].get(lab))
            for lab in labels
            if row["in_groups"].get(lab) and row["exists"].get(lab)
        )
        if CONFIG["DETAIL_ONLY_CHANGED_OR_TEXT"]:
            if not row["change_desc"] and not has_text:
                continue
        for col, key in enumerate(["name", "period", "field"], 1):
            cell = ws2.cell(ri, col, row[key])
            cell.font = NORMAL
            cell.border = THIN
        for fi, lab in enumerate(labels):
            write_val_cell(
                ws2.cell(ri, 4 + fi * 2),
                row["vals"].get(lab),
                row["bgs"].get(lab, ""),
                is_fml=row["formulas"].get(lab, False),
            )
            bg = row["bgs"].get(lab, "")
            if not row["in_groups"].get(lab):
                label = "无此人列"
            elif not row["exists"].get(lab):
                label = "无该期间"
            elif bg:
                label = color_name(bg)
            else:
                label = "无"
            cell = ws2.cell(ri, 5 + fi * 2, label)
            cell.border = THIN
            cell.font = NORMAL
            apply_bg(cell, bg)
        cell = ws2.cell(ri, 4 + len(labels) * 2, row["change_desc"])
        cell.border = THIN
        cell.font = NORMAL
        if row["change_desc"]:
            cell.fill = CHANGED_FILL
        ri += 1
        detail_n += 1
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 14
    for col in range(4, 5 + len(labels) * 2):
        ws2.column_dimensions[get_column_letter(col)].width = 20

    # --- 运行报告 ---
    wsR = out.create_sheet("运行报告")
    wsR["A1"] = "运行报告（agent/开发核对用）"
    wsR["A1"].font = TITLE_FONT
    r = 3
    wsR.cell(r, 1, "输入文件").font = BOLD
    for i, (lab, p) in enumerate(zip(labels, paths)):
        wsR.cell(r + 1 + i, 1, lab)
        wsR.cell(r + 1 + i, 2, p)
        wsR.cell(r + 1 + i, 3, f"列组={len(groups_by_label[lab])}")
    r = r + 2 + len(labels)
    wsR.cell(r, 1, "对齐存疑").font = BOLD
    if doubts:
        for i, d in enumerate(doubts):
            wsR.cell(r + 1 + i, 1, d)
    else:
        wsR.cell(r + 1, 1, "无")
    r = r + 3 + max(len(doubts), 1)
    wsR.cell(r, 1, "公式忽略清单").font = BOLD
    if result["formula_notes"]:
        for i, fn in enumerate(result["formula_notes"]):
            wsR.cell(r + 1 + i, 1, f"{fn['name']} | {fn['period']} | {fn['vals']}")
    else:
        wsR.cell(r + 1, 1, "无")
    r = r + 3 + max(len(result["formula_notes"]), 1)
    wsR.cell(r, 1, "配置摘要").font = BOLD
    wsR.cell(r + 1, 1, json.dumps(CONFIG, ensure_ascii=False))
    wsR.cell(r + 2, 1, f"明细有效行={detail_n}")
    wsR.column_dimensions["A"].width = 40
    wsR.column_dimensions["B"].width = 80
    wsR.column_dimensions["C"].width = 20

    # 回写明细行数到摘要
    # find last stats area - optional

    os.makedirs(os.path.dirname(os.path.abspath(out_path)) or ".", exist_ok=True)
    out.save(out_path)
    out.close()
    log(f"✓ 报告已写：{out_path}")
    log(
        f"  值变化={len(result['val_changes'])} 底色={len(result['color_changes'])} "
        f"付款备注={len(result['pay_remark_changes'])} 公式忽略={len(result['formula_notes'])} "
        f"明细行={detail_n}"
    )


# ----------------- 找文件 / 检验 -----------------
def list_xlsx(dirpath: str) -> List[str]:
    if not os.path.isdir(dirpath):
        return []
    out = []
    for name in sorted(os.listdir(dirpath)):
        if name.startswith("~$"):
            continue
        if name.startswith("."):
            continue
        low = name.lower()
        if not (low.endswith(".xlsx") or low.endswith(".xlsm")):
            continue
        # 跳过产出报告
        if "对比报告" in name or "对比结论" in name:
            continue
        out.append(os.path.join(dirpath, name))
    return out


def sniff_file(path: str) -> Dict[str, Any]:
    info = {
        "path": path,
        "basename": os.path.basename(path),
        "ok": False,
        "sheet": None,
        "groups": 0,
        "periods": [],
        "score": 0,
        "error": "",
    }
    try:
        wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    except Exception as e:
        info["error"] = f"打开失败: {e}"
        return info
    try:
        sn = find_sheet(wb)
        if not sn:
            # 尝试打分选最高
            best_sn, best_sc = None, -1
            for name in wb.sheetnames:
                sc = score_as_progress_sheet(wb[name])
                if sc > best_sc:
                    best_sc, best_sn = sc, name
            if best_sc >= 50:
                sn = best_sn
                info["score"] = best_sc
            else:
                info["error"] = f"未找到「{CONFIG['SHEET_NAME']}」sheet；候选最高分={best_sc}"
                wb.close()
                return info
        else:
            info["score"] = score_as_progress_sheet(wb[sn])
        ws = wb[sn]
        groups = parse_groups(ws)
        periods = get_periods(ws)
        info["ok"] = len(groups) >= 1 and len(periods) >= 1
        info["sheet"] = sn
        info["groups"] = len(groups)
        info["periods"] = periods
        if not info["ok"]:
            info["error"] = f"sheet={sn} 但列组={len(groups)} 期间={len(periods)}，不像进度表"
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return info


def inspect_dir(input_dir: str) -> List[Dict[str, Any]]:
    files = list_xlsx(input_dir)
    results = [sniff_file(p) for p in files]
    return results


# ----------------- 主流程 -----------------
def run_compare(paths: List[str], out_path: str) -> Dict[str, Any]:
    if len(paths) < 2:
        raise SystemExit("至少需要 2 个 Excel 版本才能对比。")

    labels = unique_labels(paths)
    # 按标签排序（通常 MMDD 数字序）；保持用户给定顺序更可预测
    # 用户 --files 顺序优先；--input-dir 用文件名排序
    sheets = {}
    wbs = []
    groups_by_label = {}
    row_maps = {}
    period_lists = {}

    for lab, path in zip(labels, paths):
        if not os.path.isfile(path):
            raise SystemExit(f"文件不存在: {path}")
        wb = openpyxl.load_workbook(path, data_only=False)
        wbs.append(wb)
        sn = find_sheet(wb)
        if not sn:
            best_sn, best_sc = None, -1
            for name in wb.sheetnames:
                sc = score_as_progress_sheet(wb[name])
                if sc > best_sc:
                    best_sc, best_sn = sc, name
            if best_sc >= 50:
                sn = best_sn
                log(f"⚠ {os.path.basename(path)} 无精确 sheet 名，选用「{sn}」(分={best_sc})")
            else:
                raise SystemExit(
                    f"{path} 找不到「{CONFIG['SHEET_NAME']}」sheet，且无高分候选。"
                )
        ws = wb[sn]
        sheets[lab] = ws
        groups_by_label[lab] = parse_groups(ws)
        row_maps[lab] = period_row_map(ws)
        period_lists[lab] = get_periods(ws)
        log(
            f"· {lab}: {os.path.basename(path)} | sheet={sn} | "
            f"列组={len(groups_by_label[lab])} | 期间={period_lists[lab]}"
        )
        if len(groups_by_label[lab]) == 0:
            raise SystemExit(f"{path} 解析不到任何人名列组，请检查表头行配置。")

    # 期间并集：保持第一文件顺序，其后追加
    period_union: List[str] = list(period_lists[labels[0]])
    for lab in labels[1:]:
        for p in period_lists[lab]:
            if p not in period_union:
                period_union.append(p)

    aligned, doubts = align_groups(groups_by_label)
    if doubts:
        for d in doubts:
            log(f"⚠ 对齐存疑: {d}")

    result = compare_all(labels, sheets, row_maps, aligned, period_union)
    write_report(
        out_path, labels, paths, groups_by_label, period_union, aligned, result, doubts
    )

    for wb in wbs:
        try:
            wb.close()
        except Exception:
            pass

    return {
        "labels": labels,
        "paths": paths,
        "period_union": period_union,
        "aligned": aligned,
        "doubts": doubts,
        "result": result,
        "out_path": out_path,
        "groups_by_label": {k: len(v) for k, v in groups_by_label.items()},
    }


def default_out_path(paths: List[str]) -> str:
    base_dir = os.path.dirname(os.path.abspath(paths[-1])) or WORK_OUTPUT
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(base_dir, f"追觅应收进度对比报告_{ts}.xlsx")


def main(argv=None) -> int:
    load_config()
    ap = argparse.ArgumentParser(
        description="追觅应收进度多版本对比（期间并集 / 人名对齐 / 值+底色 / 公式忽略）"
    )
    ap.add_argument("--inspect", action="store_true", help="只识别目录内文件，不跑对比")
    ap.add_argument("--input-dir", default=None, help="输入目录（默认 skill 工作区/input）")
    ap.add_argument("--files", nargs="+", help="按时间从旧到新的 Excel 路径（≥2）")
    ap.add_argument("--out", default=None, help="输出报告路径；默认落在最新源文件同目录")
    ap.add_argument("--sheet", default=None, help="覆盖目标 sheet 名")
    args = ap.parse_args(argv)

    if args.sheet:
        CONFIG["SHEET_NAME"] = args.sheet

    input_dir = args.input_dir or WORK_INPUT

    if args.inspect:
        results = inspect_dir(input_dir)
        print(json.dumps({
            "input_dir": os.path.abspath(input_dir),
            "count": len(results),
            "files": results,
        }, ensure_ascii=False, indent=2))
        ok = [r for r in results if r["ok"]]
        log(f"识别到 {len(results)} 个 xlsx，其中 {len(ok)} 个可用作进度表。")
        return 0 if results else 1

    paths: List[str] = []
    if args.files:
        paths = [os.path.abspath(p) for p in args.files]
    else:
        paths = [os.path.abspath(p) for p in list_xlsx(input_dir)]
        # 按文件名排序（通常含日期）
        paths.sort(key=lambda p: (extract_label(p, 0), os.path.basename(p)))

    if len(paths) < 2:
        log(f"需要 ≥2 个源 Excel，当前 {len(paths)}。目录={input_dir}")
        log("用法: --files a.xlsx b.xlsx [c.xlsx]  或  --input-dir 放文件的目录")
        return 2

    out_path = os.path.abspath(args.out) if args.out else default_out_path(paths)
    try:
        run_compare(paths, out_path)
    except SystemExit as e:
        log(str(e))
        return 2
    except Exception as e:
        log(f"✗ 对比失败: {e}")
        raise
    print(out_path)  # stdout 只打路径，方便 agent 抓
    return 0


if __name__ == "__main__":
    sys.exit(main())
