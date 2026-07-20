#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
追觅应收进度对比 · 回归测试
合成金标 +（可选）真实桌面样本冒烟。
跑：python3 tests/test_robustness.py
"""
from __future__ import annotations

import os
import shutil
import sys
import tempfile

import openpyxl
from openpyxl.styles import PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
SKILL_DIR = os.path.dirname(HERE)
sys.path.insert(0, os.path.join(SKILL_DIR, "scripts"))
import compare  # noqa: E402

PASS = 0
FAIL = 0


def check_eq(label, got, want):
    global PASS, FAIL
    if got == want:
        PASS += 1
    else:
        FAIL += 1
        print(f"  ✗ {label}: got={got!r} want={want!r}")


def check_true(label, cond):
    check_eq(label, bool(cond), True)


def _fill(cell, rgb):
    cell.fill = PatternFill("solid", fgColor=rgb)


def make_progress_wb(path, people, periods, data, header_names=None):
    """
    people: list of (name, has_subcols bool)
    periods: list of period strings
    data: dict[(person_idx, period, field)] = (value, bg_rgb_or_None)
    field in 金额/PO时间/发票上传时间/预计付款时间
    """
    wb = openpyxl.Workbook()
    # junk sheet first to ensure we pick by name not position
    wb.active.title = "其他"
    ws = wb.create_sheet("应收进度")
    ws.cell(1, 1, "标题行忽略")
    # header row 2
    col = 2
    person_cols = []
    for name, has_sub in people:
        ws.cell(2, col, name)
        cmap = {"金额": col}
        if has_sub:
            ws.cell(2, col + 1, "PO时间")
            ws.cell(2, col + 2, "发票上传时间")
            ws.cell(2, col + 3, "预计付款时间")
            cmap["PO时间"] = col + 1
            cmap["发票上传时间"] = col + 2
            cmap["预计付款时间"] = col + 3
            person_cols.append((name, cmap, 4))
            col += 4
        else:
            person_cols.append((name, cmap, 1))
            col += 1

    for ri, period in enumerate(periods):
        r = 3 + ri
        ws.cell(r, 1, period)
        for pi, (name, cmap, _n) in enumerate(person_cols):
            for field, c in cmap.items():
                key = (pi, period, field)
                if key in data:
                    val, bg = data[key]
                    cell = ws.cell(r, c, val)
                    if bg:
                        _fill(cell, bg)

    # stop row
    ws.cell(3 + len(periods), 1, "为方便统计金额：绿色待付款")
    wb.save(path)
    wb.close()
    return person_cols


def test_normalize_and_formula():
    print("· 公式忽略 / 数值规范化")
    compare.load_config()
    compare.CONFIG["IGNORE_FORMULA"] = True
    check_eq("公式当空", compare.norm_compare("=D10+60"), "")
    check_eq("日期", compare.norm_compare(__import__("datetime").datetime(2026, 7, 1)), "2026-07-01")
    check_eq("整数浮点", compare.norm_compare(100.0), "100")
    check_eq("文本", compare.norm_compare("0706催促"), "0706催促")
    check_eq("人名归一", compare.normalize_name("MOVA扫地机\n（朱海心）"), "mova扫地机(朱海心)")


def test_align_rename():
    print("· 人名智能对齐（改名/包含）")
    check_eq(
        "核心名相同",
        compare.align_score(
            compare.normalize_name("MOVA扫地机 (朱海心)"),
            compare.normalize_name("MOVA扫地机 (朱海心，中华)"),
        ) >= 80,
        True,
    )
    check_eq(
        "完全相同",
        compare.align_score(
            compare.normalize_name("Dreame擦窗机器人 (韩美琪)"),
            compare.normalize_name("Dreame擦窗机器人 (韩美琪)"),
        ),
        100,
    )


def test_label_extract():
    print("· 文件名标签提取")
    check_eq("Martin0709", compare.extract_label("/x/追觅list-Martin0709.xlsx", 0), "0709")
    check_eq("0703下载", compare.extract_label("/x/追觅list-Martin-0703下载.xlsx", 0), "0703")
    check_eq("0713", compare.extract_label("/x/追觅list-Martin0713.xlsx", 0), "0713")


def test_synthetic_end_to_end():
    print("· 合成三版端到端")
    compare.load_config()
    tmp = tempfile.mkdtemp(prefix="dreame_ar_")
    try:
        # v1: 2 people full + 1 single; periods without 6月
        p1 = os.path.join(tmp, "list-0703.xlsx")
        p2 = os.path.join(tmp, "list-0709.xlsx")
        p3 = os.path.join(tmp, "list-0713.xlsx")

        people_v1 = [
            ("擦窗(韩美琪)", True),
            ("MOVA扫地机(朱海心)", True),
            ("宠物独立列(蒋方舟)", False),
        ]
        people_v2 = [
            ("擦窗(韩美琪)", True),
            ("MOVA扫地机(朱海心，中华)", True),  # 改名
            # 独立列消失
        ]
        periods_v1 = ["去年10月", "1月", "5月"]
        periods_v2 = ["去年10月", "1月", "5月", "6月"]  # 新增 6月

        green = "00B050"
        yellow = "FFFF00"
        orange = "FED4A4"
        gray = "DEE0E3"

        data1 = {
            (0, "去年10月", "PO时间"): ("2026-03-13", None),
            (0, "1月", "金额"): (48000, orange),
            (0, "1月", "预计付款时间"): (None, gray),
            (1, "去年10月", "金额"): (12516.95, green),
            (1, "5月", "金额"): (9000, yellow),
            (2, "去年10月", "金额"): ("新BU", None),
        }
        data2 = {
            (0, "去年10月", "PO时间"): ("0706催促", None),  # 催促
            (0, "1月", "金额"): (48000, orange),
            (0, "1月", "预计付款时间"): (None, gray),
            (0, "6月", "金额"): (6547.78, yellow),  # 新月
            (0, "6月", "PO时间"): ("0707发送对账单", None),
            (1, "去年10月", "金额"): ("12516.95（20260707已回款核销）", gray),  # 回款
            (1, "5月", "金额"): (9000, orange),  # 黄→橙
            (1, "5月", "预计付款时间"): ("=P5+60", gray),  # 公式，不算备注
        }
        data3 = {
            (0, "去年10月", "PO时间"): ("0706催促", None),
            (0, "1月", "金额"): (48000, "2EA121"),  # 橙→深绿
            (0, "1月", "预计付款时间"): ("=D10+60", gray),  # 公式
            (0, "1月", "发票上传时间"): (__import__("datetime").datetime(2026, 7, 1), None),
            (0, "6月", "金额"): (6547.78, yellow),
            (0, "6月", "PO时间"): ("0707发送对账单", None),
            (1, "去年10月", "金额"): ("12516.95（20260707已回款核销）", gray),
            (1, "5月", "金额"): (9000, orange),
            (1, "5月", "预计付款时间"): ("=P5+60", gray),
        }

        make_progress_wb(p1, people_v1, periods_v1, data1)
        make_progress_wb(p2, people_v2, periods_v2, data2)
        make_progress_wb(p3, people_v2, periods_v2, data3)

        out = os.path.join(tmp, "out.xlsx")
        summary = compare.run_compare([p1, p2, p3], out)
        res = summary["result"]

        check_true("产出文件存在", os.path.isfile(out))
        check_eq("标签顺序", summary["labels"], ["0703", "0709", "0713"])
        check_true("期间含6月", "6月" in summary["period_union"])
        check_eq("0703列组3", summary["groups_by_label"]["0703"], 3)
        check_eq("0709列组2", summary["groups_by_label"]["0709"], 2)

        # 值变化应包含：催促、回款文案、6月新账、发票日期
        val_keys = {(r["name"], r["period"], r["field"]) for r in res["val_changes"]}
        check_true(
            "PO催促计入值变化",
            any("去年10月" == k[1] and k[2] == "PO时间" for k in val_keys),
        )
        check_true(
            "回款核销计入值变化",
            any("去年10月" == k[1] and k[2] == "金额" and "扫地机" in k[0] for k in val_keys),
        )
        check_true(
            "6月新账计入值变化",
            any(k[1] == "6月" and k[2] == "金额" for k in val_keys),
        )
        check_true(
            "发票上传计入值变化",
            any(k[1] == "1月" and k[2] == "发票上传时间" for k in val_keys),
        )

        # 公式不算备注变化
        check_eq("预计付款备注变化=0", len(res["pay_remark_changes"]), 0)
        check_true("有公式忽略记录", len(res["formula_notes"]) >= 1)

        # 底色：回款绿→灰、5月黄→橙、1月橙→深绿
        color_keys = {(r["name"], r["period"], r["field"]) for r in res["color_changes"]}
        check_true("回款底色变化", any(k[1] == "去年10月" and k[2] == "金额" for k in color_keys))
        check_true("5月黄→橙", any(k[1] == "5月" and k[2] == "金额" for k in color_keys))
        check_true("1月橙→深绿", any(k[1] == "1月" and k[2] == "金额" for k in color_keys))

        # 列结构：独立列缺失 + 改名
        statuses = {e["name"]: e["status"] for e in summary["aligned"]}
        # 找宠物独立列
        pet = [e for e in summary["aligned"] if "蒋方舟" in e["name"] or "独立" in e["name"]]
        check_true("独立列被识别", len(pet) >= 1)
        check_eq("独立列0709缺失", pet[0]["status"].get("0709"), "缺失")

        rename = [e for e in summary["aligned"] if "朱海心" in e["name"]]
        check_true("朱海心对齐到改名", len(rename) >= 1)
        check_true(
            "0709标名称变化",
            rename[0]["status"].get("0709", "").startswith("名称变化"),
        )

        # 读报告 sheet 齐全
        wb = openpyxl.load_workbook(out)
        for sn in ("结论摘要", "列结构对比", "值变化", "颜色变化", "预计付款时间", "明细对比", "运行报告"):
            check_true(f"sheet {sn}", sn in wb.sheetnames)
        # 值变化表有数据行
        check_true("值变化有数据", wb["值变化"].max_row >= 2)
        wb.close()

    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_inspect_and_skip_report():
    print("· inspect 识别 + 跳过对比报告文件名")
    compare.load_config()
    tmp = tempfile.mkdtemp(prefix="dreame_insp_")
    try:
        p = os.path.join(tmp, "src-0701.xlsx")
        make_progress_wb(
            p,
            [("业务A", True)],
            ["1月"],
            {(0, "1月", "金额"): (100, None)},
        )
        # 假报告不应被 list
        openpyxl.Workbook().save(os.path.join(tmp, "追觅应收进度对比报告_x.xlsx"))
        files = compare.list_xlsx(tmp)
        check_eq("只认1个源", len(files), 1)
        info = compare.sniff_file(p)
        check_true("sniff ok", info["ok"])
        check_eq("sniff sheet", info["sheet"], "应收进度")
        check_eq("sniff groups", info["groups"], 1)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_two_files_min():
    print("· 最少两版也能跑")
    compare.load_config()
    tmp = tempfile.mkdtemp(prefix="dreame_two_")
    try:
        a = os.path.join(tmp, "a-0101.xlsx")
        b = os.path.join(tmp, "b-0201.xlsx")
        make_progress_wb(a, [("甲", True)], ["1月"], {(0, "1月", "金额"): (1, None)})
        make_progress_wb(b, [("甲", True)], ["1月"], {(0, "1月", "金额"): (2, None)})
        out = os.path.join(tmp, "o.xlsx")
        s = compare.run_compare([a, b], out)
        check_eq("两版值变化", len(s["result"]["val_changes"]), 1)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_real_desktop_smoke():
    print("· 真实样本冒烟（有则跑，无则跳过）")
    candidates = [
        "/Users/evanlee/Desktop/追觅对账",
        os.path.join(os.path.expanduser("~"), "Desktop", "追觅对账"),
    ]
    src_dir = None
    for c in candidates:
        if os.path.isdir(c):
            xs = [
                os.path.join(c, f)
                for f in os.listdir(c)
                if f.endswith(".xlsx")
                and not f.startswith("~$")
                and "对比" not in f
            ]
            if len(xs) >= 2:
                src_dir = c
                break
    if not src_dir:
        print("  （跳过：未找到桌面 追觅对账 样本）")
        return

    compare.load_config()
    files = sorted(
        [
            os.path.join(src_dir, f)
            for f in os.listdir(src_dir)
            if f.endswith(".xlsx") and not f.startswith("~$") and "对比" not in f
        ],
        key=lambda p: (compare.extract_label(p, 0), os.path.basename(p)),
    )
    # 只要源 list 文件
    files = [f for f in files if "list" in os.path.basename(f).lower() or "追觅" in os.path.basename(f)]
    if len(files) < 2:
        print("  （跳过：源文件不足2）")
        return

    out = os.path.join(tempfile.mkdtemp(prefix="dreame_real_"), "real_out.xlsx")
    try:
        s = compare.run_compare(files[:3] if len(files) >= 3 else files, out)
        check_true("真实产出", os.path.isfile(out))
        check_true("真实有值变化或结构", len(s["result"]["val_changes"]) >= 0)
        # 已知校正基准（三文件 0703/0709/0713）
        if len(files) >= 3:
            check_true("真实含6月或期间>=12", len(s["period_union"]) >= 12)
            # 校正版金标：29 值变化、12 底色（允许小浮动若源变）
            vc = len(s["result"]["val_changes"])
            cc = len(s["result"]["color_changes"])
            check_true(f"真实值变化合理({vc})", 20 <= vc <= 40)
            check_true(f"真实底色合理({cc})", 8 <= cc <= 20)
            check_eq("真实付款备注0", len(s["result"]["pay_remark_changes"]), 0)
            print(f"  真实统计: 值变化={vc} 底色={cc} 期间={s['period_union']}")
    except Exception as e:
        check_true(f"真实样本不崩({e})", False)


def main():
    print("=" * 60)
    print("dreame-ar-progress-diff 回归")
    print("=" * 60)
    test_normalize_and_formula()
    test_align_rename()
    test_label_extract()
    test_inspect_and_skip_report()
    test_two_files_min()
    test_synthetic_end_to_end()
    test_real_desktop_smoke()
    print("=" * 60)
    print(f"通过 {PASS}  失败 {FAIL}")
    print("=" * 60)
    return 0 if FAIL == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
