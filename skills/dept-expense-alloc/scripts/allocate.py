#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
部门费用归集分摊 — 主脚本
输入当月材料文件夹 → 输出 部门科目余额表 + 利润表 + 运行报告
规则见 config/业务规则.md；金额只由本脚本计算，禁止改数凑平。
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

HERE = Path(__file__).resolve().parent
SKILL = HERE.parent
CONFIG = SKILL / "config"

try:
    import openpyxl
except ImportError:
    print("需要 openpyxl: pip install openpyxl", file=sys.stderr)
    raise SystemExit(1)


def log(msg: str):
    print(msg, file=sys.stderr)


def load_json(name: str, default=None):
    p = CONFIG / name
    if not p.is_file():
        return default if default is not None else {}
    return json.loads(p.read_text(encoding="utf-8"))


def clean_code(v) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, float):
        if abs(v - int(v)) < 1e-9:
            return str(int(v))
        return str(v)
    s = str(v).strip()
    if re.match(r"^\d+\.0+$", s):
        return s.split(".")[0]
    return s


def num(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("，", "").strip()
    if not s or s == "-":
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def load_aliases():
    return load_json("列名别名.json", {})


def load_dept_map():
    d = load_json("部门名映射.json", {})
    return {k: v for k, v in d.items() if not str(k).startswith("_")}


def load_structure():
    return load_json("表结构模板.json", {})


# 整挂：长前缀优先。仅 540101/540103 整挂项目总监；5402/5504 财务中心
DIRECT_HANG = {
    "540101": "项目总监及助理",
    "540103": "项目总监及助理",
    "5402": "财务中心",
    "5504": "财务中心",
}


def hang_dept(code: str) -> str | None:
    if not code:
        return None
    for pref, dept in sorted(DIRECT_HANG.items(), key=lambda x: -len(x[0])):
        if code == pref or code.startswith(pref):
            # 5402 不要误伤 5401*
            if pref == "5402" and not code.startswith("5402"):
                continue
            if pref.startswith("5401") and not (code == pref or code.startswith(pref)):
                continue
            return dept
    return None


def is_person_alloc_code(code: str) -> bool:
    """需要按人拆的科目（成本除 540101/103 外；销售/管理/研发）"""
    if not code:
        return False
    if hang_dept(code):
        return False
    if code.startswith("5401"):
        return True
    if code.startswith("5501") or code.startswith("5502") or code.startswith("5503"):
        return True
    return False


def list_data_files(root: Path):
    out = []
    if not root.is_dir():
        return out
    for p in root.rglob("*"):
        if not p.is_file() or p.name.startswith("~$") or p.name.startswith("."):
            continue
        if p.suffix.lower() in {".xls", ".xlsx", ".xlsm", ".csv"}:
            out.append(p)
    return sorted(out)


def guess_entity_name(path: Path) -> str | None:
    s = path.name + " " + str(path.parent)
    pairs = [
        ("甲骨易", "甲骨易"),
        ("北京", "甲骨易"),
        ("文化", "文化"),
        ("上海", "上海"),
        ("山东", "山东分公司"),
        ("湖南", "湖南分公司"),
        ("济南", "济南分公司"),
    ]
    for kw, name in pairs:
        if kw in s:
            return name
    return None


def read_balance_any(path: Path) -> dict[str, dict]:
    """U8 发生额及余额表 → {code: {name, debit, credit}} 本期发生."""
    if path.suffix.lower() == ".xls":
        import xlrd

        wb = xlrd.open_workbook(str(path))
        sh = wb.sheet_by_index(0)
        start = 0
        for r in range(min(12, sh.nrows)):
            row = [str(sh.cell_value(r, c)) for c in range(min(8, sh.ncols))]
            if any("科目编码" in x for x in row):
                start = r + 1
                if start < sh.nrows and any(
                    "借" in str(sh.cell_value(start, c)) for c in range(min(8, sh.ncols))
                ):
                    start += 1
                break
        out = {}
        for r in range(start, sh.nrows):
            code = clean_code(sh.cell_value(r, 0))
            if not code or not re.match(r"^\d+", code):
                continue
            name = str(sh.cell_value(r, 1)).strip()
            debit = num(sh.cell_value(r, 4)) if sh.ncols > 4 else 0.0
            credit = num(sh.cell_value(r, 5)) if sh.ncols > 5 else 0.0
            out[code] = {"name": name, "debit": debit, "credit": credit}
        return out

    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    sh = wb[wb.sheetnames[0]]
    rows = list(sh.iter_rows(values_only=True))
    wb.close()
    start = 0
    for i, row in enumerate(rows[:12]):
        if row and any(x and "科目编码" in str(x) for x in row):
            start = i + 1
            if start < len(rows) and any(x and "借" in str(x) for x in (rows[start] or [])):
                start += 1
            break
    out = {}
    for row in rows[start:]:
        if not row:
            continue
        code = clean_code(row[0] if len(row) > 0 else "")
        if not code or not re.match(r"^\d+", code):
            continue
        name = str(row[1] or "").strip() if len(row) > 1 else ""
        debit = num(row[4]) if len(row) > 4 else 0.0
        credit = num(row[5]) if len(row) > 5 else 0.0
        out[code] = {"name": name, "debit": debit, "credit": credit}
    return out


def _norm_header(h) -> str:
    return re.sub(r"\s+", "", str(h or "").strip())


def _find_col(headers: list, aliases: list[str]) -> int | None:
    norms = [_norm_header(h) for h in headers]
    for a in aliases:
        an = _norm_header(a)
        for i, h in enumerate(norms):
            if an and (an == h or an in h or h in an):
                return i
    return None


def read_sheet_rows(path: Path) -> list[tuple[list, list]]:
    """Return list of (headers, data_rows) per sheet/table."""
    results = []
    if path.suffix.lower() == ".csv":
        with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
            rows = list(csv.reader(f))
        if not rows:
            return []
        return [(rows[0], rows[1:])]
    if path.suffix.lower() == ".xls":
        import xlrd

        wb = xlrd.open_workbook(str(path))
        for i in range(wb.nsheets):
            sh = wb.sheet_by_index(i)
            if sh.nrows < 2:
                continue
            # find header row
            header_r = 0
            for r in range(min(8, sh.nrows)):
                vals = [str(sh.cell_value(r, c)) for c in range(sh.ncols)]
                if any(k in "".join(vals) for k in ("姓名", "申请人", "科目", "金额", "员工")):
                    header_r = r
                    break
            headers = [sh.cell_value(header_r, c) for c in range(sh.ncols)]
            data = []
            for r in range(header_r + 1, sh.nrows):
                data.append([sh.cell_value(r, c) for c in range(sh.ncols)])
            results.append((headers, data))
        return results

    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    for name in wb.sheetnames:
        sh = wb[name]
        rows = list(sh.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        header_r = 0
        for r in range(min(8, len(rows))):
            vals = [str(x or "") for x in (rows[r] or [])]
            if any(k in "".join(vals) for k in ("姓名", "申请人", "科目", "金额", "员工", "组织")):
                header_r = r
                break
        headers = list(rows[header_r] or [])
        data = [list(r or []) for r in rows[header_r + 1 :]]
        results.append((headers, data))
    wb.close()
    return results


def load_people_map(input_dir: Path, report: list) -> dict[str, str]:
    """姓名 → 部门列名"""
    aliases = load_aliases().get("人员归属表", {})
    name_as = aliases.get("姓名", ["姓名", "员工", "人员"])
    o1_as = aliases.get("组织架构1", ["组织架构1", "中心", "一级部门"])
    o2_as = aliases.get("组织架构2", ["组织架构2", "子部门", "二级部门", "组别"])
    dmap = load_dept_map()
    people = {}
    for f in list_data_files(input_dir):
        # prefer folders/names with 归属/人员
        tag = str(f).lower() + f.name
        if not any(k in tag for k in ("归属", "人员", "组织", "staff", "people")):
            # still try all files that look like 3-col maps
            pass
        try:
            for headers, data in read_sheet_rows(f):
                ci_name = _find_col(headers, name_as)
                ci_o2 = _find_col(headers, o2_as)
                ci_o1 = _find_col(headers, o1_as)
                if ci_name is None:
                    continue
                n = 0
                for row in data:
                    if ci_name >= len(row):
                        continue
                    nm = str(row[ci_name] or "").strip()
                    if not nm or nm in ("合计", "小计"):
                        continue
                    dept_raw = ""
                    if ci_o2 is not None and ci_o2 < len(row) and row[ci_o2]:
                        dept_raw = str(row[ci_o2]).strip()
                    elif ci_o1 is not None and ci_o1 < len(row) and row[ci_o1]:
                        dept_raw = str(row[ci_o1]).strip()
                    if not dept_raw:
                        continue
                    dept = dmap.get(dept_raw, dept_raw)
                    people[nm] = dept
                    n += 1
                if n:
                    report.append(f"人员归属: {f.name} 载入 {n} 人")
        except Exception as e:
            report.append(f"⚠ 读人员表失败 {f.name}: {e}")
    if not people:
        report.append("⚠ 未识别到人员归属表（姓名→部门）。按人拆科目将无法填充。")
    return people


def load_income_by_dept(input_dir: Path, people: dict, report: list) -> dict[str, float]:
    """部门 → 收入金额（贷方性质）"""
    aliases = load_aliases().get("开票明细", {})
    app_as = aliases.get("申请人", ["申请人", "业务员", "销售", "营销人员", "开票人", "姓名"])
    amt_as = aliases.get("金额", ["金额", "价税合计", "合计金额", "开票金额", "含税金额", "收入"])
    by_dept = defaultdict(float)
    unmatched = 0
    used = 0
    for f in list_data_files(input_dir):
        tag = f.name + str(f.parent)
        if not any(k in tag for k in ("收入", "开票", "发票", "微信", "支付宝", "外汇", "03_")):
            # still try if columns match
            pass
        try:
            for headers, data in read_sheet_rows(f):
                ci_app = _find_col(headers, app_as)
                ci_amt = _find_col(headers, amt_as)
                if ci_app is None or ci_amt is None:
                    continue
                for row in data:
                    if max(ci_app, ci_amt) >= len(row):
                        continue
                    person = str(row[ci_app] or "").strip()
                    amount = num(row[ci_amt])
                    if not person or amount == 0:
                        continue
                    dept = people.get(person)
                    if not dept:
                        unmatched += 1
                        continue
                    by_dept[dept] += amount
                    used += 1
                if used:
                    report.append(f"收入底稿: {f.name} 已汇总（累计有效行见运行报告统计）")
        except Exception as e:
            report.append(f"⚠ 读收入失败 {f.name}: {e}")
    if used:
        report.append(f"收入按人汇总: 有效 {used} 行，未匹配人员 {unmatched} 行")
    elif unmatched:
        report.append(f"⚠ 收入有 {unmatched} 行但人员归属未匹配")
    return dict(by_dept)


def load_person_expenses(
    input_dir: Path, people: dict, report: list
) -> dict[str, dict[str, float]]:
    """code → {dept: amount} 按人拆费用"""
    aliases = load_aliases().get("按人费用明细", {})
    name_as = aliases.get("姓名", ["姓名", "员工", "人员", "职员", "报销人", "申请人"])
    code_as = aliases.get("科目编码", ["科目编码", "编码", "科目代码"])
    cname_as = aliases.get("科目名称", ["科目名称", "费用类型", "费用项目", "科目"])
    amt_as = aliases.get("金额", ["金额", "借方", "本期发生", "发生额", "报销金额"])
    # 科目名模糊 → 编码前缀
    name_to_prefix = {
        "销售": "5501",
        "管理": "5502",
        "研发": "5503",
        "工资": "540109",
        "社保": "540111",
        "公积金": "540112",
        "差旅": "540116",
        "招待": "540117",
        "房租": "540124",
    }
    result = defaultdict(lambda: defaultdict(float))
    used = unmatched_p = unmatched_c = 0
    for f in list_data_files(input_dir):
        tag = f.name + str(f.parent)
        # skip pure balance sheets if obvious
        if "发生额" in f.name and "余额" in f.name:
            continue
        try:
            for headers, data in read_sheet_rows(f):
                ci_name = _find_col(headers, name_as)
                ci_amt = _find_col(headers, amt_as)
                ci_code = _find_col(headers, code_as)
                ci_cname = _find_col(headers, cname_as)
                if ci_name is None or ci_amt is None:
                    continue
                for row in data:
                    if max(ci_name, ci_amt) >= len(row):
                        continue
                    person = str(row[ci_name] or "").strip()
                    amount = abs(num(row[ci_amt]))
                    if not person or amount == 0 or person in ("合计", "小计"):
                        continue
                    dept = people.get(person)
                    if not dept:
                        unmatched_p += 1
                        continue
                    code = ""
                    if ci_code is not None and ci_code < len(row):
                        code = clean_code(row[ci_code])
                    if not code and ci_cname is not None and ci_cname < len(row):
                        cname = str(row[ci_cname] or "")
                        for k, pref in name_to_prefix.items():
                            if k in cname:
                                code = pref
                                break
                    if not code:
                        unmatched_c += 1
                        continue
                    if not is_person_alloc_code(code) and not hang_dept(code):
                        # still allow if looks like expense
                        if not (code.startswith("54") or code.startswith("55")):
                            continue
                    if hang_dept(code):
                        # 整挂科目不按人
                        continue
                    result[code][dept] += amount
                    used += 1
        except Exception as e:
            report.append(f"⚠ 读按人费用失败 {f.name}: {e}")
    report.append(f"按人费用: 有效 {used} 行，人未匹配 {unmatched_p}，缺科目 {unmatched_c}")
    return {c: dict(d) for c, d in result.items()}


def discover_entity_balances(input_dir: Path, report: list) -> dict[str, dict[str, dict]]:
    entity_data = {}
    for f in list_data_files(input_dir):
        ent = guess_entity_name(f)
        # also: balance-like
        try:
            peek = None
            if f.suffix.lower() == ".xls":
                import xlrd

                wb = xlrd.open_workbook(str(f))
                sh = wb.sheet_by_index(0)
                head = " ".join(str(sh.cell_value(r, c)) for r in range(min(5, sh.nrows)) for c in range(min(4, sh.ncols)))
            else:
                wb = openpyxl.load_workbook(str(f), data_only=True, read_only=True)
                sh = wb[wb.sheetnames[0]]
                rows = list(sh.iter_rows(max_row=5, max_col=4, values_only=True))
                wb.close()
                head = " ".join(str(x or "") for row in rows for x in (row or []))
            if "科目编码" not in head and "发生额" not in head and "余额" not in head:
                continue
            data = read_balance_any(f)
            if not data:
                continue
            if not ent:
                # default first free entity slots
                ent = f.stem[:20]
            if ent in entity_data:
                # merge
                for code, v in data.items():
                    if code not in entity_data[ent]:
                        entity_data[ent][code] = v
                    else:
                        entity_data[ent][code]["debit"] += v["debit"]
                        entity_data[ent][code]["credit"] += v["credit"]
            else:
                entity_data[ent] = data
            report.append(f"主体余额: {ent} ← {f.name}（{len(data)} 科目）")
        except Exception as e:
            report.append(f"⚠ 读余额失败 {f.name}: {e}")
    return entity_data


def build_workbook(
    structure: dict,
    entity_data: dict,
    income_by_dept: dict,
    person_exp: dict,
    report: list,
):
    entities = [e["name"] for e in structure.get("entities", [])]
    # include any discovered entities not in template
    for en in entity_data:
        if en not in entities:
            entities.append(en)
    depts = [d["name"] for d in structure.get("departments", [])]
    accounts = structure.get("account_rows", [])
    tol = float(structure.get("tolerance", 0.02))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "部门科目余额表"

    ws.cell(1, 1, entities[0] if entities else "")
    ws.cell(2, 1, "科目编码")
    ws.cell(2, 2, "科目名称")
    col = 3
    ent_cols = {}
    for en in entities:
        ws.cell(1, col, en)
        ws.cell(2, col, "本期发生借方")
        ws.cell(2, col + 1, "本期发生贷方")
        ent_cols[en] = (col, col + 1)
        col += 2
    total_d, total_c = col, col + 1
    ws.cell(1, total_d, "合计")
    ws.cell(2, total_d, "本期发生借方")
    ws.cell(2, total_c, "本期发生贷方")
    col += 2
    dept_cols = {}
    for d in structure.get("departments", []):
        dname = d["name"]
        parent = d.get("parent")
        if parent:
            ws.cell(1, col, parent)
        ws.cell(2, col, dname)
        dept_cols[dname] = col
        col += 1
    # extra depts from data
    for dname in list(income_by_dept.keys()) + [x for m in person_exp.values() for x in m]:
        if dname not in dept_cols:
            ws.cell(2, col, dname)
            dept_cols[dname] = col
            col += 1
    fee_col, check_col = col, col + 1
    ws.cell(2, fee_col, "费用合计")
    ws.cell(2, check_col, "核对")

    # index person_exp by code + also roll up to parent codes when needed
    def dept_alloc_for_code(code: str) -> dict[str, float]:
        out = defaultdict(float)
        if code in person_exp:
            for d, a in person_exp[code].items():
                out[d] += a
        # children roll into parent display? only exact code
        hang = hang_dept(code)
        if hang:
            # amount from entity totals later
            pass
        return out

    unflat_checks = []

    for ri, acc in enumerate(accounts):
        r = ri + 3
        code = acc.get("code") or ""
        name = acc.get("name") or ""
        ws.cell(r, 1, code)
        ws.cell(r, 2, name)

        sum_d = sum_c = 0.0
        for en, (dc, cc) in ent_cols.items():
            ed = entity_data.get(en, {}).get(code, {})
            d = float(ed.get("debit", 0) or 0)
            c = float(ed.get("credit", 0) or 0)
            if d:
                ws.cell(r, dc, round(d, 2))
            if c:
                ws.cell(r, cc, round(c, 2))
            sum_d += d
            sum_c += c
        if sum_d:
            ws.cell(r, total_d, round(sum_d, 2))
        if sum_c:
            ws.cell(r, total_c, round(sum_c, 2))

        dept_amt = defaultdict(float)

        # 收入 5101
        if code == "5101" or (code.startswith("5101") and code != "5101"):
            if code == "5101":
                for dname, a in income_by_dept.items():
                    if dname in dept_cols:
                        dept_amt[dname] += a
            # detail income lines left empty unless we have breakdown

        # 整挂
        hang = hang_dept(code)
        if hang and hang in dept_cols:
            net = sum_d if sum_d else sum_c
            if net:
                dept_amt[hang] += net

        # 按人
        if is_person_alloc_code(code):
            for dname, a in person_exp.get(code, {}).items():
                if dname in dept_cols:
                    dept_amt[dname] += a
            # 若明细无、但大类有按人汇总：尝试前缀归集到本 code 当本 code 是 4 位大类
            if len(code) == 4 and not dept_amt:
                for pc, dm in person_exp.items():
                    if pc.startswith(code):
                        for dname, a in dm.items():
                            if dname in dept_cols:
                                dept_amt[dname] += a

        fee = sum(dept_amt.values())
        for dname, a in dept_amt.items():
            if a and dname in dept_cols:
                ws.cell(r, dept_cols[dname], round(a, 2))

        if sum_c and not sum_d:
            total_side = sum_c
        else:
            total_side = sum_d
        # 有拆分意图时才严核：整挂或按人或收入
        should_check = bool(dept_amt) or hang or code == "5101"
        if should_check and (total_side or fee):
            check = round(total_side - fee, 6)
        else:
            # 无部门拆分时：费用合计空，核对=合计侧（提示尚未拆）
            check = round(total_side - fee, 6) if fee else round(total_side, 6)
            if total_side and not fee and is_person_alloc_code(code):
                report.append(f"待拆分 {code} {name}: 有发生额但无按人明细")
        if fee:
            ws.cell(r, fee_col, round(fee, 2))
        ws.cell(r, check_col, check)
        if should_check and abs(check) > tol and (total_side or fee):
            unflat_checks.append(f"{code} {name} 差额={check}")

    if unflat_checks:
        report.append(f"核对不平 {len(unflat_checks)} 行（前20）:")
        report.extend("  " + x for x in unflat_checks[:20])
    else:
        report.append("已尝试拆分的科目：核对均在容差内（或无发生额）")

    # 利润表
    ws2 = wb.create_sheet("利润表")
    ws2.cell(1, 1, "项目")
    for i, en in enumerate(entities):
        ws2.cell(1, i + 2, en)
    ws2.cell(1, len(entities) + 2, "合计")
    pl_map = [
        ("收入", "5101", "credit"),
        ("成本", "5401", "debit"),
        ("税金及附加", "5402", "debit"),
        ("销售费用", "5501", "debit"),
        ("管理费用", "5502", "debit"),
        ("研发费用", "5503", "debit"),
        ("财务费用", "5504", "debit"),
    ]
    for ri, (lab, code, side) in enumerate(pl_map):
        r = ri + 2
        ws2.cell(r, 1, lab)
        total = 0.0
        for i, en in enumerate(entities):
            ed = entity_data.get(en, {}).get(code, {})
            v = float(ed.get(side, 0) or 0)
            if v:
                ws2.cell(r, i + 2, round(v, 2))
            total += v
        if total:
            ws2.cell(r, len(entities) + 2, round(total, 2))

    return wb


def main():
    ap = argparse.ArgumentParser(description="部门费用归集分摊")
    ap.add_argument("--input-dir", required=False, default=str(SKILL / "工作区" / "input"))
    ap.add_argument("--out", default="", help="输出 xlsx")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    input_dir = Path(args.input_dir).expanduser().resolve()
    structure = load_structure()
    if not structure:
        print("缺少 config/表结构模板.json", file=sys.stderr)
        return 1

    report = []
    report.append("=== 部门费用归集分摊 ===")
    report.append(f"输入: {input_dir}")
    report.append(
        f"模板: {len(structure.get('account_rows', []))} 科目 / "
        f"{len(structure.get('departments', []))} 部门"
    )
    report.append(
        "规则: 540101/540103→项目总监及助理; 5402/5504→财务中心; "
        "其余5401*与5501/5502/5503按人; 5101收入三源按人; 左合计应=右合计"
    )

    if not input_dir.is_dir():
        report.append(f"⚠ 输入目录不存在: {input_dir}")
        out_rep = SKILL / "工作区" / "output" / "运行报告.txt"
        out_rep.parent.mkdir(parents=True, exist_ok=True)
        out_rep.write_text("\n".join(report) + "\n", encoding="utf-8")
        print("\n".join(report))
        return 2

    people = load_people_map(input_dir, report)
    entity_data = discover_entity_balances(input_dir, report)
    income_by_dept = load_income_by_dept(input_dir, people, report)
    person_exp = load_person_expenses(input_dir, people, report)

    if args.dry_run:
        out_rep = SKILL / "工作区" / "output" / "运行报告_dryrun.txt"
        out_rep.parent.mkdir(parents=True, exist_ok=True)
        out_rep.write_text("\n".join(report) + "\n", encoding="utf-8")
        print("\n".join(report))
        print(f"dry-run: {out_rep}")
        return 0

    if not entity_data:
        report.append("⚠ 未读到任何主体余额表（需含「科目编码」的发生额及余额表）。仍生成空骨架。")

    wb = build_workbook(structure, entity_data, income_by_dept, person_exp, report)
    out = Path(args.out) if args.out else SKILL / "工作区" / "output" / "部门科目余额表_产出.xlsx"
    out = out.expanduser().resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    report.append(f"已写出: {out}")
    rep_path = out.parent / "运行报告.txt"
    rep_path.write_text("\n".join(report) + "\n", encoding="utf-8")
    # 对人话摘要：不打印金额明细
    print("=== 完成 ===")
    print(f"输出: {out}")
    print(f"报告: {rep_path}")
    print(f"主体数: {len(entity_data)} 人员映射: {len(people)} 收入部门块: {len(income_by_dept)} 按人科目: {len(person_exp)}")
    print("请打开运行报告查看核对情况。金额明细见 Excel，勿外传。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
