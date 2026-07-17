#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""盘点 input 目录：按文件夹名/内容特征识别材料角色，写出运行报告（不含金额明细）。"""
from __future__ import annotations

import argparse
import json
import os
import sys
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

HERE = Path(__file__).resolve().parent
SKILL = HERE.parent
CONFIG = SKILL / "config"


def load_aliases():
    p = CONFIG / "列名别名.json"
    if p.is_file():
        return json.loads(p.read_text(encoding="utf-8"))
    return {}


def list_files(root: Path):
    out = []
    if not root.is_dir():
        return out
    for dirpath, _, files in os.walk(root):
        for f in files:
            if f.startswith("~$") or f.startswith("."):
                continue
            low = f.lower()
            if low.endswith((".xls", ".xlsx", ".xlsm", ".csv", ".txt", ".md")):
                out.append(Path(dirpath) / f)
    return sorted(out)


def guess_role(path: Path, aliases: dict) -> str:
    parts = " ".join(path.parts).lower()
    name = path.name.lower()
    folder_map = aliases.get("文件夹角色", {})
    for role, keys in folder_map.items():
        for k in keys:
            if k.lower() in parts or k.lower() in name:
                return role
    # content-ish name hints
    if "归属" in path.name or "人员" in path.name:
        return "人员归属"
    if "开票" in path.name or "发票" in path.name:
        return "收入底稿"
    if "工资" in path.name or "社保" in path.name:
        return "工资社保"
    if "部门科目" in path.name or "定稿" in path.name:
        return "定稿对照"
    if "余额" in path.name or "发生额" in path.name:
        return "主体余额"
    return "未识别"


def peek_excel(path: Path) -> dict:
    info = {"path": str(path), "sheets": []}
    try:
        if path.suffix.lower() == ".xls":
            import xlrd

            wb = xlrd.open_workbook(str(path))
            for i, n in enumerate(wb.sheet_names()):
                sh = wb.sheet_by_index(i)
                headers = []
                for r in range(min(3, sh.nrows)):
                    row = [str(sh.cell_value(r, c))[:20] for c in range(min(8, sh.ncols))]
                    headers.append(row)
                info["sheets"].append(
                    {"name": n, "rows": sh.nrows, "cols": sh.ncols, "preview": headers}
                )
        else:
            import openpyxl

            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            for n in wb.sheetnames:
                sh = wb[n]
                # rough size
                rows = sh.max_row or 0
                cols = sh.max_column or 0
                preview = []
                for i, row in enumerate(sh.iter_rows(max_row=3, max_col=min(8, cols or 8), values_only=True)):
                    preview.append([str(c)[:20] if c is not None else "" for c in row])
                info["sheets"].append(
                    {"name": n, "rows": rows, "cols": cols, "preview": preview}
                )
            wb.close()
    except Exception as e:
        info["error"] = str(e)
    return info


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input-dir", default=str(SKILL / "工作区" / "input"))
    ap.add_argument("--out", default="", help="运行报告路径，默认 input 同级 output/运行报告_盘点.txt")
    args = ap.parse_args()
    root = Path(args.input_dir).expanduser().resolve()
    aliases = load_aliases()
    files = list_files(root)

    lines = []
    lines.append(f"盘点目录: {root}")
    lines.append(f"文件数: {len(files)}")
    lines.append("")

    roles = {}
    for f in files:
        role = guess_role(f, aliases)
        roles.setdefault(role, []).append(f)
        info = peek_excel(f) if f.suffix.lower() in {".xls", ".xlsx", ".xlsm"} else {"path": str(f)}
        lines.append(f"[{role}] {f.name}")
        lines.append(f"  路径: {f}")
        if "error" in info:
            lines.append(f"  ⚠ 读取失败: {info['error']}")
        for sh in info.get("sheets", []):
            lines.append(f"  sheet「{sh['name']}」 {sh['rows']}行×{sh['cols']}列")
        lines.append("")

    required = ["主体余额", "人员归属", "收入底稿"]
    # 管理/研发按人拆需要
    recommended = ["用友按人", "工资社保", "外地代理", "定稿对照"]
    lines.append("=== 齐全度 ===")
    for r in required:
        ok = r in roles and roles[r]
        lines.append(f"  {'✓' if ok else '✗'} 关键·{r}")
    for r in recommended:
        ok = r in roles and roles[r]
        lines.append(f"  {'✓' if ok else '·'} 建议·{r}")

    missing = [r for r in required if r not in roles or not roles[r]]
    if missing:
        lines.append("")
        lines.append("⚠ 缺关键输入，技能包暂不能完整出表: " + "、".join(missing))
        lines.append("请斯佳姐按 01~07 文件夹清单补齐后重跑盘点。")
    else:
        lines.append("")
        lines.append("关键输入已齐，可进入 allocate（若按人规则科目仍缺 05/06 会单独报停）。")

    text = "\n".join(lines) + "\n"
    out = Path(args.out) if args.out else root.parent / "output" / "运行报告_盘点.txt"
    if not args.out and (root / "output").is_dir():
        out = root / "output" / "运行报告_盘点.txt"
    # default skill workspace
    if not args.out:
        out = SKILL / "工作区" / "output" / "运行报告_盘点.txt"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(text, encoding="utf-8")
    print(text)
    print(f"报告已写: {out}")
    return 0 if not missing else 2


if __name__ == "__main__":
    raise SystemExit(main())
