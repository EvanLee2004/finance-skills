#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations
import json, subprocess, sys, tempfile
from pathlib import Path
import openpyxl

HERE = Path(__file__).resolve().parent
SKILL = HERE.parent
SCRIPTS = SKILL / "scripts"
CONFIG = SKILL / "config"


def test_structure_template_complete():
    st = json.loads((CONFIG / "表结构模板.json").read_text(encoding="utf-8"))
    assert len(st["entities"]) >= 3
    assert len(st["departments"]) >= 20
    assert len(st["account_rows"]) >= 100
    codes = [a["code"] for a in st["account_rows"] if a["code"]]
    assert "5101" in codes and "5401" in codes
    assert not any(c.endswith(".0") for c in codes)


def test_business_rules_file():
    text = (CONFIG / "业务规则.md").read_text(encoding="utf-8")
    assert "540101" in text and "540103" in text
    assert "按人" in text
    assert "所有主体合计" in text or "左" in text


def _write_balance(path: Path, tag: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "发生额及余额表"
    ws["A2"] = tag
    ws["A3"] = "科目编码"
    ws["B3"] = "科目名称"
    ws["C3"] = "期初余额"
    ws["E3"] = "本期发生"
    ws["G3"] = "期末余额"
    ws["C4"] = "借方"
    ws["D4"] = "贷方"
    ws["E4"] = "借方"
    ws["F4"] = "贷方"
    ws["A5"] = "5101"
    ws["B5"] = "主营业务收入"
    ws["F5"] = 1000
    ws["A6"] = "540103"
    ws["B6"] = "翻译语言服务"
    ws["E6"] = 200
    ws["A7"] = "540109"
    ws["B7"] = "工资"
    ws["E7"] = 100
    ws["A8"] = "5402"
    ws["B8"] = "税金及附加"
    ws["E8"] = 50
    ws["A9"] = "5502"
    ws["B9"] = "管理费用"
    ws["E9"] = 80
    ws["A10"] = "5504"
    ws["B10"] = "财务费用"
    ws["E10"] = 10
    wb.save(path)


def _write_people(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["姓名", "组织架构1", "组织架构2"])
    ws.append(["张三", "营销中心", "本地化"])
    ws.append(["李四", "运营保障中心", "运营保障中心"])
    ws.append(["王五", "项目中心", "项目一组"])
    wb.save(path)


def _write_income(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["申请人", "金额"])
    ws.append(["张三", 600])
    ws.append(["张三", 400])
    wb.save(path)


def _write_person_exp(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["姓名", "科目编码", "科目名称", "金额"])
    ws.append(["王五", "540109", "工资", 100])
    ws.append(["李四", "5502", "管理费用", 80])
    wb.save(path)


def test_full_pipeline_synthetic():
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        (td / "01_主体余额表").mkdir()
        (td / "03_收入底稿").mkdir()
        (td / "04_人员归属表").mkdir()
        (td / "05_用友按人明细").mkdir()
        _write_balance(td / "01_主体余额表" / "文化_余额.xlsx", "文化")
        _write_people(td / "04_人员归属表" / "归属.xlsx")
        _write_income(td / "03_收入底稿" / "开票.xlsx")
        _write_person_exp(td / "05_用友按人明细" / "按人.xlsx")
        out = td / "out.xlsx"
        r = subprocess.run(
            [sys.executable, str(SCRIPTS / "allocate.py"), "--input-dir", str(td), "--out", str(out)],
            capture_output=True, text=True,
        )
        assert r.returncode == 0, r.stderr + r.stdout
        assert out.is_file()
        wb = openpyxl.load_workbook(out, data_only=True)
        assert "部门科目余额表" in wb.sheetnames
        assert "利润表" in wb.sheetnames
        ws = wb["部门科目余额表"]
        headers = {ws.cell(2, c).value: c for c in range(1, ws.max_column + 1)}
        assert "项目总监及助理" in headers
        assert "财务中心" in headers
        assert "本地化" in headers
        pm = headers["项目总监及助理"]
        fin = headers["财务中心"]
        loc = headers["本地化"]
        fee = headers["费用合计"]
        chk = headers["核对"]
        found = {}
        for r in range(3, ws.max_row + 1):
            code = str(ws.cell(r, 1).value)
            if code in ("540103", "5402", "5504", "5101", "540109", "5502"):
                found[code] = r
        assert "540103" in found
        r = found["540103"]
        assert float(ws.cell(r, pm).value or 0) == 200
        assert abs(float(ws.cell(r, chk).value or 0)) < 0.02
        r = found["5402"]
        assert float(ws.cell(r, fin).value or 0) == 50
        r = found["5101"]
        assert float(ws.cell(r, loc).value or 0) == 1000
        r = found["540109"]
        assert float(ws.cell(r, headers["项目一组"]).value or 0) == 100
        r = found["5502"]
        assert float(ws.cell(r, headers["运营保障中心"]).value or 0) == 80


def test_inspect_runs():
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        (td / "01_主体余额表").mkdir()
        _write_balance(td / "01_主体余额表" / "上海.xlsx", "上海")
        r = subprocess.run(
            [sys.executable, str(SCRIPTS / "inspect_inputs.py"), "--input-dir", str(td)],
            capture_output=True, text=True,
        )
        assert r.returncode in (0, 2), r.stderr + r.stdout


if __name__ == "__main__":
    test_structure_template_complete()
    test_business_rules_file()
    test_full_pipeline_synthetic()
    test_inspect_runs()
    print("OK all tests passed")
