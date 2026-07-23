"""summarize + write_report：合成明细手算金标。"""

from __future__ import annotations

import sys
from pathlib import Path

_SCRIPTS = Path(__file__).resolve().parents[1] / "scripts"
sys.path.insert(0, str(_SCRIPTS))

from summarize import (  # noqa: E402
    DEPT_DISPLAY_ORDER,
    dept_totals,
    parse_amount,
    summarize_records,
    to_display_dept,
)
from write_report import write_report  # noqa: E402


def test_parse_amount_variants():
    assert parse_amount(12345) == 12345.0
    assert parse_amount("12,345.50") == 12345.50
    assert parse_amount("(100)") == -100.0
    assert parse_amount(None) is None
    assert parse_amount("") is None


def test_local_to_multilang_display():
    assert to_display_dept("本地化") == "多语（不含运保）"
    assert to_display_dept("数据") == "数据"
    assert to_display_dept(None) == "（未匹配）"


def test_summarize_fixture_handcalc(tmp_path):
    """
    手算：
    - 张三 本地化 2026-07-22 金额 100000 → 10.00 万
    - 李四 本地化 2026-07-22 金额 50000  → 5.00 万
    - 王五 游戏   2026-07-22 金额 8700   → 0.87 万
    - 赵六 未匹配 2026-07-22 金额 10000  → 1.00 万 进（未匹配）
    总计 = 16.87；多语=15.00；游戏=0.87；未匹配=1.00
    """
    org = {
        "张三": "本地化",
        "李四": "本地化",
        "王五": "游戏",
        # 赵六 故意不进表
    }
    records = [
        {
            "销售": "张三",
            "下单日期": "2026-07-22",
            "下单预估额/本币": 100000,
        },
        {
            "销售": "李四",
            "下单日期": "2026-07-22",
            "下单预估额/本币": "50,000",
        },
        {
            "销售": "王五",
            "下单日期": "2026-07-22",
            "下单预估额/本币": 8700,
        },
        {
            "销售": "赵六",
            "下单日期": "2026-07-22",
            "下单预估额/本币": 10000,
        },
    ]
    result = summarize_records(records, org)
    assert result.detail_row_count == 4
    assert result.grand_total_wan == 16.87
    assert result.by_date["2026-07-22"]["多语（不含运保）"] == 15.0
    assert result.by_date["2026-07-22"]["游戏"] == 0.87
    assert result.by_date["2026-07-22"]["（未匹配）"] == 1.0
    assert result.by_date["2026-07-22"].get("数据", 0) == 0
    assert "赵六" in result.unmatched_sales
    assert result.amount_field_used in ("下单预估额/本币",)

    dt = dept_totals(result)
    assert dt["多语（不含运保）"] == 15.0
    assert dt["游戏"] == 0.87
    assert dt["（未匹配）"] == 1.0

    out = write_report(
        result,
        tmp_path / "out.xlsx",
        window_start="2026-07-22",
        window_end="2026-07-22",
        include_detail=True,
    )
    assert out.is_file()
    from openpyxl import load_workbook

    wb = load_workbook(out)
    assert "下单数据" in wb.sheetnames
    assert "处理日志" in wb.sheetnames
    assert "明细" in wb.sheetnames
    ws = wb["下单数据"]
    headers = [c.value for c in ws[1]]
    assert headers[0] == "日期"
    assert headers[1] == "总计"
    assert "多语（不含运保）" in headers
    assert "（未匹配）" in headers
    # row 2 values
    row2 = [c.value for c in ws[2]]
    assert row2[0] == "2026-07-22"
    assert row2[1] == 16.87
    # column order matches DEPT_DISPLAY_ORDER after 总计
    for i, name in enumerate(DEPT_DISPLAY_ORDER):
        assert headers[2 + i] == name


def test_amount_fallback_to_original_currency():
    org = {"甲": "数据"}
    records = [
        {
            "销售": "甲",
            "下单日期": "2026-01-01",
            "下单预估额/原币": 20000,  # no 本币
        }
    ]
    result = summarize_records(records, org)
    assert result.grand_total_wan == 2.0
    assert "原币" in (result.amount_field_used or "")
