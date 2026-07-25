# -*- coding: utf-8 -*-
"""覆盖台账 + 《统计区间》页：亮晶要能一眼看出"这份表算了哪几天、有没有漏"。"""
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))
import coverage  # noqa: E402


def test_first_use_has_no_gaps(tmp_path):
    """第一次用，没有"漏"的概念，别吓人。"""
    info = coverage.find_gaps(tmp_path, through=date(2026, 10, 8))
    assert info["first_use"] is True and info["gaps"] == []


def test_records_and_detects_holiday_gap(tmp_path):
    """
    核心场景：10-1 统计过，节后 10-8 才又跑（抓 10-07）。
    中间 10-2、10-5、10-6 三个工作日从没统计过 → 必须报出来。
    10-3/10-4 是周末，不报（报了全是噪音）。
    """
    coverage.record(tmp_path, [date(2026, 10, 1)], run_day=date(2026, 10, 2))
    coverage.record(tmp_path, [date(2026, 10, 7)], run_day=date(2026, 10, 8))
    gaps = coverage.find_gaps(tmp_path, through=date(2026, 10, 7))["gaps"]
    assert [d.isoformat() for d in gaps] == ["2026-10-02", "2026-10-05", "2026-10-06"]


def test_weekend_not_reported_as_gap(tmp_path):
    coverage.record(tmp_path, [date(2026, 10, 1)], run_day=date(2026, 10, 2))
    info = coverage.find_gaps(tmp_path, through=date(2026, 10, 5))
    assert date(2026, 10, 3) not in info["gaps"]      # 周六
    assert date(2026, 10, 4) not in info["gaps"]      # 周日
    assert len(info["skipped_weekend"]) == 2


def test_rerun_same_day_is_idempotent(tmp_path):
    """同一天重复跑 = 覆盖更新，不该把它算成两天。"""
    coverage.record(tmp_path, [date(2026, 10, 7)], run_day=date(2026, 10, 8))
    coverage.record(tmp_path, [date(2026, 10, 7)], run_day=date(2026, 10, 8))
    assert coverage.covered_dates(tmp_path) == {date(2026, 10, 7)}


def test_backfill_closes_the_gap(tmp_path):
    """补跑之后断档要消失——否则每天都在报同一批，她会当噪音忽略。"""
    coverage.record(tmp_path, [date(2026, 10, 1)], run_day=date(2026, 10, 2))
    coverage.record(tmp_path, [date(2026, 10, 7)], run_day=date(2026, 10, 8))
    assert len(coverage.find_gaps(tmp_path, through=date(2026, 10, 7))["gaps"]) == 3
    coverage.record(
        tmp_path,
        [date(2026, 10, 2), date(2026, 10, 5), date(2026, 10, 6)],
        run_day=date(2026, 10, 8),
    )
    assert coverage.find_gaps(tmp_path, through=date(2026, 10, 7))["gaps"] == []


def test_report_has_coverage_sheet(tmp_path):
    """产物必须有《统计区间》页，且把漏掉的日期印在里面。"""
    import openpyxl
    from summarize import SummaryResult
    from write_report import write_report

    result = SummaryResult(by_date={"2026-10-07": {"数据": 1.0}})
    out = tmp_path / "t.xlsx"
    write_report(
        result, out,
        window_start=date(2026, 10, 7), window_end=date(2026, 10, 7),
        covered_days=[date(2026, 10, 7)],
        gaps=[date(2026, 10, 2), date(2026, 10, 5)],
        run_day=date(2026, 10, 8),
    )
    wb = openpyxl.load_workbook(str(out))
    assert "统计区间" in wb.sheetnames
    text = "\n".join(
        str(c) for row in wb["统计区间"].iter_rows(values_only=True) for c in row if c
    )
    assert "2026-10-07" in text          # 算了哪天
    assert "2026-10-02" in text          # 漏了哪天
    assert "别一天一天补跑" in text        # 防重复统计的告诫
