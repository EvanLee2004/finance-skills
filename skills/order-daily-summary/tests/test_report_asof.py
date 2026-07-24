"""抓取时刻戳 + 晚跑提示：写进「下单数据」下方与「处理日志」。"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

_SCRIPTS = Path(__file__).resolve().parents[1] / "scripts"
sys.path.insert(0, str(_SCRIPTS))

from summarize import summarize_records  # noqa: E402
from write_report import write_report  # noqa: E402


def _mk_result():
    org = {"张三": "本地化"}
    records = [{"销售": "张三", "下单日期": "2026-07-23", "下单预估额/本币": 21107.13}]
    return summarize_records(records, org)


def test_asof_stamped_no_warning(tmp_path):
    res = _mk_result()
    asof = datetime(2026, 7, 24, 9, 2, 0)
    out = write_report(
        res, tmp_path / "o.xlsx", window_start="2026-07-23", window_end="2026-07-23",
        data_asof=asof, late_warning=None,
    )
    from openpyxl import load_workbook

    wb = load_workbook(out)
    data_txt = "\n".join(
        str(c.value) for row in wb["下单数据"].iter_rows() for c in row if c.value
    )
    assert "数据截至 2026-07-24 09:02:00" in data_txt
    assert "⚠" not in data_txt  # 未晚跑，无警告
    log_txt = "\n".join(str(c.value) for row in wb["处理日志"].iter_rows() for c in row if c.value)
    assert "数据截至" in log_txt
    assert "2026-07-24 09:02:00" in log_txt


def test_late_warning_rendered(tmp_path):
    res = _mk_result()
    asof = datetime(2026, 7, 24, 10, 50, 26)
    warn = "本次于 10:50 抓取，晚于每日 09:00 快照；昨日订单当天可能已被改期/改单号。"
    out = write_report(
        res, tmp_path / "o.xlsx", window_start="2026-07-23", window_end="2026-07-23",
        data_asof=asof, late_warning=warn,
    )
    from openpyxl import load_workbook

    wb = load_workbook(out)
    data_txt = "\n".join(
        str(c.value) for row in wb["下单数据"].iter_rows() for c in row if c.value
    )
    assert "⚠" in data_txt
    assert "10:50" in data_txt
    log_txt = "\n".join(str(c.value) for row in wb["处理日志"].iter_rows() for c in row if c.value)
    assert "晚跑提示" in log_txt
