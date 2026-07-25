"""写出亮晶版式 xlsx：sheet「下单数据」+「处理日志」；可选明细。"""

from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from summarize import DEPT_DISPLAY_ORDER, SummaryResult, dept_totals, row_totals


def write_report(
    result: SummaryResult,
    out_path: Path | str,
    *,
    window_start: date | str | None = None,
    window_end: date | str | None = None,
    api_row_count: int | None = None,
    include_detail: bool = False,
    data_asof: datetime | None = None,
    late_warning: str | None = None,
    extra_log: dict[str, Any] | None = None,
    covered_days: list[date] | None = None,
    gaps: list[date] | None = None,
    run_day: date | None = None,
) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    asof = data_asof or datetime.now()
    asof_text = (
        f"数据截至 {asof.strftime('%Y-%m-%d %H:%M:%S')}"
        "（智云为实时数据，昨日订单当天可能被改期，本表为该时刻快照）"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "下单数据"

    headers = ["日期", "总计", *DEPT_DISPLAY_ORDER]
    ws.append(headers)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    dates = sorted(result.by_date.keys())
    totals = row_totals(result)
    for d in dates:
        depts = result.by_date[d]
        row = [d, totals.get(d, 0.0)]
        for col in DEPT_DISPLAY_ORDER:
            row.append(depts.get(col, 0.0))
        ws.append(row)

    # number format
    for r in range(2, ws.max_row + 1):
        for c in range(2, len(headers) + 1):
            ws.cell(row=r, column=c).number_format = "#,##0.00"

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.freeze_panes = "A2"

    # 数据截至 + 晚跑提示：直接落在数据下方，亮晶打开就看得到
    ws.append([])
    asof_row = ws.max_row + 1
    ws.cell(row=asof_row, column=1, value=asof_text).font = Font(italic=True, color="808080")
    if late_warning:
        warn_row = asof_row + 1
        wc = ws.cell(row=warn_row, column=1, value=f"⚠ {late_warning}")
        wc.font = Font(bold=True, color="C00000")
        wc.alignment = Alignment(wrap_text=False, vertical="center")

    # ── 统计区间（亮晶打开第二眼就看这页：这份表到底算了哪几天）──────────
    # 2026-07-25 明昊要求：每天跑出来的产物要单独开一页说清"统计的是哪些天"，
    # 顺带把"有没有哪几天从来没统计过"（节假日漏跑）直接印出来。
    rng = wb.create_sheet("统计区间")
    title = rng.cell(row=1, column=1, value="这份表统计的是哪几天")
    title.font = Font(bold=True, size=14)
    rng.append([])
    rng.append(["运行日（哪天跑的）", (run_day or date.today()).isoformat()])
    rng.append(["统计窗口", f"{window_start or ''} ~ {window_end or ''}".strip(" ~") or "（未指定）"])

    days = list(covered_days or [])
    if not days and window_start and window_end:
        try:
            s = date.fromisoformat(str(window_start))
            e = date.fromisoformat(str(window_end))
            days = [s + timedelta(days=i) for i in range((e - s).days + 1)]
        except (TypeError, ValueError):
            days = []
    rng.append(["共统计天数", len(days)])
    rng.append([])

    rng.append(["逐日明细", "", ""])
    rng.cell(row=rng.max_row, column=1).font = Font(bold=True)
    rng.append(["下单日期", "星期", "该日合计(万元)"])
    for cell in rng[rng.max_row]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    _wk = ("周一", "周二", "周三", "周四", "周五", "周六", "周日")
    for d in days:
        key = d.isoformat()
        rng.append([key, _wk[d.weekday()], totals.get(key, 0.0)])
        rng.cell(row=rng.max_row, column=3).number_format = "#,##0.00"

    rng.append([])
    if gaps:
        _wkn = ("周一", "周二", "周三", "周四", "周五", "周六", "周日")
        warn = rng.cell(
            row=rng.max_row + 1,
            column=1,
            value=f"⚠ 注意：有 {len(gaps)} 个工作日从来没统计过（下面这些天的下单没进过任何一张表）",
        )
        warn.font = Font(bold=True, color="C00000")
        rng.append(["漏掉的日期", "星期", ""])
        for cell in rng[rng.max_row]:
            cell.font = Font(bold=True)
        for d in gaps:
            rng.append([d.isoformat(), _wkn[d.weekday()], ""])
            rng.cell(row=rng.max_row, column=1).fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
        rng.append([])
        tip = rng.cell(
            row=rng.max_row + 1,
            column=1,
            value=(
                "多半是法定假期或机器没跑——本程序只认周末、不认节假日。"
                "补法：从智云手工导出一张覆盖这几天的「下单」表，用离线模式跑一次"
                "（--from-xlsx <那张表> --no-date-filter）。"
                "⛔ 别一天一天补跑：窗口按运行日倒推、逐天跑会互相重叠，同一天被算好几遍。"
            ),
        )
        tip.font = Font(color="C00000")
    else:
        ok = rng.cell(row=rng.max_row + 1, column=1, value="✅ 从第一次统计到现在，工作日没有断档")
        ok.font = Font(bold=True, color="1F7A1F")

    rng.column_dimensions["A"].width = 46
    rng.column_dimensions["B"].width = 10
    rng.column_dimensions["C"].width = 16

    # 处理日志
    log = wb.create_sheet("处理日志")
    log.append(["项目", "内容"])
    log["A1"].font = Font(bold=True)
    log["B1"].font = Font(bold=True)
    log.append(["处理时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    log.append(["数据截至", asof.strftime("%Y-%m-%d %H:%M:%S")])
    log.append(["快照说明", "智云为实时数据，昨日订单当天可能被改期，本表为该时刻快照"])
    if late_warning:
        log.append(["⚠晚跑提示", late_warning])
    log.append(
        [
            "日期窗口",
            f"{window_start or ''}～{window_end or ''}".strip("～") or "（未指定）",
        ]
    )
    log.append(["接口行数", api_row_count if api_row_count is not None else result.detail_row_count])
    log.append(["明细行数", result.detail_row_count])
    log.append(["总计万元", result.grand_total_wan])
    log.cell(row=log.max_row, column=2).number_format = "#,##0.00"  # 不依赖行号，插行也不错位
    log.append(["金额字段", result.amount_field_used or "（未解析到）"])
    dtot = dept_totals(result)
    for name in DEPT_DISPLAY_ORDER:
        log.append([f"分部门万元·{name}", dtot.get(name, 0.0)])
        log.cell(row=log.max_row, column=2).number_format = "#,##0.00"
    unmatched = "、".join(result.unmatched_sales) if result.unmatched_sales else "无"
    log.append(["未匹配销售数量", len(result.unmatched_sales)])
    log.append(["未匹配销售名单", unmatched])
    if gaps:
        log.append(["⚠断档天数", len(gaps)])
        log.append(["⚠断档日期", "、".join(d.isoformat() for d in gaps[:20])])
        log.append([
            "怎么补",
            "从智云导一张覆盖这几天的「下单」表 → 离线模式 --from-xlsx <表> --no-date-filter；"
            "⛔ 别一天一天补跑（窗口会重叠、同一天算好几遍）",
        ])
    if extra_log:
        for k, v in extra_log.items():
            log.append([str(k), v])
    log.column_dimensions["A"].width = 22
    log.column_dimensions["B"].width = 60

    if include_detail and result.detail_rows:
        detail = wb.create_sheet("明细")
        dheaders = ["销售", "SO", "订单名称", "下单日期", "金额本币", "金额万元", "归类", "金额字段"]
        detail.append(dheaders)
        for cell in detail[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for r in result.detail_rows:
            detail.append([r.get(h, "") for h in dheaders])
        widths = {"销售": 12, "SO": 14, "订单名称": 36, "下单日期": 12, "金额本币": 12, "金额万元": 10, "归类": 16, "金额字段": 16}
        for col, h in enumerate(dheaders, start=1):
            detail.column_dimensions[get_column_letter(col)].width = widths.get(h, 14)

    wb.save(out_path)
    return out_path
