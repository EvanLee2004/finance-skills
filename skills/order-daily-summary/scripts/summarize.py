"""组织架构映射 + 万元汇总（对齐死程序 process_order 口径，主产物按亮晶版式）。

- 金额优先：下单预估额/本币 → 回退 下单预估额/原币
- 本地化 → 展示名「多语（不含运保）」
- 未匹配销售：金额进「（未匹配）」列，不进四部门合计
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

# 展示列顺序（固定）
DEPT_DISPLAY_ORDER = (
    "多语（不含运保）",
    "数据",
    "游戏",
    "其他",
    "（未匹配）",
)

# 组织架构 B 列分类 → 展示名
DEPT_DISPLAY_MAP = {
    "本地化": "多语（不含运保）",
    "多语（不含运保）": "多语（不含运保）",
    "数据": "数据",
    "游戏": "游戏",
    "其他": "其他",
}

AMOUNT_PRIMARY_KEYS = (
    "下单预估额/本币",
    "下单预估额（本币）",
    "下单预估额(本币)",
)
AMOUNT_FALLBACK_KEYS = (
    "下单预估额/原币",
    "下单预估额（原币）",
    "下单预估额(原币)",
)
SALES_KEYS = ("销售", "销售姓名")
DATE_KEYS = ("下单日期",)


@dataclass
class SummaryResult:
    """按日期×展示部门的万元表 + 元数据。"""

    # date_str (YYYY-MM-DD) -> display_dept -> amount_wan
    by_date: dict[str, dict[str, float]] = field(default_factory=dict)
    unmatched_sales: list[str] = field(default_factory=list)
    detail_row_count: int = 0
    amount_field_used: str = ""
    grand_total_wan: float = 0.0
    # 可选明细行：供 --detail
    detail_rows: list[dict[str, Any]] = field(default_factory=list)


def normalize_name(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return (
        str(value)
        .strip()
        .replace("\n", "")
        .replace("\r", "")
        .replace("（", "(")
        .replace("）", ")")
    )


def parse_amount(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    negative = text.startswith("(") and text.endswith(")")
    cleaned = (
        text.strip("()")
        .replace(",", "")
        .replace("，", "")
        .replace("￥", "")
        .replace("¥", "")
        .replace("RMB", "")
        .replace("CNY", "")
        .replace(" ", "")
    )
    try:
        number = float(cleaned)
    except ValueError:
        return None
    return -number if negative else number


def normalize_date_key(value: Any) -> str:
    """Return YYYY-MM-DD or 'N/A'."""
    if value in (None, ""):
        return "N/A"
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return "N/A"
    # common: 2026-07-22 / 2026/07/22 / 2026.07.22
    head = text[:10].replace("/", "-").replace(".", "-")
    try:
        datetime.strptime(head, "%Y-%m-%d")
        return head
    except ValueError:
        return text


def load_org_map_from_xlsx(path: Path | str, sheet_name: str | None = None) -> dict[str, str]:
    """A=销售 B=分类。与 process_order.load_org_map 语义一致。"""
    from openpyxl import load_workbook

    path = Path(path)
    wb = load_workbook(path, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    elif "组织架构" in wb.sheetnames:
        ws = wb["组织架构"]
    else:
        ws = wb.active

    org_map: dict[str, str] = {}
    for row in range(1, ws.max_row + 1):
        sales_name = normalize_name(ws.cell(row=row, column=1).value)
        department = ws.cell(row=row, column=2).value
        department_text = "" if department is None else str(department).strip()
        if not sales_name or not department_text:
            continue
        if sales_name in {"销售", "销售姓名"} and normalize_header(department_text) in {
            "部门",
            "部门/业务分类",
            "业务分类",
        }:
            continue
        org_map[sales_name] = department_text
    if not org_map:
        raise ValueError(f"组织架构表未读取到有效映射：{path}")
    return org_map


def load_org_map_from_dict(mapping: dict[str, str]) -> dict[str, str]:
    return {normalize_name(k): str(v).strip() for k, v in mapping.items() if normalize_name(k) and str(v).strip()}


def to_display_dept(raw_dept: str | None) -> str:
    if raw_dept is None or str(raw_dept).strip() == "":
        return "（未匹配）"
    raw = str(raw_dept).strip()
    return DEPT_DISPLAY_MAP.get(raw, raw if raw in DEPT_DISPLAY_ORDER else "（未匹配）")


def _pick_field(record: dict[str, Any], candidates: tuple[str, ...]) -> tuple[Any, str]:
    """Return (value, matched_key). Prefer exact then normalized header match."""
    if not record:
        return None, ""
    # exact
    for k in candidates:
        if k in record and record[k] not in (None, ""):
            return record[k], k
    # normalized keys map
    norm_map = {normalize_header(k): k for k in record.keys()}
    for k in candidates:
        nk = normalize_header(k)
        if nk in norm_map:
            orig = norm_map[nk]
            if record[orig] not in (None, ""):
                return record[orig], orig
    # empty exact present
    for k in candidates:
        if k in record:
            return record[k], k
        nk = normalize_header(k)
        if nk in norm_map:
            return record[norm_map[nk]], norm_map[nk]
    return None, ""


def pick_amount(record: dict[str, Any]) -> tuple[float | None, str]:
    val, key = _pick_field(record, AMOUNT_PRIMARY_KEYS)
    if val not in (None, ""):
        return parse_amount(val), key
    val, key = _pick_field(record, AMOUNT_FALLBACK_KEYS)
    if val not in (None, ""):
        return parse_amount(val), key
    return None, key or ""


def summarize_records(
    records: list[dict[str, Any]],
    org_map: dict[str, str],
) -> SummaryResult:
    """聚合明细 → 日期×展示部门 万元（round 2）。"""
    org = load_org_map_from_dict(org_map) if org_map else {}
    by_date: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    unmatched: set[str] = set()
    amount_field = ""
    detail_rows: list[dict[str, Any]] = []
    n = 0

    for rec in records:
        n += 1
        sales_val, _ = _pick_field(rec, SALES_KEYS)
        sales_name = normalize_name(sales_val)
        date_val, _ = _pick_field(rec, DATE_KEYS)
        date_key = normalize_date_key(date_val)
        amount, field_used = pick_amount(rec)
        if field_used and not amount_field:
            amount_field = field_used
        elif field_used and amount_field and field_used != amount_field:
            # keep primary if already set; log only first
            pass

        raw_dept = org.get(sales_name)
        if raw_dept is None:
            display = "（未匹配）"
            unmatched.add(sales_name or "(空白销售)")
        else:
            display = to_display_dept(raw_dept)
            # if raw dept not in known map and not already display, treat as unmatched
            if display == "（未匹配）" and raw_dept not in DEPT_DISPLAY_MAP:
                unmatched.add(sales_name or "(空白销售)")

        wan = round((amount or 0.0) / 10000.0, 10)  # keep precision then round at output
        by_date[date_key][display] += wan

        detail_rows.append(
            {
                "销售": sales_name,
                "下单日期": date_key,
                "金额本币": amount if amount is not None else 0.0,
                "金额万元": round(wan, 2),
                "归类": display,
                "金额字段": field_used or amount_field,
            }
        )

    # round each cell to 2
    rounded: dict[str, dict[str, float]] = {}
    grand = 0.0
    for d, depts in by_date.items():
        rounded[d] = {}
        for dept, v in depts.items():
            rv = round(v, 2)
            rounded[d][dept] = rv
            grand += rv
    grand = round(grand, 2)

    return SummaryResult(
        by_date=rounded,
        unmatched_sales=sorted(unmatched),
        detail_row_count=n,
        amount_field_used=amount_field,
        grand_total_wan=grand,
        detail_rows=detail_rows,
    )


def row_totals(result: SummaryResult) -> dict[str, float]:
    """Per-date total across all display depts including unmatched."""
    out: dict[str, float] = {}
    for d, depts in result.by_date.items():
        out[d] = round(sum(depts.values()), 2)
    return out


def dept_totals(result: SummaryResult) -> dict[str, float]:
    out: dict[str, float] = {k: 0.0 for k in DEPT_DISPLAY_ORDER}
    for depts in result.by_date.values():
        for k, v in depts.items():
            out[k] = round(out.get(k, 0.0) + v, 2)
    return out
