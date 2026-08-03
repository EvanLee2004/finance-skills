#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
把校验通过的计划写进盈亏**工作副本**（plan → validate → **execute**）。

2026-07-23 明妹口径：
  - **只改「明细」sheet**，其它 sheet 一个字节不许动
  - 可以长期用同一份副本；输出仍写**新文件**（她对照原版其它 sheet 总数验收）
  - 回填内容来自第 6 步智云判定结果，不是流转表

安全设计：
  输入 = 她给的盈亏副本（只读打开）+ 校验后的计划
  输出 = 04_产出/盈亏核算表_已回填_日期.xlsx（新文件）+ 变更清单 + 订单写入差异表
  用 OOXML 补丁只改明细格，避免 openpyxl 毁图/透视

写完立刻回读逐格比对；对不上非 0 退出。
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import shutil
import sys
from pathlib import Path
from typing import Dict, List

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
import common  # noqa: E402
from validate_plan import DERIVED, FIVE, _norm, check_one, read_ledger_rows  # noqa: E402

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass


def locate_columns(ws, aliases) -> Dict[str, int]:
    """找到『明细』的表头行、应收与五列（+SO/SOD）。返回 1-based 列号。"""
    all_rows = list(ws.iter_rows(values_only=True))
    hrow, headers = common.find_header_row(
        all_rows, "盈亏明细", ["SO", "SOD", "计提", "回款明细", "是否结账"], aliases
    )
    cols = common.resolve_columns(
        headers,
        "盈亏明细",
        ["SO", "SOD", "计提", "回款明细", "是否结账", "收款时间", "收款方式"],
        aliases,
    )
    diff_idx = common.fuzzy_find_col(
        headers, (aliases.get("盈亏明细", {}) or {}).get("差异", ["差异"])
    )
    if diff_idx is not None:
        cols["差异"] = diff_idx
    yidx = common.fuzzy_find_col(
        headers, (aliases.get("盈亏明细", {}) or {}).get("应收", ["应收金额", "应收"])
    )
    if yidx is None:
        raise ValueError("盈亏『明细』找不到应收金额列，无法执行部分回款拆行")
    cols["应收"] = yidx
    return {k: v + 1 for k, v in cols.items()}  # openpyxl 是 1-based


def write_plan(src: Path, out: Path, items: List[dict]) -> List[dict]:
    """
    把 items 的五列写进 out（out 是 src 的无损副本）。返回变更明细。

    **不用 openpyxl 保存**：实测她的真表用 openpyxl 载入再保存会丢 5 个 drawing、
    1 张内嵌图片和若干 rels（74 个部件变 59）。改用 `xlsx_patch` 只补丁目标格，
    其余部件逐字节原样写回。
    """
    import openpyxl
    import xlsx_patch

    # 先用只读模式拿列位置与改前值（不保存，纯读）
    wb = openpyxl.load_workbook(str(src), read_only=True, data_only=True)
    if "明细" not in wb.sheetnames:
        names = list(wb.sheetnames)
        wb.close()
        raise ValueError(
            f"盈亏表无『明细』sheet（明妹规定只许改明细）：现有={names}"
        )
    # 铁律：后续 patch 目标名写死「明细」，禁止调用方改成别的 sheet
    target_sheet = "明细"
    cols = locate_columns(wb[target_sheet], common.load_aliases())
    wb.close()

    before_rows = read_ledger_rows(src)
    changes: List[dict] = []
    edits: List = []
    insertions: List = []
    source_rows = sorted(
        int(it["ledger_row_ref"])
        for it in items
        if (it.get("row_operation") or {}).get("type") == "split_below"
    )

    def final_original_row(row: int) -> int:
        return int(row) + sum(1 for source in source_rows if source < int(row))

    for it in items:
        r = int(it["ledger_row_ref"])
        applied_r = final_original_row(r)
        it["_applied_row_ref"] = applied_r
        five = it.get("five_cols") or {}
        derived = it.get("derived_cols") or {}
        before = before_rows.get(r, {})
        for k in FIVE:
            v = five.get(k)
            if v is None:
                continue  # 部分核销时计提留空——留空就是留空，不写 0
            if k == "收款时间":
                v = common.norm_date(v) or v
            elif k in ("计提", "回款明细"):
                v = float(v)
            edits.append((r, cols[k], v))
        if "差异" in derived:
            if "差异" not in cols:
                raise ValueError("本次需要写差异，但盈亏『明细』找不到“差异”列")
            target_diff = float(derived["差异"])
            formula = (
                f"={xlsx_patch.col_letter(cols['应收'])}{applied_r}"
                f"-{xlsx_patch.col_letter(cols['计提'])}{applied_r}"
            )
            edits.append(
                (r, cols["差异"], xlsx_patch.FormulaValue(formula, target_diff))
            )
        sod = five.get("实收SOD") or it.get("sod")
        if sod:
            edits.append((r, cols["SOD"], sod))
        op = it.get("row_operation") or {}
        if op.get("type") == "split_below":
            edits.append((r, cols["应收"], float(op["paid_receivable"])))
            # 部分回款阶段两侧计提与业务值差异都必须留空。旧版测试表可以没有
            # “差异”列；存在时显式把源行和复制出的未回款行保持为空。
            if "差异" in cols:
                edits.append((r, cols["差异"], None))
            inserted_five = op.get("inserted_five_cols") or {}
            overrides = {
                cols["应收"]: float(op["unpaid_receivable"]),
                cols["计提"]: None,
                cols["回款明细"]: None,
                cols["是否结账"]: "否",
                cols["收款时间"]: None,
                cols["收款方式"]: None,
            }
            if "差异" in cols:
                overrides[cols["差异"]] = None
            inserted_sod = inserted_five.get("实收SOD") or sod
            if inserted_sod:
                overrides[cols["SOD"]] = inserted_sod
            insertions.append((r, overrides))
            it["_inserted_row_ref"] = applied_r + 1
        changes.append(
            {
                "案例ID": it.get("case_id"),
                "行号": applied_r,
                "SO": it.get("so"),
                "SOD": sod,
                "改前": {k: _norm(before.get(k)) for k in FIVE},
                "改后": {k: _norm(five.get(k)) for k in FIVE},
                "派生列_改前": {k: _norm(before.get(k)) for k in DERIVED if k in derived},
                "派生列_改后": {k: _norm(derived.get(k)) for k in DERIVED if k in derived},
                "派生列_公式": {
                    "差异": formula if "差异" in derived else "",
                },
                "操作": "拆分并在下方新增未回款行" if op else "更新",
                "新增行号": applied_r + 1 if op else "",
            }
        )
    xlsx_patch.patch_cells(src, out, target_sheet, edits, insertions=insertions)

    # 写完立刻自证没搞坏她的表：少一个部件都算失败
    lost = xlsx_patch.parts_diff(src, out)
    if lost:
        raise ValueError(f"写入后工作簿部件缺失（不该发生）：{lost[:5]}")
    return changes


def precheck_before_write(plan: dict, items: List[dict], src: Path) -> List[str]:
    """
    **写入前再复核一次**（2026-07-25 立）。

    为什么非有不可：`validate_plan` 复核的是"校验那一刻"的表，即使主流程随后立即写入，
    工作副本仍可能被其它进程或人工同时修改。而 `--in-place` 写的正是她天天在用的那份表：
      · 月初贴交付会**插行**、部分核销会在上方**插行** → 行号当场全部错位
      · 错位之后按旧行号写 = 把这一笔的五列写到**别人那一单**上
      · 更糟的是我们还会写 SOD 列，等于把那行的单号也覆盖掉
    而写后回读只证明"值写进去了"，证明不了"写对了行"，所以照样会报 ✓。

    两道闸：
      ① 指纹：校验时的盈亏副本 sha256 与现在不一致 → 表被动过，整批拒写
      ② 身份：拿**当前**表对每条 write 项重跑 check_one，任何一条不再是 write → 整批拒写
    任一不过就整批中止（不是跳过那一条）——因为插一行会让它**之后所有行**一起错位，
    只跳过报错的那条，等于把剩下的照样写歪。
    """
    problems: List[str] = []

    recorded = plan.get("ledger_sha256")
    if recorded:
        if plan.get("ledger_path") and Path(plan["ledger_path"]).resolve() != src.resolve():
            problems.append(
                f"这份计划是对着 {Path(plan['ledger_path']).name} 校验的，"
                f"现在要写的却是 {src.name}"
            )
        elif common.sha256_file(src) != recorded:
            problems.append(
                f"{src.name} 在「校验」之后被改动过（指纹对不上）"
                "——可能你自己填了几行、或者插了行"
            )
    else:
        print(
            "WARN: 这份计划里没有盈亏表指纹（旧版计划），跳过指纹闸；仍会逐行复核身份。",
            file=sys.stderr,
        )

    rows = read_ledger_rows(src)
    for it in items:
        ref = it.get("ledger_row_ref")
        ident = it.get("_identity") or {}
        cur = rows.get(int(ref)) if ref else None
        if cur is None:
            problems.append(f"第 {ref} 行现在不存在了（{it.get('case_id')}）")
            continue
        # 身份：这一行还是校验时那一行吗
        for key in ("SO", "SOD"):
            was, now = ident.get(key), cur.get(key)
            if was and now and was != now:
                problems.append(
                    f"第 {ref} 行已经不是原来那单了：校验时 {key}={was}，现在 {key}={now}"
                )
        # 内容：这一行还能写吗（她可能刚好自己把这行填了）
        res = check_one(it, rows)
        if res["verdict"] != "write":
            problems.append(
                f"第 {ref} 行现在不能写了（{res['verdict']}）：{res['reason']}"
            )
    return problems


def verify_written(out: Path, items: List[dict]) -> List[str]:
    """回读逐格比对——写完必须证明真写对了，而不是"保存没报错就算成"。"""
    import openpyxl
    import xlsx_patch

    rows = read_ledger_rows(out)
    problems: List[str] = []
    formula_wb = openpyxl.load_workbook(str(out), read_only=True, data_only=False)
    formula_ws = formula_wb["明细"]
    formula_cols = locate_columns(formula_ws, common.load_aliases())
    for it in items:
        r = int(it.get("_applied_row_ref") or it["ledger_row_ref"])
        five = it.get("five_cols") or {}
        row = rows.get(r)
        if row is None:
            problems.append(f"第 {r} 行写完却读不到")
            continue
        for k in FIVE:
            if five.get(k) is None:
                continue
            if _norm(row.get(k)) != _norm(five.get(k)):
                problems.append(
                    f"第 {r} 行 {k}：期望 {_norm(five.get(k))!r} 实际 {_norm(row.get(k))!r}"
                )
        derived = it.get("derived_cols") or {}
        for k in DERIVED:
            if k not in derived:
                continue
            if _norm(row.get(k)) != _norm(derived[k]):
                problems.append(
                    f"第 {r} 行 {k}：期望 {_norm(derived[k])!r} 实际 {_norm(row.get(k))!r}"
                )
            if k == "差异":
                expected_formula = (
                    f"={xlsx_patch.col_letter(formula_cols['应收'])}{r}"
                    f"-{xlsx_patch.col_letter(formula_cols['计提'])}{r}"
                )
                actual_formula = formula_ws.cell(r, formula_cols["差异"]).value
                if actual_formula != expected_formula:
                    problems.append(
                        f"第 {r} 行 差异公式：期望 {expected_formula!r} "
                        f"实际 {actual_formula!r}"
                    )
        # SOD 也是我们写进去的一列（write_plan 会写），必须一起回读证明。
        # 2026-07-25 补：旧版只比五列，SOD 写歪了回读照样报「全部一致 ✓」。
        sod = (five.get("实收SOD") or it.get("sod") or "").strip()
        if sod and _norm(row.get("SOD")) != _norm(sod):
            problems.append(
                f"第 {r} 行 SOD：期望 {sod!r} 实际 {_norm(row.get('SOD'))!r}"
            )
        op = it.get("row_operation") or {}
        if op.get("type") == "split_below":
            inserted_r = int(it.get("_inserted_row_ref") or (r + 1))
            inserted = rows.get(inserted_r)
            if inserted is None:
                problems.append(f"第 {inserted_r} 行应为新增未回款行，但写完读不到")
                continue
            expected = {
                "应收金额": op.get("unpaid_receivable"),
                "计提": None, "回款明细": None, "是否结账": "否",
                "收款时间": None, "收款方式": None, "SOD": sod,
            }
            if "差异" in formula_cols:
                expected["差异"] = None
            for key, want in expected.items():
                got = inserted.get(key)
                if _norm(got) != _norm(want):
                    problems.append(
                        f"第 {inserted_r} 行 {key}：期望 {_norm(want)!r} 实际 {_norm(got)!r}"
                    )
    formula_wb.close()
    return problems


def write_change_report(changes: List[dict], path: Path) -> None:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "变更清单"
    headers = (
        ["案例ID", "行号", "SO", "SOD", "操作", "新增行号"]
        + [f"改前_{k}" for k in FIVE]
        + [f"改后_{k}" for k in FIVE]
        + [f"改前_{k}" for k in DERIVED]
        + [f"改后_{k}" for k in DERIVED]
        + [f"公式_{k}" for k in DERIVED]
    )
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for ch in changes:
        ws.append(
            [ch["案例ID"], ch["行号"], ch["SO"], ch["SOD"], ch["操作"], ch["新增行号"]]
            + [ch["改前"][k] for k in FIVE]
            + [ch["改后"][k] for k in FIVE]
            + [ch.get("派生列_改前", {}).get(k, "") for k in DERIVED]
            + [ch.get("派生列_改后", {}).get(k, "") for k in DERIVED]
            + [ch.get("派生列_公式", {}).get(k, "") for k in DERIVED]
        )
    ws.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


def _comparison_objects(items: List[dict]) -> List[dict]:
    """把本次实际写入计划展开成需要和上传盈亏表逐项核对的订单对象。"""
    objects: List[dict] = []
    for item in items:
        five = item.get("five_cols") or {}
        derived = item.get("derived_cols") or {}
        sod = five.get("实收SOD") or item.get("sod") or ""
        applied_row = int(item.get("_applied_row_ref") or item["ledger_row_ref"])
        expected = {
            "SO": item.get("so") or "",
            "SOD": sod,
        }
        for key in FIVE:
            if five.get(key) is not None:
                expected[key] = five[key]
        for key in DERIVED:
            if key in derived:
                expected[key] = derived[key]

        operation = item.get("row_operation") or {}
        if operation.get("type") == "split_below":
            expected["应收金额"] = operation.get("paid_receivable")
            expected["计提"] = None
            expected["差异"] = None
        objects.append(
            {
                "item": item,
                "object_type": "写入订单",
                "planned_row": applied_row,
                "expected": expected,
            }
        )

        if operation.get("type") == "split_below":
            inserted_row = int(item.get("_inserted_row_ref") or (applied_row + 1))
            objects.append(
                {
                    "item": item,
                    "object_type": "拆分新增未回款行",
                    "planned_row": inserted_row,
                    "expected": {
                        "SO": item.get("so") or "",
                        "SOD": sod,
                        "应收金额": operation.get("unpaid_receivable"),
                        "计提": None,
                        "回款明细": None,
                        "是否结账": "否",
                        "收款时间": None,
                        "收款方式": None,
                        "差异": None,
                    },
                }
            )
    return objects


def _report_value(value):
    """保留数字/日期类型，标识符与说明按文本写入差异表。"""
    if value is None:
        return ""
    if isinstance(value, (int, float, dt.date, dt.datetime)):
        return value
    return str(value).strip()


def build_order_difference(
    items: List[dict],
    ledger: Path,
    *,
    hexiao_date: str = "",
) -> dict:
    """
    将本次实际写入订单与上传盈亏表的写后副本逐字段对比。

    只核对本批次实际写入的订单，不把上传表中的全部历史订单算作“多余订单”。
    """
    rows = read_ledger_rows(ledger)
    order_rows: List[dict] = []
    field_differences: List[dict] = []
    for obj in _comparison_objects(items):
        item = obj["item"]
        expected = obj["expected"]
        planned_row = obj["planned_row"]
        actual_row = rows.get(planned_row)
        actual_row_ref = planned_row if actual_row is not None else None
        row_differences: List[dict] = []

        if actual_row is None:
            row_differences.append(
                {
                    "field": "订单行",
                    "expected": f"第 {planned_row} 行",
                    "actual": "上传表中不存在",
                    "expected_value": f"第 {planned_row} 行",
                    "actual_value": "上传表中不存在",
                }
            )
        else:
            for field, wanted in expected.items():
                actual = actual_row.get(field)
                if _norm(actual) != _norm(wanted):
                    row_differences.append(
                        {
                            "field": field,
                            "expected": _norm(wanted),
                            "actual": _norm(actual),
                            "expected_value": wanted,
                            "actual_value": actual,
                        }
                    )

        for diff in row_differences:
            field_differences.append(
                {
                    "核销日期": hexiao_date,
                    "案例ID": item.get("case_id") or "",
                    "AR": item.get("ar") or "",
                    "SO": item.get("so") or "",
                    "SOD": expected.get("SOD") or "",
                    "对比对象": obj["object_type"],
                    "计划行号": planned_row,
                    "上传表行号": actual_row_ref or "",
                    "差异字段": diff["field"],
                    "计划写入值": _report_value(diff["expected_value"]),
                    "上传表写后值": _report_value(diff["actual_value"]),
                }
            )

        order_rows.append(
            {
                "核销日期": hexiao_date,
                "案例ID": item.get("case_id") or "",
                "AR": item.get("ar") or "",
                "SO": item.get("so") or "",
                "SOD": expected.get("SOD") or "",
                "对比对象": obj["object_type"],
                "计划行号": planned_row,
                "上传表行号": actual_row_ref or "",
                "对比结果": "一致" if not row_differences else "有差异",
                "差异字段数": len(row_differences),
                "差异字段": "、".join(d["field"] for d in row_differences),
                "计划计提": _report_value(expected.get("计提")),
                "上传表计提": _report_value((actual_row or {}).get("计提")),
                "计划回款明细": _report_value(expected.get("回款明细")),
                "上传表回款明细": _report_value(
                    (actual_row or {}).get("回款明细")
                ),
            }
        )
    return {
        "hexiao_date": hexiao_date,
        "ledger": str(ledger),
        "written_order_count": len(items),
        "comparison_object_count": len(order_rows),
        "matched_count": sum(1 for row in order_rows if row["对比结果"] == "一致"),
        "difference_count": len(field_differences),
        "order_rows": order_rows,
        "field_differences": field_differences,
    }


def write_order_difference_report(result: dict, path: Path) -> None:
    """生成可直接交付的写后订单差异 Excel。"""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    summary = wb.active
    summary.title = "汇总"
    summary_rows = [
        ("核销日期", result.get("hexiao_date") or ""),
        ("上传盈亏表", Path(result.get("ledger") or "").name),
        ("本次写入订单数", result.get("written_order_count", 0)),
        ("对比对象数", result.get("comparison_object_count", 0)),
        ("一致对象数", result.get("matched_count", 0)),
        ("字段差异数", result.get("difference_count", 0)),
        (
            "结论",
            "写入订单与上传表写后数据全部一致"
            if not result.get("difference_count")
            else "存在差异，请查看“字段差异”工作表",
        ),
    ]
    for key, value in summary_rows:
        summary.append([key, value])
    for cell in summary["A"]:
        cell.font = Font(bold=True)
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 48

    order_headers = [
        "核销日期", "案例ID", "AR", "SO", "SOD", "对比对象",
        "计划行号", "上传表行号", "对比结果", "差异字段数", "差异字段",
        "计划计提", "上传表计提", "计划回款明细", "上传表回款明细",
    ]
    orders = wb.create_sheet("订单对比")
    orders.append(order_headers)
    for row in result.get("order_rows") or []:
        orders.append([row.get(header, "") for header in order_headers])

    difference_headers = [
        "核销日期", "案例ID", "AR", "SO", "SOD", "对比对象",
        "计划行号", "上传表行号", "差异字段", "计划写入值", "上传表写后值",
    ]
    differences = wb.create_sheet("字段差异")
    differences.append(difference_headers)
    for row in result.get("field_differences") or []:
        differences.append([row.get(header, "") for header in difference_headers])
    if differences.max_row == 1:
        differences.append(["无差异"])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    success_fill = PatternFill("solid", fgColor="E6F4EA")
    difference_fill = PatternFill("solid", fgColor="FCE8E6")
    summary.sheet_view.showGridLines = False
    summary["B7"].fill = (
        success_fill if not result.get("difference_count") else difference_fill
    )
    summary["B7"].font = Font(bold=True)
    for ws in (orders, differences):
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for column in ws.columns:
            letter = column[0].column_letter
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 36)
            ws.column_dimensions[letter].width = max(width, 10)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in ("L", "M", "N", "O"):
        for cell in orders[column][1:]:
            cell.number_format = '#,##0.00;[Red](#,##0.00);-'
    for row in orders.iter_rows(min_row=2):
        if row[8].value == "有差异":
            for cell in row:
                cell.fill = difference_fill
        else:
            for cell in row:
                cell.fill = success_fill
    for row in differences.iter_rows(min_row=2):
        if row[0].value != "无差异":
            for cell in row:
                cell.fill = difference_fill

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


def _apply_new_file(
    src: Path,
    out: Path,
    report: Path,
    difference_report: Path,
    writable: List[dict],
    *,
    hexiao_date: str = "",
) -> int:
    """默认模式：写一份新文件，她给的副本一个字节不动（最安全，头几次现场用这个并排验）。"""
    out.parent.mkdir(parents=True, exist_ok=True)
    try:
        changes = write_plan(src, out, writable)
    except ValueError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 2

    problems = verify_written(out, writable)
    write_change_report(changes, report)
    try:
        difference_result = build_order_difference(
            writable, out, hexiao_date=hexiao_date
        )
        write_order_difference_report(difference_result, difference_report)
    except Exception as e:
        print(
            f"ERROR: 订单写入差异表生成失败：{type(e).__name__}"
            "（她给的副本没有改动）",
            file=sys.stderr,
        )
        return 2

    print(f"已写入 {len(changes)} 笔 → {out}")
    print(f"变更清单 → {report}")
    print(f"订单写入差异表 → {difference_report}")
    print(f"基线未动（她给的副本）→ {src}")
    if problems:
        print("⚠ 写后回读比对不符：", file=sys.stderr)
        for x in problems[:10]:
            print(f"  - {x}", file=sys.stderr)
        return 1
    print("写后回读逐格比对：全部一致 ✓")
    return 0


def _apply_in_place(
    src: Path,
    report: Path,
    difference_report: Path,
    writable: List[dict],
    *,
    hexiao_date: str = "",
) -> int:
    """
    就地模式（明妹要的）：她固定用同一份副本，我们直接往这份里回填，省得她天天贴。

    她只要求「直接往回写」，但"直接写她天天用的表"就没有了人工过一眼的安全网，
    所以这里把安全全做在程序里，让"直接写"和"绝不搞坏她的表"两件事同时成立：
      1. 写之前先把这份副本整份备份到 `备份/`（真出事，拿备份一还原就回来了）
      2. 先写一个临时文件、跑无损校验(部件不缺)+逐格回读比对
      3. **只有全过了才原子替换**原副本；没过就原样保留她的副本、把临时结果留着给人看
    """
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = src.parent / "备份"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup = backup_dir / f"{src.stem}_备份_{ts}{src.suffix}"
    shutil.copy2(src, backup)

    tmp = src.with_name(f".{src.stem}_写入中_{ts}{src.suffix}")
    try:
        changes = write_plan(src, tmp, writable)
    except ValueError as e:
        tmp.unlink(missing_ok=True)
        print(f"ERROR: {e}\n（原副本没动，备份在 {backup}）", file=sys.stderr)
        return 2

    problems = verify_written(tmp, writable)
    if problems:
        print(
            "⚠ 写后回读比对不符——**没有改动你的副本**（写坏的只是临时文件）：",
            file=sys.stderr,
        )
        for x in problems[:10]:
            print(f"  - {x}", file=sys.stderr)
        print(f"临时结果留在 {tmp}（给同事看）；你的副本原样；备份在 {backup}", file=sys.stderr)
        return 1

    try:
        difference_result = build_order_difference(
            writable, tmp, hexiao_date=hexiao_date
        )
        # 就地模式先在临时文件验证，但交付报告必须显示最终业务副本名。
        difference_result["ledger"] = str(src)
        write_order_difference_report(difference_result, difference_report)
    except Exception as e:
        tmp.unlink(missing_ok=True)
        print(
            f"ERROR: 订单写入差异表生成失败：{type(e).__name__}"
            f"（原副本没动，备份在 {backup}）",
            file=sys.stderr,
        )
        return 2

    try:
        tmp.replace(src)  # 要么整份换成新的，要么完全没换，不会写一半
    except OSError as e:
        difference_report.unlink(missing_ok=True)
        print(
            f"ERROR: 无法原子替换盈亏副本：{type(e).__name__}"
            f"（原副本没动，备份在 {backup}）",
            file=sys.stderr,
        )
        return 2
    write_change_report(changes, report)
    _resnapshot_sources(src)
    print(f"已就地回填 {len(changes)} 笔 → {src}")
    print(f"写前备份 → {backup}")
    print(f"变更清单 → {report}")
    print(f"订单写入差异表 → {difference_report}")
    print("写后回读逐格比对：全部一致 ✓")
    return 0


def _resnapshot_sources(ledger: Path) -> None:
    """
    就地回填成功后重新打指纹。

    否则下一次 `verify_sources verify` 必然报「盈亏表被改动」——**那是我们自己
    经校验后合法写的**，却长得跟"程序偷偷改了她的表"一模一样。
    2026-07-23 opencode 实测就踩到：AI 照 SKILL 在 apply 后跑 verify，
    当场甩出一句吓人的「校验未通过」。新指纹＝新基线，之后再变才是真异常。
    """
    ws = ledger.parent.parent
    try:
        import verify_sources

        verify_sources.do_snapshot(ws)
    except Exception as e:  # 快照失败不该让已成功的写入变成失败
        print(
            f"WARN: 重打源文件指纹失败（不影响已写入的数据）：{type(e).__name__}",
            file=sys.stderr,
        )


def _mark_review_applied(checked_p: Path) -> None:
    """写成功后把旧版待确认标记改成已应用（若存在）。"""
    for folder in (checked_p.parent, checked_p.parent.parent / "04_产出"):
        stamp = folder / "回填审核_待确认.json"
        if stamp.is_file():
            try:
                data = json.loads(stamp.read_text(encoding="utf-8"))
            except Exception:
                data = {}
            data["status"] = "applied"
            data["applied_at"] = dt.datetime.now().isoformat(timespec="seconds")
            stamp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
            return


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="把校验通过的计划写进盈亏副本的「明细」sheet")
    ap.add_argument("--checked", required=True, help="validate_plan.py 产出的校验后计划")
    ap.add_argument("--ledger", required=True, help="她的盈亏副本")
    ap.add_argument("--out", default="", help="新文件模式的输出路径（默认落 04_产出/）")
    ap.add_argument("--report", default="", help="变更清单 xlsx")
    ap.add_argument("--difference-report", default="", help="订单写入差异表 xlsx")
    ap.add_argument("--force", action="store_true", help="即使计划里有 conflict 也照写可写的那部分")
    ap.add_argument(
        "--in-place",
        action="store_true",
        help="就地写这份副本（她要的：一直用同一份）。写前自动备份、写后校验、校验过才替换",
    )
    ap.add_argument(
        "--confirmed",
        action="store_true",
        help="已废弃的兼容参数；现在校验与日清生成后可直接写入。",
    )
    args = ap.parse_args(argv)

    checked_p, src = Path(args.checked), Path(args.ledger)
    for p, name in ((checked_p, "校验后计划"), (src, "盈亏副本")):
        if not p.is_file():
            print(f"ERROR: 找不到{name} {p}", file=sys.stderr)
            return 2

    plan = json.loads(checked_p.read_text(encoding="utf-8"))
    writable = plan.get("write") or []
    conflicts = plan.get("conflict") or []
    if conflicts and not args.force:
        print(
            f"ERROR: 计划里还有 {len(conflicts)} 笔冲突没处理，先看清楚再写。\n"
            f"  （确认要跳过冲突、只写可写的那部分，就加 --force）",
            file=sys.stderr,
        )
        return 2
    if not writable:
        print("没有可写的笔（可能都已经填过了）。什么都没改。")
        return 0

    # ★ 写入前最后一道闸：她的表在「校验 → 写入」之间是否发生变化
    try:
        stale = precheck_before_write(plan, writable, src)
    except ValueError as e:
        print(f"ERROR: 写入前复核读不了表：{e}", file=sys.stderr)
        return 2
    if stale:
        print(
            "ERROR: 写入前复核没过——**一个字都没写**。\n"
            "  你的表在「出清单」之后被改动过，再按旧行号写就会写到别人那一单上。",
            file=sys.stderr,
        )
        for x in stale[:10]:
            print(f"  - {x}", file=sys.stderr)
        if len(stale) > 10:
            print(f"  …另有 {len(stale) - 10} 条", file=sys.stderr)
        print(
            "  → 怎么办：重跑 validate_plan.py + build_worklist.py 出一份新的《核销日清》，"
            "她再确认一次（几十秒的事，别绕过）。",
            file=sys.stderr,
        )
        return 2

    today = dt.date.today().strftime("%Y%m%d")
    hexiao_date = str(plan.get("hexiao_date") or "")
    normalized_date = common.norm_date(hexiao_date)
    date_tag = normalized_date.strftime("%Y%m%d") if normalized_date else today
    report = Path(args.report) if args.report else (
        src.parent.parent / "04_产出" / f"变更清单_{date_tag}.xlsx"
    )
    difference_report = Path(args.difference_report) if args.difference_report else (
        src.parent.parent / "04_产出" / f"订单写入差异_{date_tag}.xlsx"
    )
    report.parent.mkdir(parents=True, exist_ok=True)

    if args.in_place:
        rc = _apply_in_place(
            src,
            report,
            difference_report,
            writable,
            hexiao_date=hexiao_date,
        )
    else:
        out = (
            Path(args.out)
            if args.out
            else src.parent.parent / "04_产出" / f"盈亏核算表_已回填_{today}.xlsx"
        )
        rc = _apply_new_file(
            src,
            out,
            report,
            difference_report,
            writable,
            hexiao_date=hexiao_date,
        )

    if rc == 0:
        _mark_review_applied(checked_p)
    return rc


if __name__ == "__main__":
    sys.exit(main())
