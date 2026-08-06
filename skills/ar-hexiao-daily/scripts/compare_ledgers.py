#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""比较两份盈亏表“明细”：严格差额保留，业务差额按1元判定。

配对先按 SO+SOD 分组，再以计提、回款明细、是否结账为首要成本做确定性
最小差异配对。收款时间、收款方式不参与结果；参考表不得反向影响核销计划。
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import math
import re
import sys
from collections import Counter, defaultdict
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
import amount_policy  # noqa: E402


IGNORED_FIELDS = frozenset({"收款时间", "收款方式"})
CORE_FIELDS = ("计提金额", "回款明细", "是否结账")
AMOUNT_FIELD_WORDS = ("金额", "回款", "计提", "应收", "未收", "差异", "销售额")


def canonical(value: Any) -> Any:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, float):
        if math.isnan(value):
            return None
        return round(value, 10)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return ""
        if re.fullmatch(r"\d{4}-\d{1,2}-\d{1,2}(?: 00:00:00)?", text):
            return text[:10]
        return text
    return value


def is_blank(value: Any) -> bool:
    return canonical(value) in (None, "")


def is_number(value: Any) -> bool:
    return isinstance(canonical(value), (int, float)) and not isinstance(value, bool)


def exact_equal(left: Any, right: Any) -> bool:
    left, right = canonical(left), canonical(right)
    if is_blank(left) and is_blank(right):
        return True
    if is_number(left) and is_number(right):
        return amount_policy.technical_equal(left, right)
    return left == right


def is_amount_field(header: str) -> bool:
    text = str(header or "")
    return any(word in text for word in AMOUNT_FIELD_WORDS)


def business_equal(left: Any, right: Any, header: str) -> bool:
    left, right = canonical(left), canonical(right)
    if is_blank(left) and is_blank(right):
        return True
    if is_number(left) and is_number(right) and is_amount_field(header):
        return amount_policy.within_business_tolerance(left, right)
    return exact_equal(left, right)


def numeric_delta(left: Any, right: Any) -> Optional[float]:
    if not (is_number(left) and is_number(right)):
        return None
    return round(float(left) - float(right), 10)


def _header_index(headers: List[str], names: Iterable[str]) -> Optional[int]:
    def norm_header(value: Any) -> str:
        return re.sub(r"[\s（）()/_-]+", "", str(value or "")).lower()

    normalized = {norm_header(value): index for index, value in enumerate(headers)}
    for name in names:
        hit = normalized.get(norm_header(name))
        if hit is not None:
            return hit
    return None


def load_detail(path: Path) -> Tuple[List[str], List[dict], Dict[str, int]]:
    formula_wb = openpyxl.load_workbook(path, read_only=True, data_only=False, keep_links=True)
    value_wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=True)
    if "明细" not in formula_wb.sheetnames or "明细" not in value_wb.sheetnames:
        formula_wb.close()
        value_wb.close()
        raise ValueError(f"{path.name} 缺少“明细”工作表")
    formula_rows = list(formula_wb["明细"].iter_rows(values_only=True))
    value_rows = list(value_wb["明细"].iter_rows(values_only=True))
    formula_wb.close()
    value_wb.close()
    if not formula_rows:
        raise ValueError(f"{path.name} 的“明细”为空")
    headers = [str(canonical(value) or f"列{index}") for index, value in enumerate(formula_rows[0], 1)]
    indexes = {
        "SO": _header_index(headers, ["新智云单号", "SO"]),
        "SOD": _header_index(headers, ["实收金额", "SOD"]),
        "计提金额": _header_index(headers, ["计提金额", "计提"]),
        "回款明细": _header_index(headers, ["回款明细"]),
        "是否结账": _header_index(headers, ["是否结账（是/否）", "是否结账"]),
        "收款时间": _header_index(headers, ["收款时间"]),
        "收款方式": _header_index(headers, ["收款方式(支/汇/现)", "收款方式"]),
    }
    missing = [key for key in ("SO", "SOD", *CORE_FIELDS) if indexes.get(key) is None]
    if missing:
        raise ValueError(f"{path.name} 缺少比较列：{missing}")
    width = max(len(formula_rows[0]), len(value_rows[0]))
    records = []
    for offset in range(1, max(len(formula_rows), len(value_rows))):
        formulas = list(formula_rows[offset]) if offset < len(formula_rows) else []
        values = list(value_rows[offset]) if offset < len(value_rows) else []
        formulas += [None] * (width - len(formulas))
        values += [None] * (width - len(values))
        effective = [
            canonical(values[index] if isinstance(formulas[index], str) and formulas[index].startswith("=") else formulas[index])
            for index in range(width)
        ]
        if all(is_blank(value) for value in effective[:22]):
            continue
        records.append({"row": offset + 1, "effective": effective, "formula": formulas})
    return headers, records, {key: int(value) for key, value in indexes.items() if value is not None}


def identity(record: dict, indexes: Dict[str, int]) -> Tuple[str, str]:
    return (
        str(canonical(record["effective"][indexes["SO"]]) or ""),
        str(canonical(record["effective"][indexes["SOD"]]) or ""),
    )


def row_cost(left: dict, right: dict, headers: List[str], indexes: Dict[str, int]) -> int:
    cost = 0
    core_indexes = {indexes[name] for name in CORE_FIELDS}
    width = min(len(headers), len(left["effective"]), len(right["effective"]), 22)
    ignored_indexes = {indexes[name] for name in IGNORED_FIELDS if name in indexes}
    for index in range(width):
        if index in ignored_indexes:
            continue
        if not business_equal(left["effective"][index], right["effective"][index], headers[index]):
            cost += 20 if index in core_indexes else 1
    return cost


def pair_group(
    left: List[dict], right: List[dict], headers: List[str], indexes: Dict[str, int]
) -> Tuple[List[Tuple[dict, dict]], List[dict], List[dict]]:
    left, right = sorted(left, key=lambda row: row["row"]), sorted(right, key=lambda row: row["row"])
    if not left or not right:
        return [], left, right
    swapped = len(left) > len(right)
    short, long = (right, left) if swapped else (left, right)
    if len(long) <= 10:
        @lru_cache(maxsize=None)
        def solve(index: int, mask: int) -> Tuple[int, Tuple[int, ...]]:
            if index == len(short):
                return 0, ()
            best = (10**9, ())
            for candidate in range(len(long)):
                if mask & (1 << candidate):
                    continue
                tail_cost, choices = solve(index + 1, mask | (1 << candidate))
                total = row_cost(short[index], long[candidate], headers, indexes) + tail_cost
                if total < best[0]:
                    best = (total, (candidate, *choices))
            return best
        choices = solve(0, 0)[1]
    else:
        unused = set(range(len(long)))
        picked = []
        for item in short:
            candidate = min(unused, key=lambda idx: (row_cost(item, long[idx], headers, indexes), long[idx]["row"]))
            unused.remove(candidate)
            picked.append(candidate)
        choices = tuple(picked)
    used = set(choices)
    pairs = [
        (long[candidate], short[index]) if swapped else (short[index], long[candidate])
        for index, candidate in enumerate(choices)
    ]
    return (
        pairs,
        [row for index, row in enumerate(long) if swapped and index not in used],
        [row for index, row in enumerate(long) if not swapped and index not in used],
    )


def align(
    final_records: List[dict], reference_records: List[dict], headers: List[str], indexes: Dict[str, int]
) -> Tuple[List[Tuple[dict, dict]], List[dict], List[dict]]:
    final_groups, reference_groups = defaultdict(list), defaultdict(list)
    for record in final_records:
        final_groups[identity(record, indexes)].append(record)
    for record in reference_records:
        reference_groups[identity(record, indexes)].append(record)
    pairs, final_extra, reference_extra = [], [], []
    for key in sorted(set(final_groups) | set(reference_groups)):
        group_pairs, left, right = pair_group(
            final_groups.get(key, []), reference_groups.get(key, []), headers, indexes
        )
        pairs.extend(group_pairs)
        final_extra.extend(left)
        reference_extra.extend(right)
    return pairs, final_extra, reference_extra


def compare_pair(final: dict, reference: dict, headers: List[str], indexes: Dict[str, int]) -> List[dict]:
    ignored_indexes = {indexes[name] for name in IGNORED_FIELDS if name in indexes}
    rows = []
    width = min(len(headers), len(final["effective"]), len(reference["effective"]), 22)
    so, sod = identity(final, indexes)
    for index in range(width):
        if index in ignored_indexes:
            continue
        left, right = final["effective"][index], reference["effective"][index]
        exact = exact_equal(left, right)
        business = business_equal(left, right, headers[index])
        if exact and business:
            continue
        rows.append({
            "SO": so, "SOD": sod, "最终表行": final["row"], "参考表行": reference["row"],
            "字段": headers[index], "最终表值": left, "参考表值": right,
            "精确差额": numeric_delta(left, right),
            "精确结果": "一致" if exact else "有差异",
            "业务容差": float(amount_policy.BUSINESS_SETTLEMENT_TOLERANCE) if is_amount_field(headers[index]) else 0,
            "业务结果": "一致" if business else "有差异",
        })
    return rows


def read_cases(paths: List[Path]) -> List[dict]:
    cases = []
    for path in paths:
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
        target_date = payload.get("hexiao_date") or ""
        for item in payload.get("auto") or []:
            cases.append({"date": target_date, "item": item})
    return cases


def compare_cases(
    cases: List[dict], final_records: List[dict], pair_map: Dict[int, dict], indexes: Dict[str, int]
) -> List[dict]:
    by_identity = defaultdict(list)
    for record in final_records:
        by_identity[identity(record, indexes)].append(record)
    used_rows = set()
    output = []
    for case in cases:
        item = case["item"]
        key = (str(item.get("so") or ""), str(item.get("sod") or ""))
        target = item.get("five_cols") or item.get("current_values") or {}
        reusable = (item.get("row_operation") or {}).get("type") == "preserve_aggregate_tail_tolerance"
        candidates = [row for row in by_identity.get(key, []) if reusable or row["row"] not in used_rows]
        if not candidates:
            output.append({"核销日": case["date"], "SO": key[0], "SOD": key[1], "业务结果": "最终表未定位"})
            continue
        def case_cost(record: dict) -> Tuple[int, int]:
            cost = 0
            for field in CORE_FIELDS:
                expected_key = "计提" if field == "计提金额" else field
                if not business_equal(record["effective"][indexes[field]], target.get(expected_key), field):
                    cost += 1
            return cost, record["row"]
        final = min(candidates, key=case_cost)
        if not reusable:
            used_rows.add(final["row"])
        reference = pair_map.get(final["row"])
        row = {"核销日": case["date"], "SO": key[0], "SOD": key[1], "最终表行": final["row"], "参考表行": reference["row"] if reference else ""}
        if reference is None:
            row["业务结果"] = "参考表未配对"
        else:
            differences = []
            exact_differences = []
            for field in CORE_FIELDS:
                left, right = final["effective"][indexes[field]], reference["effective"][indexes[field]]
                if not exact_equal(left, right):
                    exact_differences.append(field)
                if not business_equal(left, right, field):
                    differences.append(field)
            row["精确结果"] = "一致" if not exact_differences else "有差异"
            row["精确差异字段"] = "、".join(exact_differences)
            row["业务结果"] = "一致" if not differences else "有差异"
            row["业务差异字段"] = "、".join(differences)
        output.append(row)
    return output


def _write_sheet(wb, name: str, rows: List[dict]) -> None:
    ws = wb.create_sheet(name)
    headers = list(rows[0]) if rows else ["结果"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def run(final: Path, reference: Path, output: Path, case_paths: Optional[List[Path]] = None) -> dict:
    headers, final_records, indexes = load_detail(final)
    reference_headers, reference_records, reference_indexes = load_detail(reference)
    if headers[:22] != reference_headers[:22] or indexes != reference_indexes:
        raise ValueError("两份明细表的前22列表头不一致，停止自动比较")
    pairs, final_extra, reference_extra = align(final_records, reference_records, headers, indexes)
    differences = [diff for left, right in pairs for diff in compare_pair(left, right, headers, indexes)]
    exact_differences = [row for row in differences if row["精确结果"] == "有差异"]
    business_differences = [row for row in differences if row["业务结果"] == "有差异"]
    pair_map = {left["row"]: right for left, right in pairs}
    case_rows = compare_cases(read_cases(case_paths or []), final_records, pair_map, indexes)
    presence = [
        {"仅存在于": side, "行号": row["row"], "SO": identity(row, indexes)[0], "SOD": identity(row, indexes)[1]}
        for side, records in (("最终表", final_extra), ("参考表", reference_extra)) for row in records
    ]
    summary = {
        "最终表业务行": len(final_records), "参考表业务行": len(reference_records),
        "已配对业务行": len(pairs), "最终表独有行": len(final_extra), "参考表独有行": len(reference_extra),
        "精确差异单元格": len(exact_differences), "业务差异单元格": len(business_differences),
        "一元内精确差异但业务一致": sum(1 for row in differences if row["精确结果"] == "有差异" and row["业务结果"] == "一致"),
        "本批任务数": len(case_rows), "本批业务一致": sum(1 for row in case_rows if row.get("业务结果") == "一致"),
    }
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "汇总"
    for key, value in summary.items():
        ws.append([key, value])
    ws.append(["比较说明", "收款时间、收款方式不参与；精确结果使用0.005技术容差，金额业务结果使用1.00元容差。"])
    for cell in ws["A"]:
        cell.font = Font(bold=True)
    _write_sheet(wb, "业务差异", business_differences)
    _write_sheet(wb, "精确差异", exact_differences)
    _write_sheet(wb, "本批订单", case_rows)
    _write_sheet(wb, "独有行", presence)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return summary


def main() -> int:
    parser = argparse.ArgumentParser(description="盈亏明细双层容差差异比较")
    parser.add_argument("--final", required=True, type=Path)
    parser.add_argument("--reference", required=True, type=Path)
    parser.add_argument("--case", action="append", default=[], type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()
    try:
        summary = run(args.final, args.reference, args.output, args.case)
    except (OSError, ValueError, KeyError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2
    print(json.dumps(summary, ensure_ascii=False))
    print(f"差异表 → {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
