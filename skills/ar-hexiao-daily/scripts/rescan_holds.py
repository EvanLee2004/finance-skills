#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
第8步：挂账重扫。读挂账台账 + 今日判定/数据，标记可补做；幂等（连跑两遍一致）。
只维护 工作区/03_台账/挂账台账.xlsx，不写用户表。
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
import common  # noqa: E402

HEADERS = [
    "案例ID",   # = AR|SO：一笔到账挂多个 SO 时每个 SO 独立跟踪，否则状态互相覆盖
    "AR",
    "SO",
    "挂起日",
    "E码",
    "原因",
    "复查条件",  # 写清"等什么"，月初/回满/销售修关系
    "重扫次数",
    "最近重扫日",
    "状态",
]


def ledger_path(workspace: Path) -> Path:
    d = workspace / "03_台账"
    d.mkdir(parents=True, exist_ok=True)
    return d / "挂账台账.xlsx"


def load_ledger(path: Path) -> List[dict]:
    if not path.is_file():
        return []
    import openpyxl

    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    # 允许缺列时报错
    required = ["状态"]  # 案例ID/AR 兼容老台账：老表只有 AR 时按 AR 迁移
    for r in required:
        if r not in headers:
            raise ValueError(f"挂账台账缺列 {r}；实际表头：{headers}")
    idx = {h: i for i, h in enumerate(headers)}
    out = []
    for row in rows[1:]:
        vals = list(row)
        if not any(vals):
            continue

        def g(name, default=""):
            i = idx.get(name)
            if i is None or i >= len(vals) or vals[i] is None:
                return default
            return vals[i]

        ar = str(g("AR") or "").strip()
        so = str(g("SO") or g("SO列表") or "").strip()
        cid = str(g("案例ID") or "").strip() or f"{ar or '-'}|{so or '-'}"
        out.append(
            {
                "案例ID": cid,
                "AR": ar,
                "SO": so,
                "挂起日": str(g("挂起日") or "")[:10],
                "E码": str(g("E码") or "").strip(),
                "原因": str(g("原因") or "").strip(),
                "复查条件": str(g("复查条件") or "").strip(),
                "重扫次数": int(g("重扫次数") or 0),
                "最近重扫日": str(g("最近重扫日") or "")[:10],
                "状态": str(g("状态") or "挂起").strip(),
            }
        )
    return out


def save_ledger(path: Path, rows: List[dict]) -> None:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "挂账台账"
    ws.append(HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append(
            [
                r.get("案例ID") or f"{r.get('AR') or '-'}|{r.get('SO') or '-'}",
                r.get("AR") or "",
                r.get("SO") or "",
                r.get("挂起日") or "",
                r.get("E码") or "",
                r.get("原因") or "",
                r.get("复查条件") or "",
                r.get("重扫次数") or 0,
                r.get("最近重扫日") or "",
                r.get("状态") or "挂起",
            ]
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


def revisit_condition(code: str) -> str:
    """写清"等什么"——她才知道什么时候该回来看（李尚 business-rules §8 的复查条件）。"""
    return {
        "E1": "该到账后续回满、系统给出逐单核销金额后",
        "E2": "月初把上月交付数据贴进盈亏表后",
        "E3": "拿到对应年度盈亏表，或确认老单走人工",
        "E4": "销售/财务核对超额核销并修正后",
        "E0": "补登到账流转表、或确认这笔到账在哪张渠道表",
        "E12": "人工指定这笔到账对应流转表的哪一行",
        "E8": "人工指定这个 SO 记在盈亏表的哪一行",
        "E9": "与销售核对预收余额差异后",
    }.get(code, "条件具备后（人工判断）")


def merge_from_classify(rows: List[dict], result: dict, today: str) -> List[dict]:
    """
    把今日 hold/exception 并入台账；已存在则更新原因，不重复加行。
    **主键 = 案例ID(AR|SO)**：一笔 AR 挂多个 SO 时各自独立，
    否则 SO 列表会被覆盖、且任一 SO 转 auto 就把整个 AR 误标"可补做"。
    """
    # 兼容老台账/直接传入的旧结构：没有案例ID就现场补（AR|SO，SO 兼容旧列名 SO列表）
    by_case = {}
    for r in rows:
        cid = (r.get("案例ID") or "").strip()
        if not cid:
            ar0 = (r.get("AR") or "").strip()
            so0 = (r.get("SO") or r.get("SO列表") or "").strip()
            if not ar0 and not so0:
                continue
            cid = f"{ar0 or '-'}|{so0 or '-'}"
            r["案例ID"] = cid
            r.setdefault("SO", so0)
        by_case[cid] = r
    for item in (result.get("hold") or []) + (result.get("exception") or []):
        ar = (item.get("ar") or "").strip()
        so = (item.get("so") or "").strip()
        cid = item.get("case_id") or f"{ar or '-'}|{so or '-'}"
        if cid == "-|-":
            continue
        code = item.get("code") or ""
        if cid in by_case:
            rec = by_case[cid]
            if rec.get("状态") == "已完成":
                continue
            rec["E码"] = code or rec.get("E码") or ""
            rec["原因"] = item.get("reason") or rec.get("原因") or ""
            rec["复查条件"] = revisit_condition(rec["E码"])
            rec["AR"] = ar or rec.get("AR") or ""
            rec["SO"] = so or rec.get("SO") or ""
        else:
            by_case[cid] = {
                "案例ID": cid,
                "AR": ar,
                "SO": so,
                "挂起日": today,
                "E码": code,
                "原因": item.get("reason") or "",
                "复查条件": revisit_condition(code),
                "重扫次数": 0,
                "最近重扫日": "",
                "状态": "挂起",
            }
    # 今日 auto 命中同一案例 → 可补做（按 案例ID 精确匹配，不按 AR 粗放匹配）
    auto_cases = {
        (i.get("case_id") or f"{(i.get('ar') or '-')}|{(i.get('so') or '-')}")
        for i in (result.get("auto") or [])
    }
    for cid, rec in by_case.items():
        if rec.get("状态") == "已完成":
            continue
        if cid in auto_cases:
            rec["状态"] = "可补做"
    return list(by_case.values())

def rescan(rows: List[dict], result: Optional[dict], today: str) -> List[dict]:
    """重扫：递增次数、更新日期；状态机保持幂等。"""
    if result:
        rows = merge_from_classify(rows, result, today)
    # 稳定排序保证幂等写出
    rows = sorted(rows, key=lambda r: (r.get("案例ID") or "", r.get("AR") or ""))
    for r in rows:
        if r.get("状态") == "已完成":
            continue
        r["重扫次数"] = int(r.get("重扫次数") or 0) + 1
        r["最近重扫日"] = today
    return rows


def snapshot(rows: List[dict]) -> str:
    """用于 diff 的稳定序列化（不含重扫次数与最近重扫日的可比核心？）

    幂等定义：连跑两遍后，除「重扫次数/最近重扫日」外业务字段一致；
    且第二遍相对第一遍次数 +1。验收 A7：台账 diff 为空——
    故第二次跑时若 today 相同，我们采用：同一天重复跑不重复 +1（幂等）。
    """
    core = []
    for r in sorted(rows, key=lambda x: x.get("案例ID") or ""):
        core.append(
            {
                "案例ID": r.get("案例ID"),
                "AR": r.get("AR"),
                "SO": r.get("SO"),
                "复查条件": r.get("复查条件"),
                "挂起日": r.get("挂起日"),
                "E码": r.get("E码"),
                "原因": r.get("原因"),
                "状态": r.get("状态"),
                "重扫次数": r.get("重扫次数"),
                "最近重扫日": r.get("最近重扫日"),
            }
        )
    return json.dumps(core, ensure_ascii=False, sort_keys=True)


# 本地可自行复查的码：只要盈亏表变了就能重判，不必等它再出现在当日导出里
LOCAL_RECHECKABLE = {"E2", "E3", "E8"}
# 必须有新的智云导出才能判的码（本地重扫无能为力，如实说明）
NEEDS_FRESH_EXPORT = {"E1", "E4", "E9", "E0", "E12"}


def reclassify_against_ledger(rows: List[dict], ledger, today: str) -> Dict[str, int]:
    """
    **拿当前盈亏表对挂起项重新判定**（不是等它再次出现在当日导出里）。

    这是第 8 步的核心价值：月初把交付数据贴进盈亏表后，上个月因"表里还没这个 SO"
    而挂起的笔，应当自动变成"可补做"——它们不会再出现在今天的 T-1 导出中。

    只处理本地判得动的码（E2/E3/E8：SO 在不在表、在几行）；
    E1/E4/E9 等要新导出才知道的，原样保留并在原因里说明。
    """
    stat = {"升级可补做": 0, "本地判不动": 0, "跳过": 0}
    if ledger is None:
        stat["跳过"] = len(rows)
        return stat
    for r in rows:
        if r.get("状态") in ("已完成", "可补做"):
            stat["跳过"] += 1
            continue
        code = (r.get("E码") or "").strip()
        so = (r.get("SO") or "").strip()
        if code in NEEDS_FRESH_EXPORT or not so:
            stat["本地判不动"] += 1
            continue
        if code and code not in LOCAL_RECHECKABLE:
            stat["本地判不动"] += 1
            continue
        hits = list(getattr(ledger, "so_index", {}).get(so, []) or [])
        if not hits:
            stat["跳过"] += 1
            continue
        if len(hits) > 1:
            r["E码"] = "E8"
            r["原因"] = f"该 SO 在盈亏表命中 {len(hits)} 行，需人工指定记哪一行"
            r["复查条件"] = revisit_condition("E8")
            stat["本地判不动"] += 1
            continue
        r["状态"] = "可补做"
        r["原因"] = f"盈亏表已出现该 SO（{today} 重扫发现）→ 可以补做了"
        r["复查条件"] = "已具备，去清单里做掉"
        stat["升级可补做"] += 1
    return stat


def rescan_idempotent(rows: List[dict], result: Optional[dict], today: str) -> List[dict]:
    """同一天已重扫过则不再增加次数（保证连跑两遍 diff 空）。"""
    if result:
        rows = merge_from_classify(rows, result, today)
    rows = sorted(rows, key=lambda r: (r.get("案例ID") or "", r.get("AR") or ""))
    for r in rows:
        if r.get("状态") == "已完成":
            continue
        if str(r.get("最近重扫日") or "")[:10] == today:
            # 已扫过今日：不改次数
            continue
        r["重扫次数"] = int(r.get("重扫次数") or 0) + 1
        r["最近重扫日"] = today
    return rows


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="挂账重扫（幂等）")
    ap.add_argument("--workspace", default=str(common.WORK))
    ap.add_argument("--result", default="", help="今日判定结果 json（可选）")
    ap.add_argument("--init-from-result", action="store_true", help="用判定 hold 初始化台账")
    ap.add_argument("--ledger", default="", help="盈亏核算表副本（只读）；给了就对挂起项真重判")
    ap.add_argument("--no-reclassify", action="store_true", help="只合并，不拿盈亏表重判")
    args = ap.parse_args(argv)

    common.ensure_out_dirs()
    ws = Path(args.workspace)
    path = ledger_path(ws)
    today = dt.date.today().isoformat()

    result = None
    if args.result:
        result = json.loads(Path(args.result).read_text(encoding="utf-8"))
    else:
        out_dir = ws / "04_产出"
        cands = sorted(out_dir.glob("判定结果_*.json")) if out_dir.is_dir() else []
        if cands:
            result = json.loads(cands[-1].read_text(encoding="utf-8"))

    try:
        rows = load_ledger(path)
    except ValueError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 2

    if not rows and result is None:
        # 空台账也算成功（可初始化）
        save_ledger(path, [])
        print("挂账台账为空（0 行）；状态=就绪")
        print(f"台账: {path}")
        return 0

    rows = rescan_idempotent(rows, result, today)

    # —— 真重判：拿当前盈亏表复查挂起项（第 8 步的核心，不能只 merge）——
    stat = None
    if not args.no_reclassify:
        ledger = None
        lp = Path(args.ledger) if args.ledger else None
        if lp is None:
            cand = list((ws / "02_我的表副本").glob("*盈亏*")) if (ws / "02_我的表副本").is_dir() else []
            lp = cand[0] if cand else None
        if lp and lp.is_file():
            try:
                sys.path.insert(0, str(HERE))
                from classify_hexiao import LedgerIndex

                ledger = LedgerIndex(lp)
            except Exception as e:  # 读不动就如实说，不静默
                print(f"WARN: 盈亏表读取失败，跳过重判：{e}", file=sys.stderr)
        if ledger is not None:
            stat = reclassify_against_ledger(rows, ledger, today)
        else:
            print(
                "WARN: 未提供盈亏表 → 本轮只合并、未做重判。"
                "挂账里等交付数据的笔不会被自动捞出来（用 --ledger 指一份副本）",
                file=sys.stderr,
            )

    save_ledger(path, rows)

    n_hold = sum(1 for r in rows if r.get("状态") == "挂起")
    n_ready = sum(1 for r in rows if r.get("状态") == "可补做")
    n_done = sum(1 for r in rows if r.get("状态") == "已完成")
    print(f"重扫完成 挂起={n_hold} 可补做={n_ready} 已完成={n_done} 合计={len(rows)}")
    if stat:
        print(f"本轮重判：升级可补做={stat['升级可补做']} 本地判不动={stat['本地判不动']} 跳过={stat['跳过']}")
    print(f"台账: {path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
