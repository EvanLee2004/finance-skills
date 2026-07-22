#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""盘点工作区：认出各文件角色、缺什么、表头是否对得上（不含金额明细）。"""
from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
import common  # noqa: E402


def list_files(root: Path):
    out = []
    if not root.is_dir():
        return out
    for dirpath, _, files in os.walk(root):
        for f in files:
            if f.startswith("~$") or f.startswith(".") or f == ".gitkeep":
                continue
            low = f.lower()
            if low.endswith((".xls", ".xlsx", ".xlsm", ".csv", ".json", ".txt", ".md")):
                out.append(Path(dirpath) / f)
    return sorted(out)


def guess_role(path: Path) -> str:
    name = path.name
    parts = " ".join(path.parts)
    if "挂账" in name or "台账" in name:
        return "挂账台账"
    if "回款记录" in name or "回款" in name and "记录" in name:
        return "回款记录"
    if "对账" in name:
        return "回款核销对账"
    if "核销明细" in name or "同币种" in name:
        return "核销明细"
    if "盈亏" in name:
        return "盈亏核算表"
    if "流转" in name:
        return "到账流转表"
    if "日记账" in name or "银行" in name:
        return "银行日记账"
    if name.endswith(".json") and ("夹具" in name or "fixture" in name.lower() or "取数" in name):
        return "取数夹具json"
    if "工作清单" in name or "判定" in name:
        return "产出"
    if "01_智云" in parts:
        return "智云导出(未细分)"
    if "02_我的" in parts:
        return "我的表副本(未细分)"
    return "未识别"


def peek_headers(path: Path) -> dict:
    info = {"path": str(path), "kind": path.suffix.lower()}
    try:
        if path.suffix.lower() == ".json":
            import json

            data = json.loads(path.read_text(encoding="utf-8"))
            info["json_keys"] = list(data.keys())[:20] if isinstance(data, dict) else "list"
            return info
        import openpyxl

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sheets = []
        for n in wb.sheetnames[:8]:
            sh = wb[n]
            headers = []
            for i, row in enumerate(
                sh.iter_rows(max_row=3, max_col=min(12, sh.max_column or 12), values_only=True)
            ):
                headers.append([str(c)[:24] if c is not None else "" for c in row])
            sheets.append({"name": n, "preview": headers})
        info["sheets"] = sheets
        wb.close()
    except Exception as e:
        info["error"] = str(e)
    return info


def main():
    ap = argparse.ArgumentParser(description="盘点 ar-hexiao-daily 工作区输入")
    ap.add_argument("--workspace", default=str(common.WORK), help="工作区根目录")
    ap.add_argument("--report", default="", help="运行报告输出路径")
    args = ap.parse_args()
    ws = Path(args.workspace)
    common.ensure_out_dirs()
    files = list_files(ws)
    lines = [
        f"工作区：{ws.resolve()}",
        f"发现文件数：{len(files)}",
        "",
    ]
    roles = {}
    for p in files:
        role = guess_role(p)
        roles.setdefault(role, 0)
        roles[role] += 1
        rel = p.relative_to(ws) if str(p).startswith(str(ws)) else p.name
        lines.append(f"[{role}] {rel}")
        peek = peek_headers(p)
        if "error" in peek:
            lines.append(f"  ⚠ 读失败：{peek['error']}")
        elif "json_keys" in peek:
            lines.append(f"  json keys: {peek['json_keys']}")
        elif "sheets" in peek:
            for sh in peek["sheets"][:3]:
                top = sh["preview"][0] if sh["preview"] else []
                lines.append(f"  sheet「{sh['name']}」表头预览：{top}")

    need = ["回款记录或对账/夹具", "盈亏核算表副本(判定时)", "银行日记账(步骤2)"]
    lines.append("")
    lines.append("—— 角色计数 ——")
    for k, v in sorted(roles.items()):
        lines.append(f"  {k}: {v}")
    lines.append("")
    lines.append("提示：缺文件时脚本会在对应步骤报错停下，不会猜列。")
    text = "\n".join(lines) + "\n"
    print(text)
    report = args.report or str(common.OUT_DIR / "运行报告_盘点.txt")
    Path(report).parent.mkdir(parents=True, exist_ok=True)
    Path(report).write_text(text, encoding="utf-8")
    print(f"报告已写：{report}")
    print(f"文件数={len(files)}")
    return 0


if __name__ == "__main__":
    sys.exit(main() or 0)
