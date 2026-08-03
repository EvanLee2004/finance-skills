#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
第6步：核销判定（单入口 · SOD 级 · 三栏分流）v2

2026-07-23 重写。旧版按「回款类型」把每笔到账路由到三条取数通道（对账表 / 预存明细 /
分笔），实测出了两个硬伤，都在这一版根除：

  1. **静默丢单**：预收类既被踢出对账通道、又在明细子表拿不到行，两头落空还不报错
     —— 2026-07-22 真跑丢了 3 笔到账 6.5 万，产出里一个字都没有。
     → 本版取消通道路由；每笔到账都走同一条路，且**强制 AR 覆盖率校验**，
       一笔产不出任何判定就整体报错退出，绝不静默。
  2. **把 SOD 拆行误判成"歧义"**：一个 SO 在她盈亏表占 N 行，其实是 N 个 SOD、
     每行一个金额，合计 = 本次核销额。旧版判 E8「同SO多行歧义」挂起（27 笔里 26 笔是误判）。
     → 本版引入「订单明细」表拿到逐 SOD 交付额，按 SOD 精确落行。

判定口径（明妹 2026-07-23 当面确认 + 接口实调验证）：
  · 逐 SO 本次核销额 H：有「同币种核销明细」就用它（按核销日期过滤本次）；
    **该表 0 行 = 全额核销** → H[so] = 该 SO 交付额。
  · ΣH ≤ 到账额（她的铁律：销的钱不能比到的钱多）。
  · 每个 SO 展开到 SOD：H[so] == Σ该SO的SOD交付额 → 全部 SOD 入选；
    否则在 SOD 里凑唯一子集；凑不出唯一解才挂起。
  · 盈亏表定位：(新智云单号=SO, 应收金额=该SOD交付额) → 唯一行。

红线：分笔一律 hold（禁止比例分摊）；永不写智云；本脚本不写用户任何表。
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
import common  # noqa: E402
import writeoff_duplicate_audit as WDA  # noqa: E402

TOL = 0.005  # 金额比较容差（分以下）
ROUNDING_TAIL_TOL = 0.011  # 多 SO 合计与父回款相差 1 分：作为分位尾差，不判部分回款
SUBSET_MAX_LINES = 22  # 超过这么多 SOD 就不硬凑子集，直接交人

HUIKUAN_NAMES = {
    "ar": ["回款记录ID", "回款记录编号"],
    "currency": ["原币币种", "币种"],
    "arrival_date": ["到账日期", "回款日期"],
    "amount_orig": ["到账金额/原币", "到账金额原币"],
    "amount_local": ["到账金额/本币", "到账金额本币"],
    "fee": ["手续费/原币", "手续费"],
    "huikuan_type": ["回款类型"],
    "status": ["核销状态"],
    "customer": ["开票客户", "客户名称", "客户"],
    "hexiao_date": ["核销日期"],
}


class InputError(Exception):
    """输入缺件/结构不对 —— 脚本非 0 退出，绝不带病继续。"""


class CoverageError(Exception):
    """AR 覆盖率校验没过 —— 有到账没产出任何判定，说明逻辑漏了单。"""


# ══════════════════════════════════════════════════════════════
# 一、读取 01_智云导出 的四张表 → payments
# ══════════════════════════════════════════════════════════════
def _sheet_rows(path: Path) -> Tuple[List[str], List[list]]:
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    body = [list(r) for r in rows[1:] if any(x is not None and str(x).strip() for x in r)]
    return headers, body


def _col(headers: Sequence[str], role: str, key: str, aliases: dict) -> Optional[int]:
    cands = (aliases.get(role) or {}).get(key, [key])
    return common.fuzzy_find_col(headers, cands)


def _need(headers: Sequence[str], role: str, keys: Sequence[str], aliases: dict) -> Dict[str, int]:
    out = {}
    missing = []
    for k in keys:
        i = _col(headers, role, k, aliases)
        if i is None:
            missing.append(k)
        else:
            out[k] = i
    if missing:
        raise InputError(
            f"「{role}」缺列 {missing}。实际表头：{[h for h in headers if h]}"
        )
    return out


def _get(vals: list, idx: Optional[int]) -> Any:
    if idx is None or idx >= len(vals):
        return None
    return vals[idx]


EXPORT_DATE_RE = re.compile(r"(20\d{6})")


def _export_date(path: Path) -> Optional[dt.date]:
    """从导出文件名提取 YYYYMMDD；老金标文件无日期时返回 None。"""
    m = EXPORT_DATE_RE.search(path.stem)
    if not m:
        return None
    try:
        return dt.datetime.strptime(m.group(1), "%Y%m%d").date()
    except ValueError:
        return None


def _role_files(directory: Path, *keys: str) -> List[Path]:
    return [
        p for p in sorted(directory.glob("*.xlsx"))
        if not p.name.startswith("~$") and any(k in p.name for k in keys)
    ]


def _base_export(
    files: List[Path], role: str, target_date: Optional[dt.date], required: bool = True
) -> Optional[Path]:
    """
    选当天主快照。目录里有多天文件时禁止再拿“排序第一份”冒充目标日。
    无日期文件只用于兼容单文件金标夹具。
    """
    if target_date is not None:
        exact = [p for p in files if _export_date(p) == target_date]
        if len(exact) == 1:
            return exact[0]
        if len(exact) > 1:
            raise InputError(f"「{role}」{target_date} 有多份主快照：{[p.name for p in exact]}")
        if len(files) == 1 and _export_date(files[0]) is None:
            return files[0]
        if required:
            raise InputError(f"没有「{role}」{target_date.strftime('%Y%m%d')} 的主快照")
        return None
    if len(files) == 1:
        return files[0]
    if not files:
        if required:
            raise InputError(f"01_智云导出/ 里没有「{role}」表")
        return None
    raise InputError(
        f"01_智云导出/ 里有多天「{role}」文件，必须用 --hexiao-date 指定核销日："
        f"{[p.name for p in files]}"
    )


def _eligible_snapshots(files: List[Path], target_date: Optional[dt.date], base: Optional[Path]) -> List[Path]:
    """补跑时可从更晚核销日文件还原历史子明细；无日期金标仍只读主文件。"""
    if target_date is None:
        return [base] if base else []
    out = [p for p in files if (_export_date(p) or target_date) >= target_date]
    return out or ([base] if base else [])


def find_shifted_detail_dates(workspace: Path) -> Dict[str, dict]:
    """
    找出“父回款当前筛选日 > 子表真实核销日”的历史核销明细。
    父记录发生新核销后会移动到最新核销日；本函数用于漏日补跑时还原旧日。
    """
    d = workspace / "01_智云导出"
    if not d.is_dir():
        return {}
    aliases = common.load_aliases()
    found: Dict[str, dict] = {}
    seen = set()
    for path in _role_files(d, "核销明细"):
        snapshot_date = _export_date(path)
        if snapshot_date is None:
            continue
        h, body = _sheet_rows(path)
        c = _need(h, "核销明细", ["回款记录NUM", "核销日期", "本次核销金额"], aliases)
        i_so = _col(h, "核销明细", "SO", aliases)
        for vals in body:
            ar = str(_get(vals, c["回款记录NUM"]) or "").strip()
            hd = common.norm_date(_get(vals, c["核销日期"]))
            so = str(_get(vals, i_so) or "").strip()
            amt = common.to_number(_get(vals, c["本次核销金额"]))
            if not ar or not so or amt is None or hd is None or hd >= snapshot_date:
                continue
            key = (hd, ar, so, round(float(amt), 2))
            if key in seen:
                continue
            seen.add(key)
            day = hd.isoformat()
            item = found.setdefault(
                day, {"rows": 0, "ars": set(), "sos": set(), "order_keys": set(), "sources": set()}
            )
            item["rows"] += 1
            item["ars"].add(ar)
            item["sos"].add(so)
            item["order_keys"].add(f"{ar}|{so}")
            item["sources"].add(path.name)
    return {
        day: {
            "rows": item["rows"],
            "ars": sorted(item["ars"]),
            "sos": sorted(item["sos"]),
            "order_keys": sorted(item["order_keys"]),
            "sources": sorted(item["sources"]),
        }
        for day, item in sorted(found.items())
    }


def assess_shifted_detail_dates(
    workspace: Path,
    date_from: Optional[dt.date] = None,
    date_to: Optional[dt.date] = None,
) -> Dict[str, dict]:
    """
    将被最新核销日挪走的历史子明细与该日判定结果对账；缺 AR/SO 才标记重跑。
    旧结果没有来源覆盖字段也能按 auto/hold/exception 三栏逐键检查。
    """
    found = find_shifted_detail_dates(workspace)
    out: Dict[str, dict] = {}
    for day, info in found.items():
        parsed_day = common.norm_date(day)
        if date_from and parsed_day and parsed_day < date_from:
            continue
        if date_to and parsed_day and parsed_day > date_to:
            continue
        stamp = day.replace("-", "")
        result_path = workspace / "04_产出" / f"判定结果_{stamp}.json"
        produced = set()
        if result_path.is_file():
            try:
                data = json.loads(result_path.read_text(encoding="utf-8"))
                for bucket in ("auto", "hold", "exception"):
                    for item in data.get(bucket) or []:
                        ar = str(item.get("ar") or "").strip()
                        so = str(item.get("so") or "").strip()
                        if ar and so:
                            produced.add(f"{ar}|{so}")
            except (OSError, ValueError, TypeError):
                produced = set()
        missing = sorted(set(info["order_keys"]) - produced)
        out[day] = {
            **info,
            "result": str(result_path) if result_path.is_file() else "",
            "missing_order_keys": missing,
            "needs_rerun": bool(missing),
        }
    return out


def reconcile_writeoff_details(
    payments: List[dict],
    parent_references: Dict[str, dict],
    raw_detail_rows: List[dict],
    target_date: Optional[dt.date],
) -> Tuple[List[dict], Dict[str, dict]]:
    """先按父AR纠正物理/系统重复，再形成目标日H和跨父AR累计R。"""
    raw_by_ar: Dict[str, List[dict]] = {}
    for row in raw_detail_rows:
        raw_by_ar.setdefault(row["ar"], []).append(row)

    audits: Dict[str, dict] = {}
    logical_rows: List[dict] = []
    unresolved_sos: Dict[str, List[str]] = {}
    for ar in sorted(set(parent_references) | set(raw_by_ar)):
        parent = parent_references.get(ar) or {
            "ar": ar, "amount_orig": None, "amount_local": None, "currency": ""
        }
        logical, audit = WDA.audit_parent_writeoffs(parent, raw_by_ar.get(ar, []))
        audit["ar"] = ar
        audits[ar] = audit
        if audit["status"] == "unresolved":
            for so in {
                str(row.get("so") or "").strip()
                for row in raw_by_ar.get(ar, [])
                if str(row.get("so") or "").strip()
            }:
                unresolved_sos.setdefault(so, []).append(ar)
            continue
        logical_rows.extend(logical)

    by_ar = {p["ar"]: p for p in payments}
    for p in payments:
        audit = audits.get(p["ar"]) or {
            "ar": p["ar"], "status": "normal", "comparison_basis": "unavailable",
            "raw_input_count": 0, "raw_record_count": 0, "logical_record_count": 0,
            "raw_total": 0.0, "logical_total": 0.0,
            "delta_raw": p.get("amount_local") or p.get("amount_orig"),
            "delta_dedup": p.get("amount_local") or p.get("amount_orig"),
            "duplicate_groups": [], "records": [], "ignored_record_count": 0,
            "physical_snapshot_duplicate_count": 0, "revoked_count": 0,
            "reason": "没有逐SO核销明细，沿用现有全额核销语义",
        }
        p["duplicate_writeoff_audit"] = audit
        p["writeoffs"] = {}
        p["writeoffs_local"] = {}
        p["cumulative_writeoffs"] = {}
        p["cumulative_writeoffs_local"] = {}
        p["_source_meta"]["raw_writeoff_rows"] = int(audit.get("raw_input_count") or 0)
        p["_source_meta"]["accounted_writeoff_rows"] = len(audit.get("records") or [])
        if audit.get("status") == "unresolved":
            p["_system_over_writeoff_unresolved"] = audit.get("reason") or "系统超核销无法解释"

    current_rows: List[dict] = []
    for item in logical_rows:
        p = by_ar.get(item["ar"])
        current_day = target_date if target_date is not None else (p or {}).get("hexiao_date")
        if p is not None and (current_day is None or item.get("date") == current_day):
            current_rows.append(item)
            w = p["writeoffs"]
            w[item["so"]] = round(w.get(item["so"], 0.0) + float(item["amount"]), 2)
            if item.get("amount_local") is not None:
                wl = p["writeoffs_local"]
                wl[item["so"]] = round(
                    wl.get(item["so"], 0.0) + float(item["amount_local"]), 2
                )
            if item.get("snapshot_date") and target_date and item["snapshot_date"] > target_date:
                p["_source_meta"]["historical_detail_rows"] += 1

    global_cumulative: Dict[str, float] = {}
    global_cumulative_local: Dict[str, float] = {}
    for item in logical_rows:
        so = item["so"]
        global_cumulative[so] = round(
            global_cumulative.get(so, 0.0) + float(item["amount"]), 2
        )
        if item.get("amount_local") is not None:
            global_cumulative_local[so] = round(
                global_cumulative_local.get(so, 0.0) + float(item["amount_local"]), 2
            )
    for p in payments:
        p_sos = {
            str(item.get("so") or "").strip()
            for item in raw_by_ar.get(p["ar"], [])
            if str(item.get("so") or "").strip()
        }
        inherited = sorted({ar for so in p_sos for ar in unresolved_sos.get(so, [])})
        if inherited and not p.get("_system_over_writeoff_unresolved"):
            p["_system_over_writeoff_unresolved"] = (
                "同一SO的跨父AR历史核销存在未解决超核销，累计回款不可安全计算："
                + ",".join(inherited)
            )
        for so in p_sos:
            if so in global_cumulative:
                p["cumulative_writeoffs"][so] = global_cumulative[so]
            if so in global_cumulative_local:
                p["cumulative_writeoffs_local"][so] = global_cumulative_local[so]
    return current_rows, audits


def load_exports(workspace: Path, target_date: Optional[dt.date] = None) -> List[dict]:
    """
    01_智云导出/ 四张表 → payments。缺件直接 InputError（不凑合、不猜）。

    文件按名字认，并按 target_date 选主快照。后续快照里真实核销日仍等于
    target_date 的子明细会自动回捞；其父回款和关联订单也一并补入本批。
    """
    d = workspace / "01_智云导出"
    if not d.is_dir():
        raise InputError(f"没有 {d}")
    aliases = common.load_aliases()

    hk_files = _role_files(d, "回款记录")
    xd_files = _role_files(d, "订单交付", "下单")
    mx_files = _role_files(d, "核销明细")
    sod_files = _role_files(d, "订单明细")
    p_hk = _base_export(hk_files, "回款记录", target_date, required=target_date is None)
    p_xd = _base_export(xd_files, "订单交付", target_date, required=target_date is None)
    p_mx = _base_export(mx_files, "核销明细", target_date, required=False)
    p_sod = _base_export(sod_files, "订单明细", target_date, required=False)

    if not p_hk and target_date is None:
        raise InputError(
            "01_智云导出/ 里没有「回款记录」表。\n"
            "  这一版取数只认单入口：先跑 fetch_zhiyun.py（回款记录按核销日期=T-1 + 关联子表），\n"
            "  它会一次产出 回款记录 / 订单交付 / 核销明细 / 订单明细 四份。"
        )
    if not p_xd and target_date is None:
        raise InputError(
            "01_智云导出/ 里没有「订单交付」表（每笔到账关联了哪几个 SO + 交付额）。\n"
            "  旧版的「回款核销对账」已废弃且不兼容——它拿不到 SOD、还会漏掉预收类。\n"
            "  请重新跑 fetch_zhiyun.py 取一次数。"
        )

    # ① 读取全部原始核销记录。不同核销记录NUM必须先保留；同一NUM跨快照
    #    的物理去重和明显超核销后的条件性业务纠正在父AR层统一完成。
    raw_detail_rows: List[dict] = []
    for path in _eligible_snapshots(mx_files, target_date, p_mx):
        h, body = _sheet_rows(path)
        c = _need(h, "核销明细", ["回款记录NUM", "核销日期", "本次核销金额"], aliases)
        i_record = _col(h, "核销明细", "核销记录NUM", aliases)
        i_rowid = _col(h, "核销明细", "rowid", aliases)
        i_so = _col(h, "核销明细", "SO", aliases)
        i_local = _col(h, "核销明细", "本次核销金额本币", aliases)
        i_currency = _col(h, "核销明细", "币种", aliases)
        i_rate = _col(h, "核销明细", "汇率", aliases)
        i_revoked = _col(h, "核销明细", "是否已撤销", aliases)
        snapshot_date = _export_date(path)
        for row_number, vals in enumerate(body, start=2):
            ar = str(_get(vals, c["回款记录NUM"]) or "").strip()
            hd = common.norm_date(_get(vals, c["核销日期"]))
            so = str(_get(vals, i_so) or "").strip()
            amt = common.to_number(_get(vals, c["本次核销金额"]))
            amt_local = common.to_number(_get(vals, i_local))
            if not ar or not so or amt is None:
                raise InputError(
                    f"核销明细原始行缺AR/SO/金额：{path.name} 第{row_number}行，禁止静默丢弃"
                )
            if target_date is not None and hd is not None and hd > target_date:
                continue
            raw_detail_rows.append({
                "record_id": str(_get(vals, i_record) or "").strip(),
                "rowid": str(_get(vals, i_rowid) or "").strip(),
                "ar": ar, "date": hd, "so": so, "amount": round(float(amt), 2),
                "amount_local": round(float(amt_local), 2) if amt_local is not None else None,
                "currency": str(_get(vals, i_currency) or "").strip(),
                "rate": common.to_number(_get(vals, i_rate)),
                "revoked": str(_get(vals, i_revoked) or "").strip(),
                "source": path.name, "snapshot_date": snapshot_date,
                "_input_index": len(raw_detail_rows),
            })
    detail_ars = {
        x["ar"] for x in raw_detail_rows
        if target_date is None or x.get("date") == target_date
    }
    if p_hk is None and not detail_ars:
        raise InputError(
            f"没有目标日 {target_date} 的回款主快照，也没有后续快照中的同日核销子明细"
        )

    # ② 回款记录：当天父记录 + 后续核销日文件中承载历史同日子明细的父记录。
    payments: List[dict] = []
    parent_references: Dict[str, dict] = {}
    seen_payments = set()
    for path in _eligible_snapshots(hk_files, target_date, p_hk):
        h, body = _sheet_rows(path)
        c = _need(h, "回款记录", ["AR", "核销日期"], aliases)
        for k in ["到账日期", "到账金额原币", "到账金额本币", "手续费", "原币币种",
                  "回款类型", "核销状态", "开票客户"]:
            i = _col(h, "回款记录", k, aliases)
            if i is not None:
                c[k] = i
        snapshot_date = _export_date(path)
        for vals in body:
            ar = str(_get(vals, c["AR"]) or "").strip()
            parent_date = common.norm_date(_get(vals, c["核销日期"]))
            if not ar:
                continue
            info = {
                "ar": ar,
                "hexiao_date": target_date or parent_date,
                "parent_hexiao_date": parent_date,
                "arrival_date": common.norm_date(_get(vals, c.get("到账日期"))),
                "amount_orig": common.to_number(_get(vals, c.get("到账金额原币"))),
                "amount_local": common.to_number(_get(vals, c.get("到账金额本币"))),
                "fee": common.to_number(_get(vals, c.get("手续费"))) or 0.0,
                "currency": str(_get(vals, c.get("原币币种")) or "").strip() or "人民币CNY",
                "huikuan_type": str(_get(vals, c.get("回款类型")) or "").strip(),
                "status": str(_get(vals, c.get("核销状态")) or "").strip(),
                "customer": str(_get(vals, c.get("开票客户")) or "").strip(),
                "orders": [],
                "writeoffs": {},
                "writeoffs_local": {},
                "cumulative_writeoffs": {},
                "cumulative_writeoffs_local": {},
                "_source_meta": {
                    "payment_source": path.name,
                    "payment_snapshot_date": snapshot_date,
                    "historical_detail_rows": 0,
                    "recovered_deliveries": 0,
                },
            }
            parent_references[ar] = info
            if ar in seen_payments:
                continue
            if target_date is not None and parent_date != target_date and ar not in detail_ars:
                continue
            seen_payments.add(ar)
            payments.append(info)
    if not payments:
        raise InputError(
            f"目标核销日没有可处理的回款记录：{target_date or (p_hk.name if p_hk else '')}"
        )
    by_ar = {p["ar"]: p for p in payments}
    detail_rows, duplicate_audits = reconcile_writeoff_details(
        payments, parent_references, raw_detail_rows, target_date
    )

    # ③ 订单交付：按快照时间更新，同一 AR/SO 取最新非空交付额。
    order_map: Dict[Tuple[str, str], dict] = {}
    for path in _eligible_snapshots(xd_files, target_date, p_xd):
        h, body = _sheet_rows(path)
        c = _need(h, "订单交付", ["AR", "SO", "交付额原币"], aliases)
        i_rate = _col(h, "订单交付", "汇率", aliases)
        i_cur = _col(h, "订单交付", "币种", aliases)
        i_name = _col(h, "订单交付", "订单名称", aliases)
        for vals in body:
            ar = str(_get(vals, c["AR"]) or "").strip()
            so = str(_get(vals, c["SO"]) or "").strip()
            if not ar or not so or ar not in by_ar:
                continue
            key = (ar, so)
            amt = common.to_number(_get(vals, c["交付额原币"]))
            old = order_map.get(key)
            if old is None or amt is not None:
                order_map[key] = {
                    "so": so, "deliver": amt,
                    "rate": common.to_number(_get(vals, i_rate)),
                    "currency": str(_get(vals, i_cur) or "").strip(),
                    "name": str(_get(vals, i_name) or "").strip(),
                }

    # ④ 逻辑核销明细：物理跨快照重复和系统重复纠正均已完成。
    for item in detail_rows:
        if item["ar"] not in by_ar:
            raise CoverageError(
                f"核销明细有父回款未取到：{item['ar']} / {item['so']} / {item['source']}"
            )
        p = by_ar[item["ar"]]
        if target_date is None and item["date"] is not None and p.get("hexiao_date") is not None:
            if item["date"] != p["hexiao_date"]:
                continue
        if (item["ar"], item["so"]) not in order_map:
            order_map[(item["ar"], item["so"])] = {
                "so": item["so"], "deliver": None, "rate": None,
                "currency": "", "name": "",
            }
    for (ar, _so), order in order_map.items():
        by_ar[ar]["orders"].append(order)

    # ⑤ 订单明细（SO → SOD + 逐 SOD 交付额）；同一 SOD 取后续快照最新非空值。
    sod_map: Dict[Tuple[str, str], dict] = {}
    needed_sos = {o["so"] for p in payments for o in p["orders"]}
    for path in _eligible_snapshots(sod_files, target_date, p_sod):
        h, body = _sheet_rows(path)
        c = _need(h, "订单明细", ["SO", "SOD", "交付额原币"], aliases)
        for vals in body:
            so = str(_get(vals, c["SO"]) or "").strip()
            sod = str(_get(vals, c["SOD"]) or "").strip()
            if not so or not sod or so not in needed_sos:
                continue
            amt = common.to_number(_get(vals, c["交付额原币"]))
            key = (so, sod)
            if key not in sod_map or amt is not None:
                sod_map[key] = {"sod": sod, "deliver": amt, "source": path.name}
    sod_lines: Dict[str, List[dict]] = {}
    for (so, _sod), line in sod_map.items():
        sod_lines.setdefault(so, []).append(line)

    # 订单交付额为空时，只允许用该 SO 全部、唯一且非空的 SOD 交付额合计回补。
    for p in payments:
        for order in p["orders"]:
            if order.get("deliver") is not None:
                continue
            lines = sod_lines.get(order["so"]) or []
            if lines and all(x.get("deliver") is not None for x in lines):
                total = round(sum(float(x["deliver"]) for x in lines), 2)
                if total > TOL:
                    order["deliver"] = total
                    order["delivery_source"] = "订单明细SOD合计"
                    p["_source_meta"]["recovered_deliveries"] += 1
    for p in payments:
        p["sod_lines"] = sod_lines
        p["_duplicate_writeoff_audits"] = duplicate_audits
    return payments


# ══════════════════════════════════════════════════════════════
# 二、payments → SOD 级 records（核心展开）
# ══════════════════════════════════════════════════════════════
def subset_sum_unique(
    lines: List[dict], target: float, max_lines: int = SUBSET_MAX_LINES
) -> Optional[List[dict]]:
    """
    在 SOD 行里凑出金额恰好 = target 的**唯一**子集。
    多解 / 无解 / 行太多 → None（交给人，绝不随便挑一个）。
    """
    usable = [x for x in lines if x.get("deliver") is not None]
    if not usable or len(usable) > max_lines:
        return None
    cents = [int(round(float(x["deliver"]) * 100)) for x in usable]
    tgt = int(round(target * 100))
    if tgt <= 0:
        return None
    # ways[sum] = (方案数(封顶2), 一个见证掩码)
    ways: Dict[int, Tuple[int, int]] = {0: (1, 0)}
    for i, v in enumerate(cents):
        if v <= 0:
            continue
        nxt = dict(ways)
        for s, (n, mask) in ways.items():
            s2 = s + v
            if s2 > tgt:
                continue
            n0, m0 = nxt.get(s2, (0, 0))
            nxt[s2] = (min(n0 + n, 2), m0 if n0 else (mask | (1 << i)))
        ways = nxt
    n, mask = ways.get(tgt, (0, 0))
    if n != 1:
        return None
    return [usable[i] for i in range(len(usable)) if mask & (1 << i)]


def _payment_local(p: dict, rates: Dict[str, float]) -> Tuple[Optional[float], Optional[str]]:
    """到账本币：系统给了就用系统的，没给才按汇率算。"""
    if p.get("amount_local") is not None:
        return float(p["amount_local"]), None
    return common.compute_local_amount(p.get("amount_orig"), p.get("currency") or "", rates)


def _localize_amount(
    amount_orig: Optional[float],
    p: dict,
    rates: Dict[str, float],
    *,
    explicit_local: Optional[float] = None,
    explicit_orig: Optional[float] = None,
    row_rate: Optional[float] = None,
) -> Tuple[Optional[float], Optional[str]]:
    """
    把订单/SOD原币金额换成本币。优先级：
    子核销本币 → 订单汇率 → 父回款原/本币反算 → 命令行汇率。

    父/子同时给出原币和本币时，反算只用于同币种SOD的确定性换算，
    不是在多个SO之间猜分摊。
    """
    if amount_orig is None:
        return None, "E7"
    amount = float(amount_orig)
    currency = p.get("currency") or ""
    if common.is_cny(currency):
        return round(amount, 2), None
    if explicit_local is not None:
        base_orig = common.to_number(explicit_orig)
        if base_orig is not None and abs(float(base_orig)) > TOL:
            return round(amount * float(explicit_local) / float(base_orig), 2), None
        if abs(amount) <= TOL:
            return 0.0, None
    rate = common.to_number(row_rate)
    if rate is not None and float(rate) > 0:
        return round(amount * float(rate), 2), None
    parent_orig = common.to_number(p.get("amount_orig"))
    parent_local = common.to_number(p.get("amount_local"))
    if (
        parent_orig is not None
        and parent_local is not None
        and abs(float(parent_orig)) > TOL
    ):
        return round(amount * float(parent_local) / float(parent_orig), 2), None
    return common.compute_local_amount(amount, currency, rates)


def _hold(p: dict, code: str, reason: str, so: str = "", sod: str = "", **extra) -> dict:
    rec = {
        "ar": p["ar"], "so": so, "sod": sod,
        "customer": p.get("customer") or "",
        "amount_orig": None,
        "currency": p.get("currency") or "人民币CNY",
        "hexiao_date": p.get("hexiao_date"),
        "shoukuan_date": p.get("arrival_date"),
        "status": p.get("status") or "",
        "huikuan_type": p.get("huikuan_type") or "",
        "fee": p.get("fee") or 0.0,
        "arrival_total": p.get("amount_orig"),
        "forced_code": code,
        "forced_reason": reason,
    }
    rec.update(extra)
    return rec


def _hold_each_source_order(p: dict, code: str, reason: str) -> List[dict]:
    """付款级卡点也按来源 SO 展示；有多个订单时不再只留一条 AR 总挂账。"""
    sos = set((p.get("writeoffs") or {}).keys())
    if not sos:
        sos = {str(o.get("so") or "").strip() for o in (p.get("orders") or [])}
    sos.discard("")
    if not sos:
        return [_hold(p, code, reason)]
    return [_hold(p, code, reason, so=so) for so in sorted(sos)]


def partial_split_guidance(
    latest_delivery: float,
    current_received: float,
    *,
    initial_receivable: Optional[float] = None,
    existing_received: Optional[float] = None,
) -> str:
    """按 2026-07-29 定稿口径生成部分回款拆行说明。"""
    latest = round(float(latest_delivery), 2)
    current = round(float(current_received), 2)
    existing = round(float(existing_received or 0.0), 2)
    cumulative = round(existing + current, 2)
    remaining = round(latest - cumulative, 2)

    if remaining < -TOL:
        balance = (
            f"表内历史回款 {existing:.2f} + 本次 {current:.2f} = {cumulative:.2f}，"
            f"已超过智云最新交付 {latest:.2f}；交付额下降/超收口径未定，先人工核对，禁止自动写。"
        )
    elif remaining <= TOL:
        balance = (
            f"表内历史回款 {existing:.2f} + 本次 {current:.2f} = {cumulative:.2f}，"
            f"已回满智云最新交付 {latest:.2f}。"
        )
    else:
        balance = (
            f"智云最新实际交付 {latest:.2f}，表内历史回款 {existing:.2f}，"
            f"本次实际回款 {current:.2f}；本次后累计回款 {cumulative:.2f}，"
            f"实际未收 = 最新交付 − 累计实际回款 = {remaining:.2f}。"
        )

    if initial_receivable is None:
        baseline = (
            "盈亏表原始应收基线不能被最新交付额覆盖：先汇总这张单拆行前的原始应收，"
            "未收行应收填实际未收；已收侧应收合计 = 原始应收合计 − 实际未收，"
            "保证拆行后应收合计仍等于原始基线。"
        )
    else:
        initial = round(float(initial_receivable), 2)
        paid_side = round(initial - max(remaining, 0.0), 2)
        baseline = (
            f"这张单盈亏表原始应收合计 {initial:.2f} 必须保持不变："
            f"未收行应收填 {max(remaining, 0.0):.2f}，"
            f"已收侧应收合计填 {paid_side:.2f}，两边仍合计 {initial:.2f}。"
        )

    return (
        balance
        + baseline
        + f" 本次已收行回款明细填 {current:.2f}、结账填「是」、日期和方式照到账/核销月份规则；"
        "其下新增未收行，结账填「否」，回款明细、日期、方式留空。"
        f" 累计回款未达到最新交付 {latest:.2f} 前，两行计提都留空；"
        f"只有累计回款达到 {latest:.2f} 时，最后结清行计提才填 {latest:.2f}。"
    )


def _writeoff_business_amount(
    amount: Optional[float],
    *,
    explicit_local: Optional[float] = None,
    explicit_orig: Optional[float] = None,
) -> Tuple[Optional[float], Optional[str]]:
    """
    智云核销业务金额口径。

    有“本次核销金额本币”时按同一子核销金额的原/本币比例落到当前 SOD；
    没有本币列时直接使用“本次核销金额”，不再反向要求父回款汇率，也不读取、
    扣减或分配手续费。调用方必须保证金额来自同一 SO 的核销明细或最新交付额。
    """
    value = common.to_number(amount)
    if value is None:
        return None, "E7"
    local = common.to_number(explicit_local)
    original = common.to_number(explicit_orig)
    if local is not None and original is not None and abs(float(original)) > TOL:
        return round(float(value) * float(local) / float(original), 2), None
    return round(float(value), 2), None


def _order_delivery_local(
    amount_orig: Optional[float], p: dict, rates: Dict[str, float], order: Optional[dict],
    *, explicit_local: Optional[float] = None, explicit_orig: Optional[float] = None,
) -> Tuple[Optional[float], Optional[str]]:
    """盈亏表金额统一为人民币：订单/SOD交付原币必须按订单币种和汇率换算。"""
    order = order or {}
    money_p = dict(p)
    money_p["currency"] = order.get("currency") or p.get("currency") or ""
    return _localize_amount(
        amount_orig, money_p, rates,
        explicit_local=explicit_local,
        explicit_orig=explicit_orig,
        row_rate=order.get("rate"),
    )


def expand_payment(p: dict, rates: Dict[str, float]) -> List[dict]:
    """一笔到账 → 若干 SOD 级 record。**任何情况下至少产出一条**（不许静默丢）。"""
    sod_lines: Dict[str, List[dict]] = p.get("sod_lines") or {}
    orders = p.get("orders") or []
    duplicate_audit = p.get("duplicate_writeoff_audit") or {}

    def finish(items: List[dict]) -> List[dict]:
        if duplicate_audit:
            for item in items:
                item["duplicate_writeoff_audit"] = duplicate_audit
                if duplicate_audit.get("status") == "recovered":
                    item.setdefault("warning_codes", []).append(
                        "W_SYSTEM_DUPLICATE_WRITEOFF_COLLAPSED"
                    )
        return items

    unresolved = p.get("_system_over_writeoff_unresolved")
    if unresolved:
        return finish(_hold_each_source_order(
            p,
            "E_SYSTEM_OVER_WRITEOFF_UNRESOLVED",
            f"智云核销记录需人工检查：{unresolved}",
        ))

    # 分笔回款只有在智云确实缺少逐 SO 核销金额时才 hold；已有逐 SO 明细就按明细继续判。
    if "分笔" in (p.get("huikuan_type") or "") and not (p.get("writeoffs") or {}):
        return finish(_hold_each_source_order(p, "E1", "分笔回款：系统无逐单金额，禁止比例分摊"))

    if not orders:
        return finish([_hold(
            p, "E7",
            "这笔到账在智云没关联任何下单（下单栏为空）——先去智云看看这笔回款建对了没",
        )])

    # 逐 SO 本次核销额是业务真相源。手续费字段完全不参与判定、不扣减、不分配。
    # 核销明细无行仍沿用智云业务语义：该父回款关联订单为全额核销。
    writeoffs = dict(p.get("writeoffs") or {})
    has_itemized_writeoff = bool(writeoffs)
    if writeoffs:
        H = writeoffs
        basis = "智云逐SO本次核销金额(手续费忽略)"
    else:
        H = {}
        missing_delivery_sos = sorted({
            str(o.get("so") or "").strip() for o in orders
            if o.get("deliver") is None and o.get("so")
        })
        if missing_delivery_sos:
            return finish([
                _hold(
                    p, "E7",
                    (
                        f"下单 {so} 没有交付额，且 SOD 金额不完整，判不了"
                        if so in missing_delivery_sos
                        else f"同笔回款含无交付额订单 {','.join(missing_delivery_sos)}，整笔先挂起"
                    ),
                    so=so,
                )
                for so in sorted({str(o.get("so") or "").strip() for o in orders if o.get("so")})
            ])
        for o in orders:
            H[o["so"]] = round(H.get(o["so"], 0.0) + float(o["deliver"]), 2)
        basis = "智云无逐SO明细=全额核销(手续费忽略)"

    out: List[dict] = []
    deliver_by_so: Dict[str, float] = {}
    order_by_so: Dict[str, dict] = {}
    for o in orders:
        if o.get("so"):
            order_by_so[o["so"]] = o
        if o.get("deliver") is not None:
            deliver_by_so[o["so"]] = round(deliver_by_so.get(o["so"], 0.0) + float(o["deliver"]), 2)

    for so, h in H.items():
        lines = sod_lines.get(so) or []
        chosen: Optional[List[dict]] = None
        how = ""
        if lines and all(x.get("deliver") is not None for x in lines):
            total_lines = round(sum(float(x["deliver"]) for x in lines), 2)
            if abs(total_lines - h) <= TOL:
                chosen, how = lines, "全部SOD"
            elif len(lines) == 1 and float(h) < total_lines - TOL:
                # 唯一SOD的本次核销可以只是最终/中间一笔；是否结清看截至目标日累计核销。
                chosen, how = lines, "唯一SOD累计核销"
            else:
                chosen = subset_sum_unique(lines, h)
                how = "SOD子集" if chosen else ""
        if chosen is None:
            if lines:
                cand = "、".join(
                    f"{x['sod']}={x['deliver']}" for x in lines[:12]
                ) + ("…" if len(lines) > 12 else "")
                order = order_by_so.get(so) or {}
                current_local = None
                if has_itemized_writeoff:
                    current_local, _ = _writeoff_business_amount(
                        h,
                        explicit_local=(p.get("writeoffs_local") or {}).get(so),
                        explicit_orig=h,
                    )
                else:
                    current_local, _ = _order_delivery_local(h, p, rates, order)
                default_lines = []
                for one_line in lines:
                    one = dict(one_line)
                    if has_itemized_writeoff:
                        one["deliver_local"], _ = _writeoff_business_amount(
                            one.get("deliver"),
                            explicit_local=(p.get("writeoffs_local") or {}).get(so),
                            explicit_orig=h,
                        )
                    else:
                        one["deliver_local"], _ = _order_delivery_local(
                            one.get("deliver"), p, rates, order
                        )
                    default_lines.append(one)
                out.append(_hold(
                    p, "E5",
                    f"{so} 本次核销 {h:.2f}，但它下面的 SOD 金额凑不出唯一组合"
                    f"（SOD 合计 {round(sum(float(x['deliver']) for x in lines), 2)}）。"
                    f"候选：{cand}。你指一下这次核的是哪几个 SOD；"
                    f"若你指的那个 SOD 交付额比 {h:.2f} 大，就是只回了一部分——"
                    "实际未收必须按「智云最新交付额 − 累计实际回款」算，不能拿盈亏表旧应收减；"
                    "回款明细填实际回款，计提目标更新为最新交付额，同时保持拆行后的原始应收合计不变。"
                    "按已确认规则，程序将尝试选择盈亏表中该 SO 的首个未结清 SOD；"
                    "若该行不能安全承接则继续挂账。",
                    so=so,
                    default_first_sod=True,
                    default_amount_orig=float(h),
                    default_amount_local=current_local,
                    default_sod_lines=default_lines,
                    default_match_basis=basis,
                    warning_codes=["W_DEFAULT_FIRST_SOD"],
                ))
            else:
                # 订单明细查不到 SOD → 退化成按 SO 匹配盈亏表（老路，仍可判）
                order = order_by_so.get(so) or {}
                if has_itemized_writeoff:
                    current_local, _ = _writeoff_business_amount(
                        h,
                        explicit_local=(p.get("writeoffs_local") or {}).get(so),
                        explicit_orig=h,
                    )
                else:
                    current_local, _ = _order_delivery_local(
                        h, p, rates, order,
                        explicit_local=(p.get("writeoffs_local") or {}).get(so),
                        explicit_orig=h,
                    )
                deliver_orig = deliver_by_so.get(so)
                if has_itemized_writeoff:
                    deliver_local, _ = _writeoff_business_amount(deliver_orig)
                else:
                    deliver_local, _ = _order_delivery_local(deliver_orig, p, rates, order)
                cumulative_orig = (p.get("cumulative_writeoffs") or {}).get(so)
                cumulative_local = None
                if cumulative_orig is not None:
                    if has_itemized_writeoff:
                        cumulative_local, _ = _writeoff_business_amount(
                            cumulative_orig,
                            explicit_local=(p.get("cumulative_writeoffs_local") or {}).get(so),
                            explicit_orig=cumulative_orig,
                        )
                    else:
                        cumulative_local, _ = _order_delivery_local(
                            cumulative_orig, p, rates, order,
                            explicit_local=(p.get("cumulative_writeoffs_local") or {}).get(so),
                            explicit_orig=cumulative_orig,
                        )
                out.append({
                    "ar": p["ar"], "so": so, "sod": "",
                    "customer": p.get("customer") or "",
                    "amount_orig": h,
                    "amount_local": current_local,
                    "currency": order.get("currency") or p.get("currency") or "人民币CNY",
                    "hexiao_date": p.get("hexiao_date"),
                    "shoukuan_date": p.get("arrival_date"),
                    "status": p.get("status") or "",
                    "huikuan_type": p.get("huikuan_type") or "",
                    "fee": 0.0,
                    "arrival_total": p.get("amount_orig"),
                    "deliver_local": deliver_local,
                    "cumulative_received_local": cumulative_local,
                    "match_basis": f"{basis}/无SOD(按SO匹配)",
                })
            continue
        for line in chosen:
            order = order_by_so.get(so) or {}
            line_orig = float(line["deliver"])
            unique_partial = (
                len(lines) == 1
                and len(chosen) == 1
                and float(h) < line_orig - TOL
            )
            current_orig = float(h) if unique_partial else line_orig
            if has_itemized_writeoff:
                current_local, _ = _writeoff_business_amount(
                    current_orig,
                    explicit_local=(p.get("writeoffs_local") or {}).get(so),
                    explicit_orig=h,
                )
                deliver_local, _ = _writeoff_business_amount(
                    line_orig,
                    explicit_local=(p.get("writeoffs_local") or {}).get(so),
                    explicit_orig=h,
                )
            else:
                current_local, _ = _order_delivery_local(
                    current_orig, p, rates, order,
                    explicit_local=(p.get("writeoffs_local") or {}).get(so),
                    explicit_orig=h,
                )
                deliver_local, _ = _order_delivery_local(
                    line_orig, p, rates, order,
                    explicit_local=(p.get("writeoffs_local") or {}).get(so),
                    explicit_orig=h,
                )
            cumulative_local = None
            cumulative_orig = (p.get("cumulative_writeoffs") or {}).get(so)
            if len(lines) == 1 and cumulative_orig is not None:
                if has_itemized_writeoff:
                    cumulative_local, _ = _writeoff_business_amount(
                        cumulative_orig,
                        explicit_local=(p.get("cumulative_writeoffs_local") or {}).get(so),
                        explicit_orig=cumulative_orig,
                    )
                else:
                    cumulative_local, _ = _order_delivery_local(
                        cumulative_orig, p, rates, order,
                        explicit_local=(p.get("cumulative_writeoffs_local") or {}).get(so),
                        explicit_orig=cumulative_orig,
                    )
            out.append({
                "so_all_lines": lines,  # 整段对齐消歧要用（见 LedgerIndex.positional_row）
                "ar": p["ar"], "so": so, "sod": line["sod"],
                "customer": p.get("customer") or "",
                "amount_orig": current_orig,
                "amount_local": current_local,
                "currency": (
                    line.get("currency")
                    or order.get("currency")
                    or p.get("currency")
                    or "人民币CNY"
                ),
                "hexiao_date": p.get("hexiao_date"),
                "shoukuan_date": p.get("arrival_date"),
                "status": p.get("status") or "",
                "huikuan_type": p.get("huikuan_type") or "",
                "fee": 0.0,
                "arrival_total": p.get("amount_orig"),
                "deliver_local": deliver_local,
                "cumulative_received_local": cumulative_local,
                "match_basis": f"{basis}/{how}",
            })

    if not out:  # 兜底：绝不静默丢单
        out.append(_hold(p, "E7", "这笔到账展开不出任何订单行（请把这条截图发给明昊）"))
    return finish(out)


def source_coverage(payments: List[dict], records: List[dict]) -> dict:
    """
    检查输入侧每个 AR/SO 是否都进入判定（自动、挂账或异常均算“进入”）。
    这道闸覆盖“整笔 AR 在，但其中几个订单没进计划”的漏单。
    """
    expected = set()
    for p in payments:
        sos = set((p.get("writeoffs") or {}).keys())
        if not sos:
            sos = {str(o.get("so") or "").strip() for o in (p.get("orders") or [])}
        expected.update((p.get("ar"), so) for so in sos if p.get("ar") and so)

    produced = {
        (r.get("ar"), str(r.get("so") or "").strip())
        for r in records if r.get("ar") and r.get("so")
    }
    missing = sorted(expected - produced)
    if missing:
        raise CoverageError(
            "有智云订单没有进入任何判定行（即使同一 AR 的其他订单已进入也不放行）："
            f"{[f'{ar}|{so}' for ar, so in missing]}"
        )

    historical_rows = sum(
        int((p.get("_source_meta") or {}).get("historical_detail_rows") or 0)
        for p in payments
    )
    recovered = sum(
        int((p.get("_source_meta") or {}).get("recovered_deliveries") or 0)
        for p in payments
    )
    all_audits = next(
        (p.get("_duplicate_writeoff_audits") for p in payments if p.get("_duplicate_writeoff_audits")),
        {},
    )
    raw_writeoff_rows = sum(
        int(audit.get("raw_input_count") or 0) for audit in all_audits.values()
    )
    accounted_writeoff_rows = sum(
        len(audit.get("records") or []) for audit in all_audits.values()
    )
    if raw_writeoff_rows != accounted_writeoff_rows:
        raise CoverageError(
            "原始核销记录处置覆盖不完整："
            f"读取{raw_writeoff_rows}条，只有{accounted_writeoff_rows}条有明确处置"
        )
    return {
        "expected_order_keys": len(expected),
        "produced_order_keys": len(expected) - len(missing),
        "missing_order_keys": [f"{ar}|{so}" for ar, so in missing],
        "historical_detail_rows": historical_rows,
        "recovered_delivery_orders": recovered,
        "raw_writeoff_rows": raw_writeoff_rows,
        "accounted_writeoff_rows": accounted_writeoff_rows,
        "complete": not missing,
    }


def expand_payments(payments: List[dict], rates: Optional[Dict[str, float]] = None) -> List[dict]:
    """全部到账 → records，并做 **AR + AR/SO 两级覆盖率硬校验**。"""
    rates = rates or {}
    records: List[dict] = []
    for p in payments:
        records.extend(expand_payment(p, rates))
    want = {p["ar"] for p in payments if p.get("ar")}
    got = {r.get("ar") for r in records if r.get("ar")}
    missing = sorted(want - got)
    if missing:
        raise CoverageError(
            "有到账没产出任何判定（这就是 2026-07-22 静默丢 3 笔的那类 bug，绝不放行）："
            f"{missing}"
        )
    source_coverage(payments, records)
    return records


# ══════════════════════════════════════════════════════════════
# 三、盈亏表索引
# ══════════════════════════════════════════════════════════════
class LedgerIndex:
    """盈亏表『明细』只读索引。她表里**一行 = 一个 SOD**。"""

    def __init__(self, path: Optional[Path] = None, synthetic: Optional[dict] = None):
        self.path = path
        self.so_index: Dict[str, List[int]] = {}
        self.sod_index: Dict[str, List[int]] = {}
        self.so_amount_index: Dict[Tuple[str, int], List[int]] = {}
        self.row_snapshot: Dict[int, dict] = {}
        self.cols: dict = {}
        if synthetic is not None:
            self.so_index = {k: list(v) for k, v in synthetic.get("so", {}).items()}
            self.sod_index = {k: list(v) for k, v in synthetic.get("sod", {}).items()}
            self.row_snapshot = synthetic.get("rows", {})
            for r, snap in self.row_snapshot.items():
                y = common.to_number(snap.get("yingshou"))
                so = snap.get("so") or ""
                if so and y is not None:
                    self.so_amount_index.setdefault((so, int(round(y * 100))), []).append(int(r))
            return
        if path is not None:
            self._load(path)

    def _load(self, path: Path):
        import openpyxl

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        if "明细" not in wb.sheetnames:
            names = list(wb.sheetnames)
            wb.close()
            raise ValueError(f"盈亏表无『明细』sheet：{names}")
        all_rows = list(wb["明细"].iter_rows(values_only=True))
        wb.close()

        aliases = common.load_aliases()
        hrow, headers = common.find_header_row(
            all_rows, "盈亏明细", ["SO", "SOD", "计提", "回款明细", "是否结账"], aliases
        )
        cols = common.resolve_columns(
            headers, "盈亏明细",
            ["SO", "SOD", "计提", "回款明细", "是否结账", "收款时间", "收款方式"],
            aliases,
        )
        role = aliases.get("盈亏明细", {})
        diff_idx = common.fuzzy_find_col(headers, role.get("差异", ["差异"]))
        if diff_idx is not None:
            cols["差异"] = diff_idx
        yidx = common.fuzzy_find_col(headers, role.get("应收", ["应收金额", "应收"]))
        if yidx is not None:
            cols["应收"] = yidx
        self.cols = cols

        for r_i, row in enumerate(all_rows, start=1):
            if r_i <= hrow + 1:
                continue
            vals = list(row)

            def cell(key):
                i = cols.get(key)
                return vals[i] if i is not None and i < len(vals) else None

            so_s = str(cell("SO") or "").strip()
            sod_s = str(cell("SOD") or "").strip()
            if not so_s.startswith("SO") and not sod_s.startswith("SOD"):
                continue
            if so_s.startswith("SO"):
                self.so_index.setdefault(so_s, []).append(r_i)
            if sod_s.startswith("SOD"):
                self.sod_index.setdefault(sod_s, []).append(r_i)
            yingshou = common.to_number(cell("应收"))
            if so_s.startswith("SO") and yingshou is not None:
                self.so_amount_index.setdefault(
                    (so_s, int(round(yingshou * 100))), []
                ).append(r_i)
            self.row_snapshot[r_i] = {
                "so": so_s, "sod": sod_s,
                "jiti": cell("计提"), "huikuan": cell("回款明细"),
                "chayi": cell("差异"),
                "jiezhang": cell("是否结账"), "shoukuan_time": cell("收款时间"),
                "shoukuan_way": cell("收款方式"), "yingshou": yingshou,
            }

    def positional_row(
        self, so: str, sod: str, lines: List[dict]
    ) -> Optional[Tuple[int, str, Optional[float]]]:
        """
        (SO+应收金额) 定位不到唯一行时的严格消歧：验证**整段能不能对齐**。

          她表里这个 SO 的全部行（行号升序） vs 智云该 SO 的全部 SOD（编号降序）

        为什么按这个序：两边都源自月初同一份「交付数据」导出，天然同序。

        返回 (行号, 依据, 比例)；对不齐返回 None（老老实实挂起）。依据两种：

        · `exact`  —— 金额序列逐位完全相等。实测 SO26040297（12 行）、SO26040481（10 行）。
        · `ratio`  —— 有几位不等，但**所有不等的位比值完全一致**（系统性口径差，不是行错位）。
                      实测 SO26040322：智云 477.61/661.31 vs 她表 488.64/676.58，
                      两处比值都是 0.977433，她当天就是按智云金额填的。
                      行错位不可能凑出同一个比值，所以这条判据是可证的，不是猜。

        只要行数对不上（比如她把某个 SOD 拆成了两行）→ 直接 None。
        """
        rows = sorted(self.so_index.get(so, []))
        if not rows or not lines or len(rows) != len(lines):
            return None
        ordered = sorted(lines, key=lambda x: str(x.get("sod") or ""), reverse=True)
        ratios: List[float] = []
        for r, ln in zip(rows, ordered):
            y = common.to_number((self.row_snapshot.get(r) or {}).get("yingshou"))
            d = ln.get("deliver")
            if y is None or d is None or float(y) == 0.0:
                return None
            if abs(float(y) - float(d)) > TOL:
                ratios.append(float(d) / float(y))
        kind, ratio = "exact", None
        if ratios:
            ratio = sum(ratios) / len(ratios)
            if not (0.5 < ratio < 1.5):
                return None
            if any(abs(x - ratio) > 5e-4 for x in ratios):
                return None  # 比值不一致 → 更像行错位，不敢认
            kind = "ratio"
        for r, ln in zip(rows, ordered):
            if (ln.get("sod") or "") == sod:
                return r, kind, ratio
        return None

    @staticmethod
    def _is_outstanding(snap: dict) -> bool:
        """拆行后唯一可继续承接回款的行：未结账且没有回款明细。"""
        settled = str(snap.get("jiezhang") or "").strip()
        received = common.to_number(snap.get("huikuan"))
        return settled != "是" and (received is None or abs(float(received)) <= TOL)

    def business_rows(self, so: str, sod: str, row: Optional[int] = None) -> List[int]:
        """
        返回同一 SOD 的全部拆分行。

        首次回款前盈亏表可能尚未回填 SOD，此时只能用已唯一命中的当前行作为基线；
        一旦拆行，两行都保留 SOD，后续累计只在该 SOD 内计算，绝不再把同 SO 的其它
        SOD 回款混进来。
        """
        if sod:
            rows = list(self.sod_index.get(sod, []))
            if rows:
                return sorted(rows)
        return [int(row)] if row is not None else []

    def business_totals(
        self, so: str, sod: str, row: Optional[int] = None
    ) -> Tuple[Optional[float], float, List[int]]:
        """返回该 SOD 拆分行的原始应收合计、累计已填回款、行号。"""
        receivable_values: List[float] = []
        received_values: List[float] = []
        rows = self.business_rows(so, sod, row)
        for one_row in rows:
            snap = self.row_snapshot.get(one_row) or {}
            receivable = common.to_number(snap.get("yingshou"))
            received = common.to_number(snap.get("huikuan"))
            if receivable is not None:
                receivable_values.append(float(receivable))
            if received is not None:
                received_values.append(float(received))
        initial_receivable = (
            round(sum(receivable_values), 2) if receivable_values else None
        )
        return initial_receivable, round(sum(received_values), 2), rows

    def comparison_row(
        self,
        so: str,
        sod: str = "",
        current_received: Optional[float] = None,
    ) -> Optional[int]:
        """
        在离线对比表中定位候选行；仅供差异审计，不参与日常判定。

        拆行后的同一 SOD 可能有多行；优先用“本次回款明细”唯一命中当前切片，
        其次才接受唯一 SOD / 唯一 SO。不能唯一证明就返回 None，禁止猜行。
        """
        candidates: List[int] = []
        if sod:
            candidates = list(self.sod_index.get(sod, []))
        if not candidates and so:
            candidates = list(self.so_index.get(so, []))
        if not candidates:
            return None
        if current_received is not None:
            exact = []
            for one_row in candidates:
                got = common.to_number(
                    (self.row_snapshot.get(one_row) or {}).get("huikuan")
                )
                if (
                    got is not None
                    and abs(float(got) - float(current_received)) <= TOL
                ):
                    exact.append(one_row)
            if len(exact) == 1:
                return exact[0]
        return candidates[0] if len(candidates) == 1 else None

    def so_totals(self, so: str) -> Tuple[Optional[float], float]:
        """兼容旧调用；仅用于没有 SOD 粒度的人工提示。"""
        receivable_values: List[float] = []
        received_values: List[float] = []
        for row in self.so_index.get(so, []):
            snap = self.row_snapshot.get(row) or {}
            receivable = common.to_number(snap.get("yingshou"))
            received = common.to_number(snap.get("huikuan"))
            if receivable is not None:
                receivable_values.append(float(receivable))
            if received is not None:
                received_values.append(float(received))
        return (
            round(sum(receivable_values), 2) if receivable_values else None,
            round(sum(received_values), 2),
        )

    def match(
        self, so: str, sod: str, amount: Optional[float] = None
    ) -> Tuple[Optional[int], str, List[int]]:
        """
        返回 (行号, 依据, 候选)。依据优先级：
          ① SO + 应收金额  ← 主键：回填前后都成立（她一行 = 一个 SOD，应收金额=该SOD交付额）
          ② SOD           ← 她已回填过「实收金额」列时可用（幂等重跑走这条）
          ③ SO 唯一行
        """
        if so and amount is not None:
            rows = self.so_amount_index.get((so, int(round(float(amount) * 100))), [])
            if len(rows) == 1:
                return rows[0], "SO+应收金额", rows
            if len(rows) > 1:
                return None, "E8", rows
        if sod:
            rows = self.sod_index.get(sod, [])
            if len(rows) == 1:
                return rows[0], "SOD", rows
            if len(rows) > 1:
                outstanding = [
                    r for r in rows if self._is_outstanding(self.row_snapshot.get(r) or {})
                ]
                if len(outstanding) == 1:
                    return outstanding[0], "SOD未结清行", rows
                return None, "E8", rows
        if so:
            rows = self.so_index.get(so, [])
            if len(rows) == 1:
                return rows[0], "SO", rows
            if len(rows) > 1:
                return None, "E8", rows
            return None, "E2", []
        return None, "E7", []


# ══════════════════════════════════════════════════════════════
# 四、单条判定
# ══════════════════════════════════════════════════════════════
def classify_one(
    rec: dict,
    ledger: Optional[LedgerIndex],
    rates: Dict[str, float],
    thr: float,
    year_now: int,
) -> dict:
    if rec.get("default_first_sod") and ledger is not None and rec.get("so"):
        so = str(rec.get("so") or "").strip()
        first_row = next(
            (
                row_no
                for row_no in sorted(ledger.so_index.get(so, []))
                if ledger._is_outstanding(ledger.row_snapshot.get(row_no) or {})
            ),
            None,
        )
        if first_row is not None:
            snap = ledger.row_snapshot.get(first_row) or {}
            target_sod = str(snap.get("sod") or "").strip()
            target_line = next(
                (
                    line
                    for line in (rec.get("default_sod_lines") or [])
                    if str(line.get("sod") or "").strip() == target_sod
                ),
                None,
            )
            if target_sod and target_line is not None and target_line.get("deliver_local") is not None:
                resolved = dict(rec)
                resolved.pop("forced_code", None)
                resolved.pop("forced_reason", None)
                resolved.pop("default_first_sod", None)
                resolved.update({
                    "sod": target_sod,
                    "amount_orig": rec.get("default_amount_orig"),
                    "amount_local": rec.get("default_amount_local"),
                    "deliver_local": target_line.get("deliver_local"),
                    "currency": target_line.get("currency") or rec.get("currency"),
                    "cumulative_received_local": None,
                    "so_all_lines": rec.get("default_sod_lines") or [],
                    "preferred_ledger_row": first_row,
                    "match_basis": (
                        f"{rec.get('default_match_basis') or '智云核销金额'}"
                        "/同SO首个未结清SOD默认"
                    ),
                })
                resolved_result = classify_one(resolved, ledger, rates, thr, year_now)
                resolved_result["reason"] = (
                    "按确认规则选择盈亏表中同 SO 的首个未结清 SOD。"
                    + str(resolved_result.get("reason") or "")
                )
                return resolved_result

    result = {
        "ar": rec.get("ar") or "",
        "so": rec.get("so") or "",
        "sod": rec.get("sod") or "",
        # case_id = AR × SO × SOD：她表里一行一个 SOD，粒度必须到 SOD 否则台账互相覆盖
        "case_id": "|".join(
            x for x in (rec.get("ar") or "-", rec.get("so") or "-", rec.get("sod") or "") if x
        ),
        "customer_masked": common.mask_customer(rec.get("customer") or ""),
        "flow_hits": rec.get("flow_hits"),
        "flow_locate": rec.get("flow_locate") or "",
        "flow_matched_by": rec.get("flow_matched_by") or "",
        "flow_order_suggest": rec.get("flow_order_suggest") or "",
        "flow_file": rec.get("flow_file") or "",
        "flow_sheet": rec.get("flow_sheet") or "",
        "flow_row_no": rec.get("flow_row_no"),
        "flow_order_existing": rec.get("flow_order_existing") or "",
        "huikuan_type": rec.get("huikuan_type") or "",
        "status": rec.get("status") or "",
        "match_basis": rec.get("match_basis") or "",
        "bucket": "exception",
        "code": "",
        "reason": "",
        "five_cols": {},
        "locate_hint": "",
        "current_values": {},
        "candidates": [],
        "warning_codes": list(rec.get("warning_codes") or []),
        "duplicate_writeoff_audit": rec.get("duplicate_writeoff_audit") or {},
    }
    if rec.get("so"):
        result["locate_hint"] = (
            f"在盈亏『明细』按「新智云单号」筛选：{rec['so']}"
            + (f"，找应收金额={rec.get('amount_orig')} 那行" if rec.get("amount_orig") is not None else "")
            + "（禁止用行号）"
        )

    # 展开阶段已定性的（分笔/超额/没回满/无下单…）直接落地
    forced = rec.get("forced_code")
    if forced:
        result["code"] = forced
        reason = rec.get("forced_reason") or ""
        if (
            forced == "E5"
            and ledger is not None
            and rec.get("so")
            and rec.get("partial_latest_delivery") is not None
            and rec.get("partial_current_received") is not None
        ):
            initial_receivable, existing_received = ledger.so_totals(rec["so"])
            reason = (
                f"{rec['so']} 没回满。"
                + partial_split_guidance(
                    rec["partial_latest_delivery"],
                    rec["partial_current_received"],
                    initial_receivable=initial_receivable,
                    existing_received=existing_received,
                )
                + " 当前无法唯一命中 SOD，先指明承接回款的 SOD；唯一后才允许确认写入。"
            )
        result["reason"] = reason
        result["bucket"] = "exception" if forced in ("E4", "E7", "E10", "E12", "E0") else "hold"
        return result

    if (rec.get("status") or "") == "已作废":
        result["code"] = "E7"
        result["reason"] = "核销已作废"
        return result

    amount_orig = rec.get("amount_orig")
    if amount_orig is None:
        result["code"] = "E7"
        result["reason"] = "核销金额为空"
        return result

    # 智云已经给出本币金额时直接采用，不再反向要求汇率。只有本币金额也缺失时，
    # 才按“子记录汇率 → 父回款原/本币隐含汇率 → 命令行汇率”逐级换算。
    local = common.to_number(rec.get("amount_local"))
    err = None
    if local is None:
        local, err = _localize_amount(
            float(amount_orig),
            rec,
            rates,
            row_rate=rec.get("rate"),
        )
    if err == "E6" or local is None:
        result["code"] = "E6"
        result["reason"] = (
            f"外币既没有本币核销金额，也没有可用汇率（{rec.get('currency')}）"
        )
        return result

    # 流转表三键信号（有做才判）
    flow_hits = rec.get("flow_hits")
    if flow_hits is not None:
        try:
            fh = int(flow_hits)
        except (TypeError, ValueError):
            fh = -1
        if fh == 0:
            result["code"] = "E0"
            result["reason"] = "流转表三键对不到账（0 行）"
            return result
        if fh > 1:
            result["code"] = "E12"
            result["reason"] = "同日同额同名命中多行"
            return result
    if rec.get("customer_archive_failed"):
        result["code"] = "E10"
        result["reason"] = "建档失败/搜不到客户"
        return result

    if ledger is None:
        result["bucket"] = "hold"
        result["code"] = "E2"
        result["reason"] = "未提供盈亏表，无法确认 SO 是否在明细"
        return result

    so, sod = rec.get("so") or "", rec.get("sod") or ""
    preferred_row = rec.get("preferred_ledger_row")
    if preferred_row is not None:
        row, how, cands = int(preferred_row), "同SO首个未结清SOD默认", []
    else:
        row, how, cands = ledger.match(so, sod, amount_orig)

    # 定位不到唯一行 → 用「整段逐位对齐」严格消歧（对不齐就继续挂起）
    align_note = ""
    if how in ("E8", "E2") and sod and rec.get("so_all_lines"):
        aligned = ledger.positional_row(so, sod, rec["so_all_lines"])
        if aligned is not None:
            row, kind, ratio = aligned
            how, cands = "SO整段按SOD序对齐", []
            if kind == "ratio":
                snap0 = ledger.row_snapshot.get(row) or {}
                align_note = (
                    f"⚠ 智云交付额 {amount_orig} 与你表里应收 {snap0.get('yingshou')} 不一致"
                    f"（这个 SO 每一行都差同一个比例 {ratio:.6f}），已按**智云金额**填；"
                    "口径待确认，填之前扫一眼"
                )

    if how == "E8":
        result["bucket"] = "hold"
        result["code"] = "E8"
        result["reason"] = (
            f"盈亏表里有 {len(cands)} 行同时满足 SO={so}"
            + (f" 且应收金额={amount_orig}" if amount_orig is not None else "")
            + "，分不清记哪行，你指一下"
        )
        result["candidates"] = cands
        return result
    if how == "E7":
        result["code"] = "E7"
        result["reason"] = "无单号"
        return result
    if how == "E2" or row is None:
        # 表里真找不到这单，这时才区分「跨年老单」还是「还没交付」
        y = common.year_from_so(sod or so)
        result["bucket"] = "hold"
        if y is not None and y < year_now:
            result["code"] = "E3"
            result["reason"] = f"{y} 年的老单，今年盈亏表里没有这行"
        else:
            result["code"] = "E2"
            result["reason"] = "盈亏表里还没有这张单（多半还没交付进表）"
        return result

    snap = ledger.row_snapshot.get(row, {})
    yingshou = common.to_number(snap.get("yingshou"))
    deliver = common.to_number(rec.get("deliver_local"))

    # 超额核销：本次本币 > **智云这单交付额**（不是她表应收）。
    # 2026-07-24 明妹口径：判超额的上限是智云交付额，别以她表应收为准 —— 应收可能是旧值，
    # 交付额中途变大时（初始 1 → 结算 2），拿旧应收当上限会把「本该填的 2」误报成超额。
    # 只有拿不到智云交付额时，才退回用应收当兜底上限。
    ceiling = deliver if deliver is not None else yingshou
    if (
        ceiling is not None
        and common.is_cny(rec.get("currency") or "")
        and float(local) > float(ceiling) + max(thr, TOL)
    ):
        result["code"] = "E4"
        src = "智云这单交付额" if deliver is not None else "你表里应收"
        result["reason"] = (
            f"本次核销 {local} 比{src} {round(float(ceiling), 2)} 还多，不正常，先找销售核对"
        )
        return result

    shoukuan_date = common.norm_date(rec.get("shoukuan_date"))
    hexiao_date = common.norm_date(rec.get("hexiao_date"))
    # 收款时间/方式只比较到账月和核销月，不再按回款类型分支：
    # 同月 = 到账日期 +「汇」；跨月 = 核销日期 +「冲预收」。
    r_time = common.receipt_time(shoukuan_date, hexiao_date)
    way = common.pay_way(
        rec.get("status") or "",
        shoukuan_date,
        hexiao_date,
    )
    local_f = round(float(local), 2)

    # ── 结账 / 计提（2026-07-29 交付额变动会议更新）──────────────────────────
    # 【结账】= 这笔到账给这个 SOD 下发的「任务」做完没有。任务 = 智云本次核销这个单的金额。
    #   能走到这里（bucket=auto）的行 = 核销命中 + 已交付进表 → 任务能做且已做 → 一律「是」。
    #   ⚠ **绝不能拿「回款记录整笔」的核销状态判每个小单**：
    #     “预存部分核销”只是说那笔预存余额没花完（实测 6300 核 6090、剩 210 挂预收），
    #     跟这个单本身收没收满、任务做没做完**毫无关系**。旧版拿 status∈SETTLED 判结账，
    #     把 6 个已收满的预存视频单全误判成「否」（2026-07-24 对明妹真答案实测：146/152→修后应全对）。
    #   （结账=否 只属于“没交付/没做任务”的行，那是 E2/E3，早已挂起、根本走不到这里。）
    # 【计提】只有累计实际回款达到最新实际交付额 D 才填 D；未达到时两侧计提都留空。
    # 【部分回款】当前未结清行改成“本次已回款行”，并在其正下方复制出“剩余未回款行”：
    #   U = D - (历史累计回款 + 本次回款)
    #   当前已回款切片应收 = 当前未结清行应收 - U
    #   新增未回款行应收 = U
    # 这样所有拆分行应收合计始终等于原始应收基线，且累计只在当前 SOD 内计算。
    initial_receivable, existing_received, business_rows = ledger.business_totals(so, sod, row)
    source_cumulative = common.to_number(rec.get("cumulative_received_local"))
    if source_cumulative is not None:
        # 智云核销明细是事实源：同一 SO/SOD 截至目标日的历史核销 + 本次核销，
        # 比盈亏表里是否已经写过历史回款更可靠，也避免把历史核销漏算成“本次部分回款”。
        cumulative_received = round(float(source_cumulative), 2)
        existing_received = round(max(cumulative_received - local_f, 0.0), 2)
    else:
        target_received = common.to_number(snap.get("huikuan"))
        if (
            target_received is not None
            and abs(float(target_received) - local_f) <= max(thr, TOL)
        ):
            existing_received = round(max(existing_received - float(target_received), 0.0), 2)
        cumulative_received = round(existing_received + local_f, 2)

    if deliver is not None and cumulative_received > float(deliver) + max(thr, TOL):
        result["bucket"] = "hold"
        result["code"] = "E5"
        result["reason"] = partial_split_guidance(
            float(deliver),
            local_f,
            initial_receivable=initial_receivable,
            existing_received=existing_received,
        )
        result["ledger_row_ref"] = row
        return result

    if deliver is not None and cumulative_received < float(deliver) - max(thr, TOL):
        remaining = round(float(deliver) - cumulative_received, 2)
        current_receivable = common.to_number(snap.get("yingshou"))
        if current_receivable is None:
            result["bucket"] = "hold"
            result["code"] = "E5"
            result["reason"] = "部分回款已识别，但当前未结清行没有应收金额，无法证明拆行后应收守恒。"
            result["ledger_row_ref"] = row
            return result
        paid_slice = round(float(current_receivable) - remaining, 2)
        if paid_slice < -max(thr, TOL):
            result["bucket"] = "hold"
            result["code"] = "E5"
            result["reason"] = (
                f"部分回款后实际未收 {remaining:.2f} 大于当前未结清行应收 "
                f"{float(current_receivable):.2f}，无法安全拆行；先核对交付额变化。"
            )
            result["ledger_row_ref"] = row
            return result

        paid_slice = max(paid_slice, 0.0)
        paid_side_total = (
            round(float(initial_receivable) - remaining, 2)
            if initial_receivable is not None
            else None
        )
        result["five_cols"] = {
            "计提": None,
            "回款明细": local_f,
            "是否结账": "是",
            "收款时间": r_time.isoformat() if r_time else None,
            "收款方式": way,
            "实收SOD": sod or snap.get("sod") or None,
        }
        result["row_operation"] = {
            "type": "split_below",
            "source_receivable": round(float(current_receivable), 2),
            "paid_receivable": paid_slice,
            "unpaid_receivable": remaining,
            "baseline_receivable": initial_receivable,
            "paid_side_receivable_total": paid_side_total,
            "existing_received": existing_received,
            "current_received": local_f,
            "cumulative_received": cumulative_received,
            "latest_delivery": round(float(deliver), 2),
            "business_rows": business_rows,
            "inserted_five_cols": {
                "计提": None,
                "回款明细": None,
                "是否结账": "否",
                "收款时间": None,
                "收款方式": None,
                "实收SOD": sod or snap.get("sod") or None,
            },
        }
        result["bucket"] = "auto"
        result["code"] = "E5"
        result["reason"] = partial_split_guidance(
            float(deliver),
            local_f,
            initial_receivable=initial_receivable,
            existing_received=existing_received,
        )
        result["current_values"] = {
            "计提": snap.get("jiti"),
            "回款明细": snap.get("huikuan"),
            "差异": snap.get("chayi"),
            "是否结账": str(snap.get("jiezhang") or "").strip()
            if snap.get("jiezhang") is not None
            else "",
            "收款时间": str(snap.get("shoukuan_time") or "")[:10],
            "收款方式": snap.get("shoukuan_way"),
            "实收SOD": snap.get("sod"),
        }
        result["ledger_row_ref"] = row
        return result

    jiti_target = round(float(deliver), 2) if deliver is not None else local_f

    result["five_cols"] = {
        "计提": jiti_target,
        "回款明细": local_f,
        "是否结账": "是",
        "收款时间": r_time.isoformat() if r_time else None,
        "收款方式": way,
        "实收SOD": sod or snap.get("sod") or None,
    }
    # 业务值差异只在最终结清时产生：原始应收是历史基线，计提按最新实际交付额。
    # 两者不一致时保留原始应收不动，并在“差异”列写公式 = 应收金额 - 计提金额。
    # 部分回款阶段计提留空，因此不提前写差异。
    baseline_for_difference = (
        initial_receivable if initial_receivable is not None else yingshou
    )
    if (
        baseline_for_difference is not None
        and abs(float(baseline_for_difference) - jiti_target) > max(thr, TOL)
    ):
        result["derived_cols"] = {
            "差异": round(float(baseline_for_difference) - jiti_target, 2)
        }
    # ── 交付额变化提醒（2026-07-24 明妹原话："要让他动脑子检查交付额，别以我表里的为准")──
    # 命中行后，若**智云结算额**（本次核销/交付，= amount_orig）和她表里那行的**应收金额**
    # 对不上，一律顶一个 ⚠ 到「怎么办」。以前只有 ratio 消歧路径会提醒，靠 SOD / SO 唯一行
    # 命中的（她已手动改过应收、或压根没写 SOD）就闷声按智云额填、不吭声——她要的是**每一笔
    # 都被明确检查一次**。回款明细用本次实际核销额，计提目标用智云最新交付额；只是多一句让她扫。
    # align_note 已经说过（比例差）就不重复。
    disc_note = ""
    if not align_note and baseline_for_difference is not None:
        if abs(jiti_target - float(baseline_for_difference)) > max(thr, TOL):
            disc_note = (
                f"⚠ 智云最新实际交付额 {jiti_target} 与表里原始应收 "
                f"{round(float(baseline_for_difference), 2)} 不一致：计提按最新交付额，原始应收不改，"
                f"差异列写应收减计提"
            )

    result["bucket"] = "auto"
    result["code"] = ""
    duplicate_note = ""
    audit = rec.get("duplicate_writeoff_audit") or {}
    if audit.get("status") == "recovered":
        duplicate_note = (
            "⚠ 智云疑似系统重复核销，本次每组只按一次处理；"
            f"重复组{len(audit.get('duplicate_groups') or [])}个，"
            f"原始{audit.get('raw_record_count', 0)}条，"
            f"保留{audit.get('logical_record_count', 0)}条，"
            f"忽略{audit.get('ignored_record_count', 0)}条，"
            f"折叠前差额{audit.get('delta_raw')}，折叠后差额{audit.get('delta_dedup')}"
        )
    tail = "；".join(x for x in (align_note, disc_note, duplicate_note) if x)
    result["reason"] = f"{rec.get('match_basis') or '判定'} · 定位={how}" + (
        f"；{tail}" if tail else ""
    )
    result["current_values"] = {
        "计提": snap.get("jiti"),
        "回款明细": snap.get("huikuan"),
        "差异": snap.get("chayi"),
        "是否结账": str(snap.get("jiezhang") or "").strip() if snap.get("jiezhang") is not None else "",
        "收款时间": str(snap.get("shoukuan_time") or "")[:10],
        "收款方式": snap.get("shoukuan_way"),
        "实收SOD": snap.get("sod"),
    }
    result["ledger_row_ref"] = row
    return result


def classify_records(
    records: List[dict],
    ledger: Optional[LedgerIndex] = None,
    rates: Optional[Dict[str, float]] = None,
) -> dict:
    rates = rates or {}
    thr = common.tail_threshold()
    year_now = common.current_year()
    results = [
        classify_one(rec, ledger, rates, thr, year_now)
        for rec in records
    ]

    # 同一行被两条计划命中 → 两条都别自动写（谁对谁错要人定）
    seen: Dict[int, str] = {}
    dup: Dict[int, bool] = {}
    for r in results:
        ref = r.get("ledger_row_ref")
        if r["bucket"] != "auto" or ref is None:
            continue
        if ref in seen:
            dup[ref] = True
        seen[ref] = r.get("case_id") or ""
    for r in results:
        ref = r.get("ledger_row_ref")
        if ref is not None and dup.get(ref):
            r["bucket"] = "hold"
            r["code"] = "E8"
            r["reason"] = f"盈亏表第 {ref} 行被本次两笔同时命中，你指一下各记哪行"
            r["five_cols"] = {}

    auto = [r for r in results if r["bucket"] == "auto"]
    hold = [r for r in results if r["bucket"] == "hold"]
    exc = [r for r in results if r["bucket"] == "exception"]
    return {
        "auto": auto,
        "hold": hold,
        "exception": exc,
        "counts": {
            "auto": len(auto), "hold": len(hold),
            "exception": len(exc), "total": len(results),
        },
        "e_code_dist": _dist(results),
        "ar_summary": build_ar_summary(results),
    }


# 「没交付进表」的挂起码：只有这些才让一笔到账的流转状态掉到「部分」。
# E5（部分核销）不在内 —— 钱已全核落地、她拆行即算更新（2026-07-24 明妹口径）。
_FLOW_WAIT_CODES = {"E2", "E3"}


def _flow_ready(item: dict) -> str:
    """这笔 SOD 今天能不能在盈亏表里更新：ready=能（auto/拆行/指认）；wait=不能（没交付/异常）。"""
    if item.get("bucket") == "auto":
        return "ready"
    if item.get("bucket") == "hold" and (item.get("code") or "") not in _FLOW_WAIT_CODES:
        return "ready"  # E5 拆行 / E8 指认：行在表里，她更得动
    return "wait"       # E2/E3 没交付进表；exception 数据问题 → 今天都更新不了


def build_ar_summary(results: List[dict]) -> List[dict]:
    """按 AR 汇总，派生到账流转表「是否更新应收款」三态（是 / 部分 / 空白）。"""
    try:
        from flow_ledger import derive_flow_status
    except Exception:  # pragma: no cover
        def derive_flow_status(states):
            ready = sum(1 for s in states if s in ("ready", "auto"))
            if not states or ready == 0:
                return ""
            return "是" if ready == len(states) else "部分"

    by_ar: Dict[str, List[dict]] = {}
    for r in results:
        by_ar.setdefault(r.get("ar") or "-", []).append(r)
    out = []
    for ar, items in by_ar.items():
        flow_states = [_flow_ready(i) for i in items]
        out.append({
            "ar": ar,
            "so_count": len({i.get("so") for i in items if i.get("so")}),
            "行数": len(items),
            "buckets": [i["bucket"] for i in items],
            "流转表_是否更新应收款_建议": derive_flow_status(flow_states),
            "flow_locate": next((i.get("flow_locate") for i in items if i.get("flow_locate")), ""),
            "待处理SO": sorted({i.get("so") for i in items if i["bucket"] != "auto" and i.get("so")}),
        })
    return out


def _dist(items: List[dict]) -> Dict[str, int]:
    d: Dict[str, int] = {}
    for it in items:
        c = it.get("code") or "OK"
        d[c] = d.get(c, 0) + 1
    return d


# ══════════════════════════════════════════════════════════════
# 五、夹具（单测/离线演练用）
# ══════════════════════════════════════════════════════════════
def payments_from_fixture(fixture: dict) -> List[dict]:
    """
    夹具格式（v2）：
      {"payments": [{ar, hexiao_date, arrival_date, amount_orig, ...,
                     orders:[{so, deliver}], writeoffs:{so: amt}}],
       "sod_lines": {"SO…": [{"sod": "SOD…", "deliver": 1.0}]}}
    """
    if "payments" not in fixture:
        raise InputError("夹具不是 v2 格式（需要 payments 键）")
    sod_lines = fixture.get("sod_lines") or {}
    out = []
    for p in fixture["payments"]:
        q = dict(p)
        q["hexiao_date"] = common.norm_date(p.get("hexiao_date"))
        q["arrival_date"] = common.norm_date(p.get("arrival_date"))
        q["amount_orig"] = common.to_number(p.get("amount_orig"))
        q["amount_local"] = common.to_number(p.get("amount_local"))
        q["fee"] = common.to_number(p.get("fee")) or 0.0
        q.setdefault("currency", "人民币CNY")
        q.setdefault("orders", [])
        q.setdefault("writeoffs", {})
        q["sod_lines"] = sod_lines
        out.append(q)
    return out


def serialize_result(result: dict) -> dict:
    def fix(obj):
        if isinstance(obj, dt.date):
            return obj.isoformat()
        if isinstance(obj, dict):
            return {k: fix(v) for k, v in obj.items()}
        if isinstance(obj, list):
            return [fix(x) for x in obj]
        return obj

    return fix(result)


def main(argv: Optional[Sequence[str]] = None) -> int:
    ap = argparse.ArgumentParser(description="核销判定（单入口 · SOD 级 · 三栏）")
    ap.add_argument("--workspace", default=str(common.WORK))
    ap.add_argument("--fixture", default="", help="离线夹具 JSON（v2: payments/sod_lines）")
    ap.add_argument("--ledger", default="", help="盈亏核算表副本（只读）")
    ap.add_argument("--rate", action="append", default=[], help="外币汇率 美元USD=7.0")
    ap.add_argument("--out", default="", help="判定结果 json 路径")
    ap.add_argument("--flow", default="", help="到账流转表副本（只读）；不给则扫 02_我的表副本/")
    ap.add_argument(
        "--flow-complete", action="store_true",
        help="声明当天所有渠道的流转表都已给全；只有这时才判 E0（对不到账）",
    )
    ap.add_argument(
        "--hexiao-date", default="",
        help="声明这批是哪个**核销日期**（她确认过的那天）。给了就跟数据核对，对不上直接退出",
    )
    ap.add_argument(
        "--allow-mixed-dates", action="store_true",
        help="允许一批里混着多个核销日（默认禁止：混批会让覆盖率/幂等校验和她逐行核对全失真）",
    )
    args = ap.parse_args(argv)

    ws = common.ensure_out_dirs(args.workspace)  # 解析真工作区，防产出分家
    rates = common.parse_rate_args(args.rate)
    requested_date = common.resolve_batch_date(args.hexiao_date) if args.hexiao_date else None
    if args.hexiao_date and requested_date is None:
        print(f"ERROR: 认不出 --hexiao-date {args.hexiao_date!r}", file=sys.stderr)
        return 2

    ledger = None
    if args.ledger:
        ledger = LedgerIndex(Path(args.ledger))
    else:
        cand = sorted((ws / "02_我的表副本").glob("*盈亏*")) if (ws / "02_我的表副本").is_dir() else []
        cand = [p for p in cand if not p.name.startswith("~$")]
        if cand:
            ledger = LedgerIndex(cand[0])
        else:
            print(
                "WARN: 未提供盈亏表（--ledger 或 02_我的表副本/*盈亏*）；"
                "全部单将 hold E2，不会瞎填五列",
                file=sys.stderr,
            )

    try:
        if args.fixture:
            payments = payments_from_fixture(json.loads(Path(args.fixture).read_text(encoding="utf-8")))
        else:
            payments = load_exports(ws, target_date=requested_date)
        records = expand_payments(payments, rates)
    except (InputError, ValueError) as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 2
    except CoverageError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 3

    import flow_ledger as FL

    flow = (
        FL.FlowLedger.from_paths([Path(args.flow)])
        if args.flow
        else FL.FlowLedger.from_workspace(ws)
    )
    if flow.rows:
        FL.annotate_records(records, flow, complete=args.flow_complete)
        print(f"流转表已接入：{len(flow.rows)} 行，来源 {flow.sources}")
    else:
        print(
            "WARN: 未认出到账流转表（02_我的表副本/）→ 本轮不做三键匹配，"
            "E0/E12 不判、清单不给流转定位",
            file=sys.stderr,
        )

    # ── 这一批到底是哪个核销日（2026-07-25 立）─────────────────────────
    # 产出、清单、文件名一律按**核销日**走，不按"跑的那天"走：她补跑 7-22 的批次时，
    # 文件名写成运行日就会跟今天的批次撞名/盖掉，事后也说不清哪份是哪天的。
    batch_dates = sorted({p["hexiao_date"] for p in payments if p.get("hexiao_date")})
    if len(batch_dates) > 1 and not args.allow_mixed_dates:
        print(
            "ERROR: 这批数据里混着多个核销日期："
            + "、".join(d.isoformat() for d in batch_dates)
            + "\n  一次只跑一个核销日（混批会让 AR 覆盖率、幂等校验和她逐行核对全部失真）。"
            "\n  重新取一天的数，或确实要混批就加 --allow-mixed-dates。",
            file=sys.stderr,
        )
        return 2
    hexiao_date = batch_dates[0] if batch_dates else None
    if requested_date:
        if hexiao_date is not None and requested_date != hexiao_date:
            # 她确认的是这天、数据却是那天 → 多半取数取错了日子，绝不能闷头往下判
            print(
                f"ERROR: 你确认要跑的是 {common.date_cn(requested_date)}，"
                f"但 01_智云导出/ 里的数据是 {common.date_cn(hexiao_date)} 的。\n"
                "  先确认这次到底要跑哪天，再重新取数。",
                file=sys.stderr,
            )
            return 2
        hexiao_date = hexiao_date or requested_date

    result = classify_records(records, ledger, rates)
    duplicate_audits = next(
        (
            p.get("_duplicate_writeoff_audits")
            for p in payments if p.get("_duplicate_writeoff_audits")
        ),
        {},
    )
    result["duplicate_writeoff_audits"] = duplicate_audits
    result["duplicate_writeoff_audit_sha256"] = WDA.audit_fingerprint(duplicate_audits)
    result["flow_sources"] = flow.sources
    result["business_rules"] = {
        "fee_basis": "ignored",
        "writeoff_basis": "zhiyun_current_writeoff_direct",
        "missing_rate_policy": "use_writeoff_amount_directly",
    }
    result["payment_count"] = len(payments)
    result["source_coverage"] = source_coverage(payments, records)
    result["shifted_detail_dates"] = assess_shifted_detail_dates(
        ws, date_from=requested_date, date_to=requested_date
    )
    result["hexiao_date"] = hexiao_date.isoformat() if hexiao_date else ""
    stamp = hexiao_date.strftime("%Y%m%d") if hexiao_date else dt.date.today().strftime("%Y%m%d")
    out_path = Path(args.out) if args.out else (ws / "04_产出" / f"判定结果_{stamp}.json")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(
        json.dumps(serialize_result(result), ensure_ascii=False, indent=2), encoding="utf-8"
    )
    audit_path = out_path.parent / f"核销明细重复审计_{stamp}.json"
    audit_path.write_text(
        json.dumps(
            serialize_result({
                "hexiao_date": result["hexiao_date"],
                "fingerprint": result["duplicate_writeoff_audit_sha256"],
                "parents": duplicate_audits,
            }),
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    c = result["counts"]
    print(f"核销日期：{common.date_cn(hexiao_date)}（这批算的是这一天核销的到账）")
    print(
        f"判定完成 到账 {len(payments)} 笔 → 订单行 {c['total']} 条："
        f"可填={c['auto']} 挂账={c['hold']} 异常={c['exception']}"
    )
    print(f"E码分布: {result['e_code_dist']}")
    sc = result["source_coverage"]
    print(
        f"来源覆盖：AR/SO {sc['produced_order_keys']}/{sc['expected_order_keys']}，"
        f"原始核销记录 {sc.get('raw_writeoff_rows', 0)}/"
        f"{sc.get('accounted_writeoff_rows', 0)} 有处置，"
        f"历史子核销还原 {sc['historical_detail_rows']} 行，SOD回补交付额 {sc['recovered_delivery_orders']} 单"
    )
    recovered = [a for a in duplicate_audits.values() if a.get("status") == "recovered"]
    unresolved = [a for a in duplicate_audits.values() if a.get("status") == "unresolved"]
    print(
        "系统重复核销审计："
        f"纠正父回款 {len(recovered)} 笔，"
        f"重复组 {sum(len(a.get('duplicate_groups') or []) for a in recovered)} 个，"
        f"忽略记录 {sum(int(a.get('ignored_record_count') or 0) for a in recovered)} 条，"
        f"未解决父回款 {len(unresolved)} 笔"
    )
    print(f"重复审计: {audit_path}")
    print(f"结果: {out_path}")

    late_other = {
        day: info for day, info in result["shifted_detail_dates"].items()
        if day != result["hexiao_date"] and info.get("needs_rerun")
    }
    if late_other:
        summary = "；".join(
            f"{day} 有 {info['rows']} 行（{len(info['ars'])} 笔AR）"
            for day, info in late_other.items()
        )
        print(
            "\n⚠ 发现后续快照才出现的历史核销明细：" + summary
            + "\n   必须逐日重跑这些真实核销日，生成增补《核销日清》；确认前仍禁止写表。",
            file=sys.stderr,
        )

    # 登记跑批台账：这一天判过了。没这一步就查不出"哪天从来没跑过"
    if hexiao_date is not None:
        try:
            import batch_ledger

            batch_ledger.record(
                ws, hexiao_date, "classified",
                payments=len(payments), counts=dict(c),
            )
            # 顺手把漏天报出来。**不能指望编排的 AI 记得先跑 gaps**
            # （2026-07-25 opencode 实测：它直接开跑，§0.4 的查漏天那步被跳过了）。
            # 漏天是"事后看不出来"的错，宁可每次判完都提一句。
            info = batch_ledger.find_gaps(ws, through=hexiao_date)
            others = [g for g in info["gaps"] if g != hexiao_date]
            if others:
                short = "、".join(f"{g.month}-{g.day}" for g in others[:6])
                print(
                    f"\n⚠ 还有 {len(others)} 个核销日从来没跑过：{short}"
                    f"{'…' if len(others) > 6 else ''}"
                    f"\n   跟她说一句：这几天也没跑，要不要接着补？（一天一批，从早到晚）",
                    file=sys.stderr,
                )
        except Exception as e:  # 台账写不了不该让已算对的判定失败
            print(f"WARN: 跑批台账登记失败（不影响本次判定）：{type(e).__name__}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
