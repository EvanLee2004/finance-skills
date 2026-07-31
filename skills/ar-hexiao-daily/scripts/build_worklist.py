#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
第7步：判定 + 写入前校验 → **一份**《核销日清》xlsx。

2026-07-23 明妹当面要求（原话）：
  「这问题都都给我在这一个里边去显示吧」「第一列应该是单号」
  「用 sheet 名字就在这都写成一列就行了」
所以这一版把旧的《今日工作清单》(4 页签) + 《回填审核单》(5 页签) **合成一份**：
主页签《今日清单》一行一个 SOD，**第一列单号、第二列状态**，问题类型变成一列。

还修了她当场问倒的那个 bug —— 「我明明填过了他为什么没跳过」：
旧版「今天能填」直接倒 classify 的 auto，**没过写入校验**，所以把昨天已经填过的
23 笔又列了一遍（真跑下来那 23 笔全是"已填过"，实际要填 0 笔）。
本版清单的状态**以 validate_plan 的结论为准**，跳过的就明明白白写"已填过·跳过"。

只写 04_产出/，不写用户任何 Excel。
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
import common  # noqa: E402

# 状态 → (排序权重, 颜色, 这行该干嘛)
STATUS = {
    "今天要填": (1, "C6EFCE", "按「应填」那几列填进盈亏『明细』对应行"),
    "已填过·跳过": (2, "D9D9D9", "不用管，你之前已经填过而且跟这次算的一样"),
    "冲突·需你定": (3, "FFC7CE", "这行已经填过但跟这次算的不一样，看「差异」那列，你定按哪个来"),
    "挂账待办": (4, "FFEB9C", "今天填不了，按「怎么办」处理；处理完说「重扫挂账」"),
    "异常": (5, "F8CBAD", "数据不对劲，别填，按「怎么办」先查清楚"),
}

HEADERS = [
    "单号", "状态", "怎么办",
    "SO", "到账号(AR)", "客户(打码)", "本次金额",
    "应填_计提", "应填_回款明细", "应填_业务值差异", "应填_是否结账", "应填_收款时间", "应填_收款方式", "应填_实收SOD",
    "当前_计提", "当前_回款明细", "当前_业务值差异", "当前_是否结账", "当前_收款时间", "当前_收款方式", "当前_实收SOD",
    "当前值与计划值的比较差异", "怎么找到这行", "这笔到账在流转表哪一行", "流转表单号列建议填",
    "判定依据", "码", "警告码",
]

FIVE = ["计提", "回款明细", "是否结账", "收款时间", "收款方式"]


def _norm(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (dt.date, dt.datetime)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    try:
        return f"{float(s):.2f}"
    except (TypeError, ValueError):
        return s


def _diff_text(five: dict, cur: dict, derived: dict) -> str:
    parts = []
    for k in FIVE + ["实收SOD"]:
        if five.get(k) is None:
            continue
        a, b = _norm(cur.get(k)), _norm(five.get(k))
        if a != b:
            parts.append(f"{k}：表里={a or '(空)'} → 这次算={b}")
    for k, want in derived.items():
        a, b = _norm(cur.get(k)), _norm(want)
        if a != b:
            parts.append(f"{k}：表里={a or '(空)'} → 这次算={b}")
    return "；".join(parts)


# 这些码的 reason 是**逐笔算出来的、自带操作指引**（插哪一行、候选 SOD 是哪几个…），
# 比 config 里的通用建议有用得多 → 「怎么办」直接用 reason。
SPECIFIC_CODES = {"E4", "E5", "E7", "E8", "E_SYSTEM_OVER_WRITEOFF_UNRESOLVED"}


def _row(item: dict, status: str, codes: dict, action_override: str = "") -> List[Any]:
    five = item.get("five_cols") or {}
    derived = item.get("derived_cols") or {}
    cur = item.get("current_values") or {}
    code = item.get("code") or ""
    info = codes.get(code) or {}
    sod = item.get("sod") or five.get("实收SOD") or ""
    warn = "⚠" in (item.get("reason") or "")
    if warn:
        # 判定里带⚠的（比如智云交付额和她表里应收对不上）必须顶到「怎么办」，别只埋在依据列
        action = str(item["reason"]).split("；", 1)[-1]
    elif action_override:
        action = action_override
    elif code in SPECIFIC_CODES and item.get("reason"):
        action = item["reason"]
    elif code:
        action = info.get("建议动作") or info.get("大白话") or STATUS[status][2]
    else:
        action = STATUS[status][2]
    return [
        # 整笔到账级的挂起（没 SO 也没 SOD）不能留空——退化显示到账号，否则她定位不到
        sod or item.get("so") or item.get("ar") or "",
        status,
        action,
        item.get("so") or "",
        item.get("ar") or "",
        item.get("customer_masked") or "",
        five.get("回款明细"),
        five.get("计提"), five.get("回款明细"), derived.get("差异"), five.get("是否结账"),
        five.get("收款时间"), five.get("收款方式"), five.get("实收SOD"),
        cur.get("计提"), cur.get("回款明细"), cur.get("差异"), cur.get("是否结账"),
        cur.get("收款时间"), cur.get("收款方式"), cur.get("实收SOD"),
        _diff_text(five, cur, derived),
        item.get("locate_hint") or "",
        item.get("flow_locate") or "",
        item.get("flow_order_suggest") or "",
        item.get("reason") or info.get("原因") or "",
        code,
        "、".join(item.get("warning_codes") or []),
    ]


def collect_rows(result: dict, checked: Optional[dict]) -> List[List[Any]]:
    codes = common.load_codes()
    rows: List[List[Any]] = []

    if checked:
        by_status = [
            (checked.get("write") or [], "今天要填"),
            (checked.get("skip") or [], "已填过·跳过"),
            (checked.get("conflict") or [], "冲突·需你定"),
        ]
        for items, st in by_status:
            for it in items:
                extra = ((it.get("_check") or {}).get("reason") or "") if st != "今天要填" else ""
                rows.append(_row(it, st, codes, action_override=extra or STATUS[st][2]))
    else:
        for it in result.get("auto") or []:
            rows.append(_row(it, "今天要填", codes))

    for it in result.get("hold") or []:
        rows.append(_row(it, "挂账待办", codes))
    for it in result.get("exception") or []:
        rows.append(_row(it, "异常", codes))

    rows.sort(key=lambda r: (STATUS.get(r[1], (9,))[0], str(r[4]), str(r[3]), str(r[0])))
    return rows


def build_workbook(
    result: dict,
    checked: Optional[dict],
    out_path: Path,
    flow_plan: Optional[dict] = None,
) -> Path:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()

    # ── 使用说明 ────────────────────────────────────────────
    ws0 = wb.active
    ws0.title = "先看这里"
    counts = result.get("counts") or {}
    n = {k: 0 for k in STATUS}
    for r in collect_rows(result, checked):
        n[r[1]] = n.get(r[1], 0) + 1
    fc = (flow_plan or {}).get("counts") or {}
    m_write = int(fc.get("write") or 0)
    m_hand = int(fc.get("hand") or 0)
    coverage = result.get("source_coverage") or {}
    duplicate_audits = result.get("duplicate_writeoff_audits") or {}
    recovered_audits = [
        audit for audit in duplicate_audits.values()
        if audit.get("status") == "recovered"
    ]
    pending_shifted = {
        day: info for day, info in (result.get("shifted_detail_dates") or {}).items()
        if info.get("needs_rerun") and day != result.get("hexiao_date")
    }
    # 日期抬头：她可能今天补跑上周三的批次，也可能隔天才回来确认。
    # 不把「这是哪一天的」印在最显眼的地方，她核对的就可能是另一天的账。
    hx = result.get("hexiao_date") or ""
    lines = [
        ["《核销日清》—— 一份就够，别的表都不用开"],
        [f"★ 核销日期：{common.date_cn(hx) if hx else '(数据里没有核销日期，请核对取数)'}"],
        [f"  （这批算的是**销售在这一天核销**的到账；跟钱哪天到银行、你哪天建的回款没关系）"],
        [f"  清单生成时间：{dt.datetime.now().strftime('%Y-%m-%d %H:%M')}"],
        [""],
        [f"这次一共 {result.get('payment_count', '?')} 笔到账，拆成 {counts.get('total', 0)} 个订单行。"],
        [
            f"来源完整性：AR/SO {coverage.get('produced_order_keys', '?')}/"
            f"{coverage.get('expected_order_keys', '?')}；"
            f"原始核销记录处置 {coverage.get('accounted_writeoff_rows', 0)}/"
            f"{coverage.get('raw_writeoff_rows', 0)}；"
            f"历史子核销还原 {coverage.get('historical_detail_rows', 0)} 行；"
            f"SOD 回补交付额 {coverage.get('recovered_delivery_orders', 0)} 单。"
        ],
        [
            "系统重复纠正："
            f"父回款 {len(recovered_audits)} 笔；"
            f"重复组 {sum(len(a.get('duplicate_groups') or []) for a in recovered_audits)} 个；"
            f"忽略核销记录 {sum(int(a.get('ignored_record_count') or 0) for a in recovered_audits)} 条。"
            + (
                " 智云疑似系统重复核销，本次每组只按一次处理。"
                if recovered_audits else ""
            )
        ],
        [
            "历史增补提醒：" + (
                "；".join(
                    f"{day} 尚缺 {len(info.get('missing_order_keys') or [])} 个 AR/SO"
                    for day, info in pending_shifted.items()
                )
                if pending_shifted else "无"
            )
        ],
        [""],
        ["【盈亏明细】"],
        [f"① 今天要填     {n.get('今天要填', 0):>4} 行  ← 确认后程序可写"],
        [f"② 已填过·跳过  {n.get('已填过·跳过', 0):>4} 行  ← 不用管"],
        [f"③ 冲突·需你定  {n.get('冲突·需你定', 0):>4} 行  ← 你看一眼定个方向"],
        [f"④ 挂账待办     {n.get('挂账待办', 0):>4} 行  ← 今天填不了，看「怎么办」"],
        [f"⑤ 异常         {n.get('异常', 0):>4} 行  ← 数据不对劲，先别填"],
        [""],
        ["【到账流转】"],
        [f"⑥ 确认后将自动写  {m_write:>4} 笔  ← 强三键唯一命中（单号+是否更新）"],
        [f"⑦ 须你手填      {m_hand:>4} 笔  ← 弱命中/对不上/多命中，见「流转表怎么填」"],
        [""],
        ["怎么用："],
        ["  1. 翻到《今日清单》，按「状态」列筛「今天要填」。"],
        ["  2. 一行一个单号。拿「怎么找到这行」去盈亏『明细』筛出那一行。"],
        ["  3. 照「应填_xxx」几列对照；冲突行看「差异」。"],
        ["  4. 翻《流转表怎么填》：写入方式=自动 的确认后程序写；=手填 的你自己填。"],
        ["  5. 都看完了跟我说「确认」，我再统一写：先盈亏明细，再流转可写笔。"],
        [""],
        ["※ 这份只是清单，程序还没动你的任何表。"],
        ["※ 智云永远不写。"],
    ]
    for ln in lines:
        ws0.append(ln)
    ws0["A1"].font = Font(bold=True, size=14)
    ws0.column_dimensions["A"].width = 78

    # ── 今日清单（主表）────────────────────────────────────
    ws = wb.create_sheet("今日清单")
    ws.append(HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical="center", wrap_text=True)
    rows = collect_rows(result, checked)
    for r in rows:
        ws.append(r)
    for i, r in enumerate(rows, start=2):
        color = STATUS.get(r[1], (9, "FFFFFF"))[1]
        fill = PatternFill("solid", fgColor=color)
        ws.cell(row=i, column=1).fill = fill
        ws.cell(row=i, column=2).fill = fill
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{chr(64 + len(HEADERS))}{len(rows) + 1}" if len(HEADERS) <= 26 else None
    for col, w in (("A", 16), ("B", 14), ("C", 46), ("D", 14), ("E", 14), ("F", 16), ("T", 52), ("U", 40)):
        ws.column_dimensions[col].width = w

    # ── 按到账汇总（流转表建议 + 写入方式）──────────
    summary = result.get("ar_summary") or []
    flow_by_ar = {
        (it.get("ar") or ""): it for it in ((flow_plan or {}).get("items") or [])
    }
    if summary or flow_by_ar:
        try:
            from flow_ledger import flow_status_policy
        except Exception:  # pragma: no cover
            def flow_status_policy(_s):
                return {}

        wsum = wb.create_sheet("流转表怎么填")
        wsum.append([
            "到账号(AR)", "写入方式", "本笔关联SO数", "订单行数", "『是否更新应收款』建议填",
            "单号建议", "填法", "公式策略", "颜色标注", "还没处理完的SO", "流转表定位", "说明",
        ])
        for c in wsum[1]:
            c.font = Font(bold=True)
        rows_src = summary if summary else [
            {"ar": ar, "so_count": len(it.get("so_list") or []), "行数": 0,
             "流转表_是否更新应收款_建议": it.get("updated_suggest"),
             "flow_locate": it.get("flow_locate"), "待处理SO": []}
            for ar, it in flow_by_ar.items()
        ]
        for s in rows_src:
            ar = s.get("ar") or ""
            fp = flow_by_ar.get(ar) or {}
            verdict = fp.get("verdict") or ""
            mode = "自动" if verdict == "write" else ("手填" if verdict == "hand" else (verdict or "手填"))
            raw = s.get("流转表_是否更新应收款_建议")
            if raw is None:
                raw = fp.get("updated_suggest") or ""
            pol = flow_status_policy(raw)
            wsum.append([
                ar, mode, s.get("so_count") or 0, s.get("行数") or 0,
                (raw if raw else "（空白）"),
                fp.get("order_suggest") or "",
                pol.get("填法") or "", pol.get("公式策略") or "", pol.get("颜色标注") or "",
                " ".join(s.get("待处理SO") or []),
                s.get("flow_locate") or fp.get("flow_locate") or "",
                fp.get("reason") or pol.get("人话") or ("确认后自动写" if mode == "自动" else "须手填"),
            ])
        wsum.freeze_panes = "A2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path


def _latest(out_dir: Path, pattern: str) -> Optional[Path]:
    c = [p for p in sorted(out_dir.glob(pattern)) if not p.name.startswith("~$")]
    return c[-1] if c else None


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="生成《核销日清》（一份合一，只写产出目录）")
    ap.add_argument("--workspace", default=str(common.WORK))
    ap.add_argument("--result", default="", help="判定结果 json；默认取 04_产出 最新")
    ap.add_argument(
        "--checked", default="",
        help="validate_plan 产出的写入计划_校验后.json；默认取 04_产出 最新。"
             "给了才分得出「今天要填 / 已填过跳过 / 冲突」",
    )
    ap.add_argument("--out", default="", help="清单 xlsx 路径")
    ap.add_argument("--flow-plan", default="", help="流转写入计划_校验后.json；默认取 04_产出 最新")
    ap.add_argument("--hexiao-date", default="", help="（日期以判定结果为准，收下防止链路中断）")
    args = ap.parse_args(argv)

    ws = common.ensure_out_dirs(args.workspace)  # 解析真工作区，防产出分家
    out_dir = ws / "04_产出"

    result_path = Path(args.result) if args.result else _latest(out_dir, "判定结果_*.json")
    if not result_path or not result_path.is_file():
        print("ERROR: 找不到判定结果 json，请先跑 classify_hexiao.py", file=sys.stderr)
        return 2
    result = json.loads(result_path.read_text(encoding="utf-8"))

    checked_path = Path(args.checked) if args.checked else _latest(out_dir, "写入计划_校验后*.json")
    checked = None
    if checked_path and checked_path.is_file():
        checked = json.loads(checked_path.read_text(encoding="utf-8"))
    else:
        print(
            "WARN: 没有写入计划_校验后.json —— 清单里所有可填的都会标「今天要填」，"
            "**分不出你昨天已经填过的**。请先跑 validate_plan.py 再跑本脚本。",
            file=sys.stderr,
        )

    flow_plan = None
    fp_path = Path(args.flow_plan) if args.flow_plan else _latest(out_dir, "流转写入计划*.json")
    if fp_path and Path(fp_path).is_file():
        flow_plan = json.loads(Path(fp_path).read_text(encoding="utf-8"))
    else:
        # 无计划时现场生成一次（不写用户表）
        try:
            import build_flow_plan as BFP
            flow_plan = BFP.build_plan(result)
            auto_out = out_dir / "流转写入计划_校验后.json"
            auto_out.write_text(json.dumps(flow_plan, ensure_ascii=False, indent=2), encoding="utf-8")
            fp_path = auto_out
        except Exception as e:
            print(f"WARN: 无法生成流转计划：{e}", file=sys.stderr)

    # 文件名跟**核销日**走，不跟运行日走：她补跑 7-22 时若按运行日命名，
    # 会跟今天那批撞名甚至盖掉，事后翻出来也分不清哪份是哪天的。
    hexiao_date = common.norm_date(result.get("hexiao_date"))
    stamp = hexiao_date.strftime("%Y%m%d") if hexiao_date else dt.date.today().strftime("%Y%m%d")
    today = dt.date.today().strftime("%Y%m%d")
    out_path = Path(args.out) if args.out else out_dir / f"核销日清_{stamp}.xlsx"
    build_workbook(result, checked, out_path, flow_plan=flow_plan)

    rows = collect_rows(result, checked)
    tally: Dict[str, int] = {}
    for r in rows:
        tally[r[1]] = tally.get(r[1], 0) + 1
    fc = (flow_plan or {}).get("counts") or {}
    print(f"核销日期：{common.date_cn(hexiao_date) if hexiao_date else '(未知·请核对取数)'}")
    print("《核销日清》已生成：" + " / ".join(f"{k} {v}" for k, v in tally.items()))
    print(
        f"流转：确认后自动写 {fc.get('write', 0)} · 须手填 {fc.get('hand', 0)} · 跳过 {fc.get('skip', 0)}"
    )
    print(f"结果: {out_path}")

    if hexiao_date is not None:
        try:
            import batch_ledger

            batch_ledger.record(ws, hexiao_date, "listed", counts=tally)
        except Exception as e:
            print(f"WARN: 跑批台账登记失败（不影响清单）：{type(e).__name__}", file=sys.stderr)

    (out_dir / f"运行报告_{today}.txt").write_text(
        f"核销日期 {result.get('hexiao_date') or '(未知)'}\n"
        f"清单 {out_path.name}\n"
        f"判定 {result_path.name}\n"
        f"校验 {checked_path.name if checked else '(未跑 validate_plan)'}\n"
        f"流转计划 {Path(fp_path).name if fp_path else '(无)'}\n"
        f"到账笔数 {result.get('payment_count')}\n"
        f"计数 {result.get('counts')}\nE码 {result.get('e_code_dist')}\n"
        f"清单分布 {tally}\n"
        f"流转 counts {fc}\n",
        encoding="utf-8",
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
