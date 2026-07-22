#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
银行日记账 · 收入提取汇总（独立程序 · 出纳李明妹用）
====================================================
把一份"银行日记账.xlsx"里【每个 sheet（=每个银行/现金账户）】的"收入"记录
筛出来，汇总到一张表；顺带把"疑似漏标/缺客户名"的可疑行挑出来给人核对。

本文件是**核心库 + 命令行入口**；给出纳双击用的图形界面在 run_gui.py（它 import 本文件）。

傻瓜用法（双击 exe / 命令行不带参数）＝ input/output 文件夹模式：
  第一次双击 → 在 exe 同目录自动建好 input、output 两个文件夹；
  把银行日记账放进 input，再双击 → 逐个提取，结果写进 output，弹小结并打开 output。
识别规则**全部内置写死在本脚本**（下面的 FALLBACK / FALLBACK_CURRENCY），
不依赖任何外部配置文件——发出纳的就是一个光 exe，放哪个文件夹都能跑。

收入判定：某行「类型」列含"收入"类关键词 且「金额(借方)」列有数字 → 收入。
  （日记账里常有"类型=收入但金额空"的预填行，靠"有金额"过滤掉——最容易踩的坑。）

命令行用法（图形界面见 run_gui.py）：
    python3 extract_income.py                    # 文件夹模式（同目录 input/output，和双击一致）
    python3 extract_income.py 银行日记账.xlsx      # 单文件：结果放输入同目录
    python3 extract_income.py 银行日记账.xlsx -o 收入汇总.xlsx
"""
import sys, os, argparse, datetime, subprocess
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill


def app_dir():
    """程序所在目录：打包成 exe 后取 exe 所在目录（input/output 就建在这儿）；
    源码运行时取本脚本所在目录。"""
    if getattr(sys, "frozen", False):          # PyInstaller 冻结态
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


DEFAULT_CONFIG = None   # 识别规则内置写死，不再读外部文件（保留常量供旧测试引用）

# ── 内置识别规则（写死在程序里，无外部配置文件）──────────────────────────
FALLBACK = {
    "收入判定关键词": ["收入", "到账"],
    "支出类关键词": ["手续费", "利息", "支出", "转出", "服务费", "退款", "费用"],
    "日期列表头词": ["日期", "交易日期", "记账日期"],
    "客户名列表头词": ["摘要", "客户", "对方户名", "汇款人", "收款人", "对方"],
    "金额列表头词": ["借方", "增加", "收入金额", "收方"],
    "类型列表头词": ["类型", "类别"],
    "跳过的 sheet": ["说明", "目录", "封面", "模板", "汇总表"],
}
FALLBACK_CURRENCY = {  # 关键词 => 币种（按 sheet 名/行内标记匹配；含账户名兜底）
    "美元": "美元", "美金": "美元", "USD": "美元",
    "欧元": "欧元", "EUR": "欧元", "英镑": "英镑", "GBP": "英镑",
    "港币": "港币", "港元": "港币", "HKD": "港币", "日元": "日元", "JPY": "日元",
    # 账户名兜底：PayPal 账户收款为美元，但 sheet 名不含币种词，会被误判人民币，故写死
    "PayPal": "美元", "PAYPAL": "美元", "贝宝": "美元",
}
HEADER_SCAN_ROWS = 30   # 表头往下扫这么多行，容忍账户表前面有几行标题/空行


# ── config 解析 ──────────────────────────────────────────────────────
def load_config(path):
    """读 识别规则.md：按二级标题分节，节下 '- ' 列表项收集。
    外币节支持 '关键词 => 币种'。读不到就用兜底。"""
    lists = {k: list(v) for k, v in FALLBACK.items()}
    currency = dict(FALLBACK_CURRENCY)
    if not path or not os.path.isfile(path):
        return lists, currency
    section = None
    cur_from_config = {}
    try:
        with open(path, encoding="utf-8") as f:
            for raw in f:
                line = raw.rstrip("\n")
                s = line.strip()
                if s.startswith("## "):
                    section = s[3:].strip()
                    if section == "外币 sheet":
                        cur_from_config = {}
                    elif section in lists:
                        lists[section] = []       # 命中已知节 → 清空兜底、用配置
                    continue
                if s.startswith("- "):
                    item = s[2:].strip()
                    if not item:
                        continue
                    if section == "外币 sheet":
                        if "=>" in item:
                            k, v = item.split("=>", 1)
                            cur_from_config[k.strip()] = v.strip()
                    elif section in lists:
                        lists[section].append(item)
        if cur_from_config:
            currency = cur_from_config
    except Exception as e:
        print(f"⚠ 读 识别规则.md 失败（{e}），用内置默认规则继续。")
    # 任何节被配成空 → 回退兜底，防误清空
    for k, v in FALLBACK.items():
        if not lists.get(k):
            lists[k] = list(v)
    return lists, currency


def _norm(v):
    return "" if v is None else str(v).strip()


def to_number(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("，", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def norm_date(v):
    # 日期原样保留为「日期值」（写进 Excel 才能按时间排序/筛选/计算）；
    # 不是日期的（空、文本）走归一化成字符串。
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v
    return _norm(v)


# ── 识别表头 / 列 / 币种 ──────────────────────────────────────────────
def find_header_row(ws, cfg):
    date_keys = cfg["日期列表头词"]
    type_keys = cfg["类型列表头词"]
    amt_keys = cfg["金额列表头词"]
    for r in range(1, min(HEADER_SCAN_ROWS, ws.max_row) + 1):
        cells = [_norm(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
        has_date = any(any(k in x for k in date_keys) for x in cells)
        has_type = any(any(k in x for k in type_keys) for x in cells)
        has_amt = any(any(k in x for k in amt_keys) for x in cells)
        if has_date and (has_type or has_amt):
            return r
    return None


def map_columns(ws, header_row, cfg):
    colmap = {}
    headers = [(c, _norm(ws.cell(row=header_row, column=c).value))
               for c in range(1, ws.max_column + 1)]
    logical_keys = {
        "日期": cfg["日期列表头词"], "摘要": cfg["客户名列表头词"],
        "借方": cfg["金额列表头词"], "类型": cfg["类型列表头词"],
    }
    for logical, keys in logical_keys.items():
        for c, text in headers:
            if text and any(k in text for k in keys):
                colmap[logical] = c
                break
    # 『类型』列兜底：真表里常见"类型列的表头是空的、或被写成『单位：元』"，
    # 但它稳定地紧跟在『余额』列后面。所以表头认不出类型时，取余额右边一列。
    if "类型" not in colmap:
        bal_c = None
        for c, text in headers:
            if text and "余额" in text:
                bal_c = c
                break
        if bal_c is not None and bal_c + 1 <= ws.max_column:
            colmap["类型"] = bal_c + 1
    return colmap


def detect_currency(ws, sheet_name, currency_map):
    up = sheet_name.upper()
    for kw, cur in currency_map.items():
        if kw in sheet_name or kw.upper() in up:
            return cur
    for row in ws.iter_rows(values_only=True):
        for v in row:
            s = _norm(v)
            if s.startswith("币种") and len(s) > 2:
                return s[2:]
    return "人民币"


# ── 核心提取 ──────────────────────────────────────────────────────────
def is_income(type_text, cfg):
    return any(k in type_text for k in cfg["收入判定关键词"])


def is_known_expense(type_text, cfg):
    return any(k in type_text for k in cfg["支出类关键词"])


def load_any(path):
    """统一读表：.xlsx/.xlsm 走 openpyxl；.xls 走 xlrd 读出后转成等价 openpyxl 工作簿
    （逐格搬值，日期序列号转成真正的日期）。这样出纳直接丢系统导出的原始 .xls 也能跑，
    不用先『另存为 xlsx』。下游 extract() 统一按 openpyxl 处理，无需分支。"""
    low = path.lower()
    if low.endswith((".xlsx", ".xlsm")):
        return openpyxl.load_workbook(path, data_only=True)
    if low.endswith(".xls"):
        import xlrd
        book = xlrd.open_workbook(path)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for sh in book.sheets():
            title = (sh.name or "Sheet")[:31]
            for ch in "[]:*?/\\":
                title = title.replace(ch, "_")
            ws = wb.create_sheet(title)
            for r in range(sh.nrows):
                out = []
                for c in range(sh.ncols):
                    cell = sh.cell(r, c)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            out.append(xlrd.xldate_as_datetime(cell.value, book.datemode))
                            continue
                        except Exception:
                            pass
                    out.append(None if cell.ctype == xlrd.XL_CELL_EMPTY else cell.value)
                ws.append(out)
        return wb
    return openpyxl.load_workbook(path, data_only=True)


def extract(path, cfg, currency_map):
    wb = load_any(path)
    rows_out = []      # [渠道, 日期, 客户名称, 金额, 币种]
    flags = []         # 待人工：[渠道, 位置(行号), 问题, 相关值]
    report = []        # 每 sheet 小结
    for sheet_name in wb.sheetnames:          # 保持工作簿 sheet 先后顺序
        if any(sk in sheet_name for sk in cfg["跳过的 sheet"]):
            report.append((sheet_name, "跳过", "命中『跳过的 sheet』规则"))
            continue
        ws = wb[sheet_name]
        hr = find_header_row(ws, cfg)
        if hr is None:
            report.append((sheet_name, "跳过", "无『借方/类型』式表头——像理财/证券/债券台账或说明页，无收支明细，不提取"))
            continue
        colmap = map_columns(ws, hr, cfg)
        missing = [k for k in ("类型", "借方") if k not in colmap]
        if missing:
            report.append((sheet_name, "跳过", f"缺关键列 {missing}（对不上『列表头词』），无法判定收入"))
            continue
        currency = detect_currency(ws, sheet_name, currency_map)

        n_income = 0
        for r in range(hr + 1, ws.max_row + 1):
            typ = _norm(ws.cell(row=r, column=colmap["类型"]).value)
            amount = to_number(ws.cell(row=r, column=colmap["借方"]).value)
            cust = _norm(ws.cell(row=r, column=colmap["摘要"]).value) if "摘要" in colmap else ""
            date_v = norm_date(ws.cell(row=r, column=colmap["日期"]).value) if "日期" in colmap else ""

            if is_income(typ, cfg):
                if amount is None or amount == 0:
                    continue                      # 空预填收入行 → 跳过
                rows_out.append([sheet_name, date_v, cust, amount, currency])
                n_income += 1
                if not cust:                      # 收入但没客户名 → 下游创建回款要用，挑出来
                    flags.append([sheet_name, f"第{r}行", "收入行缺客户名(摘要空)", f"{amount:,.2f} {currency}"])
            else:
                # 有进账金额、但『类型』整个空着 → 可能忘了标（真·漏标收入）。
                # 她若填了任何非收入类型（保证金/内部调拨/税费/工资社保/译费…）＝故意归类，
                # 不是漏标，不打扰她——否则全年真数据会刷出上百条误报。
                if amount and amount > 0 and not typ:
                    flags.append([sheet_name, f"第{r}行",
                                  "有进账金额但『类型』空着(可能漏标收入)",
                                  f"金额={amount:,.2f} 摘要='{cust}'"])
        report.append((sheet_name, "完成", f"收入 {n_income} 笔（币种：{currency}）"))
    return rows_out, flags, report


# ── 写输出 ────────────────────────────────────────────────────────────
def write_output(rows_out, flags, out_path, report=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "收入汇总"
    headers = ["渠道（银行/账户）", "日期", "到账客户名称", "金额", "币种"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    for row in rows_out:
        ws.append(row)
        if not _norm(row[2]):                     # 缺客户名的行标黄
            for c in range(1, 6):
                ws.cell(row=ws.max_row, column=c).fill = warn_fill
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=4).number_format = "#,##0.00"
        dcell = ws.cell(row=r, column=2)          # 日期列：真日期值显示成 2026-06-24
        if isinstance(dcell.value, (datetime.datetime, datetime.date)):
            dcell.number_format = "yyyy-mm-dd"
    for i, w in enumerate([18, 12, 34, 15, 8], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # 待人工核对（有才建）
    if flags:
        wf = wb.create_sheet("待人工核对")
        wf.append(["渠道", "位置", "问题", "相关值"])
        for cell in wf[1]:
            cell.font = Font(bold=True)
        for f in flags:
            wf.append(f)
        for i, w in enumerate([16, 10, 30, 40], start=1):
            wf.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        wf.freeze_panes = "A2"

    # 各账户处理情况（覆盖清单）：每个 sheet 到底被处理还是跳过、提取几笔、为啥跳，
    # 一张表摊清楚——她一眼确认"所有子账户都算进来了"，没有谁被悄悄漏掉。
    if report:
        wc = wb.create_sheet("各账户处理情况")
        wc.append(["账户（sheet）", "结果", "说明"])
        for cell in wc[1]:
            cell.font = Font(bold=True)
        skip_fill = PatternFill("solid", fgColor="FCE4E4")   # 跳过的标红，最扎眼
        for name, status, detail in (report or []):
            wc.append([name, status, detail])
            if status == "跳过":
                for c in range(1, 4):
                    wc.cell(row=wc.max_row, column=c).fill = skip_fill
        for i, w in enumerate([24, 8, 46], start=1):
            wc.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        wc.freeze_panes = "A2"

    wb.save(out_path)


# ── 共享辅助（CLI 和 GUI 都用）────────────────────────────────────────
def default_output_path(input_path):
    """输出默认放输入文件同目录，命名『收入汇总_<输入名>_<日期>.xlsx』。"""
    base = os.path.splitext(os.path.basename(input_path))[0]
    today = datetime.date.today().strftime("%Y%m%d")
    return os.path.join(os.path.dirname(os.path.abspath(input_path)),
                        f"收入汇总_{base}_{today}.xlsx")


def summarize(rows_out, flags):
    """把提取结果压成给人看的几行小结（CLI 打印、GUI 弹窗都用这份）。"""
    by_cur = defaultdict(lambda: [0, 0.0])
    for _, _, _, amt, cur in rows_out:
        by_cur[cur][0] += 1
        by_cur[cur][1] += amt
    lines = [f"合计提取收入 {len(rows_out)} 笔"]
    if len(by_cur) > 1:
        lines.append("（多币种，未跨币种加总）")
    for cur, (cnt, total) in by_cur.items():
        lines.append(f"  · {cur}：{cnt} 笔，小计 {total:,.2f}")
    if flags:
        lines.append(f"\n⚠ 有 {len(flags)} 处需人工核对（见结果表『待人工核对』sheet）")
    else:
        lines.append("\n全部干净，无待人工核对项 ✓")
    return "\n".join(lines)


def open_file(path):
    """跑完自动打开结果文件（各平台各法，失败静默不影响主流程）。"""
    try:
        if sys.platform.startswith("win"):
            os.startfile(path)                                  # noqa: 仅 Windows 有
        elif sys.platform == "darwin":
            subprocess.run(["open", path], check=False)
        else:
            subprocess.run(["xdg-open", path], check=False)
    except Exception:
        pass


def process(input_path, out_path=None, config_path=DEFAULT_CONFIG):
    """一站式：取内置规则 → 提取 → 写输出。返回 (rows, flags, report, out_path)。
    GUI 直接调这个；CLI 也走它。config_path 默认 None＝用内置规则。"""
    cfg, currency_map = load_config(config_path)
    if not out_path:
        out_path = default_output_path(input_path)
    rows_out, flags, report = extract(input_path, cfg, currency_map)
    write_output(rows_out, flags, out_path, report)
    return rows_out, flags, report, out_path


# ── input/output 文件夹模式（双击 exe 的主流程）──────────────────────────
def find_input_files(input_dir):
    """扫描 input，返回可处理的 .xlsx/.xlsm 完整路径列表。"""
    return scan_input(input_dir)[0]


def scan_input(input_dir):
    """扫描 input 目录，返回 (excels, ignored)：
      excels  = 可处理的 .xlsx/.xlsm 完整路径列表；
      ignored = [(文件名, 被忽略的原因)]，用于在"没找到"时告诉用户到底看到了什么，
                好判断是放错了文件夹、还是文件格式不对。
    跳过 Excel 打开时的临时锁文件 ~$、隐藏文件（不计入 ignored，避免噪音）。"""
    excels, ignored = [], []
    if not os.path.isdir(input_dir):
        return excels, ignored
    for f in sorted(os.listdir(input_dir)):
        full = os.path.join(input_dir, f)
        if f.startswith("~$") or f.startswith("."):
            continue
        if os.path.isdir(full):
            ignored.append((f, "是子文件夹，不是文件"))
            continue
        low = f.lower()
        if low.endswith((".xlsx", ".xlsm", ".xls")):   # .xls 也能直接吃（内部用 xlrd 读）
            excels.append(full)
        else:
            ignored.append((f, "不是 Excel 文件"))
    return excels, ignored


def run_folder(base_dir):
    """在 base_dir 下找 input / output（没有就建），把 input 里每个 Excel 提取到 output。
    返回 (status, input_dir, output_dir, payload)：
      status='created' 首次运行、刚建好空 input（提示用户放文件）；payload={'ignored': [...]}
      status='empty'   input 已存在但没有可处理的 Excel；payload={'ignored': [(名,原因)...]}
      status='done'    跑完，payload=results（每项 {name, ok, out_path, rows, flags, report, error}）"""
    input_dir = os.path.join(base_dir, "input")
    output_dir = os.path.join(base_dir, "output")
    first_time = not os.path.isdir(input_dir)
    os.makedirs(input_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)

    inputs, ignored = scan_input(input_dir)
    if not inputs:
        return ("created" if first_time else "empty"), input_dir, output_dir, {"ignored": ignored}

    today = datetime.date.today().strftime("%Y%m%d")
    results = []
    for path in inputs:
        name = os.path.basename(path)
        base = os.path.splitext(name)[0]
        out_path = os.path.join(output_dir, f"收入汇总_{base}_{today}.xlsx")
        item = {"name": name, "ok": False, "out_path": out_path,
                "rows": [], "flags": [], "report": [], "error": ""}
        try:
            rows, flags, report, out_path = process(path, out_path)
            item.update(ok=True, out_path=out_path, rows=rows, flags=flags, report=report)
        except PermissionError:
            item["error"] = "结果文件可能正开着（上次的『收入汇总…xlsx』在 Excel 里没关），已跳过。"
        except Exception as e:
            item["error"] = str(e)
        results.append(item)
    return "done", input_dir, output_dir, results


def summarize_folder(results):
    """把多文件批处理结果压成给人看的小结（GUI 弹窗 / CLI 打印共用）。"""
    ok = [r for r in results if r["ok"]]
    bad = [r for r in results if not r["ok"]]
    head = f"共处理 {len(results)} 个文件，成功 {len(ok)} 个"
    if bad:
        head += f"，跳过 {len(bad)} 个"
    lines = [head + "。"]
    for r in results:
        lines.append("")
        if r["ok"]:
            rep = r["report"]
            done = [x for x in rep if x[1] == "完成"]
            skipped = [x for x in rep if x[1] == "跳过"]
            sub = summarize(r["rows"], r["flags"]).replace("\n", "\n    ")
            lines.append(f"【{r['name']}】\n    " + sub)
            # 每个账户(sheet)逐个报账——处理了几个、跳过几个，一眼看全不全
            lines.append(f"    —— 各账户：处理 {len(done)} 个、跳过 {len(skipped)} 个 ——")
            for n, st, d in rep:
                lines.append(f"      · {n}：{d}" if st == "完成" else f"      ✗ {n}：跳过（{d}）")
            if skipped:
                lines.append("    ⚠ 上面带 ✗ 的账户没算进汇总，结果表『各账户处理情况』有完整清单，请核对是否该处理。")
        else:
            lines.append(f"【{r['name']}】✗ {r['error']}")
    return "\n".join(lines)


# ── 命令行入口 ────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="银行日记账收入提取汇总（ar-hexiao-daily 第2步）")
    ap.add_argument("input", nargs="?",
                    help="银行日记账 .xlsx（不填＝用同目录 input/output 文件夹批处理）")
    ap.add_argument("--input", dest="input_opt", help="同 positional input（任务书风格）")
    ap.add_argument("-o", "--output", "--out", dest="output",
                    help="输出 xlsx 或目录（目录则自动命名）")
    args = ap.parse_args()
    if getattr(args, "input_opt", None) and not args.input:
        args.input = args.input_opt
    # 若 --out 是目录，落到目录内
    if args.input and args.output and os.path.isdir(args.output):
        base = os.path.splitext(os.path.basename(args.input))[0]
        today = datetime.date.today().strftime("%Y%m%d")
        args.output = os.path.join(args.output, f"收入汇总_{base}_{today}.xlsx")
    elif args.input and args.output and not str(args.output).lower().endswith((".xlsx", ".xlsm")):
        os.makedirs(args.output, exist_ok=True)
        base = os.path.splitext(os.path.basename(args.input))[0]
        today = datetime.date.today().strftime("%Y%m%d")
        args.output = os.path.join(args.output, f"收入汇总_{base}_{today}.xlsx")

    # 文件夹模式（不带文件名）
    if not args.input:
        status, input_dir, output_dir, payload = run_folder(app_dir())
        if status in ("created", "empty"):
            print(f"扫描的 input 文件夹：{input_dir}")
            if status == "created":
                print("（首次运行，已建好 input / output）把银行日记账放进 input，再重跑。")
            else:
                print("里面没有找到 .xlsx 文件。把银行日记账放进去再重跑。")
            for name, why in payload.get("ignored", []):
                print(f"  · 忽略『{name}』——{why}")
            return
        print(summarize_folder(payload))
        print(f"\n✓ 结果都在：{output_dir}")
        return

    # 单文件模式
    if not os.path.isfile(args.input):
        sys.exit(f"找不到文件：{args.input}")
    rows_out, flags, report, out_path = process(args.input, args.output)
    print("—— 各 sheet 处理小结 ——")
    for name, status, detail in report:
        print(f"  [{name}] {status}：{detail}")
    print()
    print(summarize(rows_out, flags))
    if flags:
        print("\n待人工明细（前 10 条）：")
        for f in flags[:10]:
            print(f"    [{f[0]}] {f[1]}  {f[2]}  {f[3]}")
        if len(flags) > 10:
            print(f"    …… 还有 {len(flags) - 10} 处")
    print(f"\n✓ 已写出 → {out_path}")


if __name__ == "__main__":
    main()
