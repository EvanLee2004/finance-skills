#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
回填前人工审核单（plan → validate → **review** → 她确认 → execute）。

把「准备改什么 / 改成啥 / 依据是啥」摊成一张 Excel，给明妹一眼扫完。
**不写她的盈亏表、不写流转表**——只写 04_产出/。

明妹 2026-07-23+：AI 回填前必须过这关；确认词见 SKILL（确认/OK/可以写/按这个写）。
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
import common  # noqa: E402
from flow_ledger import flow_status_policy  # noqa: E402

FIVE = ["计提", "回款明细", "是否结账", "收款时间", "收款方式"]

# 明妹确认前，apply 必须看到这个标记
REVIEW_STAMP_NAME = "回填审核_待确认.json"


def _norm_cell(v: Any) -> str:
    if v is None:
        return "（空）"
    if isinstance(v, (dt.date, dt.datetime)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return s if s else "（空）"


def _risk_of(item: dict) -> str:
    parts: List[str] = []
    chk = item.get("_check") or {}
    if chk.get("verdict") == "conflict":
        parts.append(chk.get("reason") or "冲突")
    matched = (item.get("flow_matched_by") or "") + (item.get("matched_by") or "")
    if "弱" in matched:
        parts.append("流转表弱命中（公司名对不上）")
    five = item.get("five_cols") or {}
    if five.get("是否结账") == "否":
        parts.append("部分核销：计提留空；表上可能要你手插行")
    cur = item.get("current_values") or {}
    for k in FIVE:
        if _norm_cell(cur.get(k)) not in ("（空）",) and five.get(k) is not None:
            if _norm_cell(cur.get(k)) != _norm_cell(five.get(k)):
                parts.append(f"表里已有{k}且与拟写不同")
    if item.get("channel") == "yucun":
        parts.append("预存通道")
    return "；".join(parts) if parts else "无额外风险"


def _delta_text(item: dict) -> str:
    five = item.get("five_cols") or {}
    cur = item.get("current_values") or {}
    # 校验后计划里可能没有 current_values，用五列空对比
    bits = []
    for k in FIVE:
        new = five.get(k)
        if new is None and k == "计提" and five.get("是否结账") == "否":
            new_s = "（留空·部分核销）"
        else:
            new_s = _norm_cell(new)
        old_s = _norm_cell(cur.get(k)) if cur else "（空）"
        if old_s != new_s:
            bits.append(f"{k}: {old_s} → {new_s}")
        else:
            bits.append(f"{k}: {new_s}")
    return " | ".join(bits)


def _basis(item: dict) -> str:
    parts = [
        item.get("reason") or "",
        f"通道={item.get('channel') or '—'}",
    ]
    if item.get("sod"):
        parts.append(f"SOD={item.get('sod')}")
    if item.get("so"):
        parts.append(f"SO={item.get('so')}")
    if item.get("locate_hint"):
        parts.append(str(item.get("locate_hint")))
    chk = item.get("_check") or {}
    if chk.get("reason"):
        parts.append(f"校验={chk['reason']}")
    return "；".join(p for p in parts if p)


def _row_write(item: dict) -> List[Any]:
    five = item.get("five_cols") or {}
    cur = item.get("current_values") or {}
    return [
        item.get("case_id") or "",
        item.get("ar") or "",
        item.get("so") or "",
        item.get("sod") or five.get("实收SOD") or "",
        item.get("ledger_row_ref") or "",
        item.get("locate_hint")
        or f"在盈亏『明细』按 SO 筛选：{item.get('so') or ''}（禁止用行号）",
        _norm_cell(cur.get("计提")),
        _norm_cell(five.get("计提")) if five.get("计提") is not None else "（留空）",
        _norm_cell(cur.get("回款明细")),
        _norm_cell(five.get("回款明细")),
        _norm_cell(cur.get("是否结账")),
        _norm_cell(five.get("是否结账")),
        _norm_cell(cur.get("收款时间")),
        _norm_cell(five.get("收款时间")),
        _norm_cell(cur.get("收款方式")),
        _norm_cell(five.get("收款方式")),
        _delta_text(item),
        _basis(item),
        item.get("channel") or "",
        item.get("flow_locate") or "",
        item.get("flow_order_suggest") or "",
        _risk_of(item),
        "拟写入（等你确认）",
    ]


def _row_skip_conflict(item: dict, label: str) -> List[Any]:
    chk = item.get("_check") or {}
    return [
        item.get("case_id") or "",
        item.get("ar") or "",
        item.get("so") or "",
        item.get("sod") or "",
        item.get("ledger_row_ref") or "",
        label,
        chk.get("reason") or item.get("reason") or "",
        _basis(item),
        _risk_of(item),
    ]


def _enrich_ar_summary(summary: List[dict], checked: dict) -> List[dict]:
    """把校验结果并进按到账汇总，方便流转三态说明。"""
    write_sos = {
        (i.get("ar"), i.get("so"))
        for i in (checked.get("write") or [])
        if i.get("so")
    }
    skip_sos = {
        (i.get("ar"), i.get("so"))
        for i in (checked.get("skip") or [])
        if i.get("so")
    }
    out = []
    for s in summary:
        ar = s.get("ar")
        # 已可写或已幂等 = 本次「更新侧」
        done_so = []
        pending_so = list(s.get("待处理SO") or [])
        # 从 buckets 再推一遍
        # summary 未必有 SO 列表，尽量用 待处理SO + write 推断
        policy = flow_status_policy(s.get("流转表_是否更新应收款_建议") or "")
        out.append(
            {
                **s,
                "已更新或可写SO": sorted(
                    {so for (a, so) in write_sos | skip_sos if a == ar and so}
                ),
                "未处理完SO": pending_so,
                "是否更新_填法": policy["填法"],
                "公式策略": policy["公式策略"],
                "颜色标注": policy["颜色标注"],
                "人话": policy["人话"],
            }
        )
    return out


def _attach_current_from_ledger(checked: dict, ledger: Optional[Path]) -> dict:
    """校验计划里若缺 current_values，从盈亏明细补一眼改前值（只读）。"""
    if not ledger or not Path(ledger).is_file():
        return checked
    try:
        from validate_plan import read_ledger_rows  # noqa: WPS433
    except Exception:
        return checked
    try:
        rows = read_ledger_rows(Path(ledger))
    except Exception:
        return checked

    def fill(items: List[dict]) -> List[dict]:
        out = []
        for it in items:
            it = dict(it)
            ref = it.get("ledger_row_ref")
            if ref and not it.get("current_values"):
                row = rows.get(int(ref)) or {}
                it["current_values"] = {k: row.get(k) for k in FIVE}
            out.append(it)
        return out

    return {
        **checked,
        "write": fill(list(checked.get("write") or [])),
        "skip": fill(list(checked.get("skip") or [])),
        "conflict": fill(list(checked.get("conflict") or [])),
    }


def build_workbook(
    checked: dict,
    out_path: Path,
    ar_summary: Optional[List[dict]] = None,
) -> Path:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()

    # ---- 使用说明 ----
    wu = wb.active
    wu.title = "使用说明"
    lines = [
        "回填审核单（写表前必看）",
        "",
        "这张表只说明「准备改什么」，还没有动你的盈亏表。",
        "",
        "怎么用：",
        "1. 打开「盈亏回填·拟写入」：逐行看「改前→改后」和「判定依据」。",
        "2. 有问题的告诉同事「跳过某某 SO」或「这笔不对」。",
        "3. 没问题就回一句：确认 / OK / 可以写 / 按这个写。",
        "4. 同事确认后才会写入你盈亏工作副本的「明细」sheet（写前自动备份）。",
        "",
        "页签说明：",
        "· 盈亏回填·拟写入 —— 校验通过、准备写入的笔（核心）",
        "· 幂等跳过 —— 表里已经是同样内容，再跑也不会改",
        "· 冲突·需你定 —— 对不上，程序不会自动写",
        "· 流转表·是否更新建议 —— 到账流转表「是否更新应收款」三态（只建议，不自动写流转表）",
        "",
        "流转表三态（明妹口径）：",
        "· 全部订单都没检索到/都不能更 → 「是否更新」留空，不要公式",
        "· 全部订单都检索到并更新 → 填「是」（你表若有惯用公式可套公式）",
        "· 有的更了有的没 → 填「部分」，未更新的用红色标出",
        "",
        f"生成时间：{dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
    ]
    for i, line in enumerate(lines, start=1):
        wu.cell(i, 1, line)
        if i == 1:
            wu.cell(i, 1).font = Font(bold=True, size=14)
    wu.column_dimensions["A"].width = 90

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    write_fill = PatternFill("solid", fgColor="E2EFDA")  # 绿：拟写
    conflict_fill = PatternFill("solid", fgColor="FCE4D6")  # 橙：冲突
    skip_fill = PatternFill("solid", fgColor="D9E2F3")  # 蓝：跳过
    partial_fill = PatternFill("solid", fgColor="FFF2CC")  # 黄：部分

    # ---- 拟写入 ----
    ww = wb.create_sheet("盈亏回填·拟写入")
    headers_w = [
        "案例ID",
        "AR",
        "SO",
        "SOD",
        "计划行号(仅校验用)",
        "怎么找到这行",
        "当前_计提",
        "拟写_计提",
        "当前_回款明细",
        "拟写_回款明细",
        "当前_是否结账",
        "拟写_是否结账",
        "当前_收款时间",
        "拟写_收款时间",
        "当前_收款方式",
        "拟写_收款方式",
        "改前→改后(一眼扫)",
        "判定依据",
        "通道",
        "流转表定位",
        "流转表单号建议",
        "风险提示",
        "状态",
    ]
    ww.append(headers_w)
    for cell in ww[1]:
        cell.font = header_font
        cell.fill = header_fill
    for item in checked.get("write") or []:
        ww.append(_row_write(item))
        for cell in ww[ww.max_row]:
            cell.fill = write_fill
    ww.freeze_panes = "A2"
    ww.auto_filter.ref = f"A1:W{max(ww.max_row, 1)}"

    # ---- 跳过 ----
    ws = wb.create_sheet("幂等跳过")
    headers_s = [
        "案例ID",
        "AR",
        "SO",
        "SOD",
        "计划行号",
        "状态",
        "原因",
        "判定依据",
        "风险提示",
    ]
    ws.append(headers_s)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for item in checked.get("skip") or []:
        ws.append(_row_skip_conflict(item, "已填过且一致·跳过"))
        for cell in ws[ws.max_row]:
            cell.fill = skip_fill
    ws.freeze_panes = "A2"

    # ---- 冲突 ----
    wc = wb.create_sheet("冲突·需你定")
    wc.append(headers_s)
    for cell in wc[1]:
        cell.font = header_font
        cell.fill = header_fill
    for item in checked.get("conflict") or []:
        wc.append(_row_skip_conflict(item, "冲突·不自动写"))
        for cell in wc[wc.max_row]:
            cell.fill = conflict_fill
    wc.freeze_panes = "A2"

    # ---- 流转建议 ----
    wf = wb.create_sheet("流转表·是否更新建议")
    headers_f = [
        "AR",
        "本笔关联SO数",
        "是否更新应收款_建议",
        "是否更新_填法",
        "公式策略",
        "颜色标注",
        "已更新或可写SO",
        "未处理完SO",
        "流转表定位",
        "人话",
    ]
    wf.append(headers_f)
    for cell in wf[1]:
        cell.font = header_font
        cell.fill = header_fill
    enriched = _enrich_ar_summary(ar_summary or [], checked)
    for s in enriched:
        sug = s.get("流转表_是否更新应收款_建议") or "（空白）"
        row = [
            s.get("ar") or "",
            s.get("so_count") or 0,
            sug if sug else "（空白）",
            s.get("是否更新_填法") or "",
            s.get("公式策略") or "",
            s.get("颜色标注") or "",
            " ".join(s.get("已更新或可写SO") or []),
            " ".join(s.get("未处理完SO") or s.get("待处理SO") or []),
            s.get("flow_locate") or "",
            s.get("人话") or "",
        ]
        wf.append(row)
        fill = None
        if sug == "部分":
            fill = partial_fill
        elif sug == "是":
            fill = write_fill
        if fill:
            for cell in wf[wf.max_row]:
                cell.fill = fill
    wf.freeze_panes = "A2"

    # 列宽略收一收
    for sheet in (ww, ws, wc, wf):
        for col in sheet.columns:
            letter = col[0].column_letter
            sheet.column_dimensions[letter].width = min(28, max(10, len(str(col[0].value or "")) + 4))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path


def write_stamp(out_dir: Path, checked_path: Path, review_xlsx: Path, counts: dict) -> Path:
    """待确认标记：apply_to_copy 无 --confirmed 时据此拒绝写入。"""
    stamp = {
        "status": "pending_review",
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "checked_plan": str(checked_path.resolve()),
        "review_xlsx": str(review_xlsx.resolve()),
        "counts": counts,
        "how_to_confirm": "明妹说「确认/OK/可以写/按这个写」后，apply 加 --confirmed",
    }
    p = out_dir / REVIEW_STAMP_NAME
    p.write_text(json.dumps(stamp, ensure_ascii=False, indent=2), encoding="utf-8")
    return p


def load_ar_summary(plan_path: Path, explicit: str = "") -> List[dict]:
    def _from_obj(data) -> List[dict]:
        if isinstance(data, list):
            return data
        if isinstance(data, dict):
            return list(data.get("ar_summary") or [])
        return []

    if explicit:
        p = Path(explicit)
        if p.is_file():
            return _from_obj(json.loads(p.read_text(encoding="utf-8")))
    # 同目录最新判定结果
    folder = plan_path.parent
    candidates = sorted(folder.glob("判定结果_*.json"), key=lambda x: x.stat().st_mtime, reverse=True)
    for c in candidates:
        try:
            got = _from_obj(json.loads(c.read_text(encoding="utf-8")))
            if got:
                return got
        except Exception:
            continue
    return []


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="生成回填审核单（写前人工闸）")
    ap.add_argument("--checked", required=True, help="validate_plan 产出的校验后计划 json")
    ap.add_argument("--out", default="", help="审核单 xlsx 路径")
    ap.add_argument("--workspace", default=str(common.WORK), help="工作区（用于默认产出目录）")
    ap.add_argument("--result", default="", help="可选：判定结果 json（取 ar_summary）")
    ap.add_argument("--ledger", default="", help="可选：盈亏副本（补「当前值」列）")
    args = ap.parse_args(argv)

    checked_p = Path(args.checked)
    if not checked_p.is_file():
        print(f"ERROR: 找不到校验后计划 {checked_p}", file=sys.stderr)
        return 2

    checked = json.loads(checked_p.read_text(encoding="utf-8"))
    if args.ledger:
        checked = _attach_current_from_ledger(checked, Path(args.ledger))
    counts = checked.get("counts") or {
        "write": len(checked.get("write") or []),
        "skip": len(checked.get("skip") or []),
        "conflict": len(checked.get("conflict") or []),
    }

    ws = Path(args.workspace)
    out_dir = ws / "04_产出" if (ws / "04_产出").is_dir() else checked_p.parent
    today = dt.date.today().strftime("%Y%m%d")
    out_path = Path(args.out) if args.out else out_dir / f"回填审核单_{today}.xlsx"

    ar_summary = load_ar_summary(checked_p, args.result)
    build_workbook(checked, out_path, ar_summary=ar_summary)
    stamp = write_stamp(out_dir, checked_p, out_path, counts)

    print(
        f"回填审核单 → {out_path}\n"
        f"  拟写入={counts.get('write', 0)}  跳过={counts.get('skip', 0)}  冲突={counts.get('conflict', 0)}\n"
        f"待确认标记 → {stamp}\n"
        f"⚠ 还没写她的表。等她确认后再：apply_to_copy … --confirmed"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
