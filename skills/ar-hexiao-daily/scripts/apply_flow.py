#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
写到账流转表安全子集：仅「单号」「是否更新应收款」。

日清与写前校验通过后直接执行；--confirmed 仅为旧命令兼容参数。
只处理 verdict=write；写前备份；优先 xlsx_patch；回读比对。
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import shutil
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
import common  # noqa: E402

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass


def _resolve_flow_path(workspace: Path, file_name: str) -> Optional[Path]:
    d = workspace / "02_我的表副本"
    if not d.is_dir():
        return None
    p = d / file_name
    if p.is_file():
        return p
    # 允许只有 basename 匹配
    for cand in d.glob("*.xlsx"):
        if cand.name == file_name or cand.name.endswith(file_name):
            return cand
    return None


def _locate_cols(ws_rows, aliases) -> Tuple[int, Dict[str, int]]:
    """返回 (header_row_0based, {字段: 0-based col})。"""
    hrow, headers = common.find_header_row(
        ws_rows, "到账流转", ["日期", "公司名称", "金额", "单号"], aliases
    )
    cols = common.resolve_columns(
        headers, "到账流转", ["单号"], aliases
    )
    # 是否更新应收款 optional but required for our write
    opt = common.fuzzy_find_col(
        headers, aliases.get("到账流转", {}).get("是否更新应收款", ["是否更新应收款"])
    )
    if opt is None:
        # try direct
        opt = common.fuzzy_find_col(headers, ["是否更新应收款", "是否更新"])
    if opt is None:
        raise ValueError("流转表找不到「是否更新应收款」列")
    cols["是否更新应收款"] = opt
    # 身份列（可缺）：写入前拿来核"这一行还是当初命中的那一行吗"
    for key, fallback in (
        ("日期", ["日期", "到账日期"]),
        ("公司名称", ["公司名称", "公司", "付款方"]),
        ("金额", ["金额", "到账金额"]),
    ):
        idx = common.fuzzy_find_col(headers, aliases.get("到账流转", {}).get(key, fallback))
        if idx is not None:
            cols[key] = idx
    return hrow, cols


def precheck_flow_identity(workspace: Path, items: List[dict]) -> List[str]:
    """
    写流转表之前复核每条 write 项的行身份（2026-07-25 立）。

    准入（强三键唯一命中）是在 `build_flow_plan` 那一刻算的，到真正写入之间，
    她的到账流转表仍可能被其它进程或人工插入新行；中间插一行，
    row_no 之后的全部错位 → 单号会写到别人那笔到账上。
    所以这里拿计划里记的 (日期/公司名称/金额) 跟**当前**那一行再对一次。

    任一条对不上 → 整批不写流转（不是只跳过那一条：插行会让它之后的全错位）。
    """
    import openpyxl

    aliases = common.load_aliases()
    problems: List[str] = []
    by_file: Dict[str, List[dict]] = {}
    for it in items:
        by_file.setdefault(it.get("file") or "", []).append(it)

    for fname, group in by_file.items():
        src = _resolve_flow_path(workspace, fname)
        if not src or not src.is_file():
            problems.append(f"找不到流转文件 {fname}")
            continue
        wb = openpyxl.load_workbook(str(src), read_only=True, data_only=True)
        try:
            for sheet_name, g2 in _group_by(group, "sheet").items():
                if sheet_name not in wb.sheetnames:
                    problems.append(f"sheet 不存在 {sheet_name}")
                    continue
                rows = list(wb[sheet_name].iter_rows(values_only=True))
                try:
                    _hrow, cols = _locate_cols(rows, aliases)
                except Exception as e:
                    problems.append(f"{fname}#{sheet_name}: 列定位失败 {e}")
                    continue
                for it in g2:
                    ident = it.get("identity") or {}
                    if not ident:
                        continue  # 旧计划没记身份：靠强三键准入，不额外拦
                    r = int(it["row_no"])
                    if r - 1 >= len(rows):
                        problems.append(f"{it.get('ar')}: 第 {r} 行现在不存在了（表被删过行？）")
                        continue
                    vals = list(rows[r - 1])

                    def cell(key):
                        i = cols.get(key)
                        return vals[i] if i is not None and i < len(vals) else None

                    if "日期" in cols and ident.get("date"):
                        now = common.norm_date(cell("日期"))
                        if now is None or now.isoformat() != ident["date"]:
                            problems.append(
                                f"{it.get('ar')}: 第 {r} 行日期变了"
                                f"（计划时 {ident['date']}，现在 {now}）"
                            )
                    if "公司名称" in cols and ident.get("payer"):
                        now = str(cell("公司名称") or "").strip()
                        if now and now != ident["payer"]:
                            problems.append(f"{it.get('ar')}: 第 {r} 行付款方变了（表被插过行？）")
                    if "金额" in cols and ident.get("amount") is not None:
                        now = common.to_number(cell("金额"))
                        if now is None or abs(float(now) - float(ident["amount"])) > 0.005:
                            problems.append(
                                f"{it.get('ar')}: 第 {r} 行金额变了（表被插过行？）"
                            )
        finally:
            wb.close()
    return problems


def _group_by(items: List[dict], key: str) -> Dict[str, List[dict]]:
    g: Dict[str, List[dict]] = {}
    for it in items:
        g.setdefault(it.get(key) or "", []).append(it)
    return g


def write_flow_items(
    workspace: Path,
    items: List[dict],
    *,
    in_place: bool,
) -> Tuple[List[dict], List[str]]:
    """
    写入 write 项。返回 (changes, problems)。
    按文件分组；每文件备份后 patch。
    """
    import openpyxl
    import xlsx_patch

    aliases = common.load_aliases()
    write_items = [it for it in items if (it.get("verdict") or "") == "write"]

    # ★ 一个格子都还没动之前，先核对每条 write 项的行身份
    stale = precheck_flow_identity(workspace, write_items)
    if stale:
        return [], [
            "写入前复核没过，**流转表一个字都没写**（她的表在出计划之后被动过）：",
            *stale,
            "→ 重跑 build_flow_plan.py + build_worklist.py 出新清单，她再确认一次。",
        ]

    by_file: Dict[str, List[dict]] = {}
    for it in write_items:
        by_file.setdefault(it.get("file") or "", []).append(it)

    changes: List[dict] = []
    problems: List[str] = []
    skipped: List[dict] = []
    today = dt.datetime.now().strftime("%Y%m%d_%H%M%S")

    for fname, group in by_file.items():
        src = _resolve_flow_path(workspace, fname)
        if not src or not src.is_file():
            for it in group:
                problems.append(f"{it.get('ar')}: 找不到流转文件 {fname}")
            continue

        # 读列位置
        wb = openpyxl.load_workbook(str(src), read_only=True, data_only=True)
        # 按 sheet 分组
        by_sheet: Dict[str, List[dict]] = {}
        for it in group:
            by_sheet.setdefault(it.get("sheet") or "", []).append(it)

        all_edits_by_sheet: Dict[str, List[Tuple[int, int, object]]] = {}
        for sheet_name, g2 in by_sheet.items():
            if sheet_name not in wb.sheetnames:
                for it in g2:
                    problems.append(f"{it.get('ar')}: sheet 不存在 {sheet_name}")
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            try:
                _hrow, cols = _locate_cols(rows, aliases)
            except Exception as e:
                for it in g2:
                    problems.append(f"{it.get('ar')}: 列定位失败 {e}")
                continue
            col_order = cols["单号"] + 1  # 1-based for patch
            col_upd = cols["是否更新应收款"] + 1
            edits: List[Tuple[int, int, object]] = []
            for it in g2:
                r = int(it["row_no"])
                order_v = it.get("order_suggest")
                if order_v is None:
                    order_v = ""
                order_v = str(order_v).strip()
                # 禁止用空串覆盖已有单号（plan 应保留 existing；双保险）
                write_order = it.get("write_order")
                if write_order is None:
                    write_order = bool(order_v)
                upd_v = it.get("updated_suggest")
                if upd_v is None:
                    upd_v = ""
                upd_v = str(upd_v).strip()
                if upd_v in ("（空白）", "空白", "空"):
                    upd_v = ""
                write_updated = it.get("write_updated")
                if write_updated is None:
                    write_updated = True
                # 幂等：表里已经是这个值就别再写一遍（2026-07-25 实测：重跑会重写同样的
                # 11 笔、每次多存一个备份、还报「写入完成 11 笔」，看着像又干了活）
                cur_vals = list(rows[r - 1]) if r - 1 < len(rows) else []

                def _cur(idx0):
                    return str(cur_vals[idx0] or "").strip() if idx0 < len(cur_vals) else ""

                def _norm_lines(s):
                    return "\n".join(
                        x.strip() for x in str(s).replace("\r", "").split("\n") if x.strip()
                    )

                same_order = _norm_lines(_cur(cols["单号"])) == _norm_lines(order_v)
                same_upd = _cur(cols["是否更新应收款"]) == upd_v

                did_order = bool(write_order and order_v and not same_order)
                if did_order:
                    edits.append((r, col_order, order_v))
                do_upd = bool(write_updated and not same_upd)
                if do_upd:
                    edits.append((r, col_upd, upd_v))
                if not did_order and not do_upd:
                    skipped.append(it)
                    continue
                write_updated = do_upd
                changes.append(
                    {
                        "ar": it.get("ar"),
                        "file": fname,
                        "sheet": sheet_name,
                        "row_no": r,
                        "单号": order_v if did_order else "(未改)",
                        "是否更新应收款": upd_v if write_updated else "(未改)",
                    }
                )
            all_edits_by_sheet[sheet_name] = edits
        wb.close()

        if not any(all_edits_by_sheet.values()):
            continue  # 这份文件里的都跟表里一样，不备份不重写

        backup_dir = src.parent / "备份"
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup = backup_dir / f"{src.stem}_流转备份_{today}{src.suffix}"
        shutil.copy2(src, backup)

        # 多 sheet：依次 patch，中间用 tmp 链
        current = src
        tmps: List[Path] = []
        try:
            for sheet_name, edits in all_edits_by_sheet.items():
                tmp = src.with_name(f".{src.stem}_flow_{sheet_name}_{today}{src.suffix}")
                xlsx_patch.patch_cells(current, tmp, sheet_name, edits)
                if current != src and current in tmps:
                    pass
                tmps.append(tmp)
                current = tmp
            # 回读校验（本文件局部问题）
            local_problems: List[str] = []
            wb2 = openpyxl.load_workbook(str(current), read_only=True, data_only=True)
            for sheet_name, g2 in by_sheet.items():
                if sheet_name not in wb2.sheetnames:
                    local_problems.append(f"回读缺 sheet {sheet_name}")
                    continue
                ws = wb2[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                _hrow, cols = _locate_cols(rows, aliases)
                for it in g2:
                    r = int(it["row_no"])
                    if r - 1 >= len(rows):
                        local_problems.append(f"{it.get('ar')}: 回读行越界 {r}")
                        continue
                    vals = list(rows[r - 1])
                    got_order = str(vals[cols["单号"]] or "").strip() if cols["单号"] < len(vals) else ""
                    got_upd = (
                        str(vals[cols["是否更新应收款"]] or "").strip()
                        if cols["是否更新应收款"] < len(vals)
                        else ""
                    )
                    exp_order = str(it.get("order_suggest") or "").strip()
                    exp_upd = str(it.get("updated_suggest") or "").strip()
                    if exp_upd in ("（空白）", "空白", "空"):
                        exp_upd = ""
                    wo = it.get("write_order")
                    if wo is None:
                        wo = bool(exp_order)
                    wu = it.get("write_updated")
                    if wu is None:
                        wu = True

                    def _norm_order(s):
                        return "\n".join(
                            x.strip()
                            for x in str(s).replace("\r", "").split("\n")
                            if x.strip()
                        )

                    # 只校验实际写入的列（空单号跳过写时不要求表变成空）
                    if wo and exp_order and _norm_order(got_order) != _norm_order(exp_order):
                        local_problems.append(
                            f"{it.get('ar')} 单号回读不符：期望 {exp_order!r} 实际 {got_order!r}"
                        )
                    if wu and got_upd != exp_upd:
                        local_problems.append(
                            f"{it.get('ar')} 是否更新回读不符：期望 {exp_upd!r} 实际 {got_upd!r}"
                        )
            wb2.close()

            n_ok = len(group)
            if local_problems:
                problems.extend(local_problems)
                print(
                    f"WARN: 回读失败，原件未改；备份在 {backup}；临时 {current}",
                    file=sys.stderr,
                )
            else:
                if in_place:
                    shutil.copy2(current, src)
                    print(f"流转已就地写入 {n_ok} 笔 → {src}（备份 {backup}）")
                else:
                    out = workspace / "04_产出" / f"到账流转_已回填_{src.stem}_{today[:8]}.xlsx"
                    out.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy2(current, out)
                    print(f"流转已写入 {n_ok} 笔 → {out}（源未动 {src}；备份 {backup}）")
                for t in tmps:
                    t.unlink(missing_ok=True)
        except Exception as e:
            problems.append(f"{fname}: 写入异常 {e}")
            for t in tmps:
                t.unlink(missing_ok=True)

    if skipped:
        print(f"流转：{len(skipped)} 笔表里已经是这个值，跳过（幂等，不重复写、不多存备份）")
    return changes, problems


def write_change_report(changes: List[dict], path: Path) -> None:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "流转变更清单"
    headers = ["ar", "file", "sheet", "row_no", "单号", "是否更新应收款"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for ch in changes:
        ws.append([ch.get(h) for h in headers])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="写流转表安全子集")
    ap.add_argument("--plan", required=True, help="流转写入计划_校验后.json")
    ap.add_argument("--workspace", default=str(common.WORK))
    ap.add_argument(
        "--confirmed",
        action="store_true",
        help="已废弃的兼容参数；现在日清与写前校验通过后可直接写入",
    )
    ap.add_argument("--in-place", action="store_true", help="就地写 02_ 里的流转副本")
    ap.add_argument("--report", default="")
    args = ap.parse_args(argv)

    plan_p = Path(args.plan)
    if not plan_p.is_file():
        print(f"ERROR: 找不到流转计划 {plan_p}", file=sys.stderr)
        return 2
    plan = json.loads(plan_p.read_text(encoding="utf-8"))
    items = plan.get("items") or []
    writable = [it for it in items if it.get("verdict") == "write"]
    if not writable:
        print("流转无可自动写的笔（全是手填/跳过）。什么都没改。")
        return 0

    ws = common.resolve_workspace(args.workspace)
    changes, problems = write_flow_items(ws, items, in_place=args.in_place)
    report = Path(args.report) if args.report else (
        ws / "04_产出" / f"流转变更清单_{dt.date.today().strftime('%Y%m%d')}.xlsx"
    )
    if changes:
        write_change_report(changes, report)
        print(f"流转变更清单 → {report}")

    if problems:
        print("⚠ 流转写入问题：", file=sys.stderr)
        for p in problems[:15]:
            print(f"  - {p}", file=sys.stderr)
        return 1
    if changes:
        print(f"流转写入完成：{len(changes)} 笔")
    else:
        print("流转：没有需要改的（表里已经是这个值）。什么都没写。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
