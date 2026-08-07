#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
第 7 步 a：**写入前校验计划**（官方 plan → validate → execute 模式的中间环）。

为什么要有这一步：判定结果是"打算怎么填"，但从判定到写入之间，她的表可能已经变了
（月初贴交付会插行、部分核销会在上方插行 → **行号立刻失效**）。
直接按行号写 = 把值写到别人家的行上。所以写之前逐条复核，过不了的不写、并说清为什么。

用法：
    python3 scripts/validate_plan.py --plan 判定结果.json --ledger 盈亏副本.xlsx \
        --out 04_产出/写入计划_校验后.json
退出码：0=有可写的（或全跳过）；2=输入不可用；**1=存在冲突**（有笔需要人看）
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import sys
from pathlib import Path
from typing import Dict, List, Optional

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
import common  # noqa: E402
import amount_policy  # noqa: E402
import writeoff_duplicate_audit as WDA  # noqa: E402

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

FIVE = ["计提", "回款明细", "是否结账", "收款时间", "收款方式"]
DERIVED = ["差异"]
VALID_JIEZHANG = {"是", "否"}
VALID_WAY = {"汇", "冲预收", "支", "现"}
BUSINESS_SETTLEMENT_TOL = float(amount_policy.BUSINESS_SETTLEMENT_TOLERANCE)


def _whole_parent_gate_error(audit: dict) -> str:
    """不只信任 status；按审计事实重算整笔回款的父 AR 守恒闸。"""
    if not audit.get("is_whole_payment"):
        return ""
    try:
        detail_count = int(
            audit.get("effective_order_amount_count")
            if audit.get("effective_order_amount_count") is not None
            else (audit.get("effective_detail_count") or 0)
        )
    except (TypeError, ValueError):
        return "整笔回款的有效订单金额计数无效"
    if detail_count <= 0:
        return "整笔回款没有有效订单已核销金额或完整交付额兜底"
    basis = audit.get("comparison_basis")
    allowed = {
        "order_written_off_local", "order_written_off_original",
        "delivery_fallback_local", "delivery_fallback_original",
        # 兼容修复前已生成、但仍可能被只读复核的旧计划。
        "detail_local", "detail_original",
    }
    if basis not in allowed:
        return "整笔回款与订单金额没有可比较的金额口径"
    delta = common.to_number(
        audit.get("delta")
        if audit.get("delta") is not None
        else audit.get("delta_dedup")
    )
    threshold = common.to_number(audit.get("threshold"))
    threshold = BUSINESS_SETTLEMENT_TOL if threshold is None else abs(float(threshold))
    if delta is None:
        return "整笔回款缺少父总到账与订单金额的差额审计"
    if basis.startswith("order_written_off") or basis.startswith("delivery_fallback"):
        order_total = common.to_number(audit.get("order_amount_total"))
        parent_total = common.to_number(
            audit.get("parent_total_local")
            if basis.endswith("_local")
            else audit.get("parent_total_orig")
        )
        if order_total is None or parent_total is None:
            return "整笔回款缺少父总到账或订单金额合计"
        expected_delta = round(float(parent_total) - float(order_total), 2)
        if abs(expected_delta - float(delta)) > 0.01:
            return "整笔回款审计中的父总到账、订单金额合计与差额不守恒"
        if basis.startswith("delivery_fallback") and not audit.get("fallback_used"):
            return "整笔回款使用交付额兜底但缺少兜底审计标记"
    if abs(float(delta)) > threshold:
        return f"整笔回款的父总到账与订单金额合计差额超过{threshold:g}元"
    return ""


def duplicate_audit_error(plan: dict, item: Optional[dict] = None) -> str:
    """防止父 AR 审计或逻辑记录在判定、校验和写入之间被手工改坏。"""
    if "duplicate_writeoff_audits" not in plan:
        return ""  # 兼容纯单测/旧夹具；当前版本四件套会始终带该字段
    audits = plan.get("duplicate_writeoff_audits") or {}
    expected = str(plan.get("duplicate_writeoff_audit_sha256") or "")
    if not audits and not expected:
        return ""  # 兼容不含父 AR 审计的旧夹具/旧计划
    actual = WDA.audit_fingerprint(audits)
    if not expected or expected != actual:
        return "系统重复核销审计指纹不一致，计划可能被修改，必须重新判定"
    items = [item] if item is not None else list(plan.get("auto") or [])
    for current in items:
        ar = str(current.get("ar") or "")
        audit = audits.get(ar) or {}
        if not audit:
            return f"父回款 {ar} 缺少父AR审计，禁止进入auto"
        status = audit.get("status")
        warnings = set(current.get("warning_codes") or [])
        if status == "unresolved":
            return f"父回款 {ar} 的父AR审计未解决，却进入了auto"
        gate_error = _whole_parent_gate_error(audit)
        if gate_error:
            return f"父回款 {ar} 未通过父AR金额守恒检查：{gate_error}"
        if status == "recovered":
            if "W_SYSTEM_DUPLICATE_WRITEOFF_COLLAPSED" not in warnings:
                return f"父回款 {ar} 已做系统重复纠正，但auto缺少警告码"
        if current.get("duplicate_writeoff_audit") != audit:
            return f"父回款 {ar} 的auto行与顶层父AR审计不一致"
    return ""


def _norm(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (dt.date, dt.datetime)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    # Excel 里 300 与 300.0 是同一个数
    try:
        f = float(s)
        return f"{f:.2f}"
    except (TypeError, ValueError):
        return s


def read_ledger_rows(path: Path) -> Dict[int, dict]:
    """把盈亏『明细』整表读成 {行号: {列名: 值}}，用于逐条复核。"""
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    if "明细" not in wb.sheetnames:
        wb.close()
        raise ValueError(f"盈亏表无『明细』sheet：{wb.sheetnames}")
    ws = wb["明细"]
    aliases = common.load_aliases()
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hrow, headers = common.find_header_row(
        all_rows, "盈亏明细", ["SO", "SOD", "计提", "回款明细", "是否结账"], aliases
    )
    cols = common.resolve_columns(
        headers,
        "盈亏明细",
        ["SO", "SOD", "计提", "回款明细", "是否结账", "收款时间", "收款方式"],
        aliases,
    )
    diff_idx = common.fuzzy_find_col(
        headers, (aliases.get("盈亏明细", {}) or {}).get("差异", ["差异"])
    )
    if diff_idx is not None:
        cols["差异"] = diff_idx
    yidx = common.fuzzy_find_col(
        headers, (aliases.get("盈亏明细", {}) or {}).get("应收", ["应收金额", "应收"])
    )
    if yidx is None:
        raise ValueError("盈亏『明细』找不到应收金额列，无法校验部分回款拆行")
    cols["应收"] = yidx
    out: Dict[int, dict] = {}
    for i, row in enumerate(all_rows, start=1):
        if i <= hrow + 1:
            continue
        vals = list(row)

        def cell(key):
            idx = cols.get(key)
            return vals[idx] if idx is not None and idx < len(vals) else None

        out[i] = {
            "SO": str(cell("SO") or "").strip(),
            "SOD": str(cell("SOD") or "").strip(),
            "计提": cell("计提"),
            "回款明细": cell("回款明细"),
            "差异": cell("差异"),
            "_差异列存在": "差异" in cols,
            "是否结账": cell("是否结账"),
            "收款时间": cell("收款时间"),
            "收款方式": cell("收款方式"),
            "应收金额": cell("应收"),
        }
    return out


def _matches_identity(row: Optional[dict], so: str, sod: str) -> bool:
    if row is None:
        return False
    if so and row.get("SO") != so:
        return False
    if sod and row.get("SOD") != sod:
        return False
    return bool(so or sod)


def _matches_planned_fields(row: dict, expected: dict) -> bool:
    """现有业务行是否已等于计划目标；用于插行后的唯一重定位与幂等复核。"""
    for key in FIVE:
        if _norm(row.get(key)) != _norm(expected.get(key)):
            return False
    expected_sod = str(expected.get("实收SOD") or "").strip()
    if expected_sod and row.get("SOD") != expected_sod:
        return False
    return True


def settled_without_open_row(item: dict, rows: Dict[int, dict]) -> Optional[int]:
    """订单已有结账行且同一订单/SOD 没有未结账行时，返回稳定幂等行。"""
    so = str(item.get("so") or "").strip()
    sod = str(item.get("sod") or "").strip()
    if not so:
        return None
    candidates = [
        row_no for row_no, row in rows.items()
        if row.get("SO") == so and (not sod or row.get("SOD") == sod)
    ]
    if not candidates:
        return None
    settled = [
        row_no for row_no in sorted(candidates)
        if str(rows[row_no].get("是否结账") or "").strip() == "是"
    ]
    if not settled:
        return None
    if any(
        str(rows[row_no].get("是否结账") or "").strip() != "是"
        for row_no in candidates
    ):
        return None
    return settled[0]


def resolve_item_row(item: dict, rows: Dict[int, dict]) -> tuple[Optional[int], str]:
    """
    优先使用判定时行号；受控插行使行号失效时，按 SO/SOD 业务身份唯一重定位。

    不允许仅按 SO 猜行。同一 SO/SOD 有多行时，只有其中恰有一行已等于计划目标，
    才可用于写后幂等复核；否则继续冲突。
    """
    ref = item.get("ledger_row_ref")
    if not ref:
        return None, "判定结果里没有行号，无法定位"
    ref = int(ref)
    so = str(item.get("so") or "").strip()
    sod = str(item.get("sod") or "").strip()
    if _matches_identity(rows.get(ref), so, sod):
        return ref, ""

    candidates = [
        row_no for row_no, row in rows.items()
        if _matches_identity(row, so, sod)
    ]
    if len(candidates) == 1:
        return candidates[0], ""
    target_matches = [
        row_no for row_no in candidates
        if _matches_planned_fields(rows[row_no], item.get("five_cols") or {})
    ]
    if len(target_matches) == 1:
        return target_matches[0], ""
    if not candidates:
        return None, f"按 SO/SOD 找不到原计划第 {ref} 行对应的业务行"
    return None, (
        f"按 SO/SOD 找到 {len(candidates)} 行，无法唯一重定位"
        f"（候选行={candidates[:8]}）"
    )


def resolve_same_so_multi_sod_row(
    item: dict, rows: Dict[int, dict]
) -> tuple[Optional[int], str]:
    """多 SOD 合并行允许以任一成员 SOD 或合并后的 SOD 文本唯一重定位。"""
    op = item.get("row_operation") or {}
    ref = int(item.get("ledger_row_ref") or 0)
    so = str(op.get("so") or item.get("so") or "").strip()
    member_sods = {str(x or "").strip() for x in (op.get("member_sods") or []) if str(x or "").strip()}
    combined = str(op.get("combined_sod") or "").strip()

    def identity(row: Optional[dict]) -> bool:
        if row is None or (so and row.get("SO") != so):
            return False
        sod = str(row.get("SOD") or "").strip()
        return not sod or sod in member_sods or sod == combined

    if ref and identity(rows.get(ref)):
        return ref, ""
    candidates = [row_no for row_no, row in rows.items() if identity(row)]
    target = op.get("target_five_cols") or item.get("five_cols") or {}
    exact = [row_no for row_no in candidates if _matches_planned_fields(rows[row_no], target)]
    if len(exact) == 1:
        return exact[0], ""
    if len(candidates) == 1:
        return candidates[0], ""
    if not candidates:
        return None, "按 SO 和多 SOD 成员找不到合并核销目标行"
    return None, f"按 SO 和多 SOD 成员找到 {len(candidates)} 行，无法唯一重定位"


def _check_split_payment_chain(item: dict, rows: Dict[int, dict], ref: int) -> Optional[dict]:
    """复核逐笔分笔链的结构、金额守恒及完整幂等状态。"""
    op = item.get("row_operation") or {}
    steps = op.get("steps") or []
    if len(steps) < 2:
        return {"verdict": "conflict", "reason": "分笔回款链至少需要两笔不同父回款"}
    try:
        source = round(float(op["source_receivable"]), 2)
        initial = round(float(op["initial_cumulative"]), 2)
        latest = round(float(op["latest_delivery"]), 2)
    except (KeyError, TypeError, ValueError):
        return {"verdict": "conflict", "reason": "分笔回款链基线参数缺失或不是数字"}
    if source < 0 or initial < 0 or latest <= 0 or abs((latest - initial) - source) > 0.011:
        return {"verdict": "conflict", "reason": "分笔链起点不守恒：最新交付额-历史累计必须等于当前应收"}

    tail_error = _tail_tolerance_audit_error(op, source)
    if tail_error:
        return {"verdict": "conflict", "reason": tail_error}

    ars = [str(step.get("ar") or "").strip() for step in steps]
    if not all(ars) or len(set(ars)) != len(ars):
        return {"verdict": "conflict", "reason": "分笔链父回款号缺失或重复"}
    indexes = [step.get("index") for step in steps]
    if indexes != list(range(len(steps))):
        return {"verdict": "conflict", "reason": "分笔链序号不连续"}
    order_keys = [tuple(str(x or "") for x in (step.get("writeoff_sequence_key") or [])) for step in steps]
    if any(len(key) < 2 or not key[1] for key in order_keys) or order_keys != sorted(order_keys):
        return {"verdict": "conflict", "reason": "分笔链缺少或打乱核销记录顺序"}

    previous_remaining = source
    expected_cumulative = initial
    for index, step in enumerate(steps):
        five = step.get("five_cols") or {}
        try:
            current = round(float(step["current_received"]), 2)
            cumulative = round(float(step["cumulative_received"]), 2)
            receivable = round(float(step["receivable"]), 2)
            remaining = round(float(step["remaining_after"]), 2)
        except (KeyError, TypeError, ValueError):
            return {"verdict": "conflict", "reason": "分笔链步骤金额缺失或不是数字"}
        expected_cumulative = round(expected_cumulative + current, 2)
        if current <= 0 or abs(cumulative - expected_cumulative) > 0.011:
            return {"verdict": "conflict", "reason": "分笔链累计未按每个父回款逐笔递增"}
        expected_remaining = round(max(latest - cumulative, 0.0), 2)
        if abs(remaining - expected_remaining) > 0.011:
            return {"verdict": "conflict", "reason": "分笔链剩余应收公式不成立"}
        settled = bool(step.get("settled"))
        expected_receivable = previous_remaining if settled else round(previous_remaining - remaining, 2)
        if receivable < 0 or abs(receivable - expected_receivable) > 0.011:
            return {"verdict": "conflict", "reason": "分笔链拆行后的单步应收不守恒"}
        if settled and index != len(steps) - 1:
            return {"verdict": "conflict", "reason": "只有分笔链最后一笔允许结清"}
        if not settled and five.get("计提") is not None:
            return {"verdict": "conflict", "reason": "未累计结清的分笔不得计提"}
        if settled and common.to_number(five.get("计提")) is None:
            return {"verdict": "conflict", "reason": "累计结清的最后一笔必须计提"}
        if abs(float(five.get("回款明细") or 0) - current) > 0.011:
            return {"verdict": "conflict", "reason": "分笔链业务行回款额与父回款步骤不一致"}
        previous_remaining = remaining

    final_unpaid = op.get("final_unpaid")
    if previous_remaining > 0.011:
        if not final_unpaid or abs(float(final_unpaid.get("receivable") or 0) - previous_remaining) > 0.011:
            return {"verdict": "conflict", "reason": "分笔链未结清但缺少最终未回款承接行"}
        unpaid_five = final_unpaid.get("five_cols") or {}
        if unpaid_five.get("是否结账") != "否" or any(
            unpaid_five.get(key) is not None for key in ("计提", "回款明细", "收款时间", "收款方式")
        ):
            return {"verdict": "conflict", "reason": "分笔链最终未回款行字段不合法"}
    elif final_unpaid:
        return {"verdict": "conflict", "reason": "分笔链已结清，不应再生成未回款行"}

    row = rows.get(ref)
    current_receivable = common.to_number((row or {}).get("应收金额"))
    if current_receivable is not None and abs(float(current_receivable) - float(steps[0]["receivable"])) <= 0.011:
        so = str(item.get("so") or "").strip()
        sod = str(item.get("sod") or "").strip()
        for offset, step in enumerate(steps):
            actual = rows.get(ref + offset)
            if (
                not _matches_identity(actual, so, sod)
                or common.to_number((actual or {}).get("应收金额")) is None
                or abs(float(actual["应收金额"]) - float(step["receivable"])) > 0.011
                or not _matches_planned_fields(actual, step.get("five_cols") or {})
            ):
                return {"verdict": "conflict", "reason": "分笔回款链看似已写入，但某个父回款行不完整或被改动"}
        if final_unpaid:
            actual = rows.get(ref + len(steps))
            expected = final_unpaid.get("five_cols") or {}
            if (
                not _matches_identity(actual, so, sod)
                or common.to_number((actual or {}).get("应收金额")) is None
                or abs(float(actual["应收金额"]) - float(final_unpaid["receivable"])) > 0.011
                or not _matches_planned_fields(actual, expected)
            ):
                return {"verdict": "conflict", "reason": "分笔回款链最终未回款行不完整或被改动"}
        return {"verdict": "skip", "reason": "分笔回款链全部父回款行已完整写入，幂等跳过"}
    if current_receivable is None or abs(float(current_receivable) - source) > 0.011:
        return {"verdict": "conflict", "reason": f"分笔链源行应收已变化：表里={current_receivable} 计划基线={source}"}

    source_five = op.get("source_five_cols") or {}
    if source_five:
        if not _matches_planned_fields(row, source_five):
            return {"verdict": "conflict", "reason": "分笔链源行已不再等于判定时的聚合基线"}
        source_derived = op.get("source_derived_cols") or {}
        if "差异" in source_derived and _norm(row.get("差异")) != _norm(source_derived.get("差异")):
            return {"verdict": "conflict", "reason": "分笔链源行差异列已在判定后变化"}
        return {"verdict": "write", "reason": "聚合基线未变化，可迁移为逐父回款分笔链"}
    return None


def _tail_tolerance_audit_error(op: dict, source: float) -> str:
    audit = op.get("tail_tolerance_audit") or {}
    if not audit:
        return ""
    try:
        tolerance = abs(float(audit["tolerance"]))
        absorbed_total = round(float(audit["absorbed_total"]), 2)
        payments = list(audit["absorbed_payments"])
        original_parent_amounts = [round(float(x), 2) for x in audit["original_parent_amounts"]]
    except (KeyError, TypeError, ValueError):
        return "结清尾差审计参数缺失或不是数字"
    if abs(tolerance - BUSINESS_SETTLEMENT_TOL) > 0.001:
        return "结清尾差审计使用了错误的业务容差"
    try:
        payment_amounts = [round(float(item["amount"]), 2) for item in payments]
    except (KeyError, TypeError, ValueError):
        return "结清尾差父回款审计金额无效"
    case_ids = [str(item.get("case_id") or "") for item in payments]
    if (
        not payments
        or any(amount <= 0 for amount in payment_amounts)
        or any(not case_id for case_id in case_ids)
        or len(set(case_ids)) != len(case_ids)
        or abs(sum(payment_amounts) - absorbed_total) > 0.011
        or not (0 < absorbed_total <= tolerance)
        or abs(sum(original_parent_amounts) - source) > 0.011
    ):
        return "结清尾差审计金额或父回款清单不守恒"
    return ""


def _check_settlement_tail_aggregate(item: dict, rows: Dict[int, dict], ref: int) -> dict:
    """复核新建业务行中的1元尾差合并；真实父回款仍由审计字段守恒。"""
    op = item.get("row_operation") or {}
    row = rows.get(int(ref))
    if row is None:
        return {"verdict": "conflict", "reason": "结清尾差合并目标行在写前校验时已不存在"}
    try:
        source = round(float(op["source_receivable"]), 2)
        initial = round(float(op["initial_cumulative"]), 2)
        latest = round(float(op["latest_delivery"]), 2)
        final = round(float(op["final_cumulative"]), 2)
    except (KeyError, TypeError, ValueError):
        return {"verdict": "conflict", "reason": "结清尾差合并计划参数缺失或不是数字"}
    tail_error = _tail_tolerance_audit_error(op, source)
    if tail_error:
        return {"verdict": "conflict", "reason": tail_error}
    target = op.get("target_five_cols") or {}
    if (
        source <= 0
        or abs((latest - initial) - source) > 0.011
        or abs(final - latest) > 0.011
        or abs(float(target.get("回款明细") or 0) - source) > 0.011
        or abs(float(target.get("计提") or 0) - latest) > 0.011
        or str(target.get("是否结账") or "").strip() != "是"
    ):
        return {"verdict": "conflict", "reason": "结清尾差合并后的业务行金额不守恒"}
    current_receivable = common.to_number(row.get("应收金额"))
    if current_receivable is None or abs(float(current_receivable) - source) > 0.011:
        return {"verdict": "conflict", "reason": "结清尾差合并目标行应收已变化"}
    if _matches_planned_fields(row, target):
        return {"verdict": "skip", "reason": "结清尾差已合并写入，幂等跳过"}
    source_five = op.get("source_five_cols") or {}
    if not source_five or not _matches_planned_fields(row, source_five):
        return {"verdict": "conflict", "reason": "结清尾差合并目标行在判定后被改动"}
    source_derived = op.get("source_derived_cols") or {}
    if "差异" in source_derived and _norm(row.get("差异")) != _norm(source_derived.get("差异")):
        return {"verdict": "conflict", "reason": "结清尾差合并目标行差异列在判定后被改动"}
    return {"verdict": "write", "reason": "结清尾差审计和目标行快照一致，可以合并写入"}


def _same_so_multi_sod_error(op: dict) -> str:
    member_sods = [str(x or "").strip() for x in (op.get("member_sods") or [])]
    case_ids = [str(x or "").strip() for x in (op.get("member_case_ids") or [])]
    try:
        so_delivery = round(float(op["so_delivery"]), 2)
        current = round(float(op["current_received"]), 2)
        amounts = [round(float(x), 2) for x in (op.get("member_amounts") or [])]
        deliveries = [round(float(x), 2) for x in (op.get("member_deliveries") or [])]
    except (KeyError, TypeError, ValueError):
        return "同 SO 多 SOD 合并金额参数缺失或不是数字"
    if (
        not op.get("so") or not op.get("ar") or not op.get("target_case_id")
        or len(member_sods) < 2 or len(set(member_sods)) != len(member_sods)
        or any(not sod for sod in member_sods)
        or len(case_ids) != len(member_sods) or len(set(case_ids)) != len(case_ids)
        or any(not case_id for case_id in case_ids)
        or len(amounts) != len(member_sods) or len(deliveries) != len(member_sods)
        or any(value <= 0 for value in amounts + deliveries)
        or so_delivery <= 0
        or abs(sum(amounts) - current) > 0.011
        or abs(current - so_delivery) > 0.011
        or abs(sum(deliveries) - so_delivery) > 0.011
        or str(op.get("combined_sod") or "") != "、".join(member_sods)
        or str(op.get("target_case_id") or "") not in case_ids
    ):
        return "同 SO 多 SOD 合并计划不完整或金额不守恒"
    target = op.get("target_five_cols") or {}
    if (
        common.to_number(target.get("计提")) != so_delivery
        or common.to_number(target.get("回款明细")) != current
        or target.get("是否结账") != "是"
        or str(target.get("实收SOD") or "") != str(op.get("combined_sod") or "")
    ):
        return "同 SO 多 SOD 合并目标值与审计参数不一致"
    return ""


def _check_same_so_multi_sod_aggregate(
    item: dict, rows: Dict[int, dict], ref: int
) -> dict:
    op = item.get("row_operation") or {}
    error = _same_so_multi_sod_error(op)
    if error:
        return {"verdict": "conflict", "reason": error}
    row = rows.get(int(ref))
    if row is None:
        return {"verdict": "conflict", "reason": "同 SO 多 SOD 合并目标行已不存在"}
    so = str(op.get("so") or "").strip()
    member_sods = {str(x or "").strip() for x in (op.get("member_sods") or [])}
    combined = str(op.get("combined_sod") or "").strip()
    if row.get("SO") not in ("", so):
        return {"verdict": "conflict", "reason": "同 SO 多 SOD 合并目标行的 SO 已变化"}
    row_sod = str(row.get("SOD") or "").strip()
    if row_sod and row_sod not in member_sods and row_sod != combined:
        return {"verdict": "conflict", "reason": "目标行 SOD 不属于本次合并组"}
    target = op.get("target_five_cols") or {}
    receivable = common.to_number(row.get("应收金额"))
    if (
        receivable is not None
        and abs(float(receivable) - float(op["so_delivery"])) <= 0.011
        and _matches_planned_fields(row, target)
        and (not row.get("_差异列存在") or _norm(row.get("差异")) in ("", "None"))
    ):
        return {"verdict": "skip", "reason": "同 SO 多 SOD 已按 SO 交付金额合并写入，幂等跳过"}
    source_receivable = common.to_number(op.get("source_receivable"))
    if source_receivable is None or receivable is None or abs(float(receivable) - float(source_receivable)) > 0.011:
        return {"verdict": "conflict", "reason": "同 SO 多 SOD 合并目标行的原应收已变化"}
    if not _matches_planned_fields(row, op.get("source_five_cols") or {}):
        return {"verdict": "conflict", "reason": "同 SO 多 SOD 合并目标行在判定后被改动"}
    source_derived = op.get("source_derived_cols") or {}
    if "差异" in source_derived and _norm(row.get("差异")) != _norm(source_derived.get("差异")):
        return {"verdict": "conflict", "reason": "同 SO 多 SOD 合并目标行差异列在判定后被改动"}
    return {"verdict": "write", "reason": "同一 SO 完整多 SOD 组金额守恒，可合并写入一行"}


def _absorbed_multi_sod_error(item: dict, target: Optional[dict]) -> str:
    marker = item.get("same_so_multi_sod_absorbed") or {}
    if not marker:
        return ""
    if target is None:
        return "多 SOD 合并引用的目标不存在"
    op = target.get("row_operation") or {}
    if op.get("type") != "same_so_multi_sod_aggregate":
        return "多 SOD 合并引用的目标不是合法合并计划"
    if (
        str(marker.get("target_case_id") or "") != str(target.get("case_id") or "")
        or str(item.get("case_id") or "") not in set(op.get("member_case_ids") or [])
        or set(marker.get("member_sods") or []) != set(op.get("member_sods") or [])
        or common.to_number(marker.get("so_delivery")) != common.to_number(op.get("so_delivery"))
    ):
        return "多 SOD 合并引用与目标审计信息不一致"
    return _same_so_multi_sod_error(op)


def _absorbed_tail_item_error(item: dict, target: Optional[dict]) -> str:
    marker = item.get("tail_tolerance_absorbed") or {}
    if not marker:
        return ""
    if target is None:
        return "结清尾差引用的合并目标不存在"
    op = target.get("row_operation") or {}
    if op.get("type") not in {"split_payment_chain", "settlement_tail_aggregate"}:
        return "结清尾差引用的目标不是合法合并计划"
    audit = op.get("tail_tolerance_audit") or {}
    try:
        tolerance = abs(float(marker["tolerance"]))
        amount = round(float(marker["amount"]), 2)
        absorbed_total = round(float(marker["absorbed_total"]), 2)
    except (KeyError, TypeError, ValueError):
        return "结清尾差被合并记录的参数缺失或不是数字"
    if (
        str(marker.get("target_case_id") or "") != str(target.get("case_id") or "")
        or abs(tolerance - BUSINESS_SETTLEMENT_TOL) > 0.001
        or amount <= 0
        or not (0 < absorbed_total <= tolerance)
    ):
        return "结清尾差被合并记录的目标或容差不合法"
    matches = [
        payment for payment in (audit.get("absorbed_payments") or [])
        if str(payment.get("case_id") or "") == str(item.get("case_id") or "")
        and abs(float(payment.get("amount") or 0) - amount) <= 0.011
    ]
    if len(matches) != 1:
        return "结清尾差被合并记录与目标审计不一致"
    return ""


def _check_preserved_aggregate_tail(item: dict, rows: Dict[int, dict], ref: int) -> dict:
    """复核 1 元以内结清尾差保留聚合行的计划，任何基线变化都禁止写入。"""
    op = item.get("row_operation") or {}
    row = rows.get(int(ref))
    if row is None:
        return {"verdict": "conflict", "reason": "尾差聚合行在写前校验时已不存在"}
    try:
        source = round(float(op["source_receivable"]), 2)
        initial = round(float(op["initial_cumulative"]), 2)
        latest = round(float(op["latest_delivery"]), 2)
        final = round(float(op["final_cumulative"]), 2)
        tolerance = abs(float(op["tolerance"]))
        tail = round(float(op["tolerated_tail_amount"]), 2)
        parent_amounts = [round(float(value), 2) for value in op["parent_amounts"]]
    except (KeyError, TypeError, ValueError):
        return {"verdict": "conflict", "reason": "1元尾差聚合计划参数缺失或不是数字"}
    if abs(tolerance - BUSINESS_SETTLEMENT_TOL) > 0.001 or not (0 < tail <= tolerance):
        return {"verdict": "conflict", "reason": "1元尾差聚合计划的容差或尾差金额不合法"}
    computed_tail = round(sum(value for value in parent_amounts if value <= tolerance), 2)
    if (
        not parent_amounts
        or any(value <= 0 for value in parent_amounts)
        or abs(computed_tail - tail) > 0.011
        or abs(sum(parent_amounts) - source) > 0.011
        or abs((latest - initial) - source) > 0.011
        or abs(final - latest) > 0.011
    ):
        return {"verdict": "conflict", "reason": "1元尾差聚合计划金额不守恒"}

    current_receivable = common.to_number(row.get("应收金额"))
    current_received = common.to_number(row.get("回款明细"))
    current_accrual = common.to_number(row.get("计提"))
    if (
        current_receivable is None
        or current_received is None
        or current_accrual is None
        or abs(float(current_receivable) - source) > tolerance + 0.001
        or abs(float(current_received) - source) > tolerance + 0.001
        or abs(float(current_accrual) - latest) > tolerance + 0.001
        or str(row.get("是否结账") or "").strip() != "是"
    ):
        return {"verdict": "conflict", "reason": "现有聚合行已不满足1元以内尾差结清条件"}
    source_five = op.get("source_five_cols") or {}
    if not source_five or not _matches_planned_fields(row, source_five):
        return {"verdict": "conflict", "reason": "1元尾差聚合行在判定后被改动"}
    source_derived = op.get("source_derived_cols") or {}
    if "差异" in source_derived and _norm(row.get("差异")) != _norm(source_derived.get("差异")):
        return {"verdict": "conflict", "reason": "1元尾差聚合行差异列在判定后被改动"}
    return {
        "verdict": "skip",
        "reason": "结清尾差合计不超过1元，保留现有聚合行并幂等跳过",
    }


def resolve_split_chain_row(item: dict, rows: Dict[int, dict]) -> tuple[Optional[int], str]:
    """按完整连续链定位分笔回款链首行，避免同批插行后逐成员误定位。"""
    so = str(item.get("so") or "").strip()
    sod = str(item.get("sod") or "").strip()
    candidates = [
        row_no for row_no, row in rows.items()
        if _matches_identity(row, so, sod)
    ]

    completed = []
    for row_no in candidates:
        result = _check_split_payment_chain(item, rows, row_no)
        if result is not None and result.get("verdict") == "skip":
            completed.append(row_no)
    if len(completed) == 1:
        return completed[0], ""
    if len(completed) > 1:
        return None, (
            f"按 SO/SOD 找到 {len(completed)} 组完整分笔回款链，无法唯一定位"
            f"（链首行={completed[:8]}）"
        )

    op = item.get("row_operation") or {}
    source = common.to_number(op.get("source_receivable"))
    source_candidates = [
        row_no for row_no in candidates
        if source is not None
        and common.to_number(rows[row_no].get("应收金额")) is not None
        and abs(float(rows[row_no]["应收金额"]) - float(source)) <= 0.011
    ]
    if len(source_candidates) == 1:
        return source_candidates[0], ""
    if len(source_candidates) > 1:
        return None, (
            f"按 SO/SOD 找到 {len(source_candidates)} 行分笔链源行，无法唯一定位"
            f"（候选行={source_candidates[:8]}）"
        )
    return resolve_item_row(item, rows)


def check_one(item: dict, rows: Dict[int, dict]) -> dict:
    """
    单条复核 → {verdict: write|skip|conflict, reason}
    - write    ：行号对得上、目标格是空的、值合法 → 可以写
    - skip     ：已经填过且与计划一致 → 幂等跳过（重复跑不重复写）
    - conflict ：行号对不上 / 已填但不一致 / 值不合法 → 不写，交给人看
    """
    ref = item.get("ledger_row_ref")
    five = item.get("five_cols") or {}
    derived = item.get("derived_cols") or {}
    so, sod = (item.get("so") or "").strip(), (item.get("sod") or "").strip()
    op = item.get("row_operation") or {}
    is_multi_sod_aggregate = op.get("type") == "same_so_multi_sod_aggregate"

    if not ref:
        return {"verdict": "conflict", "reason": "判定结果里没有行号，无法定位"}
    row = rows.get(int(ref))
    if row is None:
        return {"verdict": "conflict", "reason": f"第 {ref} 行在表里不存在了（表被删过行？）"}

    # ① 行号还指着同一单吗——她插过行的话这里必然对不上
    if not is_multi_sod_aggregate and sod and row["SOD"] and row["SOD"] != sod:
        return {
            "verdict": "conflict",
            "reason": f"第 {ref} 行现在是 {row['SOD']}，不是计划里的 {sod}（表在判定之后被插过行）",
        }
    if so and row["SO"] and row["SO"] != so:
        return {
            "verdict": "conflict",
            "reason": f"第 {ref} 行现在是 {row['SO']}，不是计划里的 {so}（表被改过）",
        }

    # ② 值本身合法吗
    if five.get("是否结账") not in VALID_JIEZHANG:
        return {"verdict": "conflict", "reason": f"是否结账取值异常：{five.get('是否结账')!r}"}
    if five.get("收款方式") not in VALID_WAY:
        return {"verdict": "conflict", "reason": f"收款方式取值异常：{five.get('收款方式')!r}"}
    for k in ("计提", "回款明细"):
        v = five.get(k)
        if v is None:
            continue  # 部分核销时计提本就留空
        try:
            float(v)
        except (TypeError, ValueError):
            return {"verdict": "conflict", "reason": f"{k} 不是数字：{v!r}"}
    for k in DERIVED:
        if k not in derived:
            continue
        if not row.get("_差异列存在"):
            return {"verdict": "conflict", "reason": "本次需要写差异，但盈亏明细没有“差异”列"}
        try:
            float(derived[k])
        except (TypeError, ValueError):
            return {"verdict": "conflict", "reason": f"{k} 不是数字：{derived[k]!r}"}

    if op:
        if op.get("type") == "split_payment_chain":
            chain_result = _check_split_payment_chain(item, rows, int(ref))
            if chain_result is not None:
                return chain_result
        elif op.get("type") == "preserve_aggregate_tail_tolerance":
            return _check_preserved_aggregate_tail(item, rows, int(ref))
        elif op.get("type") == "settlement_tail_aggregate":
            return _check_settlement_tail_aggregate(item, rows, int(ref))
        elif op.get("type") == "same_so_multi_sod_aggregate":
            return _check_same_so_multi_sod_aggregate(item, rows, int(ref))
        elif op.get("type") != "split_below":
            return {"verdict": "conflict", "reason": f"未知行操作：{op.get('type')!r}"}
        if op.get("type") == "split_below" and row.get("_差异列存在") and _norm(row.get("差异")) not in ("", "None"):
            return {
                "verdict": "conflict",
                "reason": (
                    "部分回款阶段计提和业务值差异都必须留空，"
                    f"但当前差异={_norm(row.get('差异'))!r}；禁止覆盖"
                ),
            }
        if op.get("type") == "split_below":
          try:
            source = round(float(op["source_receivable"]), 2)
            paid = round(float(op["paid_receivable"]), 2)
            unpaid = round(float(op["unpaid_receivable"]), 2)
            latest = round(float(op["latest_delivery"]), 2)
            cumulative = round(float(op["cumulative_received"]), 2)
          except (KeyError, TypeError, ValueError):
            return {"verdict": "conflict", "reason": "部分回款拆行参数缺失或不是数字"}
          if paid < 0 or unpaid <= 0:
            return {"verdict": "conflict", "reason": f"拆行应收异常：已收侧={paid} 未收侧={unpaid}"}
          if abs((paid + unpaid) - source) > 0.011:
            return {"verdict": "conflict", "reason": f"拆行不守恒：{paid}+{unpaid}!={source}"}
          if abs((latest - cumulative) - unpaid) > 0.011:
            return {"verdict": "conflict", "reason": f"未回款公式不成立：{latest}-{cumulative}!={unpaid}"}
          inserted = op.get("inserted_five_cols") or {}
          if inserted.get("是否结账") != "否":
            return {"verdict": "conflict", "reason": "拆出的未回款行必须是否结账=否"}
          for key in ("计提", "回款明细", "收款时间", "收款方式"):
            if inserted.get(key) is not None:
                return {"verdict": "conflict", "reason": f"未回款行 {key} 必须留空"}

        # 写后重跑：原行已变成已收侧、下一行已是未收侧时，识别为完整幂等状态。
        # 不能继续拿拆前 source_receivable 要求当前行，否则受控拆行必然假报冲突。
          current_receivable = common.to_number(row.get("应收金额"))
          if (
            current_receivable is not None
            and abs(float(current_receivable) - paid) <= 0.011
          ):
            next_row = rows.get(int(ref) + 1)
            next_receivable = (
                common.to_number(next_row.get("应收金额"))
                if next_row is not None else None
            )
            same_next_identity = _matches_identity(next_row, so, sod)
            paid_matches = _matches_planned_fields(row, five)
            unpaid_matches = (
                next_row is not None
                and _matches_planned_fields(next_row, inserted)
                and (
                    not next_row.get("_差异列存在")
                    or _norm(next_row.get("差异")) in ("", "None")
                )
            )
            if (
                same_next_identity
                and next_receivable is not None
                and abs(float(next_receivable) - unpaid) <= 0.011
                and paid_matches
                and unpaid_matches
            ):
                return {
                    "verdict": "skip",
                    "reason": "部分回款拆行已完整写入（已收行+紧邻未收行），幂等跳过",
                }
            return {
                "verdict": "conflict",
                "reason": "当前行看似已拆分，但已收行或紧邻未收行与计划不一致",
            }
          if current_receivable is None or abs(float(current_receivable) - source) > 0.011:
            return {
                "verdict": "conflict",
                "reason": f"当前行应收已变化：表里={current_receivable} 计划基线={source}",
            }

    # ③ 目标格现在是什么
    # ⚠「是否结账」是她盈亏表**预置的默认值**：单子交付了、钱还没到的行默认就是「否」
    #   （2026-07-24 真实全表统计：「否」+回款空 = 1973 行，全是还没收款的挂账行）。
    #   所以它**不能当"这一行她已经填过"的证据**——否则每一笔新到账（表里那行本就带着「否」）
    #   都会被误判成"已填过、且和计划不一致"→ 冲突，导致自动回填对真实新数据全线失效
    #   （2026-07-24 opencode 真实 24 号数据实测：4 笔可填全被「否」挡成冲突、可写=0）。
    #   判"填没填过"只看**回款证据列**（计提/回款明细/收款时间/收款方式）；是否结账留给下面 same/diff
    #   去比对（她填了钱却没翻「是」这种真不一致，仍会照常报冲突）。
    evidence_cols = [k for k in FIVE if k != "是否结账"]
    filled = [k for k in evidence_cols if _norm(row.get(k)) not in ("", "None")]

    # 旧版部分回款已把本次回款、结账、日期和方式写入已收行，但在最终
    # 累计达到最新交付额时可能漏填计提与差异。这里仅允许补空值：其余
    # 业务字段和 SOD 必须全部等于当前计划，任何非空不一致仍保持冲突。
    if (
        "差异" in derived
        and _norm(row.get("计提")) in ("", "None")
        and common.to_number(five.get("计提")) is not None
        and _norm(row.get("差异")) in ("", "None")
    ):
        stable_keys = ["回款明细", "是否结账", "收款时间", "收款方式"]
        stable_match = all(
            _norm(row.get(key)) == _norm(five.get(key)) for key in stable_keys
        )
        planned_sod = str(five.get("实收SOD") or sod or "").strip()
        if stable_match and (not planned_sod or row.get("SOD") == planned_sod):
            return {
                "verdict": "write",
                "reason": "最终结清证据一致，仅补空白计提与业务差异",
            }

    if not filled:
        for k in DERIVED:
            if k in derived and _norm(row.get(k)) not in ("", "None", _norm(derived[k])):
                return {
                    "verdict": "conflict",
                    "reason": f"{k}: 表里={_norm(row.get(k))!r} 计划={_norm(derived[k])!r}",
                }
        return {"verdict": "write", "reason": "回款列为空，可写（是否结账为她表预置默认，不计入已填证据）"}

    # 这行她已经填过了 → 逐列比对。**两种不一致都算不一致**：
    #   ① 计划有值 ≠ 表里的值
    #   ② 计划算的是「留空」，但表里填了值 —— 2026-07-25 修：旧版 `if five.get(k) is not None`
    #      把这种整个跳过了，结果「表里计提=1000、本次算的是留空」会被判成"已填过且一致·跳过"，
    #      静默放过。而「计提到底该不该填」恰恰是明妹口径里最容易出错的一条
    #      （回款明细合计 = 交付额才可填计提），漏报等于把最该她看的那行藏起来。
    diff: List[str] = []
    for k in FIVE:
        want, got = five.get(k), _norm(row.get(k))
        if want is None:
            if got not in ("", "None"):
                diff.append(f"{k}: 表里={got!r} 本次算的是**留空**")
            continue
        if got != _norm(want):
            diff.append(f"{k}: 表里={got!r} 计划={_norm(want)!r}")
    derived_missing: List[str] = []
    for k in DERIVED:
        if k not in derived:
            continue
        want, got = derived[k], _norm(row.get(k))
        if got in ("", "None"):
            derived_missing.append(k)
        elif got != _norm(want):
            diff.append(f"{k}: 表里={got!r} 计划={_norm(want)!r}")
    if diff:
        return {
            "verdict": "conflict",
            "reason": "这行已经填过，且和本次算的不一样 → " + "；".join(diff),
        }
    if derived_missing:
        return {
            "verdict": "write",
            "reason": "五项回款字段已一致，仅补业务值差异公式：" + "、".join(derived_missing),
        }
    if not diff:
        return {"verdict": "skip", "reason": "已经填过且与本次一致（幂等跳过）"}
    raise AssertionError("不可达")


def validate(
    plan: dict,
    rows: Dict[int, dict],
    ledger_path: Optional[Path] = None,
) -> dict:
    items = [dict(it) for it in (plan.get("auto") or [])]
    by_case_id = {
        str(it.get("case_id") or ""): it
        for it in items if str(it.get("case_id") or "")
    }
    checked: List[dict] = []
    seen_rows: Dict[int, str] = {}
    for it in items:
        audit_error = duplicate_audit_error(plan, it)
        original_ref = it.get("ledger_row_ref")
        operation_type = (it.get("row_operation") or {}).get("type")
        is_split_chain = operation_type == "split_payment_chain"
        is_multi_sod_aggregate = operation_type == "same_so_multi_sod_aggregate"
        is_guarded_aggregate = operation_type in {
            "preserve_aggregate_tail_tolerance", "settlement_tail_aggregate",
            "same_so_multi_sod_aggregate",
        }
        absorbed_marker = it.get("tail_tolerance_absorbed") or {}
        multi_sod_marker = it.get("same_so_multi_sod_absorbed") or {}
        # 分笔链必须逐行复核完整性，不能因为其中已有一行结账就短路为“整单已写”。
        # 1元尾差聚合行也必须复核判定时快照，防止计划生成后被人工改动。
        settled_ref = None if (is_split_chain or is_guarded_aggregate) else settled_without_open_row(it, rows)
        if audit_error:
            res = {"verdict": "conflict", "reason": audit_error}
        elif multi_sod_marker:
            target = by_case_id.get(str(multi_sod_marker.get("target_case_id") or ""))
            marker_error = _absorbed_multi_sod_error(it, target)
            res = (
                {"verdict": "conflict", "reason": marker_error}
                if marker_error
                else {
                    "verdict": "skip",
                    "reason": "该 SOD 已并入同一 SO 的合并核销行，不重复写金额",
                }
            )
        elif absorbed_marker:
            target = by_case_id.get(str(absorbed_marker.get("target_case_id") or ""))
            marker_error = _absorbed_tail_item_error(it, target)
            res = (
                {"verdict": "conflict", "reason": marker_error}
                if marker_error
                else {
                    "verdict": "skip",
                    "reason": "本父回款属于1元以内结清尾差，审计保留并入目标行，不单独写入",
                }
            )
        elif settled_ref is not None:
            if original_ref and int(settled_ref) != int(original_ref):
                it["_relocated_from"] = int(original_ref)
            it["ledger_row_ref"] = int(settled_ref)
            res = {
                "verdict": "skip",
                "reason": "订单已写入/已结账，且不存在拆分未结账行（幂等跳过）",
            }
        else:
            if is_split_chain:
                resolved_ref, locate_error = resolve_split_chain_row(it, rows)
            elif is_multi_sod_aggregate:
                resolved_ref, locate_error = resolve_same_so_multi_sod_row(it, rows)
            else:
                resolved_ref, locate_error = resolve_item_row(it, rows)
            if locate_error:
                res = {"verdict": "conflict", "reason": locate_error}
            else:
                if int(resolved_ref) != int(original_ref):
                    it["_relocated_from"] = int(original_ref)
                    it["ledger_row_ref"] = int(resolved_ref)
                res = check_one(it, rows)
        ref = it.get("ledger_row_ref")
        # 合法分笔回款链允许多个父 AR 计划共享同一个源行；写入层会为每一笔创建
        # 独立业务行。没有同一链标记的重复行仍然冲突。
        if res["verdict"] == "write" and ref in seen_rows:
            prior = seen_rows[ref]
            current_group = it.get("split_chain_group_id") or ""
            if not current_group or prior != current_group:
                res = {
                    "verdict": "conflict",
                    "reason": f"第 {ref} 行被多笔计划同时命中，且不属于同一合法分笔回款组，需人工指定",
                }
        elif res["verdict"] == "write":
            seen_rows[ref] = it.get("split_chain_group_id") or (it.get("case_id") or "")
        item = {**it, "_check": res}
        # 校验这一刻该行的身份，写进计划带给 apply。
        # apply 拿它跟**写入那一刻**的表再对一次：对不上说明中间被插过行/删过行。
        cur = rows.get(int(ref)) if ref else None
        if cur is not None:
            item["_identity"] = {"row": int(ref), "SO": cur["SO"], "SOD": cur["SOD"]}
        checked.append(item)
    buckets = {"write": [], "skip": [], "conflict": []}
    for c in checked:
        buckets[c["_check"]["verdict"]].append(c)
    out = {
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "hexiao_date": plan.get("hexiao_date") or "",
        "counts": {k: len(v) for k, v in buckets.items()},
        "selection": {
            "mode": "all_auto",
            "selected": len(items),
            "total_auto": len(items),
        },
        "duplicate_writeoff_audits": plan.get("duplicate_writeoff_audits") or {},
        "duplicate_writeoff_audit_sha256": plan.get("duplicate_writeoff_audit_sha256") or "",
        "parent_fallback_allocations": plan.get("parent_fallback_allocations") or {},
        "business_rules": plan.get("business_rules") or {},
        **buckets,
    }
    if ledger_path is not None:
        # 盈亏副本在「校验」这一刻的指纹。apply 前会再算一次比对：
        # 不一致 = 工作副本在"校验 → 写入"之间发生变化 → 拒写并重新校验。
        out["ledger_path"] = str(ledger_path)
        out["ledger_sha256"] = common.sha256_file(ledger_path)
    return out


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="写入前校验计划（plan→validate→execute）")
    # --plan / --ledger 都可不给：不给就去工作区自己找（脏活归程序，别让 AI/她填路径）
    ap.add_argument("--plan", default="", help="判定结果 json；不给则取 04_产出 最新")
    ap.add_argument("--ledger", default="", help="盈亏核算表副本（只读）；不给则取 02_我的表副本/*盈亏*")
    ap.add_argument("--out", default="", help="校验后计划 json")
    # 防呆：同上。--workspace 还用于在没给 --out 时把结果落进正确的 04_产出/
    ap.add_argument("--workspace", default="", help="工作区根（没给 --out 时用它定产出位置）")
    ap.add_argument("--hexiao-date", default="", help="（校验日期以判定结果为准，收下防止链路中断）")
    args = ap.parse_args(argv)

    ws = common.resolve_workspace(args.workspace or None)
    out_dir = ws / "04_产出"

    def _latest(pattern: str):
        c = [p for p in sorted(out_dir.glob(pattern)) if not p.name.startswith("~$")]
        return c[-1] if c else None

    plan_p = Path(args.plan) if args.plan else (_latest("判定结果_*.json") or Path(""))
    if args.ledger:
        ledger_p = Path(args.ledger)
    else:
        cand = [
            p for p in sorted((ws / "02_我的表副本").glob("*盈亏*"))
            if not p.name.startswith(("~$", "."))
        ] if (ws / "02_我的表副本").is_dir() else []
        ledger_p = cand[0] if cand else Path("")

    if not plan_p.is_file():
        print(
            f"ERROR: 找不到判定结果{f' {plan_p}' if args.plan else f'（{out_dir} 里没有 判定结果_*.json）'}"
            "\n  先跑 classify_hexiao.py",
            file=sys.stderr,
        )
        return 2
    if not ledger_p.is_file():
        print(
            f"ERROR: 找不到盈亏表{f' {ledger_p}' if args.ledger else f'（{ws}/02_我的表副本/ 里没有 *盈亏* 文件）'}",
            file=sys.stderr,
        )
        return 2

    plan = json.loads(plan_p.read_text(encoding="utf-8"))
    try:
        rows = read_ledger_rows(ledger_p)
    except ValueError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 2

    result = validate(plan, rows, ledger_path=ledger_p)
    # 没给 --out 就落进**解析后的工作区**的 04_产出/，别落到 plan 旁边（会跟日清分家）
    out_p = Path(args.out) if args.out else (out_dir / "写入计划_校验后.json")
    out_p.parent.mkdir(parents=True, exist_ok=True)
    out_p.write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

    c = result["counts"]
    if result.get("hexiao_date"):
        print(f"核销日期：{common.date_cn(result['hexiao_date'])}")
    print(f"校验完成 可写={c['write']} 幂等跳过={c['skip']} 冲突={c['conflict']}")
    for x in result["conflict"][:10]:
        print(f"  ⚠ {x.get('case_id')}: {x['_check']['reason']}")
    if c["conflict"] > 10:
        print(f"  …另有 {c['conflict']-10} 条冲突，详见 {out_p.name}")
    print(f"计划: {out_p}")
    return 1 if c["conflict"] else 0


if __name__ == "__main__":
    sys.exit(main())
