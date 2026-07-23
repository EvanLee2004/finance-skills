# -*- coding: utf-8 -*-
"""第7步《核销日清》：一份合一、第一列单号、状态成列、幂等接上写入校验。"""
import json
import tempfile
from pathlib import Path

import openpyxl

import build_worklist as W


def _auto_item(so="SO2601", sod="SOD2601", ar="AR1"):
    return {
        "ar": ar, "so": so, "sod": sod,
        "case_id": f"{ar}|{so}|{sod}",
        "customer_masked": "测*户",
        "locate_hint": f"在盈亏『明细』按「新智云单号」筛选：{so}（禁止用行号）",
        "five_cols": {
            "计提": 100.0, "回款明细": 100.0, "是否结账": "是",
            "收款时间": "2026-07-21", "收款方式": "汇", "实收SOD": sod,
        },
        "current_values": {"计提": None, "回款明细": None, "是否结账": "",
                           "收款时间": "", "收款方式": None, "实收SOD": None},
        "reason": "全额核销/全部SOD · 定位=SO+应收金额",
        "ledger_row_ref": 100,
    }


def _sample_result():
    return {
        "auto": [_auto_item()],
        "hold": [{
            "ar": "AR2", "so": "SO2602", "sod": "", "case_id": "AR2|SO2602",
            "customer_masked": "挂*户", "code": "E1", "reason": "分笔", "candidates": [],
        }],
        "exception": [{
            "ar": "AR3", "so": "", "sod": "", "case_id": "AR3|-",
            "customer_masked": "", "code": "E7", "reason": "没关联下单",
        }],
        "counts": {"auto": 1, "hold": 1, "exception": 1, "total": 3},
        "e_code_dist": {"OK": 1, "E1": 1, "E7": 1},
        "payment_count": 3,
    }


def _cols(ws):
    headers = [c.value for c in ws[1]]
    return headers, {h: i for i, h in enumerate(headers)}


def test_single_sheet_with_status_column():
    """明妹要求：一份文件、第一列单号、问题类型做成一列。"""
    out = Path(tempfile.mkdtemp()) / "日清.xlsx"
    W.build_workbook(_sample_result(), None, out)
    wb = openpyxl.load_workbook(out)
    assert "今日清单" in wb.sheetnames and "先看这里" in wb.sheetnames
    ws = wb["今日清单"]
    headers, col = _cols(ws)
    assert headers[0] == "单号" and headers[1] == "状态"
    assert "行号" not in headers  # 她会插行，行号隔天就失效
    rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]
    assert len(rows) == 3
    assert {r[col["状态"]] for r in rows} == {"今天要填", "挂账待办", "异常"}
    first = rows[0]
    assert first[col["单号"]] == "SOD2601"
    assert "禁止用行号" in str(first[col["怎么找到这行"]])


def test_status_comes_from_validate_not_classify():
    """
    她当场问倒的那个 bug：昨天已经填过的又出现在「今天能填」。
    有写入校验结论时，状态必须以校验为准。
    """
    checked = {
        "counts": {"write": 0, "skip": 1, "conflict": 0},
        "write": [],
        "skip": [{**_auto_item(), "_check": {"verdict": "skip", "reason": "已经填过且与本次一致（幂等跳过）"}}],
        "conflict": [],
    }
    out = Path(tempfile.mkdtemp()) / "日清.xlsx"
    W.build_workbook(_sample_result(), checked, out)
    ws = openpyxl.load_workbook(out)["今日清单"]
    _, col = _cols(ws)
    states = [r[col["状态"]] for r in ws.iter_rows(min_row=2, values_only=True)]
    assert "今天要填" not in states
    assert "已填过·跳过" in states


def test_conflict_shows_diff():
    item = _auto_item()
    item["current_values"] = {"计提": 100.0, "回款明细": 100.0, "是否结账": "是",
                              "收款时间": "2026-07-21", "收款方式": "冲预收", "实收SOD": "SOD2601"}
    checked = {
        "counts": {"write": 0, "skip": 0, "conflict": 1},
        "write": [], "skip": [],
        "conflict": [{**item, "_check": {"verdict": "conflict", "reason": "已填但不一致"}}],
    }
    out = Path(tempfile.mkdtemp()) / "日清.xlsx"
    W.build_workbook(_sample_result(), checked, out)
    ws = openpyxl.load_workbook(out)["今日清单"]
    _, col = _cols(ws)
    row2 = [c.value for c in ws[2]]
    assert row2[col["状态"]] == "冲突·需你定"
    assert "收款方式" in str(row2[col["差异(一眼扫)"]])


def test_flow_sheet_has_policy():
    res = _sample_result()
    res["ar_summary"] = [{
        "ar": "AR1", "so_count": 2, "行数": 3,
        "流转表_是否更新应收款_建议": "部分",
        "待处理SO": ["SO_X"], "flow_locate": "loc",
    }]
    out = Path(tempfile.mkdtemp()) / "日清.xlsx"
    W.build_workbook(res, None, out)
    wb = openpyxl.load_workbook(out)
    assert "流转表怎么填" in wb.sheetnames
    h = [c.value for c in wb["流转表怎么填"][1]]
    assert "公式策略" in h and "颜色标注" in h


def test_hold_row_carries_action():
    out = Path(tempfile.mkdtemp()) / "日清.xlsx"
    W.build_workbook(_sample_result(), None, out)
    ws = openpyxl.load_workbook(out)["今日清单"]
    _, col = _cols(ws)
    hold = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[col["状态"]] == "挂账待办"][0]
    assert hold[col["码"]] == "E1"
    assert str(hold[col["怎么办"]]).strip()


def test_main_from_result_json(tmp_path):
    (tmp_path / "04_产出").mkdir(parents=True)
    res = tmp_path / "04_产出" / "判定结果_20260722.json"
    res.write_text(json.dumps(_sample_result(), ensure_ascii=False), encoding="utf-8")
    out = tmp_path / "out.xlsx"
    rc = W.main(["--result", str(res), "--out", str(out), "--workspace", str(tmp_path)])
    assert rc == 0
    assert out.is_file()
    assert list((tmp_path / "04_产出").glob("运行报告_*.txt"))
