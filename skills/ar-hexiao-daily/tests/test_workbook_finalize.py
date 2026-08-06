# -*- coding: utf-8 -*-
import re
import zipfile
from pathlib import Path

import openpyxl

import workbook_finalize as W
import xlsx_patch as X


def _rewrite(path: Path, transform, additions=None):
    with zipfile.ZipFile(path) as zin:
        infos = zin.infolist()
        payload = {info.filename: zin.read(info.filename) for info in infos}
    transform(payload)
    payload.update(additions or {})
    temp = path.with_name(path.stem + "_tmp.xlsx")
    with zipfile.ZipFile(temp, "w", zipfile.ZIP_DEFLATED) as zout:
        written = set()
        for info in infos:
            if info.filename in payload:
                zout.writestr(info, payload[info.filename])
                written.add(info.filename)
        for name, data in payload.items():
            if name not in written:
                zout.writestr(name, data)
    temp.replace(path)


def _book_with_calc_chain_and_external_link(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "明细"
    ws["A1"] = 1
    ws["B1"] = "=[1]历史表!A1"
    ws["C1"] = "=A1*2"
    wb.save(path)

    def transform(payload):
        sheet = payload["xl/worksheets/sheet1.xml"].decode("utf-8")
        sheet = re.sub(
            r'(<c r="B1"[^>]*><f>).*?(</f>)(?:<v>.*?</v>)?',
            r'\g<1>[1]历史表!A1\2<v>42</v>',
            sheet,
            count=1,
        )
        payload["xl/worksheets/sheet1.xml"] = sheet.encode("utf-8")

        workbook = payload["xl/workbook.xml"].decode("utf-8")
        workbook = re.sub(r'<calcPr\b[^>]*/>', "", workbook)
        workbook = workbook.replace(
            "</workbook>",
            '<externalReferences><externalReference r:id="rIdExt1"/>'
            '</externalReferences><calcPr calcMode="auto" fullCalcOnLoad="0" '
            'forceFullCalc="0" calcId="191029"/></workbook>',
        )
        payload["xl/workbook.xml"] = workbook.encode("utf-8")

        rels = payload["xl/_rels/workbook.xml.rels"].decode("utf-8")
        rels = rels.replace(
            "</Relationships>",
            '<Relationship Id="rIdExt1" Type="http://schemas.openxmlformats.org/'
            'officeDocument/2006/relationships/externalLink" '
            'Target="externalLinks/externalLink1.xml"/>'
            '<Relationship Id="rIdCalc" Type="http://schemas.openxmlformats.org/'
            'officeDocument/2006/relationships/calcChain" Target="calcChain.xml"/>'
            "</Relationships>",
        )
        payload["xl/_rels/workbook.xml.rels"] = rels.encode("utf-8")

        content = payload["[Content_Types].xml"].decode("utf-8")
        content = content.replace(
            "</Types>",
            '<Override PartName="/xl/externalLinks/externalLink1.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.externalLink+xml"/>'
            '<Override PartName="/xl/calcChain.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.calcChain+xml"/></Types>',
        )
        payload["[Content_Types].xml"] = content.encode("utf-8")

    additions = {
        "xl/calcChain.xml": (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<calcChain xmlns="http://schemas.openxmlformats.org/'
            'spreadsheetml/2006/main"><c r="B1" i="1"/><c r="C1"/></calcChain>'
        ).encode("utf-8"),
        "xl/externalLinks/externalLink1.xml": (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<externalLink xmlns="http://schemas.openxmlformats.org/'
            'spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/'
            'officeDocument/2006/relationships"><externalBook r:id="rId1"/>'
            '</externalLink>'
        ).encode("utf-8"),
        "xl/externalLinks/_rels/externalLink1.xml.rels": (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/'
            '2006/relationships"><Relationship Id="rId1" Type="http://schemas.'
            'openxmlformats.org/officeDocument/2006/relationships/externalLinkPath" '
            'Target="file:///D:/history.xlsx" TargetMode="External"/></Relationships>'
        ).encode("utf-8"),
    }
    _rewrite(path, transform, additions)


def test_constant_patch_preserves_existing_calc_chain_and_calc_flags(tmp_path):
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _book_with_calc_chain_and_external_link(source)

    result = X.patch_cells(
        source,
        output,
        "明细",
        [(1, 1, 3)],
        return_result=True,
    )

    assert result.requires_full_rebuild is False
    assert result.calc_chain_preserved is True
    assert result.dirty_cells == ("A1",)
    with zipfile.ZipFile(source) as zf:
        source_workbook = zf.read("xl/workbook.xml")
        source_chain = zf.read("xl/calcChain.xml")
    with zipfile.ZipFile(output) as zf:
        assert zf.read("xl/workbook.xml") == source_workbook
        assert zf.read("xl/calcChain.xml") == source_chain


def test_overwriting_formula_drops_calc_chain_and_requests_rebuild(tmp_path):
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _book_with_calc_chain_and_external_link(source)

    result = X.patch_cells(
        source,
        output,
        "明细",
        [(1, 2, 42)],
        return_result=True,
    )

    assert result.requires_full_rebuild is True
    assert result.calc_chain_preserved is False
    with zipfile.ZipFile(output) as zf:
        assert "xl/calcChain.xml" not in zf.namelist()
        workbook = zf.read("xl/workbook.xml").decode("utf-8")
    assert 'fullCalcOnLoad="1"' in workbook
    assert 'forceFullCalc="1"' in workbook


def test_portable_copy_freezes_external_cache_and_filters_calc_chain(tmp_path):
    source = tmp_path / "source.xlsx"
    portable = tmp_path / "portable.xlsx"
    _book_with_calc_chain_and_external_link(source)

    audit = W.create_portable_copy(source, portable)

    assert audit.frozen_external_formulas == 1
    assert audit.external_link_parts == 0
    assert audit.external_references == 0
    assert audit.external_formula_cells == 0
    assert audit.display_value_mismatches == 0
    assert audit.formula_cells == 1
    assert audit.calc_chain_present is True
    assert audit.full_calc_on_load == "0"
    assert audit.force_full_calc == "0"

    with zipfile.ZipFile(portable) as zf:
        names = set(zf.namelist())
        sheet = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")
        chain = zf.read("xl/calcChain.xml").decode("utf-8")
    assert not any(name.startswith("xl/externalLinks/") for name in names)
    b1 = re.search(r'<c r="B1".*?</c>', sheet).group(0)
    assert "<f" not in b1
    assert "<v>42</v>" in b1
    assert 'r="B1"' not in chain
    assert re.search(r'<c\b[^>]*r="C1"[^>]*i="1"|<c\b[^>]*i="1"[^>]*r="C1"', chain)


def test_finalize_rebuilds_chain_after_formula_change_and_disables_open_rebuild(tmp_path):
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _book_with_calc_chain_and_external_link(source)
    patch = X.patch_cells(
        source,
        output,
        "明细",
        [(1, 2, X.FormulaValue("=A1+10", 11))],
        return_result=True,
    )

    result = W.finalize_workbook(output, {"明细": patch})

    assert result.mode == "按公式缓存重建计算链"
    assert result.formula_cells == 2
    assert result.calc_chain_present is True
    assert result.full_calc_on_load == "0"
    assert result.force_full_calc == "0"
    with zipfile.ZipFile(output) as zf:
        chain = zf.read("xl/calcChain.xml").decode("utf-8")
        sheet = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")
    assert set(re.findall(r'<c\b[^>]*\br="([A-Z]+\d+)"', chain)) == {"B1", "C1"}
    assert re.search(r'<c r="B1"[^>]*>.*?<v>11\.0</v>.*?</c>', sheet)


def test_finalize_rejects_formula_without_cached_value(tmp_path):
    source = tmp_path / "source.xlsx"
    _book_with_calc_chain_and_external_link(source)

    def remove_cache(payload):
        sheet = payload["xl/worksheets/sheet1.xml"].decode("utf-8")
        sheet = re.sub(r'(<c r="C1"[^>]*><f>.*?</f>)<v>.*?</v>', r'\1', sheet)
        payload["xl/worksheets/sheet1.xml"] = sheet.encode("utf-8")

    _rewrite(source, remove_cache)
    try:
        W.finalize_workbook(source, {})
    except W.WorkbookFinalizeError as exc:
        assert "公式没有缓存值" in str(exc)
    else:
        raise AssertionError("公式缺少缓存值时必须拒绝交付")
