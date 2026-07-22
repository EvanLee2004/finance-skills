# -*- coding: utf-8 -*-
"""步骤2 收入提取：合成 + 样例7账户。"""
import datetime as dt
import os
import tempfile
from pathlib import Path

import openpyxl
import pytest

import extract_income as E
from conftest import BANK_XLSX

CFG, CUR = E.load_config(E.DEFAULT_CONFIG)


def _make_book(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "现金"
    ws.append(["日期", "凭证号", "摘要", "支票根号", "借方（增加）", "贷方（减少）", "余额", "类型"])
    ws.append([None, None, None, None, None, None, 0, None])
    ws.append([dt.datetime(2026, 6, 20), None, "现金收款-甲", None, 500, None, 500, "收入"])
    w = wb.create_sheet("工行")
    w.append(["日期", "凭证号", "摘要", "支票根号", "借方（增加）", "贷方（减少）", "余额", "类型"])
    w.append([dt.datetime(2026, 6, 24), None, "客户A", None, 1000.0, None, 1000, "收入"])
    w.append([dt.datetime(2026, 6, 24), None, "银行手续费", None, None, 25, 975, "手续费"])
    w.append([None, None, None, None, None, None, 975, "收入"])  # 空预填
    w.append([dt.datetime(2026, 6, 26), None, "", None, 800, None, 1775, "收入"])  # 缺客户名
    u = wb.create_sheet("美元")
    u.append(["日期", "凭证号", "摘要", "支票根号", "借方（增加）", "贷方（减少）", "余额", "类型"])
    u.append([dt.datetime(2026, 6, 24), None, "客户B", None, 100.0, None, 100, "收入"])
    s = wb.create_sheet("说明")
    s.append(["说明页", "无表头"])
    wb.save(path)


def test_extract_count_and_skip_blank():
    tmp = tempfile.mkdtemp()
    src = os.path.join(tmp, "日记账.xlsx")
    _make_book(src)
    rows, flags, report = E.extract(src, CFG, CUR)
    assert len(rows) == 4  # 现金1+工行2+美元1
    gh = [r for r in rows if r[0] == "工行"]
    assert len(gh) == 2  # 空预填跳过
    assert any(n == "说明" and st == "跳过" for n, st, _ in report)


def test_extract_currency_not_mixed():
    tmp = tempfile.mkdtemp()
    src = os.path.join(tmp, "日记账.xlsx")
    _make_book(src)
    rows, _, _ = E.extract(src, CFG, CUR)
    usd = [r for r in rows if r[0] == "美元"]
    assert usd and usd[0][4] == "美元"
    rmb = [r for r in rows if r[4] == "人民币"]
    assert len(rmb) == 3


def test_extract_flags_missing_customer():
    tmp = tempfile.mkdtemp()
    src = os.path.join(tmp, "日记账.xlsx")
    _make_book(src)
    rows, flags, _ = E.extract(src, CFG, CUR)
    assert any("缺客户名" in f[2] for f in flags)


def test_extract_write_output():
    tmp = tempfile.mkdtemp()
    src = os.path.join(tmp, "日记账.xlsx")
    out = os.path.join(tmp, "out.xlsx")
    _make_book(src)
    rows, flags, report = E.extract(src, CFG, CUR)
    E.write_output(rows, flags, out, report)
    wb = openpyxl.load_workbook(out)
    assert "收入汇总" in wb.sheetnames


def test_process_cli_style_out_dir():
    tmp = tempfile.mkdtemp()
    src = os.path.join(tmp, "日记账.xlsx")
    outdir = os.path.join(tmp, "out")
    os.makedirs(outdir)
    _make_book(src)
    # 模拟 --out 目录
    out_path = os.path.join(outdir, "收入汇总_test.xlsx")
    rows, flags, report, op = E.process(src, out_path)
    assert os.path.isfile(op)
    assert len(rows) == 4


@pytest.mark.skipif(not BANK_XLSX.is_file(), reason="无样例日记账")
def test_sample_7_accounts():
    rows, flags, report = E.extract(str(BANK_XLSX), CFG, CUR)
    assert len(rows) == 7
    # 不跨币种混加：汇总按币种
    from collections import defaultdict

    by = defaultdict(float)
    for r in rows:
        by[r[4]] += r[3]
    assert "人民币" in by or any(r[4] == "人民币" for r in rows)
