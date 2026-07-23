# -*- coding: utf-8 -*-
"""回填审核单 + 流转三态策略 + apply 人工闸。"""
import json
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "scripts"))
import build_review_sheet as R  # noqa: E402
import flow_ledger as FL  # noqa: E402


def _checked():
    return {
        "counts": {"write": 1, "skip": 1, "conflict": 1},
        "write": [
            {
                "case_id": "AR1|SO1",
                "ar": "AR1",
                "so": "SO2601",
                "sod": "SOD2601",
                "ledger_row_ref": 2,
                "locate_hint": "按 SO 筛选 SO2601（禁止用行号）",
                "five_cols": {
                    "计提": 100.0,
                    "回款明细": 100.0,
                    "是否结账": "是",
                    "收款时间": "2026-07-08",
                    "收款方式": "汇",
                },
                "current_values": {
                    "计提": None,
                    "回款明细": None,
                    "是否结账": None,
                    "收款时间": None,
                    "收款方式": None,
                },
                "reason": "SOD 精确匹配",
                "channel": "duizhang",
                "flow_locate": "流转.xlsx#Sheet1 第3行",
                "flow_order_suggest": "SO2601",
                "_check": {"verdict": "write", "reason": "五列为空，可写"},
            }
        ],
        "skip": [
            {
                "case_id": "AR2|SO2",
                "ar": "AR2",
                "so": "SO2602",
                "sod": "SOD2602",
                "ledger_row_ref": 5,
                "_check": {"verdict": "skip", "reason": "已经填过且与本次一致（幂等跳过）"},
            }
        ],
        "conflict": [
            {
                "case_id": "AR3|SO3",
                "ar": "AR3",
                "so": "SO2603",
                "sod": "",
                "ledger_row_ref": 9,
                "_check": {"verdict": "conflict", "reason": "表被插过行"},
            }
        ],
    }


def test_flow_status_policy_tri_state():
    empty = FL.flow_status_policy("")
    assert "不要公式" in empty["公式策略"]
    assert empty["填法"] == "留空"

    full = FL.flow_status_policy("是")
    assert full["填法"] == "是"

    part = FL.flow_status_policy("部分")
    assert part["填法"] == "部分"
    assert "红" in part["颜色标注"]


def test_review_workbook_sheets(tmp_path):
    out = tmp_path / "回填审核单.xlsx"
    summary = [
        {
            "ar": "AR1",
            "so_count": 2,
            "流转表_是否更新应收款_建议": "部分",
            "待处理SO": ["SO_X"],
            "flow_locate": "loc",
        },
        {
            "ar": "AR_EMPTY",
            "so_count": 1,
            "流转表_是否更新应收款_建议": "",
            "待处理SO": ["SO_Y"],
        },
    ]
    R.build_workbook(_checked(), out, ar_summary=summary)
    wb = openpyxl.load_workbook(out)
    assert {
        "使用说明",
        "盈亏回填·拟写入",
        "幂等跳过",
        "冲突·需你定",
        "流转表·是否更新建议",
    } <= set(wb.sheetnames)

    ww = wb["盈亏回填·拟写入"]
    headers = [c.value for c in ww[1]]
    assert "改前→改后(一眼扫)" in headers
    assert "判定依据" in headers
    assert "拟写_计提" in headers
    assert ww.max_row >= 2
    assert ww.cell(2, headers.index("SO") + 1).value == "SO2601"

    wf = wb["流转表·是否更新建议"]
    fh = [c.value for c in wf[1]]
    assert "公式策略" in fh
    assert "颜色标注" in fh
    # 部分行应有填法
    col_fill = fh.index("是否更新_填法") + 1
    assert any(wf.cell(r, col_fill).value == "部分" for r in range(2, wf.max_row + 1))


def test_review_main_writes_stamp(tmp_path):
    checked = tmp_path / "写入计划_校验后.json"
    checked.write_text(json.dumps(_checked(), ensure_ascii=False), encoding="utf-8")
    out = tmp_path / "审.xlsx"
    # workspace = tmp so stamp 落在同目录（无 04_产出 时用 checked 父目录）
    rc = R.main(
        [
            "--checked",
            str(checked),
            "--out",
            str(out),
            "--workspace",
            str(tmp_path),
        ]
    )
    assert rc == 0
    assert out.is_file()
    stamp = tmp_path / R.REVIEW_STAMP_NAME
    assert stamp.is_file()
    data = json.loads(stamp.read_text(encoding="utf-8"))
    assert data["status"] == "pending_review"
    assert data["counts"]["write"] == 1
