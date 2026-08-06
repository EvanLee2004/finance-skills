import compare_ledgers as C
import openpyxl


HEADERS = [
    "部门", "销售人员", "客户名称", "单号", "新智云单号", "应收金额",
    "计提金额", "回款明细", "是否结账", "收款时间", "收款方式", "备注",
]
INDEXES = {
    "SO": 4, "SOD": 11, "计提金额": 6, "回款明细": 7,
    "是否结账": 8, "收款时间": 9, "收款方式": 10,
}


def record(row, so, sod, accrual, received, settled, customer="客户"):
    values = ["部", "人", customer, "AB", so, 100.0, accrual, received, settled, "2026-07-01", "汇", sod]
    return {"row": row, "effective": values, "formula": list(values)}


def test_exact_and_business_amount_results_are_separate():
    assert not C.exact_equal(100.0, 99.994)
    assert C.business_equal(100.0, 99.994, "回款明细")
    assert not C.business_equal(100.0, 98.99, "回款明细")


def test_pairing_uses_so_sod_then_core_writeoff_fields_not_customer_hard_key():
    final = [
        record(2, "SO1", "SOD1", 100, 100, "是", customer="甲"),
        record(3, "SO1", "SOD1", None, 40, "是", customer="乙"),
    ]
    reference = [
        record(20, "SO1", "SOD1", None, 40, "是", customer="甲"),
        record(21, "SO1", "SOD1", 100, 100, "是", customer="乙"),
    ]

    pairs, final_extra, reference_extra = C.align(final, reference, HEADERS, INDEXES)

    assert not final_extra and not reference_extra
    assert {(left["row"], right["row"]) for left, right in pairs} == {(2, 21), (3, 20)}


def test_preserved_aggregate_cases_can_share_one_final_row():
    final = [record(2, "SO1", "SOD1", 100, 100, "是")]
    reference = record(20, "SO1", "SOD1", 100, 100, "是")
    cases = [
        {
            "date": "2026-07-16",
            "item": {
                "so": "SO1", "sod": "SOD1",
                "five_cols": {"计提": 100, "回款明细": 100, "是否结账": "是"},
                "row_operation": {"type": "preserve_aggregate_tail_tolerance"},
            },
        },
        {
            "date": "2026-07-16",
            "item": {
                "so": "SO1", "sod": "SOD1",
                "five_cols": {"计提": 100, "回款明细": 100, "是否结账": "是"},
                "row_operation": {"type": "preserve_aggregate_tail_tolerance"},
            },
        },
    ]

    rows = C.compare_cases(cases, final, {2: reference}, INDEXES)

    assert [row["业务结果"] for row in rows] == ["一致", "一致"]
    assert [row["最终表行"] for row in rows] == [2, 2]


def test_run_writes_dual_status_workbook(tmp_path):
    headers = [
        "部门", "销售人员", "客户名称", "单号", "新智云单号", "应收金额",
        "计提金额", "回款明细", "是否结账（是/否）", "收款时间",
        "收款方式(支/汇/现)", "列12", "列13", "列14", "列15", "列16",
        "列17", "列18", "列19", "实收金额", "未回款金额", "差异",
    ]

    def save(path, received):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "明细"
        ws.append(headers)
        ws.append(["部", "人", "客户", "AB", "SO1", 100, 100, received, "是", "2026-07-01", "汇", None, None, None, None, None, None, None, None, "SOD1", 0, 0])
        wb.save(path)

    final = tmp_path / "final.xlsx"
    reference = tmp_path / "reference.xlsx"
    output = tmp_path / "difference.xlsx"
    save(final, 100.0)
    save(reference, 99.72)

    summary = C.run(final, reference, output)

    assert summary["精确差异单元格"] == 1
    assert summary["业务差异单元格"] == 0
    assert summary["一元内精确差异但业务一致"] == 1
    wb = openpyxl.load_workbook(output, read_only=True, data_only=True)
    assert wb.sheetnames == ["汇总", "业务差异", "精确差异", "本批订单", "独有行"]
    wb.close()
