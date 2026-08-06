# -*- coding: utf-8 -*-
"""真实误判回归：本币、累计核销、手续费、分位尾差和多 SO 安全边界。"""
import datetime as dt
import zipfile

import openpyxl

import classify_hexiao as C
import fallback_allocation_ledger as FAL
import xlsx_patch


def _payment(amount=100.0, orders=None, writeoffs=None, **extra):
    value = {
        "ar": "AR_TEST",
        "hexiao_date": dt.date(2026, 7, 27),
        "arrival_date": dt.date(2026, 7, 24),
        "amount_orig": amount,
        "amount_local": amount,
        "fee": 0.0,
        "currency": "人民币CNY",
        "huikuan_type": "",
        "status": "核销成功",
        "customer": "测试客户",
        "orders": orders if orders is not None else [{"so": "SO1", "deliver": amount}],
        "writeoffs": writeoffs or {},
        "writeoffs_local": {},
        "cumulative_writeoffs": {},
        "cumulative_writeoffs_local": {},
        "sod_lines": {},
    }
    value.update(extra)
    return value


def _ledger(rows):
    so_index, sod_index = {}, {}
    for row, snap in rows.items():
        if snap.get("so"):
            so_index.setdefault(snap["so"], []).append(row)
        if snap.get("sod"):
            sod_index.setdefault(snap["sod"], []).append(row)
    return C.LedgerIndex(
        synthetic={"so": so_index, "sod": sod_index, "rows": rows}
    )


def test_foreign_currency_uses_zhiyun_local_amount_without_rate():
    """SO26030487 类：智云已有原币/本币金额，不应再报缺汇率。"""
    p = _payment(
        amount=2500.0,
        amount_local=17979.75,
        currency="美元USD",
        orders=[{"so": "SO1", "deliver": 2500.0, "rate": None}],
        writeoffs={"SO1": 2500.0},
        writeoffs_local={"SO1": 17979.75},
        sod_lines={"SO1": [{"sod": "SOD1", "deliver": 2500.0}]},
    )
    rec = C.expand_payment(p, {})[0]
    assert rec.get("forced_code") is None
    assert rec["amount_local"] == 17979.75
    assert rec["deliver_local"] == 17979.75

    result = C.classify_one(
        rec,
        _ledger({1: {"so": "SO1", "sod": "SOD1", "yingshou": 2500.0}}),
        {},
        0.0,
        2026,
    )
    assert result["bucket"] == "auto"
    assert result["code"] != "E6"
    assert result["five_cols"]["回款明细"] == 17979.75


def test_foreign_full_writeoff_without_detail_uses_order_rate_as_cny():
    """无逐SO核销明细时，USD交付额不能把原币数直接写进人民币盈亏表。"""
    p = _payment(
        amount=2500.0,
        amount_local=17979.75,
        currency="美元USD",
        orders=[{"so": "SO1", "deliver": 2500.0, "rate": 7.1919, "currency": "美元USD"}],
        sod_lines={"SO1": [{"sod": "SOD1", "deliver": 2500.0}]},
    )
    rec = C.expand_payment(p, {})[0]
    assert rec["amount_orig"] == 2500.0
    assert rec["amount_local"] == 17979.75
    assert rec["deliver_local"] == 17979.75


def test_unique_sod_uses_cumulative_writeoff_to_recognize_final_settlement():
    """SO26050128 类：本次只回尾款，但历史+本次已达到交付额，应结清而非再拆行。"""
    p = _payment(
        amount=1030.47,
        orders=[{"so": "SO1", "deliver": 4083.60}],
        writeoffs={"SO1": 1030.47},
        cumulative_writeoffs={"SO1": 4083.60},
        cumulative_writeoffs_local={"SO1": 4083.60},
        sod_lines={"SO1": [{"sod": "SOD1", "deliver": 4083.60}]},
    )
    rec = C.expand_payment(p, {})[0]
    assert rec["amount_orig"] == 1030.47
    assert rec["cumulative_received_local"] == 4083.60

    result = C.classify_one(
        rec,
        _ledger({1: {"so": "SO1", "sod": "SOD1", "yingshou": 4083.60}}),
        {},
        0.0,
        2026,
    )
    assert result["bucket"] == "auto"
    assert "row_operation" not in result
    assert result["five_cols"]["回款明细"] == 1030.47
    assert result["five_cols"]["计提"] == 4083.60
    assert result["five_cols"]["是否结账"] == "是"


def test_split_rows_use_aggregate_receivable_baseline_for_business_difference():
    """历史拆成两行后，B0 是两行应收合计；合计等于 D 时不得再造业务差异。"""
    p = _payment(
        amount=30.0,
        orders=[{"so": "SO1", "deliver": 100.0}],
        writeoffs={"SO1": 30.0},
        cumulative_writeoffs={"SO1": 100.0},
        cumulative_writeoffs_local={"SO1": 100.0},
        sod_lines={"SO1": [{"sod": "SOD1", "deliver": 100.0}]},
    )
    rec = C.expand_payment(p, {})[0]
    result = C.classify_one(
        rec,
        _ledger({
            1: {
                "so": "SO1", "sod": "SOD1", "yingshou": 30.0,
                "huikuan": None, "jiezhang": "否",
            },
            2: {
                "so": "SO1", "sod": "SOD1", "yingshou": 70.0,
                "huikuan": 70.0, "jiezhang": "是",
            },
        }),
        {},
        0.0,
        2026,
    )
    assert result["bucket"] == "auto"
    assert result["five_cols"]["计提"] == 100.0
    assert "derived_cols" not in result


def test_whole_payment_cross_month_uses_writeoff_date_and_advance_offset():
    """回款类型不再例外：整笔回款跨月也必须填核销日 + 冲预收。"""
    p = _payment(
        amount=100.0,
        huikuan_type="整笔回款",
        arrival_date=dt.date(2026, 4, 27),
        hexiao_date=dt.date(2026, 7, 27),
        orders=[{"so": "SO1", "deliver": 100.0}],
        sod_lines={"SO1": [{"sod": "SOD1", "deliver": 100.0}]},
    )
    rec = C.expand_payment(p, {})[0]
    result = C.classify_one(
        rec,
        _ledger({1: {"so": "SO1", "sod": "SOD1", "yingshou": 100.0}}),
        {},
        0.0,
        2026,
    )
    assert result["bucket"] == "auto"
    assert result["five_cols"]["收款时间"] == "2026-07-27"
    assert result["five_cols"]["收款方式"] == "冲预收"


def test_whole_payment_same_month_uses_arrival_date_and_hui():
    """回款类型不再例外：整笔回款同月按到账日 + 汇。"""
    p = _payment(
        amount=100.0,
        huikuan_type="整笔回款",
        arrival_date=dt.date(2026, 7, 24),
        hexiao_date=dt.date(2026, 7, 27),
        orders=[{"so": "SO1", "deliver": 100.0}],
        sod_lines={"SO1": [{"sod": "SOD1", "deliver": 100.0}]},
    )
    rec = C.expand_payment(p, {})[0]
    result = C.classify_one(
        rec,
        _ledger({1: {"so": "SO1", "sod": "SOD1", "yingshou": 100.0}}),
        {},
        0.0,
        2026,
    )
    assert result["bucket"] == "auto"
    assert result["five_cols"]["收款时间"] == "2026-07-24"
    assert result["five_cols"]["收款方式"] == "汇"


def test_fee_is_audited_but_not_double_allocated_when_explicit_so_writeoffs_exist():
    """有逐 SO 明细时金额原样使用；父回款费用只保留给总额/流转审计。"""
    p = _payment(
        amount=300.0,
        amount_local=None,
        fee=99.0,
        currency="美元USD",
        orders=[{"so": "SO1", "deliver": 100.0}, {"so": "SO2", "deliver": 200.0}],
        writeoffs={"SO1": 100.0, "SO2": 200.0},
        sod_lines={
            "SO1": [{"sod": "SOD1", "deliver": 100.0}],
            "SO2": [{"sod": "SOD2", "deliver": 200.0}],
        },
    )
    recs = C.expand_payment(p, {})
    assert {r["so"] for r in recs} == {"SO1", "SO2"}
    assert all(r.get("forced_code") is None for r in recs)
    assert round(sum(r["amount_orig"] for r in recs), 2) == 300.0
    assert all(r["fee"] == 99.0 for r in recs)
    assert all("不重复分配" in r["match_basis"] for r in recs)


def test_fee_never_reduces_or_allocates_writeoff_details():
    """即使逐 SO 合计看起来包含手续费，也不得扣减或重新分配。"""
    p = _payment(
        amount=300.0,
        fee=1.0,
        orders=[{"so": "SO1", "deliver": 100.0}, {"so": "SO2", "deliver": 201.0}],
        writeoffs={"SO1": 100.0, "SO2": 201.0},
        cumulative_writeoffs={"SO1": 100.0, "SO2": 201.0},
        sod_lines={
            "SO1": [{"sod": "SOD1", "deliver": 100.0}],
            "SO2": [{"sod": "SOD2", "deliver": 201.0}],
        },
    )
    recs = C.expand_payment(p, {})
    assert {r["so"] for r in recs} == {"SO1", "SO2"}
    assert all(r.get("forced_code") is None for r in recs)
    assert round(sum(r["amount_orig"] for r in recs), 2) == 301.0
    assert round(sum(r["deliver_local"] for r in recs), 2) == 301.0


def test_fee_does_not_block_or_rewrite_historical_cumulative_writeoff():
    p = _payment(
        amount=30.0,
        amount_local=None,
        fee=1.0,
        currency="美元USD",
        orders=[{"so": "SO1", "deliver": 80.0, "currency": "人民币CNY"}],
        writeoffs={"SO1": 30.0},
        cumulative_writeoffs={"SO1": 80.0},
        sod_lines={"SO1": [{"sod": "SOD1", "deliver": 80.0, "currency": "人民币CNY"}]},
    )
    rec = C.expand_payment(p, {})[0]
    assert rec.get("forced_code") is None
    assert rec["amount_local"] == 30.0
    assert rec["cumulative_received_local"] == 80.0


def test_no_writeoff_details_use_gross_parent_waterfall():
    """没有逐 SO 子明细时，净到账加手续费作为父回款总额顺序核销。"""
    p = _payment(
        amount=300.0,
        fee=1.0,
        orders=[{"so": "SO1", "deliver": 100.0}, {"so": "SO2", "deliver": 201.0}],
        sod_lines={
            "SO1": [{"sod": "SOD1", "deliver": 100.0}],
            "SO2": [{"sod": "SOD2", "deliver": 201.0}],
        },
    )
    recs = C.expand_payment(p, {})
    assert {r["so"] for r in recs} == {"SO1", "SO2"}
    assert all(r.get("forced_code") is None for r in recs)
    assert {r["so"]: r["amount_orig"] for r in recs} == {
        "SO1": 100.0,
        "SO2": 201.0,
    }
    assert all("父回款总到账" in r["match_basis"] for r in recs)


def test_writeoff_without_local_amount_or_rate_uses_writeoff_directly():
    """有本次核销金额就直接跑，不因父回款外币或缺汇率报 E6。"""
    p = _payment(
        amount=90.0,
        amount_local=None,
        fee=1.0,
        currency="美元USD",
        orders=[{"so": "SO1", "deliver": 100.0, "currency": "人民币CNY", "rate": None}],
        writeoffs={"SO1": 90.0},
        cumulative_writeoffs={"SO1": 90.0},
        sod_lines={"SO1": [{"sod": "SOD1", "deliver": 100.0, "currency": "人民币CNY"}]},
    )
    rec = C.expand_payment(p, {})[0]
    assert rec.get("forced_code") is None
    assert rec["amount_local"] == 90.0
    assert rec["deliver_local"] == 100.0
    assert rec["cumulative_received_local"] == 90.0


def test_one_cent_multi_so_shortfall_is_partial_on_next_order():
    """父回款比多 SO 交付合计少 0.01 时，按顺序把 0.01 留在当前部分订单。"""
    p = _payment(
        amount=299.99,
        orders=[{"so": "SO1", "deliver": 100.0}, {"so": "SO2", "deliver": 200.0}],
        sod_lines={
            "SO1": [{"sod": "SOD1", "deliver": 100.0}],
            "SO2": [{"sod": "SOD2", "deliver": 200.0}],
        },
    )
    recs = C.expand_payment(p, {})
    assert len(recs) == 2
    assert all(r.get("forced_code") is None for r in recs)
    assert {r["so"]: r["amount_orig"] for r in recs} == {
        "SO1": 100.0,
        "SO2": 199.99,
    }


def test_no_writeoff_details_allocate_smallest_then_partial_next():
    """没有逐 SO 子明细时，先核销较小订单，再部分核销下一单。"""
    p = _payment(
        amount=250.0,
        orders=[{"so": "SO1", "deliver": 100.0}, {"so": "SO2", "deliver": 200.0}],
        sod_lines={
            "SO1": [{"sod": "SOD1", "deliver": 100.0}],
            "SO2": [{"sod": "SOD2", "deliver": 200.0}],
        },
    )
    recs = C.expand_payment(p, {})
    assert {r["so"] for r in recs} == {"SO1", "SO2"}
    assert all(r.get("forced_code") is None for r in recs)
    assert {r["so"]: r["amount_local"] for r in recs} == {
        "SO1": 100.0,
        "SO2": 150.0,
    }


def test_next_parent_continues_from_first_outstanding_order(tmp_path):
    """首笔结清 SO1、部分核销 SO2；下一父回款必须跳过 SO1，从 SO2 续核。"""
    orders = [{"so": "SO1", "deliver": 100.0}, {"so": "SO2", "deliver": 200.0}]
    sod_lines = {
        "SO1": [{"sod": "SOD1", "deliver": 100.0}],
        "SO2": [{"sod": "SOD2", "deliver": 200.0}],
    }
    first = _payment(amount=150.0, orders=orders, sod_lines=sod_lines)
    first_recs = C.expand_payment(first, {})
    assert {r["so"]: r["amount_orig"] for r in first_recs} == {
        "SO1": 100.0,
        "SO2": 50.0,
    }
    checked = {
        "hexiao_date": "2026-07-27",
        "parent_fallback_allocations": {first["ar"]: first["_parent_fallback_allocation"]},
        "write": [{"ar": first["ar"], "so": "SO1"}, {"ar": first["ar"], "so": "SO2"}],
        "skip": [],
    }
    FAL.commit(tmp_path, checked)

    second = _payment(
        amount=100.0,
        ar="AR_NEXT",
        orders=orders,
        sod_lines=sod_lines,
        _fallback_allocation_state=FAL.load(tmp_path),
    )
    second_recs = C.expand_payment(second, {})
    assert [(r["so"], r["amount_orig"]) for r in second_recs] == [("SO2", 100.0)]
    audit = second["_parent_fallback_allocation"]
    assert audit["already_settled_sos"] == ["SO1"]
    assert audit["allocations"][1]["historical_received"] == 50.0
    assert second_recs[0]["cumulative_received_local"] == 150.0


def test_same_parent_rerun_reuses_successful_allocation(tmp_path):
    """同一父回款重跑必须复用原分配，不能把游标再次推进到后续订单。"""
    orders = [{"so": "SO1", "deliver": 100.0}, {"so": "SO2", "deliver": 200.0}]
    p = _payment(amount=150.0, orders=orders)
    original = C.expand_payment(p, {})
    FAL.commit(tmp_path, {
        "hexiao_date": "2026-07-27",
        "parent_fallback_allocations": {p["ar"]: p["_parent_fallback_allocation"]},
        "write": [{"ar": p["ar"], "so": "SO1"}, {"ar": p["ar"], "so": "SO2"}],
        "skip": [],
    })
    rerun = _payment(
        amount=150.0,
        orders=orders,
        _fallback_allocation_state=FAL.load(tmp_path),
    )
    repeated = C.expand_payment(rerun, {})
    assert [(r["so"], r["amount_orig"]) for r in repeated] == [
        (r["so"], r["amount_orig"]) for r in original
    ]
    assert rerun["_parent_fallback_allocation"]["reused_successful_allocation"] is True


def test_same_parent_partial_rerun_skips_materialized_split(tmp_path):
    """同一父回款已拆行落表后重跑，不得再次写未结账承接行。"""
    orders = [{"so": "SO1", "deliver": 200.0}]
    sod_lines = {"SO1": [{"sod": "SOD1", "deliver": 200.0}]}
    first = _payment(amount=150.0, orders=orders, sod_lines=sod_lines)
    C.expand_payment(first, {})
    FAL.commit(tmp_path, {
        "hexiao_date": "2026-07-27",
        "parent_fallback_allocations": {
            first["ar"]: first["_parent_fallback_allocation"]
        },
        "write": [{"ar": first["ar"], "so": "SO1"}],
        "skip": [],
    })

    rerun = _payment(
        amount=150.0,
        orders=orders,
        sod_lines=sod_lines,
        _fallback_allocation_state=FAL.load(tmp_path),
    )
    rec = C.expand_payment(rerun, {})[0]
    ledger = _ledger({
        10: {
            "so": "SO1", "sod": "SOD1", "yingshou": 100.0,
            "jiti": None, "huikuan": 150.0, "jiezhang": "是",
            "shoukuan_time": dt.date(2026, 7, 24), "shoukuan_way": "汇",
        },
        11: {
            "so": "SO1", "sod": "SOD1", "yingshou": 100.0,
            "jiti": None, "huikuan": None, "jiezhang": "否",
        },
    })

    result = C.classify_one(rec, ledger, {}, 0.0, 2026)

    assert result["bucket"] == "auto"
    assert result["code"] == "OK_FALLBACK_ALLOCATION_ALREADY_APPLIED"
    assert result["ledger_row_ref"] == 10
    assert "row_operation" not in result
    assert result["five_cols"]["回款明细"] == 150.0


def test_itemized_cumulative_rerun_skips_materialized_partial_slice():
    """有逐单累计真相时，已落表的部分切片不得再次占用未结账行。"""
    rec = {
        "ar": "AR_ITEMIZED", "so": "SO1", "sod": "SOD1",
        "amount_orig": 50.0, "amount_local": 50.0,
        "deliver_local": 200.0, "cumulative_received_local": 150.0,
        "itemized_cumulative_authoritative": True,
        "currency": "人民币CNY", "status": "手动核销", "customer": "测试客户",
        "hexiao_date": dt.date(2026, 8, 2),
        "shoukuan_date": dt.date(2026, 7, 31),
    }
    ledger = _ledger({
        20: {
            "so": "SO1", "sod": "SOD1", "yingshou": 100.0,
            "jiti": None, "huikuan": 100.0, "jiezhang": "是",
            "shoukuan_time": dt.date(2026, 7, 20), "shoukuan_way": "汇",
        },
        21: {
            "so": "SO1", "sod": "SOD1", "yingshou": 40.0,
            "jiti": None, "huikuan": 50.0, "jiezhang": "是",
            "shoukuan_time": dt.date(2026, 7, 31), "shoukuan_way": "汇",
        },
        22: {
            "so": "SO1", "sod": "SOD1", "yingshou": 60.0,
            "jiti": None, "huikuan": None, "jiezhang": "否",
        },
    })

    result = C.classify_one(rec, ledger, {}, 0.0, 2026)

    assert result["bucket"] == "auto"
    assert result["code"] == "OK_ITEMIZED_CUMULATIVE_ALREADY_APPLIED"
    assert result["ledger_row_ref"] == 21
    assert "row_operation" not in result
    assert result["five_cols"]["回款明细"] == 50.0


def test_fallback_history_is_added_when_later_parent_has_itemized_writeoff():
    """SO26070160 类：前一父回款兜底已写，后续逐 SO 明细必须续加到同一累计。"""
    fallback_state = {
        "version": FAL.VERSION,
        "parents": {
            "AR_FALLBACK": {
                "ar": "AR_FALLBACK",
                "hexiao_date": "2026-07-13",
                "basis": "local",
                "parent_amount": 200.0,
                "allocations": [{
                    "so": "SO1",
                    "allocated": 200.0,
                    "allocated_orig": 200.0,
                    "allocated_local": 200.0,
                }],
            }
        },
    }
    payment = _payment(
        ar="AR_ITEMIZED",
        hexiao_date=dt.date(2026, 7, 16),
        amount=94.6,
        orders=[{"so": "SO1", "deliver": 300.0}],
        writeoffs={"SO1": 94.6},
        writeoffs_local={"SO1": 94.6},
        cumulative_writeoffs={"SO1": 94.6},
        cumulative_writeoffs_local={"SO1": 94.6},
        sod_lines={"SO1": [{"sod": "SOD1", "deliver": 300.0}]},
        _fallback_allocation_state=fallback_state,
        _detailed_parent_ars=["AR_ITEMIZED"],
    )

    rec = C.expand_payment(payment, {})[0]
    assert rec["cumulative_detail_local"] == 94.6
    assert rec["cumulative_fallback_local"] == 200.0
    assert rec["cumulative_received_local"] == 294.6

    result = C.classify_one(
        rec,
        _ledger({
            10: {
                "so": "SO1", "sod": "SOD1", "yingshou": 200.0,
                "jiti": None, "huikuan": 200.0, "jiezhang": "是",
                "shoukuan_time": dt.date(2026, 7, 9), "shoukuan_way": "汇",
            },
            11: {
                "so": "SO1", "sod": "SOD1", "yingshou": 100.0,
                "jiti": None, "huikuan": None, "jiezhang": "否",
            },
        }),
        {},
        0.0,
        2026,
    )

    assert result["bucket"] == "auto"
    assert result["code"] == "E5"
    assert result["code"] != "OK_ITEMIZED_CUMULATIVE_ALREADY_APPLIED"
    assert result["row_operation"]["existing_received"] == 200.0
    assert result["row_operation"]["current_received"] == 94.6
    assert result["row_operation"]["cumulative_received"] == 294.6
    assert result["row_operation"]["unpaid_receivable"] == 5.4


def test_itemized_idempotence_does_not_borrow_different_payment_slice():
    """表内累计较大但没有本次金额切片时，不得借用另一父回款行判幂等。"""
    rec = {
        "ar": "AR_ITEMIZED", "so": "SO1", "sod": "SOD1",
        "amount_orig": 94.6, "amount_local": 94.6,
        "deliver_local": 300.0, "cumulative_received_local": 94.6,
        "itemized_cumulative_authoritative": True,
        "currency": "人民币CNY", "status": "核销成功", "customer": "测试客户",
        "hexiao_date": dt.date(2026, 7, 16),
        "shoukuan_date": dt.date(2026, 7, 15),
    }
    ledger = _ledger({
        20: {
            "so": "SO1", "sod": "SOD1", "yingshou": 200.0,
            "jiti": None, "huikuan": 200.0, "jiezhang": "是",
        },
        21: {
            "so": "SO1", "sod": "SOD1", "yingshou": 100.0,
            "jiti": None, "huikuan": None, "jiezhang": "否",
        },
    })

    result = C.classify_one(rec, ledger, {}, 0.0, 2026)

    assert result["code"] != "OK_ITEMIZED_CUMULATIVE_ALREADY_APPLIED"
    assert result["bucket"] == "hold"


def test_fallback_history_totals_ignore_entries_after_cutoff_date():
    state = {
        "version": FAL.VERSION,
        "parents": {
            "AR_OLD": {
                "hexiao_date": "2026-07-13",
                "allocations": [{
                    "so": "SO1", "allocated_orig": 200.0, "allocated_local": 200.0,
                }],
            },
            "AR_FUTURE": {
                "hexiao_date": "2026-07-20",
                "allocations": [{
                    "so": "SO1", "allocated_orig": 50.0, "allocated_local": 50.0,
                }],
            },
        },
    }

    original, local = FAL.history_totals(
        state,
        current_ar="AR_CURRENT",
        as_of_date=dt.date(2026, 7, 16),
    )

    assert original == {"SO1": 200.0}
    assert local == {"SO1": 200.0}


def test_itemized_parent_excludes_its_own_old_fallback_allocation():
    """同一父 AR 后来取得逐 SO 明细时，旧兜底只被替换一次，不得叠加双计。"""
    state = {
        "version": FAL.VERSION,
        "parents": {
            "AR_ITEMIZED": {
                "hexiao_date": "2026-07-13",
                "allocations": [{
                    "so": "SO1", "allocated_orig": 200.0, "allocated_local": 200.0,
                }],
            },
        },
    }
    payment = _payment(
        ar="AR_ITEMIZED",
        hexiao_date=dt.date(2026, 7, 16),
        amount=94.6,
        orders=[{"so": "SO1", "deliver": 300.0}],
        writeoffs={"SO1": 94.6},
        writeoffs_local={"SO1": 94.6},
        cumulative_writeoffs={"SO1": 94.6},
        cumulative_writeoffs_local={"SO1": 94.6},
        sod_lines={"SO1": [{"sod": "SOD1", "deliver": 300.0}]},
        _fallback_allocation_state=state,
        _detailed_parent_ars=["AR_ITEMIZED"],
    )

    rec = C.expand_payment(payment, {})[0]

    assert rec["cumulative_fallback_local"] is None
    assert rec["cumulative_received_local"] == 94.6


def test_tax_and_other_explicit_fees_are_included_in_gross_parent_amount():
    p = _payment(
        amount=95.0,
        fee=1.0,
        tax=2.0,
        other_fee=2.0,
        orders=[{"so": "SO1", "deliver": 100.0}],
        sod_lines={"SO1": [{"sod": "SOD1", "deliver": 100.0}]},
    )
    rec = C.expand_payment(p, {})[0]
    assert rec["amount_orig"] == 100.0
    assert p["charge_amount_orig"] == 5.0
    assert p["total_amount_orig"] == 100.0


def test_xlsx_patch_forces_excel_formula_recalculation(tmp_path):
    """写表后要求 Excel 重算，避免公式存在但缓存值仍是旧值。"""
    src = tmp_path / "source.xlsx"
    out = tmp_path / "output.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "明细"
    ws["A1"] = 1
    ws["B1"] = "=A1*2"
    wb.save(src)

    xlsx_patch.patch_cells(src, out, "明细", [(1, 1, 3)])
    with zipfile.ZipFile(out) as zf:
        workbook_xml = zf.read("xl/workbook.xml").decode("utf-8")
    assert 'calcMode="auto"' in workbook_xml
    assert 'fullCalcOnLoad="1"' in workbook_xml
    assert 'forceFullCalc="1"' in workbook_xml
