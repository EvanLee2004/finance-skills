# -*- coding: utf-8 -*-
"""
2026-07-25 两批改动的回归：

A. 写入前的两道闸（防"确认那会儿的表"≠"写入这会儿的表"）
   · 盈亏：指纹 + 逐行身份复核（apply_to_copy.precheck_before_write）
   · 流转：命中行身份复核（apply_flow.precheck_flow_identity）
   · 回读把 SOD 也纳入比对
   · 幂等不再放过「计划算留空、表里却填了值」
B. 日期口径（核销日期 vs 到账日期）与跑批台账（漏天检测）
"""
import datetime as dt
import json
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent))

import common as C  # noqa: E402
import batch_ledger as BL  # noqa: E402
import validate_plan as V  # noqa: E402
import apply_to_copy as A  # noqa: E402
import apply_flow as AF  # noqa: E402

HDR = ["部门", "销售人员", "客户名称", "单号", "新智云单号", "应收金额",
       "计提金额", "回款明细", "是否结账（是/否）", "收款时间", "收款方式(支/汇/现)", "实收金额"]


def _ledger(tmp_path, rows, name="盈亏.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "明细"
    ws.append(HDR)
    for so, sod, five in rows:
        r = ["部", "人", "客", "AB", so, 100, None, None, None, None, None, sod]
        if five:
            r[6], r[7], r[8], r[9], r[10] = (
                five.get("计提"), five.get("回款明细"), five.get("是否结账"),
                five.get("收款时间"), five.get("收款方式"),
            )
        ws.append(r)
    p = tmp_path / name
    wb.save(str(p))
    return p


def _item(row, so="SO26010001", sod="SOD26010001", **five):
    f = {"计提": 100.0, "回款明细": 100.0, "是否结账": "是",
         "收款时间": "2026-07-08", "收款方式": "汇", "实收SOD": sod}
    f.update(five)
    return {"case_id": f"AR1|{so}", "ar": "AR1", "so": so, "sod": sod,
            "ledger_row_ref": row, "five_cols": f}


# ══════════════════════════════════════════════════════════
# A1. 幂等不再放过「计划留空 vs 表里有值」
# ══════════════════════════════════════════════════════════

def test_plan_blank_but_table_filled_is_conflict(tmp_path):
    """
    本次算的是「计提留空」（没回满交付额），但她表里计提已经填了 1000。
    旧版：只比"计划非 None"的列 → 判 skip「已填过且一致」，静默放过。
    新版：必须报冲突——「计提该不该填」正是最容易错、最该她看的一条。
    """
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", {
        "计提": 1000, "回款明细": 100.0, "是否结账": "是",
        "收款时间": "2026-07-08", "收款方式": "汇",
    })])
    rows = V.read_ledger_rows(led)
    res = V.check_one(_item(2, 计提=None), rows)
    assert res["verdict"] == "conflict"
    assert "留空" in res["reason"]


def test_identical_row_still_skips(tmp_path):
    """真一致的还是要跳过——别把幂等改坏了。"""
    five = {"计提": 100.0, "回款明细": 100.0, "是否结账": "是",
            "收款时间": "2026-07-08", "收款方式": "汇"}
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", five)])
    rows = V.read_ledger_rows(led)
    assert V.check_one(_item(2), rows)["verdict"] == "skip"


def test_plan_blank_and_table_blank_is_consistent(tmp_path):
    """两边都留空 → 仍算一致，不能因为新逻辑就误报冲突。"""
    five = {"计提": None, "回款明细": 100.0, "是否结账": "是",
            "收款时间": "2026-07-08", "收款方式": "汇"}
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", five)])
    rows = V.read_ledger_rows(led)
    assert V.check_one(_item(2, 计提=None), rows)["verdict"] == "skip"


# ══════════════════════════════════════════════════════════
# A2. validate 记指纹与身份
# ══════════════════════════════════════════════════════════

def test_validate_records_fingerprint_and_identity(tmp_path):
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", None)])
    rows = V.read_ledger_rows(led)
    out = V.validate({"auto": [_item(2)], "hexiao_date": "2026-07-24"}, rows, ledger_path=led)
    assert out["ledger_sha256"] == C.sha256_file(led)
    assert out["hexiao_date"] == "2026-07-24"
    assert out["write"][0]["_identity"] == {"row": 2, "SO": "SO26010001", "SOD": "SOD26010001"}


# ══════════════════════════════════════════════════════════
# A3. 盈亏写入前的闸
# ══════════════════════════════════════════════════════════

def _checked(led, items):
    rows = V.read_ledger_rows(led)
    return V.validate({"auto": items}, rows, ledger_path=led)


def test_precheck_passes_when_untouched(tmp_path):
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", None)])
    plan = _checked(led, [_item(2)])
    assert A.precheck_before_write(plan, plan["write"], led) == []


def test_precheck_blocks_when_table_changed_after_validate(tmp_path):
    """她在「出清单 → 说确认」之间自己动了表 → 指纹对不上 → 整批拒写。"""
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", None)])
    plan = _checked(led, [_item(2)])
    wb = openpyxl.load_workbook(str(led))
    wb["明细"]["A3"] = "她后来加的一行"
    wb.save(str(led))
    problems = A.precheck_before_write(plan, plan["write"], led)
    assert problems and any("指纹对不上" in p for p in problems)


def test_precheck_blocks_when_row_became_another_order(tmp_path):
    """
    最危险的那种：她在上方插了一行 → 第 2 行现在是**另一单**了。
    按旧行号写就会把这笔的五列 + SOD 写到别人那单上，而写后回读照样报 ✓。
    """
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", None)])
    plan = _checked(led, [_item(2)])
    wb = openpyxl.load_workbook(str(led))
    wb["明细"].insert_rows(2)  # 模拟"月初贴交付插行"
    wb["明细"]["E2"] = "SO26990999"
    wb["明细"]["L2"] = "SOD26990999"
    wb.save(str(led))
    problems = A.precheck_before_write(plan, plan["write"], led)
    assert problems
    assert any("已经不是原来那单" in p or "现在不能写" in p for p in problems)


def test_apply_refuses_to_write_when_stale(tmp_path):
    """端到端：表被动过时 apply 必须非 0 退出，且一个格子都没写。"""
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", None)])
    plan = _checked(led, [_item(2)])
    checked_p = tmp_path / "写入计划_校验后.json"
    checked_p.write_text(json.dumps(plan, ensure_ascii=False, default=str), encoding="utf-8")

    wb = openpyxl.load_workbook(str(led))
    wb["明细"].insert_rows(2)
    wb["明细"]["E2"] = "SO26990999"
    wb.save(str(led))
    before = C.sha256_file(led)

    rc = A.main(["--checked", str(checked_p), "--ledger", str(led),
                 "--confirmed", "--in-place"])
    assert rc != 0
    assert C.sha256_file(led) == before  # 一个字节都没动


def test_verify_written_catches_wrong_sod(tmp_path):
    """回读比对必须覆盖 SOD——旧版只比五列，SOD 写歪了照样报「全部一致 ✓」。"""
    led = _ledger(tmp_path, [("SO26010001", "SOD_WRONG", {
        "计提": 100.0, "回款明细": 100.0, "是否结账": "是",
        "收款时间": "2026-07-08", "收款方式": "汇",
    })])
    problems = A.verify_written(led, [_item(2)])
    assert any("SOD" in p for p in problems)


# ══════════════════════════════════════════════════════════
# A4. 流转写入前的身份闸
# ══════════════════════════════════════════════════════════

FLOW_HDR = ["日期", "公司名称", "金额", "单号", "是否更新应收款"]


def _flow(tmp_path, rows):
    d = tmp_path / "02_我的表副本"
    d.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "汇总"
    ws.append(FLOW_HDR)
    for r in rows:
        ws.append(list(r))
    p = d / "到账流转单.xlsx"
    wb.save(str(p))
    return p


def _flow_item(row_no, date="2026-07-24", payer="某某公司", amount=1000.0):
    return {
        "ar": "AR1", "file": "到账流转单.xlsx", "sheet": "汇总", "row_no": row_no,
        "verdict": "write", "order_suggest": "SO26010001", "updated_suggest": "是",
        "hits": 1, "matched_by": "三键",
        "identity": {"date": date, "payer": payer, "amount": amount},
    }


def test_flow_precheck_passes_when_untouched(tmp_path):
    _flow(tmp_path, [[dt.date(2026, 7, 24), "某某公司", 1000.0, "", ""]])
    assert AF.precheck_flow_identity(tmp_path, [_flow_item(2)]) == []


def test_flow_precheck_blocks_after_insert(tmp_path):
    """她往流转表里插了一行 → 第 2 行换人了 → 单号会写到别人那笔到账上。"""
    _flow(tmp_path, [
        [dt.date(2026, 7, 23), "另一家公司", 555.0, "", ""],
        [dt.date(2026, 7, 24), "某某公司", 1000.0, "", ""],
    ])
    problems = AF.precheck_flow_identity(tmp_path, [_flow_item(2)])
    assert problems


def test_flow_write_aborts_whole_batch_on_identity_mismatch(tmp_path):
    """一条对不上就整批不写——插一行会让它之后的全部错位，只跳过那条等于把剩下的写歪。"""
    src = _flow(tmp_path, [
        [dt.date(2026, 7, 23), "另一家公司", 555.0, "", ""],
        [dt.date(2026, 7, 24), "某某公司", 1000.0, "", ""],
    ])
    before = C.sha256_file(src)
    changes, problems = AF.write_flow_items(tmp_path, [_flow_item(2)], in_place=True)
    assert changes == []
    assert problems
    assert C.sha256_file(src) == before


# ══════════════════════════════════════════════════════════
# B. 日期口径与跑批台账
# ══════════════════════════════════════════════════════════

def test_resolve_batch_date_is_absolute():
    today = dt.date(2026, 7, 25)  # 周六
    assert C.resolve_batch_date("yesterday", today) == dt.date(2026, 7, 24)
    assert C.resolve_batch_date("2026-07-20", today) == dt.date(2026, 7, 20)
    assert C.resolve_batch_date("今天", today) == today
    assert C.resolve_batch_date("", today) is None


def test_prev_workday_skips_weekend():
    """周一跑时"昨天"是周日，销售周末不核销 → 默认值该给上周五。"""
    assert C.prev_workday(dt.date(2026, 7, 27)) == dt.date(2026, 7, 24)  # 周一 → 上周五
    assert C.prev_workday(dt.date(2026, 7, 24)) == dt.date(2026, 7, 23)


def test_date_cn_has_weekday():
    assert C.date_cn("2026-07-24") == "2026-07-24（周五）"
    assert C.date_cn(None) == "(未知)"


def test_batch_ledger_records_and_finds_gaps(tmp_path):
    """
    核心诉求：她请假/周末跳过的那几天，程序要**主动报出来**，
    而不是每次闷头跑"昨天"、漏掉的天永远没人管。
    """
    BL.record(tmp_path, dt.date(2026, 7, 20), "applied", payments=3)
    BL.record(tmp_path, dt.date(2026, 7, 23), "applied", payments=5)
    info = BL.find_gaps(tmp_path, through=dt.date(2026, 7, 24))
    gaps = [d.isoformat() for d in info["gaps"]]
    assert gaps == ["2026-07-21", "2026-07-22", "2026-07-24"]


def test_batch_ledger_skips_weekend_by_default(tmp_path):
    BL.record(tmp_path, dt.date(2026, 7, 23), "applied", payments=1)
    info = BL.find_gaps(tmp_path, through=dt.date(2026, 7, 27))
    gaps = [d.isoformat() for d in info["gaps"]]
    assert "2026-07-25" not in gaps  # 周六
    assert "2026-07-26" not in gaps  # 周日
    assert "2026-07-24" in gaps and "2026-07-27" in gaps
    assert len(info["skipped_weekend"]) == 2


def test_empty_batch_counts_as_done(tmp_path):
    """那天真的没人核销 = 处理完了，不该一直被当成漏天反复提醒。"""
    BL.record(tmp_path, dt.date(2026, 7, 23), "classified", payments=0)
    info = BL.find_gaps(tmp_path, through=dt.date(2026, 7, 23))
    assert info["gaps"] == []


def test_fetched_only_still_counts_as_gap(tmp_path):
    """只取了数没判过 = 跑了一半被打断，仍算没处理。"""
    BL.record(tmp_path, dt.date(2026, 7, 23), "classified", payments=1)
    BL.record(tmp_path, dt.date(2026, 7, 24), "fetched", payments=2)
    info = BL.find_gaps(tmp_path, through=dt.date(2026, 7, 24))
    assert [d.isoformat() for d in info["gaps"]] == ["2026-07-24"]


def test_stage_never_goes_backwards(tmp_path):
    """重跑判定不该把「已写表·收工」降级成「判过了」。"""
    BL.record(tmp_path, dt.date(2026, 7, 23), "applied", payments=1)
    rec = BL.record(tmp_path, dt.date(2026, 7, 23), "classified", payments=1)
    assert rec["stage"] == "applied"


def test_suggest_date_prefers_earliest_gap(tmp_path):
    """有空档就先补最早的那天，而不是继续跑昨天（否则窟窿越积越大）。"""
    BL.record(tmp_path, dt.date(2026, 7, 20), "applied", payments=1)
    s = BL.suggest_date(tmp_path, today=dt.date(2026, 7, 24))
    assert s["date"] == dt.date(2026, 7, 21)
    assert "没跑过" in s["reason"]


def test_first_use_has_no_gaps(tmp_path):
    """第一次用不该从年初开始报一堆空档。"""
    info = BL.find_gaps(tmp_path, through=dt.date(2026, 7, 24))
    assert info["first_use"] is True and info["gaps"] == []


# ══════════════════════════════════════════════════════════
# C. 补多天：agent 多跑几遍（不并批），顺序从早到晚
# ══════════════════════════════════════════════════════════

def test_gaps_are_sorted_earliest_first(tmp_path):
    """
    补多天必须**从早到晚**：同一订单可能分两天核销、填她表里同一行，
    先写早的、晚的那天才看得到"已填一半"接着算；顺序反了两笔会抢同一行双双挂起。
    """
    BL.record(tmp_path, dt.date(2026, 7, 20), "applied", payments=1)
    info = BL.find_gaps(tmp_path, through=dt.date(2026, 7, 24))
    gaps = info["gaps"]
    assert gaps == sorted(gaps)
    assert gaps[0] == dt.date(2026, 7, 21)


def test_run_days_one_by_one_clears_gaps(tmp_path):
    """一天一天跑完，空档逐个消失；中途停下剩余的仍留在台账里（不会丢）。"""
    BL.record(tmp_path, dt.date(2026, 7, 20), "applied", payments=1)
    assert len(BL.find_gaps(tmp_path, through=dt.date(2026, 7, 23))["gaps"]) == 3

    BL.record(tmp_path, dt.date(2026, 7, 21), "applied", payments=2)  # 跑第 1 天
    left = BL.find_gaps(tmp_path, through=dt.date(2026, 7, 23))["gaps"]
    assert [d.isoformat() for d in left] == ["2026-07-22", "2026-07-23"]

    BL.record(tmp_path, dt.date(2026, 7, 22), "classified", payments=0)  # 第 2 天空批也算跑过
    left = BL.find_gaps(tmp_path, through=dt.date(2026, 7, 23))["gaps"]
    assert [d.isoformat() for d in left] == ["2026-07-23"]


def test_suggest_walks_forward_day_by_day(tmp_path):
    """suggest_date 每次都指向"还没跑的最早一天"，正好就是 agent 下一遍该跑的那天。"""
    BL.record(tmp_path, dt.date(2026, 7, 20), "applied", payments=1)
    assert BL.suggest_date(tmp_path, today=dt.date(2026, 7, 24))["date"] == dt.date(2026, 7, 21)
    BL.record(tmp_path, dt.date(2026, 7, 21), "applied", payments=1)
    assert BL.suggest_date(tmp_path, today=dt.date(2026, 7, 24))["date"] == dt.date(2026, 7, 22)
