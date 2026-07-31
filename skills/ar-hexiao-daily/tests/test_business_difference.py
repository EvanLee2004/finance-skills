# -*- coding: utf-8 -*-
"""业务值差异：判定、校验、OOXML公式缓存、日清展示端到端回归。"""
import datetime as dt
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))

import apply_to_copy as A  # noqa: E402
import build_worklist as W  # noqa: E402
import classify_hexiao as C  # noqa: E402
import validate_plan as V  # noqa: E402


HDR = [
    "部门", "销售人员", "客户名称", "单号", "新智云单号", "应收金额",
    "计提金额", "回款明细", "是否结账（是/否）", "收款时间",
    "收款方式(支/汇/现)", "实收金额", "差异",
]


def _ledger(tmp_path, *, with_difference=True, rows=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "明细"
    ws.append(HDR if with_difference else HDR[:-1])
    for so, sod, receivable, five, difference in rows or []:
        row = [
            "部", "人", "客", "AB", so, receivable,
            None, None, None, None, None, sod,
        ]
        if with_difference:
            row.append(difference)
        if five:
            row[6:11] = [
                five.get("计提"), five.get("回款明细"), five.get("是否结账"),
                five.get("收款时间"), five.get("收款方式"),
            ]
        ws.append(row)
    path = tmp_path / ("盈亏_有差异.xlsx" if with_difference else "盈亏_无差异.xlsx")
    wb.save(path)
    return path


def _item(row=2, so="SO1", sod="SOD1", *, amount=110.0, difference=-10.0):
    item = {
        "case_id": f"AR1|{so}|{sod}",
        "ar": "AR1",
        "so": so,
        "sod": sod,
        "ledger_row_ref": row,
        "five_cols": {
            "计提": amount,
            "回款明细": amount,
            "是否结账": "是",
            "收款时间": "2026-07-08",
            "收款方式": "汇",
            "实收SOD": sod,
        },
        "current_values": {
            "计提": None, "回款明细": None, "差异": None,
            "是否结账": "", "收款时间": "", "收款方式": None,
            "实收SOD": sod,
        },
    }
    if difference is not None:
        item["derived_cols"] = {"差异": difference}
    return item


def _split_item(row=2, so="SO_SPLIT", sod="SOD_SPLIT"):
    item = _item(row, so, sod, amount=50.0, difference=None)
    item["five_cols"]["计提"] = None
    item["row_operation"] = {
        "type": "split_below",
        "source_receivable": 100.0,
        "paid_receivable": 40.0,
        "unpaid_receivable": 60.0,
        "baseline_receivable": 100.0,
        "paid_side_receivable_total": 40.0,
        "existing_received": 0.0,
        "current_received": 50.0,
        "cumulative_received": 50.0,
        "latest_delivery": 110.0,
        "inserted_five_cols": {
            "计提": None, "回款明细": None, "是否结账": "否",
            "收款时间": None, "收款方式": None, "实收SOD": sod,
        },
    }
    return item


def _synthetic_ledger(receivable=713.24):
    return C.LedgerIndex(synthetic={
        "so": {"SO1": [1]},
        "sod": {},
        "rows": {1: {
            "so": "SO1", "sod": "", "yingshou": receivable,
            "jiti": None, "huikuan": None, "chayi": None,
            "jiezhang": "", "shoukuan_time": None, "shoukuan_way": None,
        }},
    })


def _record(amount=300.0, delivery=765.77, cumulative=765.77):
    return {
        "ar": "AR1", "so": "SO1", "sod": "SOD1",
        "amount_orig": amount, "amount_local": amount,
        "deliver_local": delivery,
        "cumulative_received_local": cumulative,
        "currency": "人民币CNY", "status": "手动核销",
        "hexiao_date": dt.date(2026, 7, 22),
        "shoukuan_date": dt.date(2026, 7, 21),
    }


def test_mismatch_uses_latest_delivery_and_business_difference():
    result = C.classify_one(_record(), _synthetic_ledger(), {}, 0.0, 2026)
    assert result["five_cols"]["计提"] == 765.77
    assert result["five_cols"]["回款明细"] == 300.0
    assert result["five_cols"]["是否结账"] == "是"
    assert result["derived_cols"]["差异"] == -52.53


def test_partial_does_not_write_accrual_or_difference():
    result = C.classify_one(
        _record(amount=300.0, delivery=765.77, cumulative=300.0),
        _synthetic_ledger(),
        {},
        0.0,
        2026,
    )
    assert result["row_operation"]["type"] == "split_below"
    assert result["five_cols"]["计提"] is None
    assert "derived_cols" not in result


def test_difference_write_skip_and_conflict(tmp_path):
    five = {
        "计提": 110.0, "回款明细": 110.0, "是否结账": "是",
        "收款时间": dt.date(2026, 7, 8), "收款方式": "汇",
    }
    missing = _ledger(
        tmp_path, rows=[("SO1", "SOD1", 100.0, five, None)]
    )
    assert V.check_one(_item(), V.read_ledger_rows(missing))["verdict"] == "write"

    correct = _ledger(
        tmp_path, rows=[("SO1", "SOD1", 100.0, five, -10.0)]
    )
    assert V.check_one(_item(), V.read_ledger_rows(correct))["verdict"] == "skip"

    wrong = _ledger(
        tmp_path, rows=[("SO1", "SOD1", 100.0, five, -9.0)]
    )
    assert V.check_one(_item(), V.read_ledger_rows(wrong))["verdict"] == "conflict"


def test_missing_difference_column_only_blocks_required_difference(tmp_path):
    ledger = _ledger(
        tmp_path,
        with_difference=False,
        rows=[("SO1", "SOD1", 100.0, None, None)],
    )
    rows = V.read_ledger_rows(ledger)
    assert V.check_one(_item(difference=None, amount=100.0), rows)["verdict"] == "write"
    result = V.check_one(_item(), rows)
    assert result["verdict"] == "conflict"
    assert "没有“差异”列" in result["reason"]


def test_ooxml_formula_and_cache_round_trip(tmp_path):
    ledger = _ledger(
        tmp_path, rows=[("SO1", "SOD1", 100.0, None, None)]
    )
    out = tmp_path / "公式缓存.xlsx"
    item = _item()
    A.write_plan(ledger, out, [item])
    assert A.verify_written(out, [item]) == []
    formula_wb = openpyxl.load_workbook(out, data_only=False)
    assert formula_wb["明细"].cell(2, 13).value == "=F2-G2"
    formula_wb.close()
    value_wb = openpyxl.load_workbook(out, data_only=True)
    assert value_wb["明细"].cell(2, 13).value == -10.0
    value_wb.close()


def test_formula_uses_final_row_after_split(tmp_path):
    ledger = _ledger(tmp_path, rows=[
        ("SO_SPLIT", "SOD_SPLIT", 100.0, None, None),
        ("SO1", "SOD1", 100.0, None, None),
    ])
    split = _split_item()
    difference = _item(row=3)
    out = tmp_path / "插行后公式.xlsx"
    A.write_plan(ledger, out, [split, difference])
    assert A.verify_written(out, [split, difference]) == []
    formula_wb = openpyxl.load_workbook(out, data_only=False)
    assert formula_wb["明细"].cell(4, 13).value == "=F4-G4"
    assert formula_wb["明细"].cell(2, 13).value is None
    assert formula_wb["明细"].cell(3, 13).value is None
    formula_wb.close()


def test_worklist_has_three_distinct_difference_columns(tmp_path):
    item = _item()
    result = {
        "auto": [item], "hold": [], "exception": [],
        "counts": {"auto": 1, "hold": 0, "exception": 0, "total": 1},
        "e_code_dist": {"OK": 1}, "payment_count": 1,
    }
    checked = {
        "write": [{**item, "_check": {"verdict": "write", "reason": "可写"}}],
        "skip": [], "conflict": [],
        "counts": {"write": 1, "skip": 0, "conflict": 0},
    }
    out = tmp_path / "核销日清.xlsx"
    W.build_workbook(result, checked, out)
    ws = openpyxl.load_workbook(out)["今日清单"]
    headers = [c.value for c in ws[1]]
    assert "应填_业务值差异" in headers
    assert "当前_业务值差异" in headers
    assert "当前值与计划值的比较差异" in headers
