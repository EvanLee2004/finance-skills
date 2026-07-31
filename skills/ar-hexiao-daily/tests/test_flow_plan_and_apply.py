# -*- coding: utf-8 -*-
"""流转写入计划 + 确认后 apply_flow / apply_all 闸门。"""
from __future__ import annotations

import hashlib
import json
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "scripts"))

import build_flow_plan as BFP  # noqa: E402
import apply_flow as AF  # noqa: E402
import apply_all as AA  # noqa: E402
import apply_to_copy as AC  # noqa: E402
import build_worklist as BW  # noqa: E402


def _sha(p: Path) -> str:
    return hashlib.sha256(p.read_bytes()).hexdigest()


def _flow_xlsx(path: Path, rows):
    """rows: list of (date, payer, amount, order, updated)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "明细"
    ws.append(["日期", "公司名称", "金额", "单号", "是否更新应收款", "收款形式"])
    for r in rows:
        ws.append(list(r))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return path


def _result_with_flow_items(items_meta):
    """
    items_meta: list of dicts with ar, so, flow_* fields, bucket
    """
    auto, hold = [], []
    for m in items_meta:
        it = {
            "ar": m["ar"],
            "so": m["so"] if "so" in m else "SO1",
            "sod": m.get("sod") or "SOD1",
            "bucket": m.get("bucket") or "auto",
            "code": m.get("code") or "",
            "flow_hits": m.get("flow_hits"),
            "flow_matched_by": m.get("flow_matched_by") or "",
            "flow_locate": m.get("flow_locate") or "",
            "flow_file": m.get("flow_file") or "",
            "flow_sheet": m.get("flow_sheet") or "明细",
            "flow_row_no": m.get("flow_row_no"),
            "flow_order_suggest": m.get("flow_order_suggest") or "",
            "flow_order_existing": m.get("flow_order_existing") or "",
            "five_cols": {"回款明细": 100, "是否结账": "是"},
        }
        (auto if it["bucket"] == "auto" else hold).append(it)
    # ar_summary
    ars = {}
    for m in items_meta:
        ars.setdefault(m["ar"], []).append(m)
    summary = []
    for ar, group in ars.items():
        summary.append({
            "ar": ar,
            "so_count": len(group),
            "行数": len(group),
            "buckets": ["auto"] * len(group),
            "流转表_是否更新应收款_建议": group[0].get("updated") or "是",
            "flow_locate": group[0].get("flow_locate") or "",
            "待处理SO": [],
        })
    return {
        "auto": auto,
        "hold": hold,
        "exception": [],
        "counts": {"auto": len(auto), "hold": len(hold), "exception": 0, "total": len(auto) + len(hold)},
        "ar_summary": summary,
        "payment_count": len(ars),
    }


def test_plan_strong_write_weak_hand():
    fname = "流转A.xlsx"
    result = _result_with_flow_items([
        {
            "ar": "AR1",
            "so": "SO1",
            "flow_hits": 1,
            "flow_matched_by": "三键",
            "flow_file": fname,
            "flow_sheet": "明细",
            "flow_row_no": 2,
            "flow_order_suggest": "SO1",
            "updated": "是",
        },
        {
            "ar": "AR2",
            "so": "SO2",
            "flow_hits": 1,
            "flow_matched_by": "日期+金额(名字不符)",
            "flow_file": fname,
            "flow_sheet": "明细",
            "flow_row_no": 3,
            "flow_order_suggest": "SO2",
            "updated": "是",
        },
        {
            "ar": "AR3",
            "so": "SO3",
            "flow_hits": 0,
            "flow_matched_by": "未在现有流转表中找到",
            "updated": "",
        },
        {
            "ar": "AR4",
            "so": "SO4",
            "flow_hits": 2,
            "flow_matched_by": "三键",
            "updated": "部分",
        },
    ])
    plan = BFP.build_plan(result)
    by = {it["ar"]: it for it in plan["items"]}
    assert by["AR1"]["verdict"] == "write"
    assert by["AR2"]["verdict"] == "hand"
    assert by["AR3"]["verdict"] == "hand"
    assert by["AR4"]["verdict"] == "hand"
    assert plan["counts"]["write"] == 1
    assert plan["counts"]["hand"] >= 3


def test_flow_plan_always_covers_all_results_without_named_so_filter():
    result = _result_with_flow_items([
        {
            "ar": "AR1",
            "so": "SO_KEEP",
            "flow_hits": 1,
            "flow_matched_by": "三键",
            "flow_file": "流转A.xlsx",
            "flow_sheet": "明细",
            "flow_row_no": 2,
        },
        {
            "ar": "AR2",
            "so": "SO_OTHER",
            "flow_hits": 1,
            "flow_matched_by": "三键",
            "flow_file": "流转A.xlsx",
            "flow_sheet": "明细",
            "flow_row_no": 3,
        },
    ])
    plan = BFP.build_plan(result)
    assert {it["ar"] for it in plan["items"]} == {"AR1", "AR2"}
    assert plan["selection"] == {
        "mode": "all_results",
    }


def test_apply_flow_rejects_without_confirmed(tmp_path):
    plan = {"items": [{"ar": "AR1", "verdict": "write", "file": "x.xlsx", "sheet": "明细", "row_no": 2,
                       "order_suggest": "SO1", "updated_suggest": "是"}], "counts": {"write": 1}}
    p = tmp_path / "plan.json"
    p.write_text(json.dumps(plan), encoding="utf-8")
    rc = AF.main(["--plan", str(p), "--workspace", str(tmp_path)])
    assert rc == 2


def test_apply_flow_strong_write_and_readback(tmp_path):
    ws = tmp_path / "ws"
    (ws / "02_我的表副本").mkdir(parents=True)
    (ws / "04_产出").mkdir(parents=True)
    flow_path = ws / "02_我的表副本" / "流转测.xlsx"
    _flow_xlsx(flow_path, [
        ("2026-07-22", "甲公司", 100, "", ""),
    ])
    before = _sha(flow_path)
    plan = {
        "items": [{
            "ar": "AR1",
            "verdict": "write",
            "file": "流转测.xlsx",
            "sheet": "明细",
            "row_no": 2,
            "order_suggest": "SO26010001",
            "updated_suggest": "是",
            "matched_by": "三键",
            "hits": 1,
        }],
        "counts": {"write": 1, "hand": 0, "skip": 0},
    }
    plan_p = ws / "04_产出" / "流转写入计划_校验后.json"
    plan_p.write_text(json.dumps(plan, ensure_ascii=False), encoding="utf-8")

    # 无 confirmed 哈希不变
    rc = AF.main(["--plan", str(plan_p), "--workspace", str(ws)])
    assert rc == 2
    assert _sha(flow_path) == before

    rc = AF.main([
        "--plan", str(plan_p), "--workspace", str(ws),
        "--confirmed", "--in-place",
    ])
    assert rc == 0
    assert _sha(flow_path) != before
    wb = openpyxl.load_workbook(str(flow_path), read_only=True, data_only=True)
    row = list(wb["明细"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
    wb.close()
    assert row[3] == "SO26010001"
    assert row[4] == "是"


def test_apply_flow_skips_hand(tmp_path):
    ws = tmp_path / "ws"
    (ws / "02_我的表副本").mkdir(parents=True)
    flow_path = ws / "02_我的表副本" / "流转测.xlsx"
    _flow_xlsx(flow_path, [("2026-07-22", "甲", 100, "", "")])
    before = _sha(flow_path)
    plan = {
        "items": [{
            "ar": "AR1", "verdict": "hand", "file": "流转测.xlsx", "sheet": "明细",
            "row_no": 2, "order_suggest": "SO9", "updated_suggest": "是",
        }],
        "counts": {"write": 0, "hand": 1},
    }
    plan_p = tmp_path / "plan.json"
    plan_p.write_text(json.dumps(plan), encoding="utf-8")
    rc = AF.main(["--plan", str(plan_p), "--workspace", str(ws), "--confirmed", "--in-place"])
    assert rc == 0
    assert _sha(flow_path) == before


def test_worklist_shows_flow_mk(tmp_path):
    result = _result_with_flow_items([
        {
            "ar": "AR1", "so": "SO1", "flow_hits": 1, "flow_matched_by": "三键",
            "flow_file": "f.xlsx", "flow_row_no": 2, "flow_order_suggest": "SO1", "updated": "是",
        },
        {
            "ar": "AR2", "so": "SO2", "flow_hits": 0, "flow_matched_by": "未找到", "updated": "",
        },
    ])
    plan = BFP.build_plan(result)
    out = tmp_path / "日清.xlsx"
    BW.build_workbook(result, None, out, flow_plan=plan)
    wb = openpyxl.load_workbook(str(out), read_only=True, data_only=True)
    assert "先看这里" in wb.sheetnames
    text = "\n".join(
        str(r[0]) for r in wb["先看这里"].iter_rows(values_only=True) if r and r[0]
    )
    assert "确认后将自动写" in text
    assert "须你手填" in text
    assert "流转表怎么填" in wb.sheetnames
    headers = [c for c in next(wb["流转表怎么填"].iter_rows(min_row=1, max_row=1, values_only=True))]
    assert "写入方式" in headers
    wb.close()


def test_apply_all_rejects_without_confirmed(tmp_path):
    led = tmp_path / "盈亏.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "明细"
    ws.append(["新智云单号", "应收金额", "计提金额", "回款明细", "是否结账（是/否）",
               "收款时间", "收款方式(支/汇/现)", "实收金额"])
    ws.append(["SO1", 100, None, None, None, None, None, "SOD1"])
    wb.save(str(led))
    checked = tmp_path / "checked.json"
    checked.write_text(json.dumps({"write": [], "skip": [], "conflict": []}), encoding="utf-8")
    rc = AA.main([
        "--checked", str(checked), "--ledger", str(led), "--workspace", str(tmp_path),
    ])
    assert rc == 2


def test_apply_all_resnapshots_after_all_writes_succeed(tmp_path, monkeypatch):
    """最终指纹必须在盈亏、流转和跑批台账都成功后统一刷新。"""
    checked = tmp_path / "checked.json"
    checked.write_text(
        json.dumps({"hexiao_date": "2026-07-27", "write": [], "skip": [], "conflict": []}),
        encoding="utf-8",
    )
    flow_plan = tmp_path / "flow.json"
    flow_plan.write_text(json.dumps({"items": []}), encoding="utf-8")
    ledger = tmp_path / "ledger.xlsx"
    ledger.write_bytes(b"placeholder")
    calls = []

    monkeypatch.setattr(AA.apply_to_copy, "main", lambda _args: 0)
    monkeypatch.setattr(AA.apply_flow, "main", lambda _args: 0)
    monkeypatch.setattr(AA, "_record_done", lambda *args, **kwargs: None)
    monkeypatch.setattr(AA, "_resnapshot_sources", lambda workspace: calls.append(Path(workspace)))

    rc = AA.main([
        "--checked", str(checked),
        "--flow-plan", str(flow_plan),
        "--ledger", str(ledger),
        "--workspace", str(tmp_path),
        "--confirmed",
        "--in-place",
        "--flow-in-place",
    ])

    assert rc == 0
    assert calls == [tmp_path]


def test_plan_empty_so_preserves_existing_order_and_exact_strong():
    """so 列表空时 order_suggest 必须保留 flow_order_existing；伪造「三键*」不得 write。"""
    result = _result_with_flow_items([
        {
            "ar": "AR_KEEP",
            "so": "",  # 无 SO
            "flow_hits": 1,
            "flow_matched_by": "三键",
            "flow_file": "流转A.xlsx",
            "flow_sheet": "明细",
            "flow_row_no": 2,
            "flow_order_existing": "EXISTING_SO",
            "flow_order_suggest": "",
            "updated": "是",
        },
        {
            "ar": "AR_FAKE",
            "so": "SO9",
            "flow_hits": 1,
            "flow_matched_by": "三键(伪造变体)",
            "flow_file": "流转A.xlsx",
            "flow_sheet": "明细",
            "flow_row_no": 3,
            "flow_order_suggest": "SO9",
            "updated": "是",
        },
    ])
    # 手工补 existing（helper 未传的字段）
    for it in result["auto"] + result["hold"]:
        if it["ar"] == "AR_KEEP":
            it["flow_order_existing"] = "EXISTING_SO"
            it["so"] = ""
    plan = BFP.build_plan(result)
    by = {it["ar"]: it for it in plan["items"]}
    assert by["AR_KEEP"]["order_suggest"] == "EXISTING_SO"
    assert by["AR_KEEP"]["verdict"] == "write"
    assert by["AR_KEEP"].get("write_order") is True
    assert by["AR_FAKE"]["verdict"] == "hand", by["AR_FAKE"]
    assert "非强三键" in (by["AR_FAKE"].get("reason") or "") or "伪造" in (
        by["AR_FAKE"].get("matched_by") or ""
    )


def test_apply_flow_empty_order_does_not_wipe_existing(tmp_path):
    """order_suggest 空 + write_order false 时不得清空表上已有单号。"""
    ws = tmp_path / "ws"
    (ws / "02_我的表副本").mkdir(parents=True)
    flow_path = ws / "02_我的表副本" / "流转测.xlsx"
    _flow_xlsx(flow_path, [
        ("2026-07-22", "甲公司", 100, "EXISTING_SO", ""),
    ])
    plan = {
        "items": [{
            "ar": "AR1",
            "verdict": "write",
            "file": "流转测.xlsx",
            "sheet": "明细",
            "row_no": 2,
            "order_suggest": "",
            "updated_suggest": "是",
            "write_order": False,
            "write_updated": True,
            "matched_by": "三键",
            "hits": 1,
        }],
        "counts": {"write": 1},
    }
    plan_p = tmp_path / "plan.json"
    plan_p.write_text(json.dumps(plan), encoding="utf-8")
    rc = AF.main([
        "--plan", str(plan_p), "--workspace", str(ws),
        "--confirmed", "--in-place",
    ])
    assert rc == 0
    wb = openpyxl.load_workbook(str(flow_path), read_only=True, data_only=True)
    row = list(wb["明细"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
    wb.close()
    assert row[3] == "EXISTING_SO"  # 单号未抹
    assert row[4] == "是"


def test_apply_all_skips_flow_when_ledger_fails(tmp_path):
    """盈亏计划有冲突且无 force → 失败，流转文件哈希不变。"""
    ws = tmp_path / "ws"
    (ws / "02_我的表副本").mkdir(parents=True)
    (ws / "04_产出").mkdir(parents=True)
    flow_path = ws / "02_我的表副本" / "流转测.xlsx"
    _flow_xlsx(flow_path, [("2026-07-22", "甲", 100, "", "")])
    before = _sha(flow_path)

    led = ws / "02_我的表副本" / "盈亏.xlsx"
    wb = openpyxl.Workbook()
    w = wb.active
    w.title = "明细"
    w.append(["部门", "销售人员", "客户名称", "单号", "新智云单号", "应收金额",
              "计提金额", "回款明细", "是否结账（是/否）", "收款时间", "收款方式(支/汇/现)", "实收金额"])
    w.append(["部", "人", "客", "AB", "SO1", 100, 50, 50, "是", "2026-07-01", "汇", "SOD1"])
    wb.save(str(led))

    # conflict 计划（有 conflict 无 force）
    checked = {
        "write": [{
            "case_id": "AR1|SO1|SOD1", "ar": "AR1", "so": "SO1", "sod": "SOD1",
            "ledger_row_ref": 2,
            "five_cols": {"计提": 100.0, "回款明细": 100.0, "是否结账": "是",
                          "收款时间": "2026-07-08", "收款方式": "汇", "实收SOD": "SOD1"},
        }],
        "skip": [],
        "conflict": [{"case_id": "x"}],
    }
    checked_p = ws / "04_产出" / "写入计划_校验后.json"
    checked_p.write_text(json.dumps(checked), encoding="utf-8")
    flow_plan = {
        "items": [{
            "ar": "AR1", "verdict": "write", "file": "流转测.xlsx", "sheet": "明细",
            "row_no": 2, "order_suggest": "SO1", "updated_suggest": "是",
        }],
        "counts": {"write": 1},
    }
    fp = ws / "04_产出" / "流转写入计划_校验后.json"
    fp.write_text(json.dumps(flow_plan), encoding="utf-8")

    rc = AA.main([
        "--checked", str(checked_p),
        "--flow-plan", str(fp),
        "--ledger", str(led),
        "--workspace", str(ws),
        "--confirmed",
        "--in-place",
        "--flow-in-place",
    ])
    assert rc != 0
    assert _sha(flow_path) == before
