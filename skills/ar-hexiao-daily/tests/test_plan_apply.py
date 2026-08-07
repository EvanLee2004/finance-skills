# -*- coding: utf-8 -*-
"""plan → validate → execute：写入前校验 + 写副本 + 写后回读。"""
import datetime as dt
import json
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent))
from conftest import LEDGER_FULL  # noqa: E402

import validate_plan as V  # noqa: E402
import apply_to_copy as A  # noqa: E402

HDR = ["部门", "销售人员", "客户名称", "单号", "新智云单号", "应收金额",
       "计提金额", "回款明细", "是否结账（是/否）", "收款时间", "收款方式(支/汇/现)", "实收金额",
       "差异"]


def _ledger(tmp_path, rows):
    """造一张最小盈亏表：表头 + 若干行。rows=[(SO, SOD, 五列值 or None)]"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "明细"
    ws.append(HDR)
    for so, sod, five in rows:
        r = ["部", "人", "客", "AB", so, 100, None, None, None, None, None, sod, None]
        if five:
            r[6], r[7], r[8], r[9], r[10] = (
                five.get("计提"), five.get("回款明细"), five.get("是否结账"),
                five.get("收款时间"), five.get("收款方式"),
            )
        ws.append(r)
    p = tmp_path / "盈亏.xlsx"
    wb.save(str(p))
    return p


def _ledger_without_difference(tmp_path, rows):
    """兼容旧测试表：结构与 _ledger 相同，但没有“差异”列。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "明细"
    ws.append(HDR[:-1])
    for so, sod, five in rows:
        r = ["部", "人", "客", "AB", so, 100, None, None, None, None, None, sod]
        if five:
            r[6], r[7], r[8], r[9], r[10] = (
                five.get("计提"), five.get("回款明细"), five.get("是否结账"),
                five.get("收款时间"), five.get("收款方式"),
            )
        ws.append(r)
    p = tmp_path / "旧版盈亏_无差异列.xlsx"
    wb.save(str(p))
    return p


def _item(row, so="SO26010001", sod="SOD26010001", **five):
    f = {"计提": 100.0, "回款明细": 100.0, "是否结账": "是",
         "收款时间": "2026-07-08", "收款方式": "汇", "实收SOD": sod}
    f.update(five)
    return {"case_id": f"AR1|{so}", "ar": "AR1", "so": so, "sod": sod,
            "ledger_row_ref": row, "five_cols": f}


def _split_item(row=2, so="SO26010001", sod="SOD26010001"):
    item = _item(row, so=so, sod=sod, 计提=None, 回款明细=50.0, 是否结账="是")
    item["row_operation"] = {
        "type": "split_below",
        "source_receivable": 100.0,
        "paid_receivable": 40.0,
        "unpaid_receivable": 60.0,
        "baseline_receivable": 100.0,
        "paid_side_receivable_total": 40.0,
        "existing_received": 0.0,
        "current_received": 50.0,
        "cumulative_received": 50.0,
        "latest_delivery": 110.0,
        "inserted_five_cols": {
            "计提": None, "回款明细": None, "是否结账": "否",
            "收款时间": None, "收款方式": None, "实收SOD": sod,
        },
    }
    return item


# ---------- 校验 ----------

def test_empty_row_is_writable(tmp_path):
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", None)])
    rows = V.read_ledger_rows(led)
    assert V.check_one(_item(2), rows)["verdict"] == "write"


def test_default_jiezhang_no_is_still_writable(tmp_path):
    """她表未收款的行**预置「是否结账=否」**、回款列全空 → 这是"还没填"，必须可写。
    2026-07-24 真实 24 号数据实测 bug：validate 把预置的「否」当"已填过"，4 笔可填全被误判冲突、
    可写=0。修复=判"填没填过"只看回款证据列（计提/回款明细/收款时间/收款方式），是否结账不算证据。"""
    five = {"是否结账": "否"}  # 只有预置默认，回款四列全空
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", five)])
    rows = V.read_ledger_rows(led)
    # 计划要把它翻「否→是」+填钱 → 应判"可写"，不是冲突
    assert V.check_one(_item(2), rows)["verdict"] == "write"


def test_partial_row_keeps_jiezhang_no_is_writable(tmp_path):
    """部分回款：计划**保持结账「否」**、计提留空、只填回款；行本就预置「否」→ 仍可写。"""
    five = {"是否结账": "否"}
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", five)])
    rows = V.read_ledger_rows(led)
    res = V.check_one(_item(2, 计提=None, 是否结账="否", 回款明细=60.0), rows)
    assert res["verdict"] == "write"


def test_already_filled_same_is_skip(tmp_path):
    five = {"计提": 100.0, "回款明细": 100.0, "是否结账": "是",
            "收款时间": dt.date(2026, 7, 8), "收款方式": "汇"}
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", five)])
    rows = V.read_ledger_rows(led)
    assert V.check_one(_item(2), rows)["verdict"] == "skip"


def test_validate_treats_fully_settled_order_as_idempotent_even_if_old_plan_differs(tmp_path):
    five = {"计提": 100.0, "回款明细": 100.0, "是否结账": "是",
            "收款时间": dt.date(2026, 7, 8), "收款方式": "汇"}
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", five)])
    item = _item(2, 回款明细=90.0)

    checked = V.validate({"auto": [item]}, V.read_ledger_rows(led))

    assert checked["counts"] == {"write": 0, "skip": 1, "conflict": 0}
    assert "已结账" in checked["skip"][0]["_check"]["reason"]


def test_validate_does_not_skip_closed_row_when_open_split_row_remains(tmp_path):
    closed = {"计提": None, "回款明细": 60.0, "是否结账": "是",
              "收款时间": dt.date(2026, 7, 8), "收款方式": "汇"}
    opened = {"计提": None, "回款明细": None, "是否结账": "否",
              "收款时间": None, "收款方式": None}
    led = _ledger(tmp_path, [
        ("SO26010001", "SOD26010001", closed),
        ("SO26010001", "SOD26010001", opened),
    ])
    item = _item(2, 计提=None, 回款明细=50.0)

    checked = V.validate({"auto": [item]}, V.read_ledger_rows(led))

    assert checked["counts"]["skip"] == 0
    assert checked["counts"]["conflict"] == 1


def test_same_five_with_missing_business_difference_is_writable(tmp_path):
    five = {"计提": 110.0, "回款明细": 110.0, "是否结账": "是",
            "收款时间": dt.date(2026, 7, 8), "收款方式": "汇"}
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", five)])
    item = _item(2, 计提=110.0, 回款明细=110.0)
    item["derived_cols"] = {"差异": -10.0}
    assert V.check_one(item, V.read_ledger_rows(led))["verdict"] == "write"


def test_wrong_existing_business_difference_is_conflict(tmp_path):
    five = {"计提": 110.0, "回款明细": 110.0, "是否结账": "是",
            "收款时间": dt.date(2026, 7, 8), "收款方式": "汇"}
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", five)])
    wb = openpyxl.load_workbook(str(led))
    wb["明细"].cell(2, 13).value = -9.0
    wb.save(str(led))
    item = _item(2, 计提=110.0, 回款明细=110.0)
    item["derived_cols"] = {"差异": -10.0}
    res = V.check_one(item, V.read_ledger_rows(led))
    assert res["verdict"] == "conflict" and "差异" in res["reason"]


def test_correct_existing_business_difference_is_skip(tmp_path):
    five = {"计提": 110.0, "回款明细": 110.0, "是否结账": "是",
            "收款时间": dt.date(2026, 7, 8), "收款方式": "汇"}
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", five)])
    wb = openpyxl.load_workbook(str(led))
    wb["明细"].cell(2, 13).value = -10.0
    wb.save(str(led))
    item = _item(2, 计提=110.0, 回款明细=110.0)
    item["derived_cols"] = {"差异": -10.0}
    assert V.check_one(item, V.read_ledger_rows(led))["verdict"] == "skip"


def test_missing_difference_column_conflicts_only_when_plan_needs_it(tmp_path):
    led = _ledger_without_difference(
        tmp_path, [("SO26010001", "SOD26010001", None)]
    )
    rows = V.read_ledger_rows(led)
    assert V.check_one(_item(2), rows)["verdict"] == "write"

    item = _item(2, 计提=110.0, 回款明细=110.0)
    item["derived_cols"] = {"差异": -10.0}
    res = V.check_one(item, rows)
    assert res["verdict"] == "conflict"
    assert "没有“差异”列" in res["reason"]


def test_partial_conflicts_instead_of_overwriting_existing_difference(tmp_path):
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", None)])
    wb = openpyxl.load_workbook(str(led))
    wb["明细"].cell(2, 13).value = 9.9
    wb.save(str(led))
    res = V.check_one(_split_item(), V.read_ledger_rows(led))
    assert res["verdict"] == "conflict"
    assert "部分回款阶段" in res["reason"]


def test_already_filled_different_is_conflict(tmp_path):
    """她填的和我们算的不一样 → 绝不覆盖，交给人（跨月冲预收那条就是这么抓出来的）。"""
    five = {"计提": 100.0, "回款明细": 100.0, "是否结账": "是",
            "收款时间": dt.date(2026, 7, 8), "收款方式": "冲预收"}
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", five)])
    rows = V.read_ledger_rows(led)
    res = V.check_one(_item(2), rows)
    assert res["verdict"] == "conflict" and "收款方式" in res["reason"]


def test_reference_or_named_marker_cannot_bypass_existing_value_conflict(tmp_path):
    """即使输入里伪造旧版参考表标记，校验器也不能覆盖已有非空值。"""
    old = {
        "计提": 105.0,
        "回款明细": 105.0,
        "是否结账": "是",
        "收款时间": dt.date(2026, 7, 27),
        "收款方式": "冲预收",
    }
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", old)])
    item = _item(2)
    item["match_basis"] = "用户指定权威盈亏表"
    item["authoritative_reference"] = {"row": 18, "path": "上传表.xlsx"}
    item["authoritative_correction"] = {
        "scope": "SO",
        "so": "SO26010001",
        "reference_row": 18,
        "target_row": 2,
    }
    assert V.check_one(item, V.read_ledger_rows(led))["verdict"] == "conflict"


def test_row_shifted_is_conflict(tmp_path):
    """她插过行 → 行号指到别人家去了，必须拦住。"""
    led = _ledger(tmp_path, [("SO_OTHER", "SOD_OTHER", None)])
    rows = V.read_ledger_rows(led)
    res = V.check_one(_item(2), rows)
    assert res["verdict"] == "conflict" and "不是计划里的" in res["reason"]


def test_bad_value_is_conflict(tmp_path):
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", None)])
    rows = V.read_ledger_rows(led)
    assert V.check_one(_item(2, 是否结账="也许"), rows)["verdict"] == "conflict"
    assert V.check_one(_item(2, 收款方式="随便"), rows)["verdict"] == "conflict"
    assert V.check_one(_item(2, 计提="一百块"), rows)["verdict"] == "conflict"


def test_missing_row_ref_is_conflict(tmp_path):
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", None)])
    rows = V.read_ledger_rows(led)
    it = _item(2)
    it["ledger_row_ref"] = None
    assert V.check_one(it, rows)["verdict"] == "conflict"


def test_two_plans_same_row_conflict(tmp_path):
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", None)])
    rows = V.read_ledger_rows(led)
    plan = {"auto": [_item(2), _item(2, so="SO26010001")]}
    res = V.validate(plan, rows)
    assert res["counts"]["write"] == 1 and res["counts"]["conflict"] == 1


def _split_chain_items(final_unpaid=None, row=2):
    first = _item(row, so="SO_SPLIT", sod="SOD_SPLIT", 计提=None, 回款明细=40.0)
    second = _item(row, so="SO_SPLIT", sod="SOD_SPLIT", 计提=100.0, 回款明细=60.0)
    first["ar"], first["case_id"] = "AR1", "AR1|SO_SPLIT|SOD_SPLIT"
    second["ar"], second["case_id"] = "AR2", "AR2|SO_SPLIT|SOD_SPLIT"
    steps = [
        {
            "index": 0, "case_id": first["case_id"], "ar": "AR1", "so": "SO_SPLIT", "sod": "SOD_SPLIT",
            "writeoff_sequence_key": ["2026-07-22", "HX1", "1", "AR1", "SO_SPLIT"],
            "current_received": 40.0, "cumulative_received": 40.0,
            "receivable": 40.0, "remaining_after": 60.0, "settled": False,
            "five_cols": dict(first["five_cols"]), "derived_cols": {},
        },
        {
            "index": 1, "case_id": second["case_id"], "ar": "AR2", "so": "SO_SPLIT", "sod": "SOD_SPLIT",
            "writeoff_sequence_key": ["2026-07-22", "HX2", "2", "AR2", "SO_SPLIT"],
            "current_received": 60.0, "cumulative_received": 100.0,
            "receivable": 60.0, "remaining_after": 0.0, "settled": True,
            "five_cols": dict(second["five_cols"]), "derived_cols": {},
        },
    ]
    op = {
        "type": "split_payment_chain", "source_receivable": 100.0,
        "initial_cumulative": 0.0, "latest_delivery": 100.0,
        "steps": steps, "final_unpaid": final_unpaid,
    }
    for index, item in enumerate((first, second)):
        item["row_operation"] = op
        item["split_chain_group_id"] = f"split-payment-chain|{row}|SO_SPLIT|SOD_SPLIT"
        item["split_chain_index"] = index
        item["split_chain_count"] = 2
    return [first, second]


def test_same_row_split_chain_validates_and_writes_each_parent_separately(tmp_path):
    led = _ledger(tmp_path, [("SO_SPLIT", "SOD_SPLIT", None)])
    rows = V.read_ledger_rows(led)
    checked = V.validate({"auto": _split_chain_items()}, rows)
    assert checked["counts"] == {"write": 2, "skip": 0, "conflict": 0}

    out = tmp_path / "分笔逐行.xlsx"
    A.write_plan(led, out, checked["write"])
    assert A.verify_written(out, checked["write"]) == []
    ws = openpyxl.load_workbook(str(out), data_only=True)["明细"]
    assert ws.cell(2, 8).value == 40.0
    assert ws.cell(3, 8).value == 60.0
    assert ws.cell(2, 7).value is None
    assert ws.cell(3, 7).value == 100.0

    rerun = V.validate({"auto": _split_chain_items()}, V.read_ledger_rows(out))
    assert rerun["counts"] == {"write": 0, "skip": 2, "conflict": 0}


def test_legacy_aggregate_row_can_migrate_to_split_chain(tmp_path):
    """旧版把两笔父回款合在一行时，只在聚合基线完全未变的前提下允许拆开。"""
    aggregate = {
        "计提": 100.0, "回款明细": 100.0, "是否结账": "是",
        "收款时间": "2026-07-08", "收款方式": "汇",
    }
    ledger = _ledger(tmp_path, [("SO_SPLIT", "SOD_SPLIT", aggregate)])
    items = _split_chain_items()
    for item in items:
        item["row_operation"]["source_five_cols"] = {
            **aggregate, "实收SOD": "SOD_SPLIT",
        }
        item["row_operation"]["source_derived_cols"] = {"差异": None}

    checked = V.validate({"auto": items}, V.read_ledger_rows(ledger))

    assert checked["counts"] == {"write": 2, "skip": 0, "conflict": 0}
    out = tmp_path / "旧聚合迁移为分笔链.xlsx"
    A.write_plan(ledger, out, checked["write"])
    assert A.verify_written(out, checked["write"]) == []
    ws = openpyxl.load_workbook(str(out), data_only=True)["明细"]
    assert [ws.cell(row, 8).value for row in (2, 3)] == [40.0, 60.0]


def test_one_yuan_tail_aggregate_is_skip_and_snapshot_change_is_conflict():
    five = {
        "计提": 211464.0, "回款明细": 211464.0, "是否结账": "是",
        "收款时间": "2026-07-15", "收款方式": "汇", "实收SOD": "SOD26020257",
    }
    operation = {
        "type": "preserve_aggregate_tail_tolerance",
        "source_receivable": 211464.0,
        "initial_cumulative": 0.0,
        "latest_delivery": 211464.0,
        "final_cumulative": 211464.0,
        "parent_amounts": [0.12, 211463.88],
        "tolerated_tail_amount": 0.12,
        "tolerance": 1.0,
        "source_five_cols": dict(five),
        "source_derived_cols": {"差异": None},
    }
    item = {
        "so": "SO26020257", "sod": "SOD26020257", "ledger_row_ref": 2,
        "five_cols": dict(five), "row_operation": operation,
    }
    row = {
        "SO": "SO26020257", "SOD": "SOD26020257", "应收金额": 211464.0,
        "计提": 211464.0, "回款明细": 211464.0, "差异": None,
        "_差异列存在": True, "是否结账": "是", "收款时间": "2026-07-15",
        "收款方式": "汇",
    }

    checked = V.validate({"auto": [item]}, {2: row})
    assert checked["counts"] == {"write": 0, "skip": 1, "conflict": 0}
    assert "不超过1元" in checked["skip"][0]["_check"]["reason"]

    changed = {2: {**row, "回款明细": 211463.0}}
    rejected = V.validate({"auto": [item]}, changed)
    assert rejected["counts"] == {"write": 0, "skip": 0, "conflict": 1}
    assert "被改动" in rejected["conflict"][0]["_check"]["reason"]


def test_new_one_yuan_tail_aggregate_writes_one_row_and_skips_absorbed_parent(tmp_path):
    led = _ledger(tmp_path, [("SO_TAIL", "SOD_TAIL", None)])
    target_five = {
        "计提": 100.0, "回款明细": 100.0, "是否结账": "是",
        "收款时间": "2026-07-16", "收款方式": "汇", "实收SOD": "SOD_TAIL",
    }
    operation = {
        "type": "settlement_tail_aggregate",
        "source_receivable": 100.0,
        "initial_cumulative": 0.0,
        "latest_delivery": 100.0,
        "final_cumulative": 100.0,
        "target_case_id": "AR_MAIN|SO_TAIL|SOD_TAIL",
        "target_five_cols": dict(target_five),
        "target_derived_cols": {},
        "source_five_cols": {
            "计提": None, "回款明细": None, "是否结账": None,
            "收款时间": None, "收款方式": None, "实收SOD": "SOD_TAIL",
        },
        "source_derived_cols": {"差异": None},
        "tail_tolerance_audit": {
            "tolerance": 1.0,
            "absorbed_total": 0.12,
            "absorbed_payments": [{
                "case_id": "AR_SMALL|SO_TAIL|SOD_TAIL",
                "ar": "AR_SMALL", "amount": 0.12,
                "writeoff_sequence_key": ["2026-07-16", "HX1"],
            }],
            "original_parent_amounts": [0.12, 99.88],
            "target_case_id": "AR_MAIN|SO_TAIL|SOD_TAIL",
        },
    }
    target = {
        "case_id": "AR_MAIN|SO_TAIL|SOD_TAIL", "ar": "AR_MAIN",
        "so": "SO_TAIL", "sod": "SOD_TAIL", "ledger_row_ref": 2,
        "five_cols": dict(target_five), "row_operation": operation,
    }
    absorbed = {
        "case_id": "AR_SMALL|SO_TAIL|SOD_TAIL", "ar": "AR_SMALL",
        "so": "SO_TAIL", "sod": "SOD_TAIL", "ledger_row_ref": 2,
        "five_cols": {},
        "tail_tolerance_absorbed": {
            "case_id": "AR_SMALL|SO_TAIL|SOD_TAIL", "ar": "AR_SMALL",
            "amount": 0.12, "target_case_id": "AR_MAIN|SO_TAIL|SOD_TAIL",
            "absorbed_total": 0.12, "tolerance": 1.0,
        },
    }

    checked = V.validate({"auto": [absorbed, target]}, V.read_ledger_rows(led))

    assert checked["counts"] == {"write": 1, "skip": 1, "conflict": 0}
    out = tmp_path / "尾差合并.xlsx"
    changes = A.write_plan(led, out, checked["write"])
    assert A.verify_written(out, checked["write"]) == []
    assert changes[0]["操作"] == "结清尾差合并写入"
    ws = openpyxl.load_workbook(str(out), data_only=True)["明细"]
    assert ws.max_row == 2
    assert ws.cell(2, 8).value == 100.0


def test_legacy_aggregate_row_changed_after_classify_is_rejected(tmp_path):
    aggregate = {
        "计提": 100.0, "回款明细": 100.0, "是否结账": "是",
        "收款时间": "2026-07-08", "收款方式": "汇",
    }
    ledger = _ledger(tmp_path, [("SO_SPLIT", "SOD_SPLIT", aggregate)])
    items = _split_chain_items()
    for item in items:
        item["row_operation"]["source_five_cols"] = {
            **aggregate, "实收SOD": "SOD_SPLIT",
        }
        item["row_operation"]["source_derived_cols"] = {"差异": None}
    changed = openpyxl.load_workbook(str(ledger))
    changed["明细"].cell(2, 8).value = 99.0
    changed.save(str(ledger))

    checked = V.validate({"auto": items}, V.read_ledger_rows(ledger))

    assert checked["counts"] == {"write": 0, "skip": 0, "conflict": 2}

def test_split_chain_rerun_finds_whole_chain_after_other_insert_shifts_source_row(tmp_path):
    led = _ledger(tmp_path, [
        ("SO_OTHER", "SOD_OTHER", None),
        ("SO_SPLIT", "SOD_SPLIT", None),
    ])
    other = _split_item(row=2, so="SO_OTHER", sod="SOD_OTHER")
    chain = _split_chain_items(row=3)
    original_plan = {"auto": [other, *chain]}
    checked = V.validate(original_plan, V.read_ledger_rows(led))
    assert checked["counts"] == {"write": 3, "skip": 0, "conflict": 0}

    out = tmp_path / "前序插行后的分笔链.xlsx"
    A.write_plan(led, out, checked["write"])
    rerun = V.validate(original_plan, V.read_ledger_rows(out))
    assert rerun["counts"] == {"write": 0, "skip": 3, "conflict": 0}


def test_split_chain_still_unpaid_keeps_each_parent_and_adds_one_open_row(tmp_path):
    led = _ledger(tmp_path, [("SO_SPLIT", "SOD_SPLIT", None)])
    items = _split_chain_items()
    op = items[0]["row_operation"]
    op["steps"][0].update({
        "current_received": 30.0, "cumulative_received": 30.0,
        "receivable": 30.0, "remaining_after": 70.0,
    })
    op["steps"][0]["five_cols"]["回款明细"] = 30.0
    op["steps"][1].update({
        "current_received": 20.0, "cumulative_received": 50.0,
        "receivable": 20.0, "remaining_after": 50.0, "settled": False,
    })
    op["steps"][1]["five_cols"].update({"计提": None, "回款明细": 20.0})
    op["final_unpaid"] = {
        "receivable": 50.0,
        "five_cols": {
            "计提": None, "回款明细": None, "是否结账": "否",
            "收款时间": None, "收款方式": None, "实收SOD": "SOD_SPLIT",
        },
    }
    items[0]["five_cols"] = dict(op["steps"][0]["five_cols"])
    items[1]["five_cols"] = dict(op["steps"][1]["five_cols"])

    checked = V.validate({"auto": items}, V.read_ledger_rows(led))
    assert checked["counts"] == {"write": 2, "skip": 0, "conflict": 0}
    out = tmp_path / "分笔未结清.xlsx"
    A.write_plan(led, out, checked["write"])
    assert A.verify_written(out, checked["write"]) == []
    ws = openpyxl.load_workbook(str(out), data_only=True)["明细"]
    assert [ws.cell(row, 8).value for row in (2, 3, 4)] == [30.0, 20.0, None]
    assert [ws.cell(row, 6).value for row in (2, 3, 4)] == [30.0, 20.0, 50.0]
    assert [ws.cell(row, 7).value for row in (2, 3, 4)] == [None, None, None]
    assert ws.cell(4, 9).value == "否"


def test_split_chain_plan_tampering_is_rejected(tmp_path):
    led = _ledger(tmp_path, [("SO_SPLIT", "SOD_SPLIT", None)])
    items = _split_chain_items()
    items[0]["row_operation"]["steps"][1]["cumulative_received"] = 99.0
    checked = V.validate({"auto": items}, V.read_ledger_rows(led))
    assert checked["counts"]["conflict"] == 2


def test_split_plan_validates_receivable_invariants(tmp_path):
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", None)])
    rows = V.read_ledger_rows(led)
    assert V.check_one(_split_item(), rows)["verdict"] == "write"
    bad = _split_item()
    bad["row_operation"]["unpaid_receivable"] = 59.0
    assert V.check_one(bad, rows)["verdict"] == "conflict"


# ---------- 写入 ----------

def test_apply_writes_and_verifies(tmp_path):
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", None)])
    out = tmp_path / "已回填.xlsx"
    changes = A.write_plan(led, out, [_item(2)])
    assert len(changes) == 1
    assert A.verify_written(out, [_item(2)]) == []
    ws = openpyxl.load_workbook(str(out))["明细"]
    assert ws.cell(2, 7).value == 100.0          # 计提
    assert ws.cell(2, 9).value == "是"            # 是否结账
    assert ws.cell(2, 12).value == "SOD26010001"  # 实收金额列存 SOD


def test_apply_writes_difference_as_formula_with_cached_value(tmp_path):
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", None)])
    item = _item(2, 计提=110.0, 回款明细=110.0)
    item["derived_cols"] = {"差异": -10.0}
    out = tmp_path / "差异公式.xlsx"
    A.write_plan(led, out, [item])
    assert A.verify_written(out, [item]) == []
    formula_wb = openpyxl.load_workbook(str(out), data_only=False)
    assert formula_wb["明细"].cell(2, 13).value == "=F2-G2"
    formula_wb.close()
    value_wb = openpyxl.load_workbook(str(out), data_only=True)
    assert value_wb["明细"].cell(2, 13).value == -10.0
    value_wb.close()


def test_difference_formula_uses_all_existing_split_rows(tmp_path):
    open_row = {"是否结账": "否"}
    paid_row = {
        "回款明细": 80.0, "是否结账": "是",
        "收款时间": dt.date(2026, 7, 8), "收款方式": "汇",
    }
    ledger = _ledger(tmp_path, [
        ("SO_SPLIT_DIFF", "SOD_SPLIT_DIFF", open_row),
        ("SO_SPLIT_DIFF", "SOD_SPLIT_DIFF", paid_row),
    ])
    wb = openpyxl.load_workbook(str(ledger))
    wb["明细"].cell(2, 6).value = 20.0
    wb["明细"].cell(3, 6).value = 80.0
    wb.save(str(ledger))
    item = _item(
        3, so="SO_SPLIT_DIFF", sod="SOD_SPLIT_DIFF",
        计提=80.0, 回款明细=80.0,
    )
    item["derived_cols"] = {"差异": 20.0}
    out = tmp_path / "拆分行合计差异公式.xlsx"

    assert V.check_one(item, V.read_ledger_rows(ledger))["verdict"] == "write"
    A.write_plan(ledger, out, [item])

    assert A.verify_written(out, [item]) == []
    formula_wb = openpyxl.load_workbook(str(out), data_only=False)
    assert formula_wb["明细"].cell(3, 13).value == "=SUM(F2,F3)-G3"
    formula_wb.close()
    value_wb = openpyxl.load_workbook(str(out), data_only=True)
    assert value_wb["明细"].cell(3, 13).value == 20.0
    value_wb.close()


def test_large_difference_formula_cache_keeps_exact_cents(tmp_path):
    """公式缓存不得用默认 :g 的六位有效数字把 15479.75 写成 15479.8。"""
    led = _ledger(tmp_path, [("SO26030487", "SOD26030654", None)])
    wb = openpyxl.load_workbook(str(led))
    wb["明细"].cell(2, 6).value = 17979.75
    wb.save(str(led))
    wb.close()

    item = _item(
        2,
        so="SO26030487",
        sod="SOD26030654",
        计提=2500.0,
        回款明细=2500.0,
    )
    item["derived_cols"] = {"差异": 15479.75}
    out = tmp_path / "大金额差异公式缓存.xlsx"

    A.write_plan(led, out, [item])

    assert A.verify_written(out, [item]) == []
    formula_wb = openpyxl.load_workbook(str(out), data_only=False)
    assert formula_wb["明细"].cell(2, 13).value == "=F2-G2"
    formula_wb.close()
    value_wb = openpyxl.load_workbook(str(out), data_only=True)
    assert value_wb["明细"].cell(2, 13).value == 15479.75
    value_wb.close()


def test_difference_formula_uses_final_row_after_earlier_split(tmp_path):
    led = _ledger(tmp_path, [
        ("SO_SPLIT", "SOD_SPLIT", None),
        ("SO_DIFF", "SOD_DIFF", None),
    ])
    split = _split_item(row=2, so="SO_SPLIT", sod="SOD_SPLIT")
    diff = _item(3, so="SO_DIFF", sod="SOD_DIFF", 计提=110.0, 回款明细=110.0)
    diff["derived_cols"] = {"差异": -10.0}
    out = tmp_path / "插行后差异公式.xlsx"
    A.write_plan(led, out, [split, diff])
    assert A.verify_written(out, [split, diff]) == []

    formula_wb = openpyxl.load_workbook(str(out), data_only=False)
    assert formula_wb["明细"].cell(4, 13).value == "=F4-G4"
    formula_wb.close()
    value_wb = openpyxl.load_workbook(str(out), data_only=True)
    assert value_wb["明细"].cell(4, 13).value == -10.0
    value_wb.close()


def test_apply_never_touches_source(tmp_path):
    """写的是新文件，她给的副本必须一个字节都不动。"""
    import hashlib
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", None)])
    before = hashlib.sha256(led.read_bytes()).hexdigest()
    A.write_plan(led, tmp_path / "out.xlsx", [_item(2)])
    assert hashlib.sha256(led.read_bytes()).hexdigest() == before


def test_in_place_writes_into_copy_and_backs_up(tmp_path):
    """就地模式（明妹要的）：直接写进她那份副本，并留一份写前备份、不留临时文件。"""
    import hashlib

    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", None)])
    before = hashlib.sha256(led.read_bytes()).hexdigest()
    checked = tmp_path / "checked.json"
    checked.write_text(
        json.dumps({"write": [_item(2)], "skip": [], "conflict": []},
                   ensure_ascii=False, default=str),
        encoding="utf-8",
    )
    rc = A.main(["--checked", str(checked), "--ledger", str(led),
                 "--report", str(tmp_path / "r.xlsx"), "--in-place", "--confirmed"])
    assert rc == 0
    # 副本被就地改了
    ws = openpyxl.load_workbook(str(led))["明细"]
    assert ws.cell(2, 7).value == 100.0            # 计提
    assert ws.cell(2, 9).value == "是"             # 是否结账
    # 写前备份存在，且等于写前内容（真出事能还原）
    backups = list((led.parent / "备份").glob("盈亏_备份_*.xlsx"))
    assert len(backups) == 1
    assert hashlib.sha256(backups[0].read_bytes()).hexdigest() == before
    # 成功后不留临时文件
    assert not list(led.parent.glob(".盈亏_写入中_*"))


def test_partial_leaves_jiti_empty(tmp_path):
    """部分核销：计提留空就是留空，不许写成 0（写 0 会被当成已计提）。"""
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", None)])
    out = tmp_path / "out.xlsx"
    A.write_plan(led, out, [_item(2, 计提=None, 是否结账="否")])
    ws = openpyxl.load_workbook(str(out))["明细"]
    assert ws.cell(2, 7).value is None
    assert ws.cell(2, 9).value == "否"


def test_apply_partial_inserts_unpaid_row_below_and_preserves_other_fields(tmp_path):
    led = _ledger(tmp_path, [
        ("SO26010001", "SOD26010001", None),
        ("SO_OTHER", "SOD_OTHER", None),
    ])
    out = tmp_path / "split.xlsx"
    item = _split_item()
    A.write_plan(led, out, [item])
    assert A.verify_written(out, [item]) == []
    ws = openpyxl.load_workbook(str(out), data_only=False)["明细"]
    assert ws.max_row == 4
    assert ws.cell(2, 6).value == 40.0
    assert ws.cell(2, 7).value is None
    assert ws.cell(2, 8).value == 50.0
    assert ws.cell(2, 9).value == "是"


    assert ws.cell(2, 13).value is None
    assert ws.cell(3, 1).value == ws.cell(2, 1).value
    assert ws.cell(3, 5).value == "SO26010001"
    assert ws.cell(3, 6).value == 60.0
    assert ws.cell(3, 7).value is None
    assert ws.cell(3, 8).value is None
    assert ws.cell(3, 9).value == "否"
    assert ws.cell(3, 10).value is None
    assert ws.cell(3, 11).value is None
    assert ws.cell(3, 12).value == "SOD26010001"
    assert ws.cell(3, 13).value is None
    assert ws.cell(4, 5).value == "SO_OTHER"


def test_same_so_multi_sod_aggregate_validates_writes_once_and_is_idempotent(tmp_path):
    ledger = _ledger(tmp_path, [("SO_MULTI", "SODA", None)])
    wb = openpyxl.load_workbook(str(ledger))
    wb["明细"].cell(2, 6).value = 40.0
    wb.save(str(ledger))
    target_five = {
        "计提": 100.0, "回款明细": 100.0, "是否结账": "是",
        "收款时间": "2026-08-05", "收款方式": "汇", "实收SOD": "SODA、SODB",
    }
    operation = {
        "type": "same_so_multi_sod_aggregate",
        "source_receivable": 40.0,
        "source_five_cols": {
            "计提": None, "回款明细": None, "是否结账": None,
            "收款时间": None, "收款方式": None, "实收SOD": "SODA",
        },
        "source_derived_cols": {"差异": None},
        "so": "SO_MULTI", "ar": "AR1", "so_delivery": 100.0,
        "current_received": 100.0, "combined_sod": "SODA、SODB",
        "member_sods": ["SODA", "SODB"],
        "member_case_ids": ["AR1|SO_MULTI|SODA", "AR1|SO_MULTI|SODB"],
        "member_amounts": [40.0, 60.0], "member_deliveries": [40.0, 60.0],
        "writeoff_sequence_key": ["2026-08-06", "HX1", "RID1", "AR1", "SO_MULTI"],
        "target_case_id": "AR1|SO_MULTI|SODA",
        "target_five_cols": dict(target_five), "target_derived_cols": {},
    }
    target = {
        "case_id": "AR1|SO_MULTI|SODA", "ar": "AR1", "so": "SO_MULTI", "sod": "SODA",
        "ledger_row_ref": 2, "five_cols": dict(target_five), "derived_cols": {},
        "row_operation": operation,
    }
    absorbed = {
        "case_id": "AR1|SO_MULTI|SODB", "ar": "AR1", "so": "SO_MULTI", "sod": "SODB",
        "ledger_row_ref": 2, "five_cols": {}, "derived_cols": {},
        "same_so_multi_sod_absorbed": {
            "target_case_id": target["case_id"], "group_id": "same-so-multi-sod|2|SO_MULTI|AR1",
            "member_sods": ["SODA", "SODB"], "so_delivery": 100.0,
        },
    }
    plan = {"auto": [target, absorbed]}

    checked = V.validate(plan, V.read_ledger_rows(ledger))
    assert checked["counts"] == {"write": 1, "skip": 1, "conflict": 0}

    out = tmp_path / "同SO多SOD合并.xlsx"
    A.write_plan(ledger, out, checked["write"])
    assert A.verify_written(out, checked["write"]) == []
    row = V.read_ledger_rows(out)[2]
    assert row["应收金额"] == 100.0
    assert row["回款明细"] == 100.0
    assert row["SOD"] == "SODA、SODB"

    rerun = V.validate(plan, V.read_ledger_rows(out))
    assert rerun["counts"] == {"write": 0, "skip": 2, "conflict": 0}


def test_old_ledger_without_difference_column_still_writes_normal_payment(tmp_path):
    led = _ledger_without_difference(
        tmp_path, [("SO26010001", "SOD26010001", None)]
    )
    out = tmp_path / "旧版普通回款.xlsx"
    item = _item(2)
    A.write_plan(led, out, [item])
    assert A.verify_written(out, [item]) == []
    ws = openpyxl.load_workbook(str(out), data_only=True)["明细"]
    assert ws.cell(2, 7).value == 100.0
    assert ws.cell(2, 8).value == 100.0


def test_change_report_includes_business_difference_value_and_formula(tmp_path):
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", None)])
    item = _item(2, 计提=110.0, 回款明细=110.0)
    item["derived_cols"] = {"差异": -10.0}
    out = tmp_path / "有差异.xlsx"
    changes = A.write_plan(led, out, [item])
    report = tmp_path / "变更清单.xlsx"
    A.write_change_report(changes, report)
    ws = openpyxl.load_workbook(str(report), data_only=False)["变更清单"]
    headers = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]
    values = dict(zip(headers, row))
    assert values["改后_差异"] == "-10.00"
    assert values["公式_差异"] == "=F2-G2"


def test_order_difference_report_matches_written_order(tmp_path):
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", None)])
    item = _item(2)
    out = tmp_path / "写后盈亏.xlsx"
    A.write_plan(led, out, [item])

    result = A.build_order_difference(
        [item], out, hexiao_date="2026-07-31"
    )
    assert result["written_order_count"] == 1
    assert result["comparison_object_count"] == 1
    assert result["matched_count"] == 1
    assert result["difference_count"] == 0
    assert result["order_rows"][0]["对比结果"] == "一致"

    report = tmp_path / "订单写入差异_20260731.xlsx"
    A.write_order_difference_report(result, report)
    wb = openpyxl.load_workbook(str(report), data_only=True)
    assert wb.sheetnames == ["汇总", "订单对比", "字段差异"]
    summary = dict(wb["汇总"].iter_rows(values_only=True))
    assert summary["本次写入订单数"] == 1
    assert summary["字段差异数"] == 0
    assert summary["不参与差异判定"] == "收款方式、收款时间"
    assert wb["字段差异"]["A2"].value == "无差异"


def test_order_difference_report_lists_actual_field_difference(tmp_path):
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", None)])
    item = _item(2)
    out = tmp_path / "写后被改动.xlsx"
    A.write_plan(led, out, [item])
    wb = openpyxl.load_workbook(str(out))
    wb["明细"].cell(2, 8).value = 99
    wb.save(str(out))

    result = A.build_order_difference(
        [item], out, hexiao_date="2026-07-31"
    )
    assert result["matched_count"] == 0
    assert result["difference_count"] == 1
    assert result["field_differences"][0]["差异字段"] == "回款明细"
    assert result["field_differences"][0]["计划写入值"] == 100.0
    assert result["field_differences"][0]["上传表写后值"] == 99


def test_order_difference_ignores_payment_time_and_method(tmp_path):
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", None)])
    item = _item(2)
    out = tmp_path / "写后收款信息不同.xlsx"
    A.write_plan(led, out, [item])
    wb = openpyxl.load_workbook(str(out))
    ws = wb["明细"]
    ws.cell(2, 10).value = "2026-08-01"
    ws.cell(2, 11).value = "冲预收"
    wb.save(str(out))

    result = A.build_order_difference(
        [item], out, hexiao_date="2026-07-31"
    )

    assert result["matched_count"] == 1
    assert result["difference_count"] == 0
    assert result["field_differences"] == []
    assert result["ignored_difference_fields"] == ["收款方式", "收款时间"]
    assert result["order_rows"][0]["对比结果"] == "一致"


def test_order_difference_includes_split_inserted_row(tmp_path):
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", None)])
    item = _split_item()
    out = tmp_path / "部分回款写后.xlsx"
    A.write_plan(led, out, [item])

    result = A.build_order_difference(
        [item], out, hexiao_date="2026-07-31"
    )
    assert result["written_order_count"] == 1
    assert result["comparison_object_count"] == 2
    assert result["matched_count"] == 2
    assert result["difference_count"] == 0
    assert {row["对比对象"] for row in result["order_rows"]} == {
        "写入订单",
        "拆分新增未回款行",
    }


def test_verify_catches_wrong_write(tmp_path):
    """回读比对必须真能发现写错——否则这道保险等于没有。"""
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", None)])
    out = tmp_path / "out.xlsx"
    A.write_plan(led, out, [_item(2)])
    wb = openpyxl.load_workbook(str(out))
    wb["明细"].cell(2, 7).value = 999          # 人为篡改
    wb.save(str(out))
    assert A.verify_written(out, [_item(2)]) != []


def test_verify_still_checks_payment_time_and_method(tmp_path):
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", None)])
    out = tmp_path / "收款信息被改动.xlsx"
    item = _item(2)
    A.write_plan(led, out, [item])
    wb = openpyxl.load_workbook(str(out))
    ws = wb["明细"]
    ws.cell(2, 10).value = "2026-08-01"
    ws.cell(2, 11).value = "冲预收"
    wb.save(str(out))

    problems = A.verify_written(out, [item])

    assert any("收款时间" in problem for problem in problems)
    assert any("收款方式" in problem for problem in problems)


def test_main_writes_after_precheck_without_confirmed(tmp_path):
    """日清和写前校验通过后直接写工作副本，旧参数仅兼容。"""
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", None)])
    checked = tmp_path / "checked.json"
    checked.write_text(
        json.dumps({"write": [_item(2)], "skip": [], "conflict": []},
                   ensure_ascii=False, default=str),
        encoding="utf-8",
    )
    rc = A.main(["--checked", str(checked), "--ledger", str(led),
                 "--out", str(tmp_path / "o.xlsx"), "--report", str(tmp_path / "r.xlsx")])
    assert rc == 0
    assert (tmp_path / "o.xlsx").exists()


def test_main_refuses_when_conflicts(tmp_path):
    """有冲突没处理就想写 → 必须拒绝（除非显式 --force）。"""
    led = _ledger(tmp_path, [("SO26010001", "SOD26010001", None)])
    checked = tmp_path / "checked.json"
    checked.write_text(json.dumps({"write": [_item(2)], "skip": [], "conflict": [_item(3)]},
                                  ensure_ascii=False, default=str), encoding="utf-8")
    rc = A.main(["--checked", str(checked), "--ledger", str(led),
                 "--out", str(tmp_path / "o.xlsx"), "--report", str(tmp_path / "r.xlsx"),
                 "--confirmed"])
    assert rc == 2


def test_main_creates_order_difference_report_named_by_hexiao_date(tmp_path):
    workspace = tmp_path / "工作区"
    ledger_dir = workspace / "02_我的表副本"
    output_dir = workspace / "04_产出"
    ledger_dir.mkdir(parents=True)
    output_dir.mkdir(parents=True)
    led = _ledger(ledger_dir, [("SO26010001", "SOD26010001", None)])
    checked = output_dir / "写入计划_校验后.json"
    checked.write_text(
        json.dumps(
            {
                "hexiao_date": "2026-07-31",
                "write": [_item(2)],
                "skip": [],
                "conflict": [],
            },
            ensure_ascii=False,
            default=str,
        ),
        encoding="utf-8",
    )
    out = output_dir / "写后盈亏.xlsx"
    rc = A.main(
        [
            "--checked", str(checked),
            "--ledger", str(led),
            "--out", str(out),
            "--confirmed",
        ]
    )
    assert rc == 0
    difference_report = output_dir / "订单写入差异_20260731.xlsx"
    assert difference_report.is_file()
    wb = openpyxl.load_workbook(str(difference_report), data_only=True)
    summary = dict(wb["汇总"].iter_rows(values_only=True))
    assert summary["字段差异数"] == 0


@pytest.mark.skipif(not LEDGER_FULL.is_file(), reason="无真实全年盈亏表")
def test_real_ledger_structure_survives_write(tmp_path):
    """写她的真表副本：透视表/外链/公式必须全须全尾。"""
    import zipfile
    out = tmp_path / "已回填.xlsx"
    rows = V.read_ledger_rows(LEDGER_FULL)
    target = next(r for r, v in rows.items() if v["SO"].startswith("SO"))
    it = _item(target, so=rows[target]["SO"], sod=rows[target]["SOD"] or "SOD_X")
    A.write_plan(LEDGER_FULL, out, [it])

    def n(p, key):
        with zipfile.ZipFile(p) as z:
            return sum(1 for x in z.namelist() if key.lower() in x.lower())
    for key in ("pivotTable", "pivotCache", "externalLink"):
        assert n(out, key) == n(LEDGER_FULL, key), f"{key} 写完少了"
    assert A.verify_written(out, [it]) == []
