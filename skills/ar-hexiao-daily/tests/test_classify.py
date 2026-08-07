# -*- coding: utf-8 -*-
"""
步骤6 判定 v2（单入口 · SOD 级）：展开、E 码、覆盖率硬校验、SOD 子集与整段对齐。
"""
import datetime as dt

import pytest

import classify_hexiao as C
from conftest import GOLD_DIR


# ── 小工具 ────────────────────────────────────────────────
def _pay(ar="AR_T", amount=100.0, orders=None, writeoffs=None, **kw):
    p = {
        "ar": ar,
        "hexiao_date": dt.date(2026, 7, 22),
        "arrival_date": dt.date(2026, 7, 21),
        "amount_orig": amount,
        "amount_local": amount,
        "fee": 0.0,
        "currency": "人民币CNY",
        "huikuan_type": "",
        "status": "手动核销",
        "customer": "测试客户甲",
        "orders": orders if orders is not None else [{"so": "SO26010001", "deliver": amount}],
        "writeoffs": writeoffs or {},
        "sod_lines": {},
    }
    p.update(kw)
    return p


def _led(rows: dict):
    """rows = {行号: {"so","sod","yingshou", …}} → 合成索引。"""
    so, sod = {}, {}
    for r, snap in rows.items():
        if snap.get("so"):
            so.setdefault(snap["so"], []).append(r)
        if snap.get("sod"):
            sod.setdefault(snap["sod"], []).append(r)
    return C.LedgerIndex(synthetic={"so": so, "sod": sod, "rows": rows})


# ══════════════════════════════════════════════════════════
# 一、subset_sum_unique
# ══════════════════════════════════════════════════════════
def test_subset_unique_hit():
    lines = [{"sod": "A", "deliver": 10.0}, {"sod": "B", "deliver": 25.5}, {"sod": "C", "deliver": 4.5}]
    got = C.subset_sum_unique(lines, 30.0)  # 25.5+4.5
    assert sorted(x["sod"] for x in got) == ["B", "C"]


def test_subset_multiple_solutions_returns_none():
    """两组都能凑出 → 不许随便挑一个。"""
    lines = [{"sod": "A", "deliver": 10.0}, {"sod": "B", "deliver": 10.0}, {"sod": "C", "deliver": 20.0}]
    assert C.subset_sum_unique(lines, 20.0) is None


def test_subset_no_solution_returns_none():
    lines = [{"sod": "A", "deliver": 10.0}, {"sod": "B", "deliver": 25.0}]
    assert C.subset_sum_unique(lines, 12.0) is None


def test_subset_too_many_lines_gives_up():
    lines = [{"sod": f"S{i}", "deliver": float(i + 1)} for i in range(40)]
    assert C.subset_sum_unique(lines, 3.0) is None


def test_subset_uses_cents_not_float():
    """0.1+0.2 这种浮点坑不许把命中判成不命中。"""
    lines = [{"sod": "A", "deliver": 0.1}, {"sod": "B", "deliver": 0.2}]
    assert C.subset_sum_unique(lines, 0.3) is not None


# ══════════════════════════════════════════════════════════
# 二、展开：一笔到账 → SOD 级 records
# ══════════════════════════════════════════════════════════
def test_no_writeoff_means_full_settle():
    """无逐SO金额时，父回款足够则按交付额升序把全部订单核销。"""
    p = _pay(amount=5800.0, orders=[
        {"so": "SO1", "deliver": 1450.0}, {"so": "SO2", "deliver": 1740.0}, {"so": "SO3", "deliver": 2610.0},
    ])
    recs = C.expand_payment(p, {})
    assert len(recs) == 3
    assert sum(r["amount_orig"] for r in recs) == 5800.0
    assert all("父回款总到账按交付额从小到大" in r["match_basis"] for r in recs)


def test_parent_fallback_stops_after_partial_smallest_order():
    """父回款小于最小交付额时，只给最小订单做部分回款。"""
    p = _pay(
        amount=80.0,
        orders=[
            {"so": "SO_BIG", "deliver": 200.0},
            {"so": "SO_SMALL", "deliver": 100.0},
        ],
        sod_lines={
            "SO_BIG": [{"sod": "SOD_BIG", "deliver": 200.0}],
            "SO_SMALL": [{"sod": "SOD_SMALL", "deliver": 100.0}],
        },
    )
    recs = C.expand_payment(p, {})
    assert [(r["so"], r["amount_orig"]) for r in recs] == [("SO_SMALL", 80.0)]
    assert p["_parent_fallback_allocation"]["zero_sos"] == ["SO_BIG"]
    coverage = C.source_coverage([p], recs)
    assert coverage["expected_order_keys"] == 1
    assert coverage["linked_order_keys"] == 2
    assert coverage["zero_allocation_order_keys"] == ["AR_T|SO_BIG"]


def test_parent_fallback_multiple_orders_are_waterfall_allocated():
    """多个订单按交付额从小到大：先全额，余额不足的当前单部分，后续为零。"""
    p = _pay(
        amount=275.0,
        orders=[
            {"so": "SO_300", "deliver": 300.0},
            {"so": "SO_050", "deliver": 50.0},
            {"so": "SO_200", "deliver": 200.0},
            {"so": "SO_100", "deliver": 100.0},
        ],
        sod_lines={
            "SO_300": [{"sod": "SOD_300", "deliver": 300.0}],
            "SO_050": [{"sod": "SOD_050", "deliver": 50.0}],
            "SO_200": [{"sod": "SOD_200", "deliver": 200.0}],
            "SO_100": [{"sod": "SOD_100", "deliver": 100.0}],
        },
    )
    recs = C.expand_payment(p, {})
    assert [(r["so"], r["amount_orig"]) for r in recs] == [
        ("SO_050", 50.0),
        ("SO_100", 100.0),
        ("SO_200", 125.0),
    ]
    audit = p["_parent_fallback_allocation"]
    assert audit["partial_sos"] == ["SO_200"]
    assert audit["zero_sos"] == ["SO_300"]
    assert audit["unallocated_parent_amount"] == 0.0


def test_parent_fallback_equal_deliveries_keep_source_order():
    """交付额相同时按智云关联顺序处理，避免重跑时漂移。"""
    p = _pay(
        amount=150.0,
        orders=[
            {"so": "SO_FIRST", "deliver": 100.0},
            {"so": "SO_SECOND", "deliver": 100.0},
        ],
        sod_lines={
            "SO_FIRST": [{"sod": "SOD_FIRST", "deliver": 100.0}],
            "SO_SECOND": [{"sod": "SOD_SECOND", "deliver": 100.0}],
        },
    )
    recs = C.expand_payment(p, {})
    assert [(r["so"], r["amount_orig"]) for r in recs] == [
        ("SO_FIRST", 100.0),
        ("SO_SECOND", 50.0),
    ]


def test_parent_fallback_surplus_is_left_unallocated():
    """父回款超过全部交付额时，订单全部核销，余额保留在父回款而不超写。"""
    p = _pay(
        amount=350.0,
        orders=[
            {"so": "SO1", "deliver": 100.0},
            {"so": "SO2", "deliver": 200.0},
        ],
    )
    recs = C.expand_payment(p, {})
    assert {r["so"]: r["amount_orig"] for r in recs} == {
        "SO1": 100.0,
        "SO2": 200.0,
    }
    assert p["_parent_fallback_allocation"]["unallocated_parent_amount"] == 50.0


def test_parent_fallback_skips_order_already_settled_in_ledger():
    """顺序台账还没有历史时，也要先跳过盈亏表已结账且无未结账拆分行的订单。"""
    p = _pay(
        amount=50.0,
        orders=[
            {"so": "SO_CLOSED", "deliver": 100.0},
            {"so": "SO_OPEN", "deliver": 200.0},
        ],
        sod_lines={
            "SO_CLOSED": [{"sod": "SOD_CLOSED", "deliver": 100.0}],
            "SO_OPEN": [{"sod": "SOD_OPEN", "deliver": 200.0}],
        },
        _ledger_settled_sos=["SO_CLOSED"],
    )

    recs = C.expand_payment(p, {})

    assert [(r["so"], r["amount_orig"]) for r in recs] == [("SO_OPEN", 50.0)]
    audit = p["_parent_fallback_allocation"]
    assert audit["already_settled_sos"] == ["SO_CLOSED"]
    closed = next(x for x in audit["allocations"] if x["so"] == "SO_CLOSED")
    assert closed["status"] == "ledger_already_settled"
    assert closed["ledger_settled_bootstrap"] is True


def test_writeoff_overrides_deliver():
    """有核销明细就以它为准（部分核销场景）。"""
    p = _pay(amount=1000.0,
             orders=[{"so": "SO1", "deliver": 900.0}, {"so": "SO2", "deliver": 500.0}],
             writeoffs={"SO1": 700.0, "SO2": 300.0})
    recs = C.expand_payment(p, {})
    assert sum(r["amount_orig"] for r in recs) == 1000.0
    assert all("本次核销金额" in r["match_basis"] for r in recs)


def test_sod_expansion_full():
    """一个 SO 拆 N 个 SOD，合计=核销额 → N 行全出（旧版误判成 E8 歧义的那类）。"""
    p = _pay(amount=300.0, orders=[{"so": "SO1", "deliver": 300.0}])
    p["sod_lines"] = {"SO1": [
        {"sod": "SOD3", "deliver": 100.0}, {"sod": "SOD2", "deliver": 120.0}, {"sod": "SOD1", "deliver": 80.0},
    ]}
    recs = C.expand_payment(p, {})
    assert len(recs) == 3
    assert {r["sod"] for r in recs} == {"SOD1", "SOD2", "SOD3"}
    assert sum(r["amount_orig"] for r in recs) == 300.0
    assert all(r["so_delivery_local"] == 300.0 for r in recs)
    assert all(r["all_sods"] == ["SOD1", "SOD2", "SOD3"] for r in recs)


def test_sod_expansion_subset():
    p = _pay(amount=180.0, orders=[{"so": "SO1", "deliver": 300.0}], writeoffs={"SO1": 180.0})
    p["sod_lines"] = {"SO1": [
        {"sod": "SOD1", "deliver": 100.0}, {"sod": "SOD2", "deliver": 120.0}, {"sod": "SOD3", "deliver": 80.0},
    ]}
    recs = C.expand_payment(p, {})
    assert {r["sod"] for r in recs} == {"SOD1", "SOD3"}


def test_sod_ambiguous_holds_e5():
    p = _pay(amount=100.0, orders=[{"so": "SO1", "deliver": 300.0}], writeoffs={"SO1": 100.0})
    p["sod_lines"] = {"SO1": [
        {"sod": "SOD1", "deliver": 100.0}, {"sod": "SOD2", "deliver": 100.0}, {"sod": "SOD3", "deliver": 100.0},
    ]}
    recs = C.expand_payment(p, {})
    assert len(recs) == 1 and recs[0]["forced_code"] == "E5"
    assert "SOD1" in recs[0]["forced_reason"]  # 候选要摆出来给她挑


def test_no_sod_falls_back_to_so():
    """订单明细查不到 SOD → 退化按 SO 匹配，仍然可判，不是丢单。"""
    p = _pay(amount=50.0, orders=[{"so": "SO1", "deliver": 50.0}])
    recs = C.expand_payment(p, {})
    assert len(recs) == 1 and recs[0]["sod"] == "" and recs[0]["amount_orig"] == 50.0


def test_fenbi_without_detail_uses_parent_waterfall():
    p = _pay(huikuan_type="分笔回款")
    recs = C.expand_payment(p, {})
    assert len(recs) == 1
    assert recs[0].get("forced_code") is None
    assert recs[0]["amount_orig"] == 100.0


def test_fee_is_included_in_parent_total_without_itemized_writeoff():
    p = _pay(
        amount=97.0,
        fee=1.62,
        orders=[{"so": "SO1", "deliver": 100.0}],
    )
    recs = C.expand_payment(p, {})
    assert recs[0].get("forced_code") is None
    assert recs[0]["amount_orig"] == 98.62
    assert recs[0]["fee"] == 1.62


def test_fee_completes_parent_waterfall_amount():
    """没有逐单金额时，手续费加回净到账形成总到账并参与顺序核销。"""
    p = _pay(
        amount=298.38,
        fee=1.62,
        orders=[
            {"so": "SO1", "deliver": 100.0},
            {"so": "SO2", "deliver": 200.0},
        ],
        sod_lines={
            "SO1": [{"sod": "SOD1", "deliver": 100.0}],
            "SO2": [{"sod": "SOD2", "deliver": 200.0}],
        },
    )
    recs = C.expand_payment(p, {})
    assert {r["so"] for r in recs} == {"SO1", "SO2"}
    assert all(r.get("forced_code") is None for r in recs)
    assert round(sum(r["amount_orig"] for r in recs), 2) == 300.0
    by_so = {r["so"]: r for r in recs}
    assert by_so["SO1"]["amount_orig"] == 100.0
    assert by_so["SO2"]["amount_orig"] == 200.0
    assert all("父回款总到账" in r["match_basis"] for r in recs)


def test_no_orders_is_e7_not_silent_drop():
    p = _pay(orders=[])
    recs = C.expand_payment(p, {})
    assert len(recs) == 1 and recs[0]["forced_code"] == "E7"


def test_explicit_writeoff_is_not_capped_by_parent_arrival():
    p = _pay(amount=100.0, orders=[{"so": "SO1", "deliver": 500.0}], writeoffs={"SO1": 500.0})
    recs = C.expand_payment(p, {})
    assert recs[0].get("forced_code") is None
    assert recs[0]["amount_local"] == 500.0


def test_explicit_partial_writeoff_reaches_partial_settlement_logic():
    """部分回款必须由逐 SO 核销明细证明，不能用父到账额代替。"""
    p = _pay(
        amount=8729.35,
        orders=[{"so": "SO1", "deliver": 8946.89}],
        writeoffs={"SO1": 8729.35},
        cumulative_writeoffs={"SO1": 8729.35},
        sod_lines={"SO1": [{"sod": "SOD1", "deliver": 8946.89}]},
    )
    recs = C.expand_payment(p, {})
    assert recs[0].get("forced_code") is None
    assert recs[0]["amount_local"] == 8729.35
    assert recs[0]["deliver_local"] == 8946.89
    assert recs[0]["cumulative_received_local"] == 8729.35


def test_fx_order_delivery_without_rate_is_not_written_as_original_currency():
    p = _pay(currency="美元USD", amount_local=None)
    recs = C.expand_payment(p, {})
    assert recs[0].get("forced_code") is None
    # 盈亏表只写人民币。无逐SO本币金额且订单也没有汇率时，禁止把美元原币直接当本币。
    assert recs[0]["amount_local"] is None
    assert recs[0]["amount_orig"] == 100.0


# ══════════════════════════════════════════════════════════
# 三、AR 覆盖率硬校验（2026-07-22 静默丢 3 笔的防回归）
# ══════════════════════════════════════════════════════════
def test_every_payment_produces_at_least_one_record():
    payments = [_pay(ar="A1"), _pay(ar="A2", orders=[]), _pay(ar="A3", huikuan_type="分笔回款")]
    recs = C.expand_payments(payments, {})
    assert {r["ar"] for r in recs} == {"A1", "A2", "A3"}


def test_coverage_error_raised_when_ar_lost(monkeypatch):
    """人为制造丢单 → 必须炸，绝不静默放行。"""
    def _drop(p, rates):
        return [] if p["ar"] == "A2" else [{"ar": p["ar"], "so": "SO1", "sod": "", "amount_orig": 1.0}]

    monkeypatch.setattr(C, "expand_payment", _drop)
    with pytest.raises(C.CoverageError) as e:
        C.expand_payments([_pay(ar="A1"), _pay(ar="A2")], {})
    assert "A2" in str(e.value)


# ══════════════════════════════════════════════════════════
# 四、盈亏表定位
# ══════════════════════════════════════════════════════════
def test_match_prefers_so_plus_amount():
    """主键 = SO+应收金额：回填前后都成立。"""
    led = _led({
        1: {"so": "SO1", "sod": "", "yingshou": 100.0},
        2: {"so": "SO1", "sod": "", "yingshou": 200.0},
    })
    assert led.match("SO1", "", 200.0)[0] == 2
    assert led.match("SO1", "", 100.0)[0] == 1


def test_match_falls_back_to_sod_then_so():
    led = _led({5: {"so": "SO1", "sod": "SODX", "yingshou": None}})
    assert led.match("SO1", "SODX", 99.0)[1] == "SOD"
    led2 = _led({7: {"so": "SO2", "sod": "", "yingshou": None}})
    assert led2.match("SO2", "", None)[1] == "SO"


def test_match_multi_same_amount_is_e8():
    led = _led({
        1: {"so": "SO1", "sod": "", "yingshou": 18.7},
        2: {"so": "SO1", "sod": "", "yingshou": 18.7},
    })
    row, how, cands = led.match("SO1", "", 18.7)
    assert how == "E8" and row is None and len(cands) == 2


def test_match_multi_same_amount_uses_unique_sod_within_candidates():
    """同 SO 同金额时，实收金额列已有唯一 SOD 就应安全消歧。"""
    led = _led({
        1: {"so": "SO1", "sod": "SOD1", "yingshou": 18.7},
        2: {"so": "SO1", "sod": "SOD2", "yingshou": 18.7},
    })
    assert led.match("SO1", "SOD1", 18.7) == (1, "SO+应收金额+SOD", [1, 2])
    assert led.match("SO1", "SOD2", 18.7) == (2, "SO+应收金额+SOD", [1, 2])


def test_match_multi_same_amount_does_not_use_sod_outside_amount_candidates():
    """SOD 即使存在，也不能跨出当前金额候选集选另一行。"""
    led = _led({
        1: {"so": "SO1", "sod": "SOD1", "yingshou": 18.7},
        2: {"so": "SO1", "sod": "SOD2", "yingshou": 18.7},
        3: {"so": "SO1", "sod": "SOD3", "yingshou": 99.0},
    })
    row, how, cands = led.match("SO1", "SOD3", 18.7)
    assert row is None and how == "E8" and cands == [1, 2]


def test_match_multi_same_amount_duplicate_sod_uses_only_outstanding_row():
    """同金额同 SOD 的拆分行只有一个未结清承接行时，沿用既有安全规则。"""
    led = _led({
        1: {
            "so": "SO1", "sod": "SOD1", "yingshou": 18.7,
            "huikuan": 18.7, "jiezhang": "是",
        },
        2: {
            "so": "SO1", "sod": "SOD1", "yingshou": 18.7,
            "huikuan": None, "jiezhang": "否",
        },
    })
    assert led.match("SO1", "SOD1", 18.7) == (
        2, "SO+应收金额+SOD未结清行", [1, 2]
    )


def test_match_multi_same_amount_duplicate_sod_multiple_open_rows_stays_e8():
    """同金额同 SOD 仍有多个未结清候选时必须继续挂起。"""
    led = _led({
        1: {
            "so": "SO1", "sod": "SOD1", "yingshou": 18.7,
            "huikuan": None, "jiezhang": "否",
        },
        2: {
            "so": "SO1", "sod": "SOD1", "yingshou": 18.7,
            "huikuan": None, "jiezhang": "否",
        },
    })
    row, how, cands = led.match("SO1", "SOD1", 18.7)
    assert row is None and how == "E8" and cands == [1, 2]


def test_positional_alignment_resolves_equal_amounts():
    """整段逐位对齐（SOD 降序 ↔ 行号升序）能严格消掉等额歧义。"""
    led = _led({
        10: {"so": "SO1", "sod": "", "yingshou": 72.98},
        11: {"so": "SO1", "sod": "", "yingshou": 18.70},
        12: {"so": "SO1", "sod": "", "yingshou": 18.70},
        13: {"so": "SO1", "sod": "", "yingshou": 25.38},
    })
    lines = [
        {"sod": "SOD9", "deliver": 72.98}, {"sod": "SOD8", "deliver": 18.70},
        {"sod": "SOD7", "deliver": 18.70}, {"sod": "SOD6", "deliver": 25.38},
    ]
    assert led.positional_row("SO1", "SOD8", lines) == (11, "exact", None)
    assert led.positional_row("SO1", "SOD7", lines) == (12, "exact", None)


def test_positional_alignment_accepts_systematic_ratio():
    """
    智云交付额与她表里应收整段差**同一个比例**（实测 SO26040322 = 0.977433）→
    是口径差不是行错位，可以对齐；行错位凑不出同一个比值。
    """
    led = _led({
        1: {"so": "SO1", "sod": "", "yingshou": 1240.14},
        2: {"so": "SO1", "sod": "", "yingshou": 514.35},
        3: {"so": "SO1", "sod": "", "yingshou": 488.64},
        4: {"so": "SO1", "sod": "", "yingshou": 676.58},
    })
    lines = [
        {"sod": "SOD27", "deliver": 1240.14}, {"sod": "SOD26", "deliver": 514.35},
        {"sod": "SOD25", "deliver": 477.61}, {"sod": "SOD24", "deliver": 661.31},
    ]
    row, kind, ratio = led.positional_row("SO1", "SOD25", lines)
    assert (row, kind) == (3, "ratio")
    assert abs(ratio - 0.977433) < 1e-5


def test_positional_alignment_refuses_inconsistent_ratios():
    """比例各不相同 → 更像行错位，必须拒绝。"""
    led = _led({
        1: {"so": "SO1", "sod": "", "yingshou": 100.0},
        2: {"so": "SO1", "sod": "", "yingshou": 200.0},
    })
    lines = [{"sod": "SOD2", "deliver": 90.0}, {"sod": "SOD1", "deliver": 150.0}]
    assert led.positional_row("SO1", "SOD2", lines) is None


def test_positional_alignment_refuses_when_sequence_mismatch():
    """她把某个 SOD 拆成了两行 → 对不齐 → 返回 None，老实挂起。"""
    led = _led({
        10: {"so": "SO1", "sod": "", "yingshou": 202.48},
        11: {"so": "SO1", "sod": "", "yingshou": 199.78},
        12: {"so": "SO1", "sod": "", "yingshou": 450.00},
    })
    lines = [{"sod": "SOD80", "deliver": 402.26}, {"sod": "SOD79", "deliver": 450.00}]
    assert led.positional_row("SO1", "SOD80", lines) is None


# ══════════════════════════════════════════════════════════
# 五、单条判定 E 码
# ══════════════════════════════════════════════════════════
def _rec(so="SO26010001", sod="SOD26010001", amount=100.0, **kw):
    r = {
        # deliver_local = 该 SOD 的智云交付额；真实 record 一定带（计提口径要用）
        "deliver_local": amount,
        "ar": "AR_T", "so": so, "sod": sod, "amount_orig": amount,
        "currency": "人民币CNY", "status": "手动核销", "fee": 0,
        "customer": "测试客户甲",
        "hexiao_date": dt.date(2026, 7, 22), "shoukuan_date": dt.date(2026, 7, 21),
    }
    r.update(kw)
    return r


def test_classify_equal_amount_rows_use_existing_unique_sod():
    """SO26060458 类：同额 SOD 已在实收金额列时应分别定位，不得误报 E8。"""
    led = _led({
        10: {"so": "SO1", "sod": "SOD1", "yingshou": 100.0},
        11: {"so": "SO1", "sod": "SOD2", "yingshou": 100.0},
        12: {"so": "SO1", "sod": "SOD3", "yingshou": 50.0},
    })
    first = C.classify_one(_rec("SO1", "SOD1", 100.0), led, {}, 0.0, 2026)
    second = C.classify_one(_rec("SO1", "SOD2", 100.0), led, {}, 0.0, 2026)
    third = C.classify_one(_rec("SO1", "SOD3", 50.0), led, {}, 0.0, 2026)
    assert (first["bucket"], first["ledger_row_ref"]) == ("auto", 10)
    assert (second["bucket"], second["ledger_row_ref"]) == ("auto", 11)
    assert (third["bucket"], third["ledger_row_ref"]) == ("auto", 12)


def test_auto_happy_path():
    led = _led({3: {"so": "SO26017777", "sod": "", "yingshou": 200.0}})
    r = C.classify_one(_rec("SO26017777", "SOD26017777", 200.0), led, {}, 0.0, 2026)
    assert r["bucket"] == "auto"
    assert r["five_cols"] == {
        "计提": 200.0, "回款明细": 200.0, "是否结账": "是",
        "收款时间": "2026-07-21", "收款方式": "汇", "实收SOD": "SOD26017777",
    }
    assert "禁止用行号" in r["locate_hint"]


def test_forced_hold_becomes_idempotent_auto_when_order_is_already_settled():
    led = _led({
        3: {
            "so": "SO_CLOSED", "sod": "SOD_CLOSED", "yingshou": 100.0,
            "jiti": 100.0, "huikuan": 100.0, "jiezhang": "是",
            "shoukuan_time": dt.date(2026, 7, 21), "shoukuan_way": "汇",
        },
    })
    rec = _rec(
        "SO_CLOSED", "SOD_CLOSED", 100.0,
        forced_code="E_SYSTEM_OVER_WRITEOFF_UNRESOLVED",
        forced_reason="旧判定本应挂起",
    )

    result = C.classify_one(rec, led, {}, 0.0, 2026)

    assert result["bucket"] == "auto"
    assert result["code"] == "OK_ALREADY_SETTLED"
    assert result["ledger_row_ref"] == 3


def test_parent_writeoff_mismatch_cannot_be_hidden_by_settled_precheck():
    led = _led({
        3: {
            "so": "SO_CLOSED", "sod": "SOD_CLOSED", "yingshou": 100.0,
            "jiti": 100.0, "huikuan": 100.0, "jiezhang": "是",
            "shoukuan_time": dt.date(2026, 7, 21), "shoukuan_way": "汇",
        },
    })
    rec = _rec(
        "SO_CLOSED", "SOD_CLOSED", 100.0,
        forced_code="E_PARENT_WRITEOFF_MISMATCH",
        forced_reason="整笔父回款金额守恒检查未通过",
    )

    result = C.classify_one(rec, led, {}, 0.0, 2026)

    assert result["bucket"] == "exception"
    assert result["code"] == "E_PARENT_WRITEOFF_MISMATCH"


def test_settled_precheck_does_not_skip_when_open_split_row_exists():
    led = _led({
        3: {"so": "SO_PART", "sod": "SOD_PART", "yingshou": 60.0,
            "huikuan": 60.0, "jiezhang": "是"},
        4: {"so": "SO_PART", "sod": "SOD_PART", "yingshou": 40.0,
            "huikuan": None, "jiezhang": "否"},
    })

    assert led.settled_without_open_row("SO_PART", "SOD_PART") is None


def test_receipt_time_is_arrival_date_same_month():
    """她填的收款时间 = 到账日（同月）。7-17 到账、7-22 核销 → 填 7-17。"""
    led = _led({1: {"so": "SO1", "sod": "", "yingshou": 10.0}})
    r = C.classify_one(
        _rec("SO1", "SOD1", 10.0, shoukuan_date=dt.date(2026, 7, 17), hexiao_date=dt.date(2026, 7, 22)),
        led, {}, 0.0, 2026,
    )
    assert r["five_cols"]["收款时间"] == "2026-07-17"
    assert r["five_cols"]["收款方式"] == "汇"


def test_cross_month_is_chongyushou():
    led = _led({1: {"so": "SO1", "sod": "", "yingshou": 10.0}})
    r = C.classify_one(
        _rec("SO1", "SOD1", 10.0, shoukuan_date=dt.date(2026, 6, 26), hexiao_date=dt.date(2026, 7, 8)),
        led, {}, 0.0, 2026,
    )
    assert r["five_cols"]["收款方式"] == "冲预收"


def test_prepaid_type_no_longer_forces_chongyushou():
    """2026-07-23 口径修正：预存回款同月核销 → 「汇」（旧版错填冲预收，15 笔全错）。"""
    led = _led({1: {"so": "SO1", "sod": "", "yingshou": 10.0}})
    r = C.classify_one(
        _rec("SO1", "SOD1", 10.0, huikuan_type="预存回款", status="核销成功"),
        led, {}, 0.0, 2026,
    )
    assert r["five_cols"]["收款方式"] == "汇"


def test_prepaid_partial_status_still_settles_each_sod():
    """2026-07-24 明妹「核销状态判断逻辑澄清会」：回款记录整笔「预存部分核销」= 那笔预存
    余额没花完（6300 核 6090 剩 210），跟每个小单收没收满**无关**。每个已核销+已交付的单
    → 结账「是」、计提按本单是否收满判。旧版拿 status∈SETTLED 判结账，把 6 个已收满的
    预存视频单全误判「否」（对明妹真答案实测 146/152，修后 152/152）。"""
    led = _led({1: {"so": "SO1", "sod": "SOD1", "yingshou": 420.0}})
    r = C.classify_one(
        _rec("SO1", "SOD1", 420.0, huikuan_type="预存回款", status="预存部分核销"),  # 本次核销=交付420=收满
        led, {}, 0.0, 2026,
    )
    assert r["bucket"] == "auto"
    assert r["five_cols"]["是否结账"] == "是"
    assert r["five_cols"]["计提"] == 420.0


def test_settle_yes_but_jiti_empty_when_not_full():
    """8块/10块 例（明妹澄清会原话）：本次核销 8 < 交付 10 → 这笔到账的任务做完=结账「是」，
    但整单没回满 → 计提留空（2 块挂应收）。结账与计提是两个独立判据。"""
    led = _led({1: {"so": "SO1", "sod": "SOD1", "yingshou": 10.0}})
    r = C.classify_one(
        _rec("SO1", "SOD1", 8.0, deliver_local=10.0),  # 本次核销 8、交付 10 → 没回满
        led, {}, 0.0, 2026,
    )
    assert r["bucket"] == "auto"
    assert r["five_cols"]["是否结账"] == "是"
    assert r["five_cols"]["计提"] is None
    assert r["row_operation"]["paid_receivable"] == 8.0
    assert r["row_operation"]["unpaid_receivable"] == 2.0


@pytest.mark.parametrize("amount,delta", [(99.72, 0.28), (99.00, 1.00)])
def test_one_yuan_business_tail_settles_without_creating_unpaid_row(amount, delta):
    led = _led({1: {"so": "SO1", "sod": "SOD1", "yingshou": 100.0}})
    r = C.classify_one(
        _rec(
            "SO1", "SOD1", amount,
            deliver_local=100.0,
            cumulative_received_local=amount,
        ),
        led, {}, 0.0, 2026,
    )

    assert r["bucket"] == "auto"
    assert "row_operation" not in r
    assert r["five_cols"]["计提"] == 100.0
    assert r["five_cols"]["回款明细"] == amount
    assert r["settlement_tolerance_audit"]["exact_delta"] == delta
    assert r["settlement_tolerance_audit"]["business_equal"] is True
    assert "W_SETTLEMENT_TAIL_TOLERATED" in r["warning_codes"]


def test_business_tail_above_one_yuan_still_creates_unpaid_row():
    led = _led({1: {"so": "SO1", "sod": "SOD1", "yingshou": 100.0}})
    r = C.classify_one(
        _rec(
            "SO1", "SOD1", 98.99,
            deliver_local=100.0,
            cumulative_received_local=98.99,
        ),
        led, {}, 0.0, 2026,
    )

    assert r["row_operation"]["type"] == "split_below"
    assert r["row_operation"]["unpaid_receivable"] == 1.01


def test_repeated_partial_only_counts_current_sod():
    """重复部分回款命中唯一未结清行，且不混入同 SO 的其它 SOD。"""
    led = _led({
        1: {"so": "SO1", "sod": "SOD1", "yingshou": 14000.0, "huikuan": 15000.0, "jiezhang": "是"},
        2: {"so": "SO1", "sod": "SOD1", "yingshou": 16000.0, "huikuan": None, "jiezhang": "否"},
        3: {"so": "SO1", "sod": "SOD2", "yingshou": 999.0, "huikuan": 999.0, "jiezhang": "是"},
    })
    r = C.classify_one(
        _rec("SO1", "SOD1", 10000.0, deliver_local=31000.0),
        led, {}, 0.0, 2026,
    )
    assert r["ledger_row_ref"] == 2
    assert r["row_operation"]["existing_received"] == 15000.0
    assert r["row_operation"]["paid_receivable"] == 10000.0
    assert r["row_operation"]["unpaid_receivable"] == 6000.0


def test_cross_year_only_after_ledger_miss():
    """2025 的单**在表里有行**就正常填；只有表里找不到才判 E3。"""
    led = _led({1: {"so": "SO25120734", "sod": "", "yingshou": 52200.0}})
    r = C.classify_one(_rec("SO25120734", "SOD25121039", 52200.0), led, {}, 0.0, 2026)
    assert r["bucket"] == "auto", r
    r2 = C.classify_one(_rec("SO25080089", "SOD25080128", 1.0), led, {}, 0.0, 2026)
    assert r2["code"] == "E3" and r2["bucket"] == "hold"


def test_missing_so_is_e2():
    led = _led({1: {"so": "SO26010000", "sod": "", "yingshou": 1.0}})
    r = C.classify_one(_rec("SO26999999", "SOD26999999", 10.0), led, {}, 0.0, 2026)
    assert r["code"] == "E2" and r["bucket"] == "hold"


def test_void_status_e7():
    r = C.classify_one(_rec(status="已作废"), C.LedgerIndex(), {}, 0.0, 2026)
    assert r["code"] == "E7" and r["bucket"] == "exception"


def test_excess_over_deliver_is_e4_not_over_yingshou():
    """超额核销的上限用**智云交付额**，不是她表应收（2026-07-24 明妹口径）。

    · 本次核销 500 > 智云交付 300 → 真超额 E4。
    · 本次核销 500 == 智云交付 500、但她表应收才 50（旧值）→ **不是超额**，
      是「交付额变大了」→ 照样按智云额填、顶个 ⚠（见下一条），绝不误报 E4。
    """
    led = _led({1: {"so": "SO1", "sod": "", "yingshou": 100.0}})
    r = C.classify_one(_rec("SO1", "SOD1", 100.0), led, {}, 0.0, 2026)
    assert r["bucket"] == "auto"

    over = _led({1: {"so": "SO1", "sod": "SOD1", "yingshou": 50.0}})
    r_over = C.classify_one(_rec("SO1", "SOD1", 500.0, deliver_local=300.0), over, {}, 0.0, 2026)
    assert r_over["code"] == "E4"

    stale = _led({1: {"so": "SO1", "sod": "SOD1", "yingshou": 50.0}})
    r_stale = C.classify_one(_rec("SO1", "SOD1", 500.0, deliver_local=500.0), stale, {}, 0.0, 2026)
    assert r_stale["bucket"] == "auto"        # 核销==交付，正常
    assert r_stale["five_cols"]["回款明细"] == 500.0


def test_delivery_amount_changed_flags_warning():
    """2026-07-24 明妹口述 107/422：交付额中途变过，她表里应收还是旧值。
    程序靠 SOD 命中那行后，要按**智云额**填、并顶个 ⚠ 让她扫一眼，不能闷声填。
    这是「让他动脑子检查交付额、别以我表里为准」的落码。"""
    # 她表：SO1 那行应收=3242（旧值），智云这次交付/核销都是 408
    led = _led({5: {"so": "SO1", "sod": "SOD1", "yingshou": 3242.0}})
    r = C.classify_one(_rec("SO1", "SOD1", 408.0, deliver_local=408.0), led, {}, 0.0, 2026)
    assert r["bucket"] == "auto"
    assert r["five_cols"]["回款明细"] == 408.0     # 按智云额填，不是她表 3242
    assert "⚠" in r["reason"]
    assert "3242" in r["reason"] and "408" in r["reason"]


def test_flow_signals():
    led = _led({1: {"so": "SO1", "sod": "", "yingshou": 100.0}})
    assert C.classify_one(_rec("SO1", "SOD1", flow_hits=0), led, {}, 0.0, 2026)["code"] == "E0"
    assert C.classify_one(_rec("SO1", "SOD1", flow_hits=3), led, {}, 0.0, 2026)["code"] == "E12"
    assert C.classify_one(
        _rec("SO1", "SOD1", customer_archive_failed=True), led, {}, 0.0, 2026
    )["code"] == "E10"


def test_no_ledger_never_auto():
    r = C.classify_one(_rec(), None, {}, 0.0, 2026)
    assert r["bucket"] == "hold" and r["code"] == "E2"
    res = C.classify_records([_rec()], None, {})
    assert res["counts"]["auto"] == 0


def test_same_row_hit_twice_both_held():
    """不同 SOD 误命中同一行仍必须挂起。"""
    led = _led({1: {"so": "SO1", "sod": "", "yingshou": 100.0}})
    res = C.classify_records([_rec("SO1", "SODA", 100.0), _rec("SO1", "SODB", 100.0)], led, {})
    assert res["counts"]["auto"] == 0
    assert all(h["code"] == "E8" for h in res["hold"])


def test_same_physical_writeoff_multi_sod_same_row_aggregates_by_so_delivery():
    """完整多 SOD 属于同一物理核销记录时，共用一行并按 SO 交付金额写一次。"""
    led = _led({
        1: {
            "so": "SO1", "sod": "SODA", "yingshou": 40.0,
            "jiti": None, "huikuan": None, "jiezhang": "否",
            "shoukuan_time": None, "shoukuan_way": None, "chayi": None,
        }
    })
    common = {
        "ar": "AR1",
        "so_delivery_local": 100.0,
        "all_sods": ["SODA", "SODB"],
        "writeoff_sequence_key": ["2026-08-06", "HX1", "RID1", "AR1", "SO1"],
    }
    first = _rec("SO1", "SODA", 40.0, deliver_local=40.0, **common)
    second = _rec("SO1", "SODB", 60.0, deliver_local=60.0, **common)

    result = C.classify_records([first, second], led, {})

    assert result["counts"] == {"auto": 2, "hold": 0, "exception": 0, "total": 2}
    target = next(item for item in result["auto"] if item.get("row_operation"))
    absorbed = next(item for item in result["auto"] if item.get("same_so_multi_sod_absorbed"))
    operation = target["row_operation"]
    assert operation["type"] == "same_so_multi_sod_aggregate"
    assert operation["so_delivery"] == 100.0
    assert target["five_cols"] == {
        "计提": 100.0, "回款明细": 100.0, "是否结账": "是",
        "收款时间": "2026-07-21", "收款方式": "汇",
        "实收SOD": "SODA、SODB",
    }
    assert absorbed["five_cols"] == {}
    assert "W_SAME_SO_MULTI_SOD_AGGREGATE" in target["warning_codes"]


def test_same_row_multi_sod_from_different_writeoffs_still_holds():
    led = _led({1: {"so": "SO1", "sod": "SODA", "yingshou": 40.0}})
    first = _rec(
        "SO1", "SODA", 40.0, ar="AR1", deliver_local=40.0,
        so_delivery_local=100.0, all_sods=["SODA", "SODB"],
        writeoff_sequence_key=["2026-08-06", "HX1", "RID1", "AR1", "SO1"],
    )
    second = _rec(
        "SO1", "SODB", 60.0, ar="AR2", deliver_local=60.0,
        so_delivery_local=100.0, all_sods=["SODA", "SODB"],
        writeoff_sequence_key=["2026-08-06", "HX2", "RID2", "AR2", "SO1"],
    )

    result = C.classify_records([first, second], led, {})

    assert result["counts"]["auto"] == 0
    assert result["counts"]["hold"] == 2


def test_same_so_sod_distinct_ar_builds_sequential_split_chain():
    """同一 SO/SOD 的不同父回款逐笔拆行，不能合并成一次回款。"""
    led = _led({1: {"so": "SO1", "sod": "SOD1", "yingshou": 100.0}})
    first = _rec(
        "SO1", "SOD1", 50.0, ar="AR1", deliver_local=100.0,
        cumulative_received_local=50.0,
        writeoff_sequence_key=["2026-07-22", "HX1", "1", "AR1", "SO1"],
    )
    second = _rec(
        "SO1", "SOD1", 50.0, ar="AR2", deliver_local=100.0,
        cumulative_received_local=100.0,
        writeoff_sequence_key=["2026-07-22", "HX2", "2", "AR2", "SO1"],
    )
    res = C.classify_records([first, second], led, {})
    assert res["counts"]["auto"] == 2
    assert res["counts"]["hold"] == 0
    assert {x["split_chain_group_id"] for x in res["auto"]} == {"split-payment-chain|1|SO1|SOD1"}
    assert all("W_SPLIT_PAYMENT_SEQUENTIAL" in x["warning_codes"] for x in res["auto"])
    op = res["auto"][0]["row_operation"]
    assert op["type"] == "split_payment_chain"
    assert [step["ar"] for step in op["steps"]] == ["AR1", "AR2"]
    assert [step["receivable"] for step in op["steps"]] == [50.0, 50.0]
    assert op["steps"][0]["five_cols"]["计提"] is None
    assert op["steps"][1]["five_cols"]["计提"] == 100.0
    assert op["final_unpaid"] is None


def test_settled_snapshot_split_chain_uses_each_parent_payment_amount():
    """已结清快照带回的是历史合计，拆分步骤仍必须保留各父回款本次额。"""
    led = _led({
        1: {
            "so": "SO1", "sod": "SOD1", "yingshou": 7200.0,
            "jiti": 22200.0, "huikuan": 7200.0, "jiezhang": "是",
        },
    })
    group = []
    for index, (ar, amount, cumulative) in enumerate(
        (("AR1", 3000.0, 18000.0), ("AR2", 4200.0, 22200.0)), start=1
    ):
        group.append({
            "ar": ar, "so": "SO1", "sod": "SOD1", "case_id": f"{ar}|SO1|SOD1",
            "five_cols": {
                "计提": 22200.0, "回款明细": 7200.0, "是否结账": "是",
                "收款时间": "2026-07-15", "收款方式": "冲预收", "实收SOD": "SOD1",
            },
            "derived_cols": {},
            "split_payment_source": {
                "amount_local": amount, "cumulative_local": cumulative,
                "delivery_local": 22200.0,
                "writeoff_sequence_key": ["2026-07-15", f"HX{index}", str(index), ar, "SO1"],
            },
        })

    operation, error = C._make_split_payment_chain(1, group, led, C.TOL)

    assert error == ""
    steps = operation["steps"]
    assert [step["five_cols"]["回款明细"] for step in steps] == [3000.0, 4200.0]


@pytest.mark.parametrize("tail", [0.12, 1.0])
def test_settled_aggregate_kept_when_parent_tail_is_within_one_yuan(tail):
    """结清尾差不超过1元时，即使来自独立父AR，也不把已有聚合行拆开。"""
    delivery = 211464.0
    led = _led({
        1: {
            "so": "SO26020257", "sod": "SOD26020257", "yingshou": delivery,
            "jiti": delivery, "huikuan": delivery, "jiezhang": "是",
            "shoukuan_time": dt.date(2026, 7, 15), "shoukuan_way": "汇",
        },
    })
    records = [
        _rec(
            "SO26020257", "SOD26020257", tail, ar="AR_SMALL",
            amount_local=tail, deliver_local=delivery, cumulative_received_local=tail,
            writeoff_sequence_key=["2026-07-16", "HX1", "1", "AR_SMALL", "SO26020257"],
        ),
        _rec(
            "SO26020257", "SOD26020257", delivery - tail, ar="AR_MAIN",
            amount_local=delivery - tail, deliver_local=delivery, cumulative_received_local=delivery,
            writeoff_sequence_key=["2026-07-16", "HX2", "2", "AR_MAIN", "SO26020257"],
        ),
    ]

    result = C.classify_records(records, led, {})

    assert result["counts"] == {"auto": 2, "hold": 0, "exception": 0, "total": 2}
    operations = [item["row_operation"] for item in result["auto"]]
    assert all(op["type"] == "preserve_aggregate_tail_tolerance" for op in operations)
    assert all(op["tolerated_tail_amount"] == tail for op in operations)
    assert all(item["five_cols"]["回款明细"] == delivery for item in result["auto"])
    assert all("不为该尾差单独拆行" in item["reason"] for item in result["auto"])


def test_parent_tail_above_one_yuan_still_builds_split_chain():
    delivery = 211464.0
    tail = 1.01
    led = _led({
        1: {
            "so": "SO1", "sod": "SOD1", "yingshou": delivery,
            "jiti": delivery, "huikuan": delivery, "jiezhang": "是",
            "shoukuan_time": dt.date(2026, 7, 15), "shoukuan_way": "汇",
        },
    })
    records = [
        _rec(
            "SO1", "SOD1", tail, ar="AR_SMALL", deliver_local=delivery,
            amount_local=tail, cumulative_received_local=tail,
            writeoff_sequence_key=["2026-07-16", "HX1", "1", "AR_SMALL", "SO1"],
        ),
        _rec(
            "SO1", "SOD1", delivery - tail, ar="AR_MAIN", deliver_local=delivery,
            amount_local=delivery - tail, cumulative_received_local=delivery,
            writeoff_sequence_key=["2026-07-16", "HX2", "2", "AR_MAIN", "SO1"],
        ),
    ]

    result = C.classify_records(records, led, {})

    assert result["counts"]["auto"] == 2
    assert result["auto"][0]["row_operation"]["type"] == "split_payment_chain"


def test_new_split_chain_absorbs_one_yuan_tail_into_final_business_row():
    led = _led({1: {"so": "SO1", "sod": "SOD1", "yingshou": 100.0}})
    records = [
        _rec(
            "SO1", "SOD1", 0.12, ar="AR_SMALL", deliver_local=100.0,
            amount_local=0.12, cumulative_received_local=0.12,
            writeoff_sequence_key=["2026-07-16", "HX1", "1", "AR_SMALL", "SO1"],
        ),
        _rec(
            "SO1", "SOD1", 99.88, ar="AR_MAIN", deliver_local=100.0,
            amount_local=99.88, cumulative_received_local=100.0,
            writeoff_sequence_key=["2026-07-16", "HX2", "2", "AR_MAIN", "SO1"],
        ),
    ]

    result = C.classify_records(records, led, {})

    assert result["counts"] == {"auto": 2, "hold": 0, "exception": 0, "total": 2}
    absorbed = next(item for item in result["auto"] if item.get("tail_tolerance_absorbed"))
    target = next(item for item in result["auto"] if item.get("row_operation"))
    assert absorbed["five_cols"] == {}
    assert target["row_operation"]["type"] == "settlement_tail_aggregate"
    assert target["five_cols"]["回款明细"] == 100.0
    audit = target["row_operation"]["tail_tolerance_audit"]
    assert audit["absorbed_total"] == 0.12
    assert audit["original_parent_amounts"] == [0.12, 99.88]


def test_multiple_small_parents_are_not_absorbed_when_total_exceeds_one_yuan():
    led = _led({1: {"so": "SO1", "sod": "SOD1", "yingshou": 1.2}})
    records = [
        _rec(
            "SO1", "SOD1", 0.60, ar="AR1", deliver_local=1.2,
            amount_local=0.60, cumulative_received_local=0.60,
            writeoff_sequence_key=["2026-07-16", "HX1", "1", "AR1", "SO1"],
        ),
        _rec(
            "SO1", "SOD1", 0.60, ar="AR2", deliver_local=1.2,
            amount_local=0.60, cumulative_received_local=1.2,
            writeoff_sequence_key=["2026-07-16", "HX2", "2", "AR2", "SO1"],
        ),
    ]

    result = C.classify_records(records, led, {})

    assert all(not item.get("tail_tolerance_absorbed") for item in result["auto"])
    assert result["auto"][0]["row_operation"]["type"] == "split_payment_chain"


def test_materialized_split_chain_is_found_when_both_parents_hit_chain_tail():
    led = _led({
        1: {
            "so": "SO1", "sod": "SOD1", "yingshou": 3000.0,
            "jiti": None, "huikuan": 3000.0, "jiezhang": "是",
            "shoukuan_time": dt.date(2026, 7, 15), "shoukuan_way": "冲预收",
        },
        2: {
            "so": "SO1", "sod": "SOD1", "yingshou": 4200.0,
            "jiti": 22200.0, "huikuan": 4200.0, "jiezhang": "是",
            "shoukuan_time": dt.date(2026, 7, 15), "shoukuan_way": "冲预收",
        },
    })
    group = []
    for index, (ar, amount, cumulative) in enumerate(
        (("AR1", 3000.0, 18000.0), ("AR2", 4200.0, 22200.0)), start=1
    ):
        group.append({
            "ar": ar, "so": "SO1", "sod": "SOD1", "case_id": f"{ar}|SO1|SOD1",
            "five_cols": {
                "计提": None, "回款明细": 3000.0, "是否结账": "是",
                "收款时间": "2026-07-15", "收款方式": "冲预收", "实收SOD": "SOD1",
            },
            "derived_cols": {},
            "split_payment_source": {
                "amount_local": amount, "cumulative_local": cumulative,
                "delivery_local": 22200.0,
                "writeoff_sequence_key": ["2026-07-15", f"HX{index}", str(index), ar, "SO1"],
            },
        })

    operation, error = C._make_split_payment_chain(2, group, led, C.TOL)

    assert error == ""
    assert operation["materialized_chain_start"] == 1
    assert operation["steps"][1]["five_cols"]["计提"] == 22200.0


def test_split_payment_chain_without_record_id_is_held():
    led = _led({1: {"so": "SO1", "sod": "SOD1", "yingshou": 100.0}})
    records = [
        _rec("SO1", "SOD1", 40.0, ar="AR1", deliver_local=100.0,
             cumulative_received_local=40.0),
        _rec("SO1", "SOD1", 60.0, ar="AR2", deliver_local=100.0,
             cumulative_received_local=100.0),
    ]
    res = C.classify_records(records, led, {})
    assert res["counts"]["auto"] == 0
    assert all(item["code"] == "E8" for item in res["hold"])
    assert all("缺少核销记录NUM" in item["reason"] for item in res["hold"])


def test_five_parent_payments_are_five_logical_steps_not_one_aggregate():
    led = _led({1: {"so": "SO1", "sod": "SOD1", "yingshou": 100.0}})
    records = [
        _rec(
            "SO1", "SOD1", 20.0, ar=f"AR{index}", deliver_local=100.0,
            cumulative_received_local=20.0 * index,
            writeoff_sequence_key=["2026-07-22", f"HX{index}", str(index), f"AR{index}", "SO1"],
        )
        for index in range(1, 6)
    ]
    result = C.classify_records(records, led, {})
    assert result["counts"] == {"auto": 5, "hold": 0, "exception": 0, "total": 5}
    operation = result["auto"][0]["row_operation"]
    assert [step["ar"] for step in operation["steps"]] == ["AR1", "AR2", "AR3", "AR4", "AR5"]
    assert [step["current_received"] for step in operation["steps"]] == [20.0] * 5
    assert [step["five_cols"]["计提"] for step in operation["steps"]] == [None, None, None, None, 100.0]


def test_case_id_is_sod_level():
    led = _led({1: {"so": "SO1", "sod": "", "yingshou": 100.0}})
    r = C.classify_one(_rec("SO1", "SOD1", 100.0), led, {}, 0.0, 2026)
    assert r["case_id"] == "AR_T|SO1|SOD1"


def test_fee_logic_has_no_named_order_overrides():
    """费用按父回款总额处理，但不得通过白名单或参考表点名改变核销金额。"""
    from pathlib import Path

    src = Path(C.__file__).read_text(encoding="utf-8")
    for bad in (
        "fee_gross_sos",
        "--fee-gross-so",
        "--fee-gross-ar",
        "--authoritative-ledger",
        "--authoritative-correct-so",
    ):
        assert bad not in src
    assert "_fee_net_payment" not in src
    assert "_fee_net_writeoffs" not in src
    assert "_allocate_weighted_cents" not in src
    assert "net_arrival_plus_explicit_fees_taxes" in src


# ══════════════════════════════════════════════════════════
# 六、真实金标端到端（有本地测试数据才跑）
# ══════════════════════════════════════════════════════════
GOLD_EXPORTS = GOLD_DIR / "01_智云导出"
GOLD_LEDGER = GOLD_DIR / "02_我的表副本" / "2026年盈亏核算表1-12月（副本）.xlsx"


@pytest.mark.skipif(not GOLD_LEDGER.is_file(), reason="无本地金标数据（真实数据不进仓库）")
def test_gold_20260722_end_to_end():
    """
    2026-07-22 真实 13 笔：判定结果必须与明妹当天手工填的**逐格一致**。
    这份数据抓出过三个真 bug（静默丢单 / SOD 拆行误判 / 冲预收规则错），是主回归闸。
    """
    import datetime as _dt

    from openpyxl import load_workbook

    payments = C.load_exports(GOLD_DIR)
    assert len(payments) == 13
    records = C.expand_payments(payments, {})
    assert len(records) == 157
    ledger = C.LedgerIndex(GOLD_LEDGER)
    res = C.classify_records(records, ledger, {})

    assert res["counts"]["auto"] == 143, res["e_code_dist"]
    assert res["counts"]["exception"] == 0
    assert res["e_code_dist"].get("E8") is None  # 同SO多行歧义已被 SOD 化解
    assert {p["ar"] for p in payments} == {r["ar"] for r in records}  # 一笔都没丢

    wb = load_workbook(GOLD_LEDGER, read_only=True, data_only=True)
    rows = list(wb["明细"].iter_rows(values_only=True))
    wb.close()
    hdr = list(rows[0])
    col = {
        "计提": hdr.index("计提金额"), "回款明细": hdr.index("回款明细"),
        "是否结账": hdr.index("是否结账（是/否）"), "收款时间": hdr.index("收款时间"),
        "收款方式": hdr.index("收款方式(支/汇/现)"), "实收SOD": hdr.index("实收金额"),
        "SO": hdr.index("新智云单号"),
    }

    def norm(v):
        if v is None:
            return ""
        if isinstance(v, (_dt.date, _dt.datetime)):
            return v.strftime("%Y-%m-%d")
        s = str(v).strip()
        try:
            return f"{float(s):.2f}"
        except (TypeError, ValueError):
            return s

    mismatched = []
    for it in res["auto"]:
        real = rows[it["ledger_row_ref"] - 1]
        assert str(real[col["SO"]]).strip() == it["so"]
        for k, v in it["five_cols"].items():
            if v is not None and norm(real[col[k]]) != norm(v):
                mismatched.append((it["case_id"], k, norm(real[col[k]]), norm(v)))
    assert not mismatched, mismatched[:10]


# ══════════════════════════════════════════════════════════
# 七、计提口径（2026-07-23 明妹原话确认）
# ══════════════════════════════════════════════════════════
def test_jiti_only_when_order_fully_paid():
    """
    明妹原话：「一个订单的金额分成多次回款，只有回款明细金额加起来等于交付金额之后，
    才可以填写计提金额」。→ 回满才填计提；没回满**计提留空、但结账仍是「是」**。
    """
    led = _led({1: {"so": "SO1", "sod": "", "yingshou": 500.0}})
    full = C.classify_one(_rec("SO1", "SOD1", 500.0, deliver_local=500.0), led, {}, 0.0, 2026)
    assert full["five_cols"]["计提"] == 500.0
    assert full["five_cols"]["是否结账"] == "是"

    part = C.classify_one(_rec("SO1", "SOD1", 200.0, deliver_local=500.0), led, {}, 0.0, 2026)
    assert part["five_cols"]["计提"] is None
    assert part["five_cols"]["回款明细"] == 200.0
    assert part["five_cols"]["是否结账"] == "是"


def test_jiti_basis_is_zhiyun_deliver_not_ledger_yingshou():
    """
    「回满」的基准是**智云交付额**，不是她表里的应收金额。
    实测 SO26040322 行2567：她表应收 488.64、回款 477.61、智云交付 477.61 → 她填了计提。
    """
    led = _led({1: {"so": "SO1", "sod": "", "yingshou": 488.64}})
    r = C.classify_one(_rec("SO1", "SOD1", 477.61, deliver_local=477.61), led, {}, 0.0, 2026)
    assert r["bucket"] == "auto"
    assert r["five_cols"]["计提"] == 477.61
    assert r["derived_cols"]["差异"] == 11.03


def test_receivable_mismatch_uses_latest_delivery_and_cumulative_writeoff():
    """原始应收只作基线；累计核销达到最新交付额时正常结清并生成业务值差异。"""
    led = _led({1: {
        "so": "SO1", "sod": "", "yingshou": 713.24,
        "chayi": None,
    }})
    r = C.classify_one(
        _rec(
            "SO1",
            "SOD1",
            300.0,
            deliver_local=765.77,
            cumulative_received_local=765.77,
        ),
        led,
        {},
        0.0,
        2026,
    )
    assert r["bucket"] == "auto"
    assert r["five_cols"]["计提"] == 765.77
    assert r["five_cols"]["回款明细"] == 300.0
    assert r["five_cols"]["是否结账"] == "是"
    assert r["derived_cols"]["差异"] == -52.53
