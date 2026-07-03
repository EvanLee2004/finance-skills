#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""社保公积金合并与在职台账 · 回归测试
合成用例（金标，覆盖单行表头/2行合并表头/上海"企业部分"/湖南组合台账 四种真实结构）
+ 真实数据冒烟（薪酬底稿/ 在才跑，缺则跳过，见 tests 末尾）。
跑：python3 tests/test_robustness.py
"""
import os
import sys
import tempfile

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(os.path.dirname(HERE), "scripts"))
import merge_insurance as m  # noqa
import openpyxl as opx

PASS = 0
FAIL = 0


def check_eq(label, got, want):
    global PASS, FAIL
    if got == want:
        PASS += 1
    else:
        FAIL += 1
        print(f"  ✗ {label}: got={got!r} want={want!r}")


def check_true(label, cond):
    check_eq(label, bool(cond), True)


# ----------------- 合成：甲骨易式 2行合并表头社保台账 -----------------
def build_social_2row(wb, sheet_name):
    ws = wb.create_sheet(sheet_name)
    ws.append([f"XX{sheet_name}月份社保台账"])
    ws.append(["序号", "部门", "岗位", "姓名", "医疗生育基数", "工伤基数", "养老失业基数",
               "单位部分", None, None, None, None, None, "个人部分", None, None, None, "五险合计"])
    ws.merge_cells("H2:M2")  # 单位部分
    ws.merge_cells("N2:Q2")  # 个人部分
    ws.append([None, None, None, None, None, None, None,
               "医疗(9%)", "生育(0.8%)", "养老(16%)", "失业(0.5%)", "工伤(0.3%)", "小计",
               "养老(8%)", "医疗(2%+3)", "失业(0.5%)", "小计", None])
    people = [("张三", 644.58, 57.30, 1145.92, 35.81, 21.49, 1905.10, 572.96, 146.24, 35.81, 755.01, 2660.11),
              ("李四", 644.58, 57.30, 1145.92, 35.81, 21.49, 1905.10, 572.96, 146.24, 35.81, 755.01, 2660.11)]
    for i, (name, med, mat, pen, une, inj, sub_u, pen2, med2, une2, sub_p, total) in enumerate(people, 1):
        ws.append([i, "测试部门", "测试岗位", name, 7162, 7162, 7162,
                   med, mat, pen, une, inj, sub_u, pen2, med2, une2, sub_p, total])
    ws.append(["合计"] + [None] * 6 +
              [sum(p[6] for p in people), None, None, None, None, None, None, None, None, None])
    return ws


def build_fund_1row(wb, sheet_name, people):
    """people: [(姓名, 单位月缴存额), ...]，甲骨易式单行表头公积金台账。"""
    ws = wb.create_sheet(sheet_name)
    ws.append(["序号", "姓名", "证件类型", "证件号码", "部门", "缴存基数", "单位月缴存额", "个人月缴存额", "月缴存额", "备注"])
    for i, (name, amt) in enumerate(people, 1):
        ws.append([i, name, "身份证", "110000000000000000", "测试部门", 7000, amt, amt, amt * 2, None])
    return ws


def build_social_shanghai(wb, sheet_name):
    """上海式：'企业部分'(非'单位部分') 2行合并表头。"""
    ws = wb.create_sheet(sheet_name)
    ws.append(["上海测试社保"])
    ws.append(["序号", "姓名", "缴存基数", "企业部分", None, None, None, None, "个人部分", None, None, None, "五险合计"])
    ws.merge_cells("D2:H2")
    ws.merge_cells("I2:L2")
    ws.append([None, None, None, "养老(16%)", "医疗+生育(9%)", "工伤(0.2%)", "失业(0.5%)", "小计",
                "养老(8%)", "医疗(2%)", "失业(0.5%)", "小计", None])
    ws.append([1, "王五", 6000, 960, 540, 12, 30, 1542, 480, 120, 30, 630, 2172])
    return ws


def build_combined_hunan(wb, sheet_name):
    """湖南式：社保+公积金合并在同一张表，社保用显式'单位小计'字面量、公积金组内嵌套'单位'子列。"""
    ws = wb.create_sheet(sheet_name)
    ws.append(["湖南测试五险一金台账"])
    ws.append(["部门", "姓名", "基数", "工伤", None, "失业", None, "养老", None, "医疗", None,
               "单位小计", "个人小计", "五险合计", "公积金", None, None, "五险一金合计单位"])
    ws.merge_cells("D2:E2")
    ws.merge_cells("F2:G2")
    ws.merge_cells("H2:I2")
    ws.merge_cells("J2:K2")
    ws.merge_cells("O2:Q2")  # "公积金" 位于第15列(O)，跨"基数/单位/个人"3个子列(O:Q)
    ws.append([None, None, None, "单位", "个人", "单位", "个人", "单位", "个人", "单位", "个人",
               None, None, None, "基数", "单位", "个人", None])
    ws.append(["测试部门", "赵六", 8000, 72, 24, 56, 24, 1280, 640, 720, 160, 2128, 848, 2976,
               7000, 350, 350, 2478])
    return ws


def run_synthetic_tests():
    print("== 合成金标：4种真实表头结构的单位金额列识别 ==")
    wb = opx.Workbook()
    wb.remove(wb.active)

    ws1 = build_social_2row(wb, "社保2row")
    social_map, diag1 = m.parse_social_sheet(ws1)
    check_eq("2行合并表头-张三单位社保小计", social_map.get("张三"), 1905.10)
    check_eq("2行合并表头-李四单位社保小计", social_map.get("李四"), 1905.10)
    check_true("2行合并表头-诊断信息提到'单位部分'", "单位部分" in diag1 or "小计" in diag1)

    ws2 = build_fund_1row(wb, "公积金1row", [("张三", 350.0), ("李四", 350.0)])
    fund_map, diag2 = m.parse_fund_sheet(ws2)
    check_eq("单行表头-张三单位公积金", fund_map.get("张三"), 350.0)

    ws3 = build_social_shanghai(wb, "上海社保")
    social_sh, diag3 = m.parse_social_sheet(ws3)
    check_eq("上海'企业部分'-王五单位社保小计", social_sh.get("王五"), 1542.0)

    ws4 = build_combined_hunan(wb, "湖南组合")
    social_hn, fund_hn, diag4 = m.parse_combined_sheet(ws4)
    check_eq("湖南组合台账-赵六单位社保小计", social_hn.get("赵六"), 2128.0)
    check_eq("湖南组合台账-赵六单位公积金", fund_hn.get("赵六"), 350.0)

    # 列识别不认位置只认内容：把湖南表整体右移3列，结果应该完全一样
    wb2 = opx.Workbook()
    wb2.remove(wb2.active)
    ws5 = wb2.create_sheet("shifted")
    for row in ws4.iter_rows(values_only=True):
        ws5.append([None, None, None] + list(row))
    for rng in list(ws4.merged_cells.ranges):
        ws5.merge_cells(start_row=rng.min_row, start_column=rng.min_col + 3,
                         end_row=rng.max_row, end_column=rng.max_col + 3)
    social_shift, fund_shift, _ = m.parse_combined_sheet(ws5)
    check_eq("列整体右移后仍能识别-单位社保", social_shift.get("赵六"), 2128.0)
    check_eq("列整体右移后仍能识别-单位公积金", fund_shift.get("赵六"), 350.0)


def run_month_normalize_tests():
    print("== 月份sheet名标准化 ==")
    check_eq("2026-06", m.normalize_month_token("2026-06"), "202606")
    check_eq("202606", m.normalize_month_token("202606"), "202606")
    check_eq("2026-6(不补零)", m.normalize_month_token("2026-6"), "202606")
    check_eq("202606甲骨易工资(带后缀)", m.normalize_month_token("202606甲骨易工资"), "202606")
    check_eq("2026-01减少", m.normalize_month_token("2026-01减少"), "202601")


def run_service_status_tests():
    print("== 在职状态推算（发工资=在职） ==")
    rows = [
        dict(姓名="A", 基本工资=1000, 月份="202601", 所属主体="X"),
        dict(姓名="A", 基本工资=1000, 月份="202602", 所属主体="X"),
        dict(姓名="A", 基本工资=None, 月份="202603", 所属主体="X"),  # 最新月无工资 -> 疑似离职
        dict(姓名="B", 基本工资=1000, 月份="202601", 所属主体="X"),
        dict(姓名="B", 基本工资=None, 月份="202602", 所属主体="X"),
        dict(姓名="B", 基本工资=1000, 月份="202603", 所属主体="X"),  # 最新月有工资 -> 在职，中间断档不影响判定
    ]
    out = m.compute_service_status(rows)
    a3 = [r for r in out if r["姓名"] == "A" and r["月份"] == "202603"][0]
    check_eq("A在职月份(2个月有工资)", a3["在职月份"], 2)
    check_true("A最新月无工资标'离职'", "离职" in a3["人员状态"])
    b3 = [r for r in out if r["姓名"] == "B" and r["月份"] == "202603"][0]
    check_eq("B在职月份(2个月有工资,含断档)", b3["在职月份"], 2)
    check_eq("B最新月有工资不应标离职", b3["人员状态"], "")


def run_real_smoke_tests():
    """真实数据冒烟：用湖南分公司真实数据链路，核对真实已知数值。缺文件则跳过。"""
    print("== 真实数据冒烟（缺文件自动跳过） ==")
    payroll_out = os.path.abspath(os.path.join(
        HERE, "..", "..", "..", "..", "技能", "社保公积金合并与在职台账", "测试数据", "工资底表_202606.xlsx"))
    combined = os.path.abspath(os.path.join(
        HERE, "..", "..", "..", "..", "技能", "社保公积金合并与在职台账", "测试数据", "薪酬底稿",
        "6. 湖南", "湖南分公司五险一金台账202606.xlsx"))
    if not (os.path.isfile(payroll_out) and os.path.isfile(combined)):
        print("  （跳过：本机没有真实测试数据，属正常，不算失败）")
        return
    with tempfile.TemporaryDirectory() as td:
        out = os.path.join(td, "out.xlsx")
        m.run("湖南分公司", "202606", payroll_out, None, None, combined, None, out)
        wb = opx.load_workbook(out, data_only=True)
        ws = wb["薪酬明细"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        header = [c.value for c in ws[1]]
        idx = {h: i for i, h in enumerate(header)}
        xieke = [r for r in rows if r[idx["姓名"]] == "谢柯"]
        check_true("真实数据-谢柯有匹配到记录", len(xieke) == 1)
        if xieke:
            check_eq("真实数据-谢柯单位社保小计(与源台账核对)", xieke[0][idx["单位社保"]], 1133.01)
            check_eq("真实数据-谢柯单位公积金(与源台账核对)", xieke[0][idx["单位住房公积金"]], 105.0)
        check_eq("真实数据-82人全部写入薪酬明细", len(rows), 82)


if __name__ == "__main__":
    run_synthetic_tests()
    run_month_normalize_tests()
    run_service_status_tests()
    run_real_smoke_tests()
    print(f"\n{'='*40}\nPASS={PASS} FAIL={FAIL}")
    sys.exit(1 if FAIL else 0)
