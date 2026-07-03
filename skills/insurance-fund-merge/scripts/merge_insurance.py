#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
社保公积金合并与在职台账
把各主体每月的社保台账 + 公积金台账（或湖南这类"五险一金"合并台账）里的【单位部分】金额，
按姓名匹配进当月工资表，追加进跨月累计的『薪酬明细』，并按"发工资=在职"的口径
自动推算每个人的在职月份数/人员状态，同时维护一张只增不删的组织架构名册、拆出编外人员。

输入是一张干净工资表，这个技能把社保/公积金"单位缴纳部分"合并进去、再叠加累计台账逻辑。

用法：
  python3 merge_insurance.py --inspect --input-dir DIR
      只认文件、猜社保/公积金/组合台账及月份覆盖，不跑。

  python3 merge_insurance.py --entity 甲骨易 --month 202606 \
      --payroll <该主体该月工资表.xlsx> \
      --insurance <社保台账.xlsx> --fund <公积金台账.xlsx> \
      [--master <累计薪酬台账.xlsx>] [--out <输出.xlsx>]

  湖南这类"五险一金"社保+公积金合并在同一份文件里的，用 --combined 代替 --insurance/--fund：
  python3 merge_insurance.py --entity 湖南分公司 --month 202606 \
      --payroll <工资表.xlsx> --combined <五险一金台账.xlsx> [--master ...] [--out ...]

规则（在职判定口径/编外人员名单/新增主体怎么配）在 config/业务规则.md、config/主体配置.json，改表不改码。

⚠ 数据敏感：含身份证号/工资/社保金额 PII，只在本地跑、结果别外传。
"""
import os
import re
import sys
import json
import argparse
from collections import defaultdict, Counter

import openpyxl

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

HERE = os.path.dirname(os.path.abspath(__file__))
SKILL_DIR = os.path.dirname(HERE)
CONFIG_DIR = os.path.join(SKILL_DIR, "config")
WORK_INPUT = os.path.join(SKILL_DIR, "工作区", "input")
WORK_OUTPUT = os.path.join(SKILL_DIR, "工作区", "output")

MASTER_DETAIL_COLS = ["序号", "姓名", "地区", "组织架构1", "组织架构2", "基本工资",
                       "单位社保", "单位住房公积金", "个人所得税", "实发工资",
                       "月份", "在职月份", "人员状态", "所属主体"]
EXTRA_COLS = ["月份", "姓名", "基本工资", "单位社保", "单位住房公积金", "所属主体", "备注"]

CONFIG = {
    "IN_SERVICE_RULE": "基本工资非空即视为该月在职",
    "HEADER_MAX_SCAN": 6,
}


def log(msg):
    print(msg, file=sys.stderr)


# ----------------- 配置加载 -----------------
def load_entity_config():
    p = os.path.join(CONFIG_DIR, "主体配置.json")
    if os.path.isfile(p):
        try:
            with open(p, encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            log(f"⚠ 读 主体配置.json 失败({e})，用空配置。")
    return {"已配置主体": {}, "未配置主体": []}


def load_excluded_names():
    """编外人员名单：她口头维护的固定名单，见 config/编外人员名单.json。"""
    p = os.path.join(CONFIG_DIR, "编外人员名单.json")
    if os.path.isfile(p):
        try:
            with open(p, encoding="utf-8") as f:
                d = json.load(f)
            return set(d.get("姓名列表", []))
        except Exception as e:
            log(f"⚠ 读 编外人员名单.json 失败({e})，按空名单处理。")
    return set()


# ----------------- 通用工具 -----------------
def _s(v):
    return "" if v is None else str(v).strip()


def to_number(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    s = str(v).strip().replace(",", "").replace("，", "")
    if s in ("", "-", "#N/A", "NA", "nan", "None", "#REF!", "#VALUE!"):
        return None
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def normalize_month_token(sheet_name):
    """把各种花样的月份 sheet 名（2026-01 / 202601 / 2026-1 / 202606甲骨易工资 等）标准化成 YYYYMM；抓不到返回 None。"""
    digits_dash = re.findall(r"(\d{4})[-/]?(\d{1,2})(?!\d)", sheet_name)
    for y, m in digits_dash:
        if 2000 <= int(y) <= 2099 and 1 <= int(m) <= 12:
            return f"{y}{int(m):02d}"
    m6 = re.search(r"(\d{6})", sheet_name)
    if m6:
        y, mm = m6.group(1)[:4], m6.group(1)[4:6]
        if 2000 <= int(y) <= 2099 and 1 <= int(mm) <= 12:
            return f"{y}{mm}"
    return None


def find_month_sheet(wb, yyyymm):
    """按月份找 sheet；同月多个候选（如「202606甲骨易工资」还混进汇总表）时，优先纯月份命名的那个。"""
    exact = [s for s in wb.sheetnames if s.strip() in (yyyymm, f"{yyyymm[:4]}-{yyyymm[4:]}", f"{yyyymm[:4]}-{int(yyyymm[4:])}")]
    if exact:
        return exact[0]
    cands = [s for s in wb.sheetnames if normalize_month_token(s) == yyyymm]
    if not cands:
        return None
    cands.sort(key=len)  # 命名越短越像"就是这个月份本身"，越长越像"月份+其他说明"
    return cands[0]


# ----------------- 表头识别（处理2行合并表头 + 单行表头两种情况，通用不认位置只认内容） -----------------
def _merged_ranges_in_row(ws, row_idx):
    out = []
    for rng in ws.merged_cells.ranges:
        if rng.min_row == row_idx and rng.max_row == row_idx:
            out.append((rng.min_col, rng.max_col))
    return out


def _row_values_filled(ws, row_idx, width):
    """读一行的值，并把该行内的横向合并单元格向右补齐（合并单元格 openpyxl 只在左上角有值）。"""
    vals = [ws.cell(row_idx, c + 1).value for c in range(width)]
    for c0, c1 in _merged_ranges_in_row(ws, row_idx):
        top = ws.cell(row_idx, c0).value
        for c in range(c0, c1 + 1):
            if 1 <= c <= width:
                vals[c - 1] = top
    return vals


def locate_header_block(ws, name_aliases=("姓名",), max_scan=None, min_width_hint=3):
    """
    找到"姓名"所在的表头行(anchor)，再往下看后续 1~2 行是不是还是表头(该行"姓名"列为空)，
    把这些行的文本按列纵向拼接成 combined_labels，返回 (data_start_row_1based, combined_labels, anchor_row_values)。
    找不到返回 None。
    """
    max_scan = max_scan or CONFIG["HEADER_MAX_SCAN"]
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_col == 0:
        return None
    width = max_col
    for anchor in range(1, min(max_scan, max_row) + 1):
        row_vals = _row_values_filled(ws, anchor, width)
        cells = [_s(v) for v in row_vals]
        w = len(cells)
        while w > 0 and cells[w - 1] == "":
            w -= 1
        if w < min_width_hint:
            continue
        name_idx = next((i for i, c in enumerate(cells[:w]) if c in name_aliases), None)
        if name_idx is None:
            continue
        # 找到了"姓名"所在行；再往下最多看2行，只要该行同一列为空就当成子表头继续拼接
        header_rows = [row_vals[:w]]
        r = anchor + 1
        while r <= min(anchor + 2, max_row):
            sub_vals = _row_values_filled(ws, r, w)
            if _s(sub_vals[name_idx]) == "":
                header_rows.append(sub_vals)
                r += 1
            else:
                break
        combined = []
        for i in range(w):
            parts = [_s(hr[i]) for hr in header_rows if _s(hr[i])]
            combined.append("".join(parts))
        return dict(data_start=r, combined=combined, name_idx=name_idx, width=w,
                    anchor_row=anchor, header_rows=header_rows)
    return None


def find_col_by_keywords(combined, must_all=(), must_not=(), prefer_leftmost=True):
    idxs = []
    for i, label in enumerate(combined):
        if all(k in label for k in must_all) and not any(k in label for k in must_not):
            idxs.append(i)
    if not idxs:
        return None
    return idxs[0] if prefer_leftmost else idxs[-1]


def find_social_unit_col(combined):
    """单位/企业部分小计列。'单位'（多数主体）或'企业'（上海用'企业部分'）二选一命中，排除公积金列。"""
    idx = find_col_by_keywords(combined, must_all=["单位", "小计"], must_not=["公积金"])
    if idx is None:
        idx = find_col_by_keywords(combined, must_all=["企业", "小计"], must_not=["公积金"])
    return idx


def find_fund_unit_col(combined):
    """单位公积金缴存列。独立公积金台账用'单位'+'缴存'；湖南组合台账用'公积金'+'单位'。"""
    idx = find_col_by_keywords(combined, must_all=["单位", "缴存"])
    if idx is None:
        idx = find_col_by_keywords(combined, must_all=["公积金", "单位"])
    return idx


# ----------------- 解析社保 / 公积金 / 组合五险一金 sheet -----------------
def parse_social_sheet(ws):
    """返回 {姓名: 单位社保小计} + 诊断信息。列识别：含'单位'或'企业'且含'小计'（排除公积金）的最靠左列。"""
    hdr = locate_header_block(ws)
    if hdr is None:
        return {}, "没识别到含'姓名'的表头，跳过"
    combined, name_idx, width, data_start = hdr["combined"], hdr["name_idx"], hdr["width"], hdr["data_start"]
    unit_idx = find_social_unit_col(combined)
    if unit_idx is None:
        return {}, f"没找到'单位/企业部分小计'列（表头：{combined}）"
    out = {}
    r = data_start
    max_row = ws.max_row or 0
    blank_streak = 0
    while r <= max_row:
        row = [ws.cell(r, c + 1).value for c in range(width)]
        name = _s(row[name_idx])
        if name and name not in ("合计", "总计", "小计"):
            v = to_number(row[unit_idx])
            if v is not None:
                out[name] = out.get(name, 0.0) + v  # 同名多行(如跨部门)累加，极少见但兜底
            blank_streak = 0
        else:
            if all(c is None for c in row):
                blank_streak += 1
                if blank_streak >= 10:
                    break
        r += 1
    return out, f"识别列：单位社保小计=第{unit_idx + 1}列({combined[unit_idx]})"


def parse_fund_sheet(ws):
    """返回 {姓名: 单位月缴存额}。列识别：含'单位'且含'缴存'的最靠左列。"""
    hdr = locate_header_block(ws)
    if hdr is None:
        return {}, "没识别到含'姓名'的表头，跳过"
    combined, name_idx, width, data_start = hdr["combined"], hdr["name_idx"], hdr["width"], hdr["data_start"]
    unit_idx = find_fund_unit_col(combined)
    if unit_idx is None:
        return {}, f"没找到'单位月缴存额'列（表头：{combined}）"
    out = {}
    r = data_start
    max_row = ws.max_row or 0
    blank_streak = 0
    while r <= max_row:
        row = [ws.cell(r, c + 1).value for c in range(width)]
        name = _s(row[name_idx])
        if name and name not in ("合计", "总计", "小计"):
            v = to_number(row[unit_idx])
            if v is not None:
                out[name] = out.get(name, 0.0) + v
            blank_streak = 0
        else:
            if all(c is None for c in row):
                blank_streak += 1
                if blank_streak >= 10:
                    break
        r += 1
    return out, f"识别列：单位月缴存额=第{unit_idx + 1}列({combined[unit_idx]})"


def parse_combined_sheet(ws):
    """湖南这类"五险一金"合并台账：社保单位小计 + 公积金单位缴存 同一张表。"""
    hdr = locate_header_block(ws)
    if hdr is None:
        return {}, {}, "没识别到含'姓名'的表头，跳过"
    combined, name_idx, width, data_start = hdr["combined"], hdr["name_idx"], hdr["width"], hdr["data_start"]
    social_idx = find_social_unit_col(combined)
    fund_idx = find_fund_unit_col(combined)
    diag = f"识别列：单位社保小计=第{(social_idx or -1) + 1}列，单位公积金=第{(fund_idx or -1) + 1}列（表头：{combined}）"
    social, fund = {}, {}
    r = data_start
    max_row = ws.max_row or 0
    blank_streak = 0
    while r <= max_row:
        row = [ws.cell(r, c + 1).value for c in range(width)]
        name = _s(row[name_idx])
        if name and name not in ("合计", "总计", "小计"):
            if social_idx is not None:
                v = to_number(row[social_idx])
                if v is not None:
                    social[name] = social.get(name, 0.0) + v
            if fund_idx is not None:
                v = to_number(row[fund_idx])
                if v is not None:
                    fund[name] = fund.get(name, 0.0) + v
            blank_streak = 0
        else:
            if all(c is None for c in row):
                blank_streak += 1
                if blank_streak >= 10:
                    break
        r += 1
    return social, fund, diag


# ----------------- 工资表（可以是清洗过的干净底表，也可以是原始工资表） -----------------
_PAYROLL_ALIAS = {
    "姓名": ["姓名"],
    "地区": ["地区"],
    "组织架构1": ["组织架构1", "组织架构-1", "部门"],
    "组织架构2": ["组织架构2", "组织架构-2", "岗位"],
    "基本工资": ["基本工资"],
    "个人所得税": ["个人所得税", "个税"],
    "实发工资": ["实发工资"],
}


def _payroll_like_sheets(wb):
    """全表扫一遍，返回所有『含姓名+基本工资表头』的 sheet 名——不止一个时不能瞎猜是哪个主体的。"""
    out = []
    for sn in wb.sheetnames:
        hdr = locate_header_block(wb[sn])
        if hdr and any("基本工资" in c for c in hdr["combined"]):
            out.append(sn)
    return out


def load_payroll_month(path, sheet_hint_month=None, sheet_override=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_override:
        if sheet_override not in wb.sheetnames:
            wb.close()
            raise ValueError(f"--payroll-sheet 指定的「{sheet_override}」在工资表里不存在，现有：{wb.sheetnames}")
        sheet_name = sheet_override
    else:
        candidates = _payroll_like_sheets(wb)
        if len(candidates) == 0:
            wb.close()
            raise ValueError("工资表里没找到任何可识别的 sheet（需含姓名+基本工资表头）")
        elif len(candidates) == 1:
            sheet_name = candidates[0]
        else:
            # 多个像工资表的 sheet（常见于把多个主体堆在一个文件里的"大杂烩"工作簿，
            # 或者是某些清洗工具的输出——"主表/待人工核实/匹配成功"三个 sheet列结构相同、互为子集）。
            month_matched = [sn for sn in candidates if sheet_hint_month and normalize_month_token(sn) == sheet_hint_month]
            if len(month_matched) == 1:
                sheet_name = month_matched[0]
            elif "主表" in candidates:
                # 「主表」通常是完整表，「待人工核实/匹配成功」只是它的子集视图，优先用完整表。
                sheet_name = "主表"
            else:
                wb.close()
                raise ValueError(
                    f"工资表里有{len(candidates)}个像工资表的 sheet（{candidates}），"
                    f"月份『{sheet_hint_month}』也没能唯一锁定，别瞎猜——用 --payroll-sheet 明确指定是哪一个"
                )
    ws = wb[sheet_name]
    hdr = locate_header_block(ws)
    if hdr is None:
        wb.close()
        raise ValueError(f"工资表 sheet「{sheet_name}」没识别到含'姓名'的表头")
    combined, name_idx, width, data_start = hdr["combined"], hdr["name_idx"], hdr["width"], hdr["data_start"]
    col = {}
    for field, aliases in _PAYROLL_ALIAS.items():
        idx = None
        for a in aliases:
            idx = find_col_by_keywords(combined, must_all=[a]) if a not in combined else combined.index(a)
            if idx is not None:
                break
        col[field] = idx
    out = []
    r = data_start
    max_row = ws.max_row or 0
    blank_streak = 0
    while r <= max_row:
        row = [ws.cell(r, c + 1).value for c in range(width)]
        name = _s(row[name_idx]) if name_idx is not None else ""
        if name and name not in ("合计", "总计", "小计"):
            rec = {"姓名": name}
            for field in _PAYROLL_ALIAS:
                idx = col.get(field)
                if idx is None:
                    rec[field] = None
                elif field in ("基本工资", "个人所得税", "实发工资"):
                    rec[field] = to_number(row[idx])
                else:
                    rec[field] = _s(row[idx]) or None
            out.append(rec)
            blank_streak = 0
        else:
            if all(c is None for c in row):
                blank_streak += 1
                if blank_streak >= 10:
                    break
        r += 1
    wb.close()
    return out, sheet_name


# ----------------- 累计主表（跨月薪酬明细 / 组织架构新 / 编外人员） -----------------
def load_master(path):
    detail, org, extra = [], {}, []
    if path and os.path.isfile(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        if "薪酬明细" in wb.sheetnames:
            ws = wb["薪酬明细"]
            hdr = locate_header_block(ws, name_aliases=("姓名",))
            if hdr:
                combined, name_idx, width, data_start = hdr["combined"], hdr["name_idx"], hdr["width"], hdr["data_start"]
                col = {c: (combined.index(c) if c in combined else None) for c in MASTER_DETAIL_COLS}
                r = data_start
                while r <= (ws.max_row or 0):
                    row = [ws.cell(r, c + 1).value for c in range(width)]
                    name = _s(row[name_idx]) if name_idx is not None else ""
                    if name:
                        rec = {}
                        for c in MASTER_DETAIL_COLS:
                            idx = col.get(c)
                            rec[c] = row[idx] if idx is not None else None
                        detail.append(rec)
                    r += 1
        if "组织架构新" in wb.sheetnames:
            ws = wb["组织架构新"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                name = _s(row[0])
                if name:
                    org[name] = dict(地区=row[1] if len(row) > 1 else None,
                                      组织架构1=row[2] if len(row) > 2 else None,
                                      组织架构2=row[3] if len(row) > 3 else None)
        if "编外人员" in wb.sheetnames:
            ws = wb["编外人员"]
            hdr_row = [_s(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])]
            idx = {c: hdr_row.index(c) for c in EXTRA_COLS if c in hdr_row}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and idx.get("姓名") is not None and row[idx["姓名"]]:
                    extra.append({c: (row[idx[c]] if c in idx else None) for c in EXTRA_COLS})
        wb.close()
    return detail, org, extra


# ----------------- 在职状态推算 -----------------
def compute_service_status(detail_rows):
    """
    口径（明妹口述、录屏演示确认，见 config/业务规则.md）：某月『基本工资』非空 = 该月在职。
    在职月份 = 该人历史所有月份里『基本工资』非空的月数。
    人员状态：仅在"该人有过非空月份、但最新一个已处理月份基本工资为空"时标『离职』，供人工核对，
    ⚠ 不是权威离职记录——入职当月/离职当月/补发工资等边界情形未验证，人工最终确认。
    """
    by_name = defaultdict(list)
    for rec in detail_rows:
        by_name[rec["姓名"]].append(rec)
    months_all = sorted(set(rec["月份"] for rec in detail_rows if rec.get("月份")))
    for name, rows in by_name.items():
        rows.sort(key=lambda r: r.get("月份") or "")
        in_service_months = sum(1 for r in rows if to_number(r.get("基本工资")) is not None)
        latest = rows[-1] if rows else None
        status = ""
        if latest is not None and to_number(latest.get("基本工资")) is None and in_service_months > 0:
            status = "离职(推算，待核实)"
        for r in rows:
            r["在职月份"] = in_service_months
            r["人员状态"] = status
    return detail_rows


# ----------------- 主流程 -----------------
def run(entity, month, payroll_path, insurance_path, fund_path, combined_path,
        master_path, out_path, payroll_sheet=None):
    excluded_names = load_excluded_names()
    warns = []

    # 1) 社保/公积金单位部分
    if combined_path:
        wb = openpyxl.load_workbook(combined_path, data_only=True)
        sn = find_month_sheet(wb, month)
        if sn is None:
            log(f"✗ 组合台账「{combined_path}」里没找到月份 {month} 对应的 sheet，现有：{wb.sheetnames}")
            sys.exit(2)
        social_map, fund_map, diag = parse_combined_sheet(wb[sn])
        wb.close()
        warns.append(f"组合台账[{sn}]: {diag}")
    else:
        social_map, fund_map = {}, {}
        if insurance_path:
            wb = openpyxl.load_workbook(insurance_path, data_only=True)
            sn = find_month_sheet(wb, month)
            if sn is None:
                log(f"✗ 社保台账「{insurance_path}」里没找到月份 {month} 对应的 sheet，现有：{wb.sheetnames}")
                sys.exit(2)
            social_map, diag = parse_social_sheet(wb[sn])
            wb.close()
            warns.append(f"社保台账[{sn}]: {diag}")
        if fund_path:
            wb = openpyxl.load_workbook(fund_path, data_only=True)
            sn = find_month_sheet(wb, month)
            if sn is None:
                log(f"✗ 公积金台账「{fund_path}」里没找到月份 {month} 对应的 sheet，现有：{wb.sheetnames}")
                sys.exit(2)
            fund_map, diag = parse_fund_sheet(wb[sn])
            wb.close()
            warns.append(f"公积金台账[{sn}]: {diag}")

    # 2) 工资表
    try:
        payroll_rows, matched_sheet = load_payroll_month(payroll_path, month, sheet_override=payroll_sheet)
    except ValueError as e:
        log(f"✗ {e}")
        sys.exit(2)
    warns.append(f"工资表命中 sheet「{matched_sheet}」，{len(payroll_rows)} 人")

    # 3) 累计主表（组织架构新 / 历史薪酬明细 / 编外人员）
    old_detail, org_master, extra_rows = load_master(master_path)
    # 幂等：先剔除本次要重新写入的 月份+主体，防止重复跑重复累加
    old_detail = [r for r in old_detail if not (r.get("月份") == month and r.get("所属主体") == entity)]
    extra_rows = [r for r in extra_rows if not (r.get("月份") == month and r.get("所属主体") == entity)]

    # 4) 逐人处理：拆编外人员 / 待人工核实 / 正常合并
    new_detail_rows = []
    to_review = []
    new_org_entries = []
    for i, p in enumerate(payroll_rows, start=1):
        name = p["姓名"]
        social = social_map.get(name)
        fund = fund_map.get(name)
        if name in excluded_names:
            extra_rows.append(dict(月份=month, 姓名=name, 基本工资=p.get("基本工资"),
                                    单位社保=social, 单位住房公积金=fund, 所属主体=entity,
                                    备注="config/编外人员名单.json 里标记的编外人员"))
            continue
        # 组织架构：只增不删，已存在的人一律沿用旧记录里的地区/组织架构（权威、不被本月工资表覆盖）
        if name in org_master:
            region = org_master[name].get("地区") or p.get("地区")
            org1 = org_master[name].get("组织架构1") or p.get("组织架构1")
            org2 = org_master[name].get("组织架构2") or p.get("组织架构2")
        else:
            region, org1, org2 = p.get("地区"), p.get("组织架构1"), p.get("组织架构2")
            org_master[name] = dict(地区=region, 组织架构1=org1, 组织架构2=org2)
            new_org_entries.append(name)
        rec = dict(序号=i, 姓名=name, 地区=region, 组织架构1=org1, 组织架构2=org2,
                   基本工资=p.get("基本工资"), 单位社保=social, 单位住房公积金=fund,
                   个人所得税=p.get("个人所得税"), 实发工资=p.get("实发工资"),
                   月份=month, 在职月份=None, 人员状态="", 所属主体=entity)
        new_detail_rows.append(rec)
        if social is None or fund is None:
            missing = []
            if social is None:
                missing.append("单位社保未匹配到")
            if fund is None:
                missing.append("单位公积金未匹配到")
            to_review.append((name, "；".join(missing)))

    # 社保/公积金台账里出现、但工资表里没有的人（编外/离职/劳务等）单独列出，不当0处理
    payroll_names = {p["姓名"] for p in payroll_rows}
    for name in sorted(set(social_map) | set(fund_map)):
        if name not in payroll_names and name not in excluded_names:
            to_review.append((name, "社保/公积金台账里有此人，但当月工资表没有——可能是编外/离职/劳务，人工核实"))

    all_detail = old_detail + new_detail_rows
    all_detail = compute_service_status(all_detail)
    # 只重排"这个月"的记录用于展示；历史记录状态字段也已被上面同步刷新
    new_detail_rows = [r for r in all_detail if r.get("月份") == month and r.get("所属主体") == entity]

    for x in warns:
        log(f"· {x}")
    log(f"· 本月{entity} {month}：{len(new_detail_rows)} 人写入薪酬明细；新增组织架构 {sum(1 for x in new_org_entries if x)} 人；"
        f"待人工核实 {len(to_review)} 条")

    write_output(out_path, all_detail, org_master, extra_rows, to_review, entity, month, warns)
    return out_path


# ----------------- 输出 -----------------
def _style(ws, ncol):
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    hfont = Font(name="等线", size=10.5, bold=True)
    hfill = PatternFill("solid", fgColor="BDD7EE")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(*[Side(style="thin", color="999999")] * 4)
    for i in range(1, ncol + 1):
        c = ws.cell(1, i)
        c.font = hfont
        c.fill = hfill
        c.alignment = align
        c.border = border
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"


def build_summary_wide(all_detail):
    """按『薪酬汇总N月』真实历史产物的列结构（组织架构1/2/地区/在职月份/人员状态/姓名/基本工资汇总/单位社保汇总/
    单位住房公积金汇总/总计/[各月×(基本工资/单位社保/单位住房公积金/汇总)]）重建宽表，纯数据，不做颜色仿真。"""
    months = sorted(set(r["月份"] for r in all_detail if r.get("月份")))
    by_name = defaultdict(dict)
    meta = {}
    for r in all_detail:
        by_name[r["姓名"]][r["月份"]] = r
        meta[r["姓名"]] = (r.get("组织架构1"), r.get("组织架构2"), r.get("地区"), r.get("在职月份"), r.get("人员状态"))
    rows = []
    for name in sorted(by_name):
        org1, org2, region, svc_months, status = meta[name]
        base_sum = sum((to_number(by_name[name].get(m, {}).get("基本工资")) or 0) for m in months)
        social_sum = sum((to_number(by_name[name].get(m, {}).get("单位社保")) or 0) for m in months)
        fund_sum = sum((to_number(by_name[name].get(m, {}).get("单位住房公积金")) or 0) for m in months)
        row = dict(组织架构1=org1, 组织架构2=org2, 地区=region, 在职月份=svc_months, 人员状态=status,
                   姓名=name, 基本工资汇总=round(base_sum, 2), 单位社保汇总=round(social_sum, 2),
                   单位住房公积金汇总=round(fund_sum, 2), 总计=round(base_sum + social_sum + fund_sum, 2))
        for m in months:
            rec = by_name[name].get(m, {})
            b, s, f = to_number(rec.get("基本工资")), to_number(rec.get("单位社保")), to_number(rec.get("单位住房公积金"))
            row[f"{m}基本工资"] = b
            row[f"{m}单位社保"] = s
            row[f"{m}单位住房公积金"] = f
            row[f"{m}汇总"] = round((b or 0) + (s or 0) + (f or 0), 2) if (b, s, f) != (None, None, None) else None
        rows.append(row)
    cols = ["组织架构1", "组织架构2", "地区", "在职月份", "人员状态", "姓名",
            "基本工资汇总", "单位社保汇总", "单位住房公积金汇总", "总计"]
    for m in months:
        cols += [f"{m}基本工资", f"{m}单位社保", f"{m}单位住房公积金", f"{m}汇总"]
    return rows, cols


def write_output(out_path, all_detail, org_master, extra_rows, to_review, entity, month, warns):
    import pandas as pd
    detail_df = pd.DataFrame(all_detail, columns=MASTER_DETAIL_COLS)
    detail_df = detail_df.sort_values(["月份", "所属主体", "姓名"], kind="stable").reset_index(drop=True)
    detail_df["序号"] = range(1, len(detail_df) + 1)

    org_rows = [dict(姓名=n, 地区=v.get("地区"), 组织架构1=v.get("组织架构1"), 组织架构2=v.get("组织架构2"))
                for n, v in sorted(org_master.items())]
    org_df = pd.DataFrame(org_rows, columns=["姓名", "地区", "组织架构1", "组织架构2"])

    extra_df = pd.DataFrame(extra_rows, columns=EXTRA_COLS)

    review_df = pd.DataFrame(to_review, columns=["姓名", "问题"])

    summary_rows, summary_cols = build_summary_wide(all_detail)
    summary_df = pd.DataFrame(summary_rows, columns=summary_cols)

    report_lines = [("处理主体", entity, ""), ("处理月份", month, ""),
                     ("薪酬明细累计人次", len(all_detail), ""),
                     ("组织架构名册人数", len(org_master), ""),
                     ("待人工核实条数", len(to_review), "")]
    for w in warns:
        report_lines.append(("诊断", w, ""))
    report_df = pd.DataFrame(report_lines, columns=["项目", "值", "备注"])

    os.makedirs(os.path.dirname(os.path.abspath(out_path)) or ".", exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        detail_df.to_excel(w, sheet_name="薪酬明细", index=False)
        org_df.to_excel(w, sheet_name="组织架构新", index=False)
        extra_df.to_excel(w, sheet_name="编外人员", index=False)
        review_df.to_excel(w, sheet_name="待人工核实", index=False)
        summary_df.to_excel(w, sheet_name="薪酬汇总", index=False)
        report_df.to_excel(w, sheet_name="运行报告", index=False)
        for sn, df in [("薪酬明细", detail_df), ("组织架构新", org_df), ("编外人员", extra_df),
                       ("待人工核实", review_df), ("薪酬汇总", summary_df), ("运行报告", report_df)]:
            _style(w.sheets[sn], len(df.columns))
    log(f"· 已写出 → {out_path}（6个sheet：薪酬明细 / 组织架构新 / 编外人员 / 待人工核实 / 薪酬汇总 / 运行报告）")
    return out_path


# ----------------- inspect 模式 -----------------
def inspect_mode(input_dir):
    ecfg = load_entity_config()
    print(f"识别输入目录：{input_dir}")
    if not os.path.isdir(input_dir):
        print("  目录不存在")
        return
    for fn in sorted(os.listdir(input_dir)):
        if not fn.lower().endswith((".xlsx",)) or fn.startswith("~$") or fn.startswith(".~"):
            continue
        path = os.path.join(input_dir, fn)
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
        except Exception as e:
            print(f"  {fn}: 打不开({e})")
            continue
        print(f"\n  {fn}  [{len(wb.sheetnames)} sheets]")
        sample_sheets = wb.sheetnames[:3]
        for sn in sample_sheets:
            ws = wb[sn]
            hdr = locate_header_block(ws)
            if hdr is None:
                print(f"    - {sn}: 未识别到姓名表头")
                continue
            combined = hdr["combined"]
            has_social = find_social_unit_col(combined) is not None
            has_fund = find_fund_unit_col(combined) is not None
            month = normalize_month_token(sn)
            kind = []
            if has_social:
                kind.append("含社保单位小计列")
            if has_fund:
                kind.append("含公积金单位列")
            print(f"    - {sn}（推测月份{month}）: {', '.join(kind) or '未识别到单位金额列'}")
        wb.close()
    known = list(ecfg.get("已配置主体", {}).keys())
    unknown = next((v for k, v in ecfg.items() if k.startswith("未配置主体")), [])
    print(f"\n已在 config/主体配置.json 校验过真实数据的主体：{known}")
    print(f"格式未校验、跑之前先 --inspect 确认识别列对不对：{unknown}")


def main():
    ap = argparse.ArgumentParser(description="社保公积金合并与在职台账")
    ap.add_argument("--entity")
    ap.add_argument("--month")
    ap.add_argument("--payroll")
    ap.add_argument("--payroll-sheet", dest="payroll_sheet",
                     help="工资表里有多个像工资表的sheet、月份又锁不定时，明确指定用哪个sheet")
    ap.add_argument("--insurance")
    ap.add_argument("--fund")
    ap.add_argument("--combined")
    ap.add_argument("--master")
    ap.add_argument("--out")
    ap.add_argument("--input-dir", default=WORK_INPUT)
    ap.add_argument("--inspect", action="store_true")
    a = ap.parse_args()

    if a.inspect:
        inspect_mode(a.input_dir)
        return

    missing = [n for n, v in [("--entity", a.entity), ("--month", a.month), ("--payroll", a.payroll)] if not v]
    if missing:
        log(f"✗ 缺必填参数：{missing}")
        sys.exit(2)
    if not a.combined and not (a.insurance and a.fund):
        log("✗ 要么传 --combined（湖南这类合并台账），要么同时传 --insurance 和 --fund")
        sys.exit(2)

    out = a.out or os.path.join(WORK_OUTPUT, f"薪酬台账_{a.entity}_{a.month}.xlsx")
    if os.path.isdir(out):
        out = os.path.join(out, f"薪酬台账_{a.entity}_{a.month}.xlsx")

    run(a.entity, a.month, a.payroll, a.insurance, a.fund, a.combined, a.master, out,
        payroll_sheet=a.payroll_sheet)


if __name__ == "__main__":
    main()
