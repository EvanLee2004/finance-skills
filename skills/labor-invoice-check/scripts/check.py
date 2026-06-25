#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
劳务发票核对（第一阶段·支付前）
把「待支付清单（国内个人）」和「个人发票统计台账」按身份证号核对：
  按身份证号把同一人多张发票的「合计金额」求和 → 与应付金额比 →
  实习生/外国人豁免、≤800放行、>800缺票或未开票标黄 → 出三表（主核对表/不付名单/可付名单）。

用法：
  python3 check.py --inspect [--input-dir DIR]      # 只认文件、看表头，不跑
  python3 check.py --list 清单.xlsx --invoice 发票.xlsx [--out 结果.xlsx]
  缺 --list/--invoice 时按内容从 --input-dir(默认 工作区/input) 自动认。

规则（门槛/容差/实习生关键字/匹配键）在 config/业务规则.md，改表不改码。
"""
import os
import re
import sys
import json
import argparse
from collections import defaultdict

import pandas as pd
import openpyxl

# Windows GBK 终端下 print 含 ✓ 等符号会 UnicodeEncodeError（数据其实已写好、只是末尾打印崩、看着像失败）。
# 统一把标准输出设成 UTF-8、容错，彻底避免这个吓人的报错。
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

HERE = os.path.dirname(os.path.abspath(__file__))
SKILL_DIR = os.path.dirname(HERE)
CONFIG_DIR = os.path.join(SKILL_DIR, "config")
WORK_INPUT = os.path.join(SKILL_DIR, "工作区", "input")
WORK_OUTPUT = os.path.join(SKILL_DIR, "工作区", "output")

# ----------------- 默认配置（config/业务规则.md 可覆盖） -----------------
CONFIG = {
    "THRESHOLD": 800.0,        # 开票门槛：应付 > 此值才要求开票
    "TOLERANCE": 1.0,          # 容差：|开票总额-应付| <= 此值 视为票款一致
    "INTERN_KEYWORDS": ["Intern", "实习"],
}
HIDE_DEFAULT = ["开户名", "账号", "开户行", "开户支行", "电话", "支付渠道", "支付日期", "身份证号"]


def log(msg):
    print(msg, file=sys.stderr)


# ----------------- 认列别名 -----------------
def load_aliases():
    p = os.path.join(CONFIG_DIR, "列名别名.json")
    if os.path.isfile(p):
        try:
            with open(p, encoding="utf-8") as f:
                d = json.load(f)
            return d.get("待支付清单_列别名", {}), d.get("发票统计_列别名", {})
        except Exception as e:
            log(f"⚠ 读 列名别名.json 失败({e})，用内置默认。")
    return {}, {}

_LIST_ALIAS_DEFAULT = {
    "供应商姓名": ["供应商姓名", "姓名", "供应商", "收款人"],
    "应付金额": ["应付金额", "应付", "金额", "应付款"],
    "备注": ["备注", "类型", "身份", "说明"],
    "身份证号": ["身份证号/护照号", "身份证号", "身份证", "证件号", "护照号", "纳税人识别号"],
}
_INV_ALIAS_DEFAULT = {
    "开票人姓名": ["销售方信息名称", "开票人", "销售方名称", "姓名"],
    "身份证号": ["销售方信息纳税人识别号", "纳税人识别号", "身份证号", "证件号"],
    "发票金额": ["合计金额（元）", "合计金额(元)", "合计金额", "价税合计"],
}


def load_rules():
    """从 config/业务规则.md 的『可调参数』表读门槛/容差/实习生关键字 → 覆盖默认。缺/解析失败不崩。"""
    p = os.path.join(CONFIG_DIR, "业务规则.md")
    if not os.path.isfile(p):
        return
    try:
        text = open(p, encoding="utf-8").read()
    except Exception as e:
        log(f"⚠ 读 业务规则.md 失败({e})，用内置默认。")
        return
    for line in text.splitlines():
        if not line.strip().startswith("|"):
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if len(cells) < 2:
            continue
        key, val = cells[0], cells[1]
        num = re.search(r"-?\d+(?:\.\d+)?", val)
        if key == "开票门槛" and num:
            CONFIG["THRESHOLD"] = float(num.group())
        elif key == "容差" and num:
            CONFIG["TOLERANCE"] = float(num.group())
        elif key == "实习生关键字" and val and "|" not in val:
            CONFIG["INTERN_KEYWORDS"] = [w.strip() for w in re.split(r"[、,，/]", val) if w.strip()]


# ----------------- 工具 -----------------
def norm_id(v):
    """身份证号归一化：去空格、大写（末位 x→X）、去不可见字符。空→''。"""
    if v is None:
        return ""
    s = str(v).strip().upper()
    s = re.sub(r"\s+", "", s)
    return s


def to_number(v):
    """金额转数字，兼容字符串(带逗号)、None。无法转→None。
       datetime/time（'金额变日期'的坏单元格）→ None 并算异常，不瞎猜。"""
    import datetime as _dt
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, (_dt.datetime, _dt.date, _dt.time)):
        return None
    s = str(v).strip().replace(",", "").replace("，", "")
    if s in ("", "-", "#N/A", "NA", "nan", "None"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def is_foreign_name(name):
    """姓名不含中文字符 = 外国人（纯英文名）。"""
    return bool(name) and not re.search(r"[一-鿿]", str(name))


def _find_col(header, aliases):
    """在 header(list) 里按别名找列索引；找不到返回 None。"""
    for a in aliases:
        if a in header:
            return header.index(a)
    return None


def _header_row_with(ws, required_groups, max_scan=6):
    """在 ws 前 max_scan 行里找一行表头：该行需含 required_groups 里**每一组**的至少一个别名。
       返回该行 0-based 索引；找不到返回 None。"""
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if all(any(a in cells for a in group) for group in required_groups):
            return ri
    return None


def _pick_sheet(wb, required_groups):
    """按**列特征**认出目标 sheet（无视 sheet 名/顺序/其它干扰 sheet）：
       挑表头含所有 required_groups 列的 sheet；多个匹配选数据行最多的。
       返回 (sheet名, 表头0-based行号) 或 (None, None)。"""
    best = None  # (rows, name, header_idx)
    for name in wb.sheetnames:
        hr = _header_row_with(wb[name], required_groups)
        if hr is None:
            continue
        try:
            nrows = wb[name].max_row or 0
        except Exception:
            nrows = 0
        if best is None or nrows > best[0]:
            best = (nrows, name, hr)
    return (best[1], best[2]) if best else (None, None)


# ----------------- 读两张表 -----------------
def read_list(path, list_alias):
    """读待支付清单 → list[dict(name,pay,note,idno)]。
       按列特征认 sheet（需含『供应商姓名』+『应付金额』），自动定位表头行，无视其它干扰 sheet。"""
    name_al = list_alias.get("供应商姓名", _LIST_ALIAS_DEFAULT["供应商姓名"])
    pay_al = list_alias.get("应付金额", _LIST_ALIAS_DEFAULT["应付金额"])
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet, hr = _pick_sheet(wb, [name_al, pay_al])
    warns = []
    if sheet is None:  # 没认出含关键列的 sheet → 退回第一个、表头第一行，并告警
        sheet, hr = wb.sheetnames[0], 0
        warns.append(f"清单未按列认出目标 sheet（需含『供应商姓名』+『应付金额』），暂用首个 sheet「{sheet}」")
    ws = wb[sheet]
    rows = list(ws.iter_rows(min_row=hr + 1, values_only=True))
    wb.close()
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx = {}
    for key, default in _LIST_ALIAS_DEFAULT.items():
        aliases = list_alias.get(key, default)
        c = _find_col(header, aliases)
        idx[key] = c
        if c is None and key in ("供应商姓名", "应付金额", "身份证号"):
            warns.append(f"清单缺关键列「{key}」(别名 {aliases})")
    SUMMARY = {"合计", "总计", "小计", "共计", "合 计", "总 计", "total", "Total", "TOTAL"}
    out = []
    for r in rows[1:]:
        if idx["供应商姓名"] is None or r[idx["供应商姓名"]] in (None, ""):
            continue
        nm = str(r[idx["供应商姓名"]]).strip()
        if nm.replace(":", "").replace("：", "") in SUMMARY:   # 跳过表底"合计/总计"汇总行
            continue
        out.append(dict(
            name=str(r[idx["供应商姓名"]]).strip(),
            pay=to_number(r[idx["应付金额"]]) if idx["应付金额"] is not None else None,
            note=str(r[idx["备注"]] or "") if idx["备注"] is not None else "",
            idno=norm_id(r[idx["身份证号"]]) if idx["身份证号"] is not None else "",
        ))
    return out, header, warns


def read_invoices(path, inv_alias):
    """读发票台账 → 按身份证号求和合计金额；同时记一份按姓名(兜底/展示)。
       按列特征认台账 sheet（需含『纳税人识别号』+『合计金额』），**无视 Sheet4/财务核对/Sheet2 等
       手工底稿等干扰 sheet**，也不依赖它叫不叫 Sheet1。"""
    id_al = inv_alias.get("身份证号", _INV_ALIAS_DEFAULT["身份证号"])
    amt_al = inv_alias.get("发票金额", _INV_ALIAS_DEFAULT["发票金额"])
    wb = openpyxl.load_workbook(path, data_only=True)
    warns = []
    sheet, hr = _pick_sheet(wb, [id_al, amt_al])   # 台账=唯一同时含 纳税人识别号+合计金额 的 sheet
    if sheet is None:
        sheet, hr = ("Sheet1" if "Sheet1" in wb.sheetnames else wb.sheetnames[0]), 0
        warns.append(f"发票台账未按列认出目标 sheet（需含『纳税人识别号』+『合计金额』），暂用「{sheet}」")
    ws = wb[sheet]
    rows = list(ws.iter_rows(min_row=hr + 1, values_only=True))
    wb.close()
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    ci_id = _find_col(header, inv_alias.get("身份证号", _INV_ALIAS_DEFAULT["身份证号"]))
    ci_amt = _find_col(header, inv_alias.get("发票金额", _INV_ALIAS_DEFAULT["发票金额"]))
    ci_nm = _find_col(header, inv_alias.get("开票人姓名", _INV_ALIAS_DEFAULT["开票人姓名"]))
    if ci_id is None:
        warns.append(f"发票表缺「身份证号(纳税人识别号)」列(别名 {inv_alias.get('身份证号', _INV_ALIAS_DEFAULT['身份证号'])})")
    if ci_amt is None:
        warns.append(f"发票表缺「合计金额」列(别名 {inv_alias.get('发票金额', _INV_ALIAS_DEFAULT['发票金额'])})")
    sum_by_id = defaultdict(float)
    cnt_by_id = defaultdict(int)
    sum_by_name = defaultdict(float)
    for r in rows[1:]:
        idv = norm_id(r[ci_id]) if ci_id is not None else ""
        amt = to_number(r[ci_amt]) if ci_amt is not None else None
        amt = amt or 0.0
        if idv:
            sum_by_id[idv] += amt
            cnt_by_id[idv] += 1
        if ci_nm is not None and r[ci_nm]:
            sum_by_name[str(r[ci_nm]).strip()] += amt
    return dict(sum_by_id=sum_by_id, cnt_by_id=cnt_by_id, sum_by_name=sum_by_name), header, warns


# ----------------- 核对核心 -----------------
def classify(clist, inv):
    """对每个清单行定状态。返回 list[dict]。纯函数、确定性。"""
    TH = CONFIG["THRESHOLD"]
    TOL = CONFIG["TOLERANCE"]
    INTERN = CONFIG["INTERN_KEYWORDS"]
    out = []
    for c in clist:
        name, pay, note, idno = c["name"], c["pay"], c["note"], c["idno"]
        rec = dict(供应商姓名=name, 应付金额=pay, 备注=note,
                   开票总额=None, 差额=None, 状态="", 核对提示="")
        # 实习生豁免
        if any(k in note for k in INTERN):
            rec["状态"] = "豁免-实习生"
            out.append(rec); continue
        # 外国人豁免（纯英文名）
        if is_foreign_name(name):
            rec["状态"] = "豁免-外国人"
            out.append(rec); continue
        # 应付金额异常
        if pay is None:
            rec["状态"] = "标黄-待人工"
            rec["核对提示"] = "应付金额缺失/非数字，人工核"
            out.append(rec); continue
        # ≤门槛放行
        if pay <= TH:
            rec["状态"] = "可付"
            rec["核对提示"] = f"≤{TH:g}不要求开票"
            out.append(rec); continue
        # >门槛：按身份证号求和发票
        matched_cnt = inv["cnt_by_id"].get(idno, 0) if idno else 0
        inv_sum = round(inv["sum_by_id"].get(idno, 0.0), 2) if idno else 0.0
        rec["开票总额"] = inv_sum
        rec["差额"] = round(inv_sum - pay, 2)
        if inv_sum <= 0:
            rec["状态"] = "标黄-未开票"
            if matched_cnt > 0:
                rec["核对提示"] = "身份证匹到票但合计为0，疑串票/录入0，人工核"
            elif not idno:
                rec["核对提示"] = "无身份证号(疑外籍/护照)，人工核"
            else:
                rec["核对提示"] = "台账无此人发票"
        elif inv_sum >= pay - TOL:
            rec["状态"] = "可付"
            if rec["差额"] > TOL:
                rec["核对提示"] = "开票多于应付(多开)，多开差异留第二轮"
        else:
            rec["状态"] = "标黄-缺票"
            rec["核对提示"] = f"差{abs(rec['差额']):.2f}，待补票"
        out.append(rec)
    return out


# ----------------- 输出 -----------------
def _style(ws, ncol, yellow_row_idx):
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    hfont = Font(name="等线", size=10.5, bold=True)
    hfill = PatternFill("solid", fgColor="BDD7EE")
    yfill = PatternFill("solid", fgColor="FFFF00")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(*[Side(style="thin", color="999999")] * 4)
    for i in range(1, ncol + 1):
        c = ws.cell(1, i)
        c.font = hfont; c.fill = hfill; c.alignment = align; c.border = border
    for ri in yellow_row_idx:  # 1-based 数据行号(含表头偏移)
        for i in range(1, ncol + 1):
            ws.cell(ri, i).fill = yfill
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"


def write_output(out_path, records):
    df = pd.DataFrame(records, columns=["供应商姓名", "应付金额", "备注", "开票总额", "差额", "状态", "核对提示"])
    flagged = df[df["状态"].str.startswith("标黄")].copy()
    payable = df[~df["状态"].str.startswith("标黄")].copy()
    # 运行报告
    from collections import Counter
    cnt = Counter(df["状态"])
    report = pd.DataFrame(
        [(k, v) for k, v in sorted(cnt.items())] + [("合计", len(df)), ("标黄合计", len(flagged))],
        columns=["状态", "人数"])
    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="主核对表", index=False)
        flagged.to_excel(w, sheet_name="不付名单(催票)", index=False)
        payable.to_excel(w, sheet_name="可付名单", index=False)
        report.to_excel(w, sheet_name="运行报告", index=False)
        # 主核对表：标黄行高亮
        yellow_rows = [i + 2 for i, st in enumerate(df["状态"]) if str(st).startswith("标黄")]
        _style(w.sheets["主核对表"], len(df.columns), yellow_rows)
        _style(w.sheets["不付名单(催票)"], len(flagged.columns), list(range(2, len(flagged) + 2)))
        _style(w.sheets["可付名单"], len(payable.columns), [])
    return out_path, df, flagged, payable, report


# ----------------- 认文件 -----------------
def _scan(path):
    """粗看一个 xlsx 是清单还是发票台账。"""
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return None
    txt = " ".join(wb.sheetnames)
    head = ""
    try:
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            head += " ".join(str(c) for c in row if c is not None)
    except Exception:
        pass
    wb.close()
    blob = txt + " " + head
    if "销售方信息纳税人识别号" in blob or ("发票" in blob and "合计金额" in blob):
        return "invoice"
    if "应付金额" in blob or "供应商姓名" in blob or "国内个人" in blob:
        return "list"
    return None


def find_inputs(input_dir):
    lst = inv = None
    if os.path.isdir(input_dir):
        for fn in sorted(os.listdir(input_dir)):
            if not fn.lower().endswith((".xlsx", ".xls")) or fn.startswith("~$"):
                continue
            kind = _scan(os.path.join(input_dir, fn))
            if kind == "list" and not lst:
                lst = os.path.join(input_dir, fn)
            elif kind == "invoice" and not inv:
                inv = os.path.join(input_dir, fn)
    return lst, inv


def inspect_mode(input_dir):
    lst, inv = find_inputs(input_dir)
    print(f"识别输入目录：{input_dir}")
    print(f"  待支付清单 = {lst or '（没认到，需手动 --list）'}")
    print(f"  发票台账   = {inv or '（没认到，需手动 --invoice）'}")
    la, ia = load_aliases()
    if lst:
        rows, hdr, w = read_list(lst, la)
        print(f"  清单：{len(rows)} 人；表头={hdr}")
        for x in w:
            print(f"    ⚠ {x}")
    if inv:
        d, hdr, w = read_invoices(inv, ia)
        print(f"  发票：{len(d['sum_by_id'])} 个身份证；表头={hdr}")
        for x in w:
            print(f"    ⚠ {x}")


# ----------------- 主流程 -----------------
def run(list_path, inv_path, out_path):
    load_rules()
    la, ia = load_aliases()
    clist, lhdr, lwarn = read_list(list_path, la)
    inv, ihdr, iwarn = read_invoices(inv_path, ia)
    for x in lwarn + iwarn:
        log(f"⚠ {x}")
    log(f"· 清单 {len(clist)} 人；发票台账 {len(inv['sum_by_id'])} 个身份证、"
        f"{sum(inv['cnt_by_id'].values())} 张票")
    log(f"· 门槛>{CONFIG['THRESHOLD']:g} 才要票｜容差≤{CONFIG['TOLERANCE']:g}｜匹配键=身份证号")
    records = classify(clist, inv)
    out_path, df, flagged, payable, report = write_output(out_path, records)
    log("· 结果：")
    for _, r in report.iterrows():
        log(f"    {r['状态']}: {r['人数']}")
    log(f"· 已写出 → {out_path}")
    return out_path


def main():
    load_rules()
    ap = argparse.ArgumentParser(description="劳务发票核对（第一阶段·支付前）")
    ap.add_argument("--list", dest="list_path")
    ap.add_argument("--invoice", dest="inv_path")
    ap.add_argument("--out")
    ap.add_argument("--input-dir", default=WORK_INPUT)
    ap.add_argument("--inspect", action="store_true")
    a = ap.parse_args()
    if a.inspect:
        inspect_mode(a.input_dir); return
    lp, ip = a.list_path, a.inv_path
    if not lp or not ip:
        flp, fip = find_inputs(a.input_dir)
        lp = lp or flp; ip = ip or fip
    if not lp or not ip:
        log("✗ 没找齐输入：需要『待支付清单』和『发票台账』两个 xlsx。"
            "放进 工作区/input/ 或用 --list/--invoice 指定。")
        sys.exit(2)
    out = a.out
    if not out:
        base = os.path.dirname(os.path.abspath(lp))
        out = os.path.join(base, "劳务发票核对结果.xlsx")
    run(lp, ip, out)


if __name__ == "__main__":
    main()
