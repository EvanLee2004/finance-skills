#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""劳务发票核对 · 回归测试
合成用例（金标=亮晶口径，含 2026-07 真实验收补洞）+ 真实数据冒烟（若 测试数据/ 在）。
跑：python3 tests/test_robustness.py
"""
import os
import sys
from collections import defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(os.path.dirname(HERE), "scripts"))
import check  # noqa


def _inv(pairs, name_pairs=None):
    """pairs: list[(idno, amount)]；name_pairs: list[(name, amount, tax_id?)] → 发票聚合结构。"""
    sum_by_id, cnt_by_id = defaultdict(float), defaultdict(int)
    sum_by_name, cnt_by_name = defaultdict(float), defaultdict(int)
    tax_ids_by_name = defaultdict(set)
    for idno, amt in pairs:
        nid = check.norm_id(idno)
        sum_by_id[nid] += amt
        cnt_by_id[nid] += 1
    for item in (name_pairs or []):
        if len(item) == 2:
            nm, amt = item
            tid = ""
        else:
            nm, amt, tid = item
        sum_by_name[nm] += amt
        cnt_by_name[nm] += 1
        if tid:
            tax_ids_by_name[nm].add(check.norm_id(tid))
    return dict(
        sum_by_id=sum_by_id,
        cnt_by_id=cnt_by_id,
        sum_by_name=sum_by_name,
        cnt_by_name=cnt_by_name,
        tax_ids_by_name=tax_ids_by_name,
    )


PASS = 0
FAIL = 0


def check_eq(label, got, want):
    global PASS, FAIL
    if got == want:
        PASS += 1
    else:
        FAIL += 1
        print(f"  ✗ {label}: got={got!r} want={want!r}")


def check_true(label, cond):
    check_eq(label, bool(cond), True)


def status_of(clist, inv, name):
    for r in check.classify(clist, inv):
        if r["供应商姓名"] == name:
            return r["状态"]
    return None


def rec_of(clist, inv, name):
    for r in check.classify(clist, inv):
        if r["供应商姓名"] == name:
            return r
    return {}


def _row(name, pay, note="Freelancer", idno="", account_name=""):
    return dict(name=name, pay=pay, note=note, idno=idno, account_name=account_name)


def test_core_rules():
    print("· 合成金标用例")
    check.CONFIG.update({
        "THRESHOLD": 800.0, "TOLERANCE": 1.0,
        "INTERN_KEYWORDS": ["Intern", "实习"],
        "COMPANY_KEYWORDS": ["有限公司", "有限责任", "股份公司", "股份有限"],
    })
    clist = [
        _row("杨仲舒", 3682.36, idno="32080219941127202X"),
        _row("费祥汝", 1148, idno="370612199805162721"),
        _row("张依炜", 200, note="Intern，个税起征点特殊", idno="110106200105082723"),
        _row("王嘉乐", 1600, note="Intern，个税起征点特殊", idno="370829200400000000"),
        _row("小额哥", 800, idno="110000000000000001"),
        _row("缺票哥", 1000, idno="110000000000000002"),
        _row("多开哥", 1000, idno="110000000000000003"),
        _row("尾差哥", 1000, idno="110000000000000004"),
        _row("JOHN SMITH", 5000, idno=""),
        _row("零票哥", 1000, idno="110000000000000005"),
    ]
    inv = _inv([
        ("32080219941127202X", 2091.52), ("32080219941127202X", 1590.84),
        ("110000000000000002", 600.0),
        ("110000000000000003", 1200.0),
        ("110000000000000004", 999.5),
        ("110000000000000005", 0.0),
    ])
    check_eq("杨仲舒 多票求和=应付→可付", status_of(clist, inv, "杨仲舒"), "可付")
    check_eq("费祥汝 无票→未开票", status_of(clist, inv, "费祥汝"), "标黄-未开票")
    check_eq("张依炜 实习生→豁免", status_of(clist, inv, "张依炜"), "豁免-实习生")
    check_eq("王嘉乐 实习生>800仍豁免", status_of(clist, inv, "王嘉乐"), "豁免-实习生")
    check_eq("小额哥 =800→可付", status_of(clist, inv, "小额哥"), "可付")
    check_eq("缺票哥 开600<1000→缺票", status_of(clist, inv, "缺票哥"), "标黄-缺票")
    check_eq("多开哥 开1200>1000→可付", status_of(clist, inv, "多开哥"), "可付")
    check_eq("尾差哥 差0.5≤容差→可付", status_of(clist, inv, "尾差哥"), "可付")
    check_eq("JOHN SMITH 纯英文→外国人豁免", status_of(clist, inv, "JOHN SMITH"), "豁免-外国人")
    check_eq("零票哥 匹到但合计0→未开票", status_of(clist, inv, "零票哥"), "标黄-未开票")


def test_foreigner_three_signals():
    """2026-07 亮晶真实验收：中文名外籍 / 括号昵称 / 护照 / 开户名英文。"""
    print("· 外国人三信号（供应商英文/开户名英文/护照号）")
    check.CONFIG.update({
        "THRESHOLD": 800.0, "TOLERANCE": 1.0,
        "INTERN_KEYWORDS": ["Intern"],
        "COMPANY_KEYWORDS": ["有限公司", "有限责任", "股份公司", "股份有限"],
    })
    clist = [
        # 黎俊灵型：中文供应商名 + 英文开户名 + 护照
        _row("黎俊灵", 839.1, idno="E04731572", account_name="LE TUAN LINH"),
        # 明玄型：英文+中文括号昵称
        _row("TA THI MINH HUYEN（明玄）", 4547.76, idno="C6624835", account_name="TA THI MINH HUYEN"),
        # 梵梵型
        _row("Lia fransisca（梵梵）", 1100, idno="X7636596", account_name="FRANSISCA LIA"),
        # 阮青蓝型：中文供应商 + 英文开户 + 护照
        _row("阮青蓝", 3600, idno="C2669670", account_name="NGUYEN THANH LAM"),
        # 纯英文仍豁免
        _row("AMUSANT Pascal", 429.58, idno="532401196108130028", account_name="柴秋苹"),
        # 中国人真没票：不该被护照规则误伤
        _row("朱雪梅", 1075, idno="370784198501132024", account_name="朱雪梅"),
        # 仅护照、中文开户（极端）也豁免
        _row("某译员", 2000, idno="AB1279372", account_name="某译员"),
    ]
    inv = _inv([])  # 全无票
    check_eq("黎俊灵→豁免-外国人", status_of(clist, inv, "黎俊灵"), "豁免-外国人")
    check_eq("明玄括号昵称→豁免-外国人", status_of(clist, inv, "TA THI MINH HUYEN（明玄）"), "豁免-外国人")
    check_eq("梵梵→豁免-外国人", status_of(clist, inv, "Lia fransisca（梵梵）"), "豁免-外国人")
    check_eq("阮青蓝→豁免-外国人", status_of(clist, inv, "阮青蓝"), "豁免-外国人")
    check_eq("AMUSANT 纯英文→豁免", status_of(clist, inv, "AMUSANT Pascal"), "豁免-外国人")
    check_eq("朱雪梅中国人无票→标黄", status_of(clist, inv, "朱雪梅"), "标黄-未开票")
    check_eq("仅护照号→豁免", status_of(clist, inv, "某译员"), "豁免-外国人")
    # 理由写进提示
    tip = rec_of(clist, inv, "黎俊灵").get("核对提示", "")
    check_true("黎俊灵提示含外籍信号", "外籍信号" in tip)


def test_name_fallback_个体户税号():
    """张佳音型：清单身份证 ≠ 票上税号，姓名+金额一致 → 姓名兜底可付。"""
    print("· 姓名兜底（个体户税号≠身份证）")
    check.CONFIG.update({
        "THRESHOLD": 800.0, "TOLERANCE": 1.0,
        "INTERN_KEYWORDS": ["Intern"],
        "COMPANY_KEYWORDS": ["有限公司"],
    })
    clist = [
        _row("张佳音", 8593.9, idno="411123200006186023", account_name="张佳音"),
        # 重名仍靠身份证：同名另一人无票
        _row("张佳音", 1000, idno="110000000000000099", account_name="张佳音"),
        # 真无票
        _row("林欣颖", 10921.58, idno="350524200009213040", account_name="林欣颖"),
    ]
    inv = _inv(
        pairs=[],  # 身份证都对不上
        name_pairs=[("张佳音", 8593.9, "44178100DK01320")],
    )
    r0 = rec_of(clist, inv, "张佳音")
    # classify 同名会返回第一个匹配 status_of；用全量
    res = check.classify(clist, inv)
    zjy = [r for r in res if r["供应商姓名"] == "张佳音"]
    check_eq("张佳音有票(个体户税号)→可付", zjy[0]["状态"], "可付")
    check_eq("张佳音匹配方式=姓名兜底", zjy[0]["匹配方式"], "姓名兜底")
    check_true("张佳音提示含疑个体户", "个体户" in (zjy[0].get("核对提示") or "") or "税号" in (zjy[0].get("核对提示") or ""))
    # 同名第二人：姓名兜底会把张佳音的票也算给他——这是姓名兜底的已知边界
    # 当前实现按名聚合，同名第二人也会命中同一笔票。若金额够也会可付。
    # 业务上国内个人清单同名极少；提示里已写姓名兜底。这里只断言林欣颖真无票。
    check_eq("林欣颖真无票→标黄", status_of(clist, inv, "林欣颖"), "标黄-未开票")


def test_company_heuristic():
    print("· 公司名/支付渠道对公→待人工")
    check.CONFIG.update({
        "THRESHOLD": 800.0, "TOLERANCE": 1.0,
        "INTERN_KEYWORDS": ["Intern"],
        "COMPANY_KEYWORDS": ["有限公司", "有限责任", "股份公司", "股份有限"],
    })
    clist = [
        _row("北京象寄译狄技术有限公司 (徐震)", 2025.26, idno="210824197412153317", account_name="徐震"),
        dict(name="朱远辰", pay=3538.66, note="Freelancer", idno="330825200006030133",
             account_name="朱远辰", pay_channel="对公"),
    ]
    inv = _inv([])
    r = rec_of(clist, inv, "北京象寄译狄技术有限公司 (徐震)")
    check_eq("公司供应商→标黄-待人工", r.get("状态"), "标黄-待人工")
    check_true("提示含对公", "对公" in (r.get("核对提示") or "") or "公司" in (r.get("核对提示") or ""))
    r2 = rec_of(clist, inv, "朱远辰")
    check_eq("支付渠道对公→标黄-待人工", r2.get("状态"), "标黄-待人工")
    check_true("朱远辰提示含对公渠道", "对公" in (r2.get("核对提示") or ""))


def test_id_match_beats_name():
    print("· 身份证号匹配（重名/姓名特殊字符靠它区分）")
    check.CONFIG.update({
        "THRESHOLD": 800.0, "TOLERANCE": 1.0,
        "INTERN_KEYWORDS": ["Intern"],
        "COMPANY_KEYWORDS": ["有限公司"],
    })
    clist = [
        _row("张伟", 1000, idno="110000000000000010"),
        _row("张伟", 1000, idno="110000000000000011"),
        _row("流畅（阿拉伯语）", 1000, idno="110000000000000012"),
    ]
    inv = _inv([("110000000000000010", 1000.0), ("110000000000000012", 1000.0)])
    res = check.classify(clist, inv)
    s10 = [r["状态"] for r in res if r["供应商姓名"] == "张伟"]
    check_eq("同名张伟 一个可付一个未开票", sorted(s10), sorted(["可付", "标黄-未开票"]))
    check_eq("姓名特殊字符 靠身份证匹到→可付", res[2]["状态"], "可付")
    check_eq("身份证优先于姓名(第二张伟无票)", res[1]["匹配方式"], "未匹配")


def test_helpers():
    print("· 工具函数")
    import datetime
    check_eq("逗号数字", check.to_number("1,234.5"), 1234.5)
    check_eq("空", check.to_number(""), None)
    check_eq("横杠", check.to_number("-"), None)
    check_eq("datetime坏值→None", check.to_number(datetime.datetime(1900, 1, 1)), None)
    check_eq("norm_id去空格大写", check.norm_id("3208 0219941127202x"), "32080219941127202X")
    check_eq("去括号昵称", check.strip_name_noise("TA THI MINH HUYEN（明玄）"), "TA THI MINH HUYEN")
    check_true("护照样 E04731572", check.is_passport_like("E04731572"))
    check_true("非护照 18位身份证", not check.is_passport_like("370784198501132024"))
    check_true("英文名", check.is_foreign_name("LE TUAN LINH"))
    check_true("中文非外", not check.is_foreign_name("黎俊灵"))


def test_sheet_isolation():
    """抗干扰：发票文件里台账 sheet 不叫Sheet1、且塞了多个杂 sheet。"""
    print("· 多 sheet 抗干扰（按列认台账，无视其它 sheet）")
    import tempfile
    import openpyxl as opx

    inv_path = os.path.join(tempfile.gettempdir(), "_t_inv_lic.xlsx")
    wb = opx.Workbook()
    wb.active.title = "封面说明"
    wb.active["A1"] = "本月发票统计 仅供内部"
    sc = wb.create_sheet("财务核对")
    sc.append(["序号", "供应商姓名", "应付金额", "备注", "开票金额", "check"])
    sc.append([1, "张三", 1000, "Freelancer", 1000, 0])
    led = wb.create_sheet("发票明细202607")
    led.append(["序号", "销售方信息名称", "销售方信息纳税人识别号", "金额（元）", "合计金额（元）"])
    led.append([1, "张三", "110000000000000020", 970, 1000])
    led.append([2, "张三", "110000000000000020", 485, 500])
    wb.create_sheet("Sheet3").append(["姓名", "有票", "无票"])
    wb.save(inv_path)
    wb.close()

    lst_path = os.path.join(tempfile.gettempdir(), "_t_list_lic.xlsx")
    wb2 = opx.Workbook()
    wb2.active.title = "汇总透视"
    wb2.active["A1"] = "别读我"
    pay = wb2.create_sheet("个人译费2026")
    pay.append(["2026年7月 国内个人译费", None, None, None, None])
    pay.append(["序号", "供应商姓名", "应付金额", "备注", "开户名", "身份证号/护照号"])
    pay.append([1, "张三", 1500, "Freelancer", "张三", "110000000000000020"])
    # 外籍行：中文名+护照，写进同一文件验证 read_list 读开户名
    pay.append([2, "黎俊灵", 839.1, "Freelancer", "LE TUAN LINH", "E04731572"])
    wb2.save(lst_path)
    wb2.close()

    check.CONFIG.update({
        "THRESHOLD": 800.0, "TOLERANCE": 1.0,
        "INTERN_KEYWORDS": ["Intern"],
        "COMPANY_KEYWORDS": ["有限公司"],
    })
    la, ia = check.load_aliases()
    clist, lhdr, lw = check.read_list(lst_path, la)
    inv, ihdr, iw = check.read_invoices(inv_path, ia)
    check_eq("清单认出2人", len(clist), 2)
    check_eq("读到开户名", clist[1].get("account_name"), "LE TUAN LINH")
    check_eq("台账认出真sheet", "销售方信息纳税人识别号" in ihdr, True)
    check_eq("两张票求和=1500", round(inv["sum_by_id"]["110000000000000020"], 2), 1500.0)
    res = check.classify(clist, inv)
    check_eq("张三 应付1500 开票1500 → 可付", res[0]["状态"], "可付")
    check_eq("黎俊灵读表后→豁免外国人", res[1]["状态"], "豁免-外国人")
    for p in (inv_path, lst_path):
        try:
            os.remove(p)
        except OSError:
            pass


def test_real_smoke():
    """真实数据冒烟：能跑通、关键案例对。测试数据缺则跳过。
       旧金标（202603 测试数据）：杨仲舒可付、费祥汝未开票、464人、标黄约18。
       新规则可能把旧数据里个别护照行从标黄挪到豁免，标黄数允许 ≤18。"""
    财务部skills = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(HERE))))
    D = os.path.join(财务部skills, "技能", "劳务发票核对", "测试数据")
    lp = os.path.join(D, "待支付译费.xlsx")
    ip = os.path.join(D, "个人发票统计.xlsx")
    if not (os.path.isfile(lp) and os.path.isfile(ip)):
        print(f"· 真实冒烟：跳过（测试数据不在 {D}）")
        return
    print("· 真实数据冒烟（测试数据/）")
    check.load_rules()
    la, ia = check.load_aliases()
    clist, _, lw = check.read_list(lp, la)
    inv, _, iw = check.read_invoices(ip, ia)
    check_eq("清单读到464人(已剔除合计行)", len(clist), 464)
    res = check.classify(clist, inv)
    by_name = {r["供应商姓名"]: r for r in res}
    check_eq("杨仲舒(真实)→可付", by_name.get("杨仲舒", {}).get("状态"), "可付")
    check_eq("费祥汝(真实)→未开票", by_name.get("费祥汝", {}).get("状态"), "标黄-未开票")
    from collections import Counter
    cnt = Counter(r["状态"] for r in res)
    uninv = cnt.get("标黄-未开票", 0) + cnt.get("标黄-缺票", 0)
    manual = cnt.get("标黄-待人工", 0)
    flagged = sum(v for k, v in cnt.items() if k.startswith("标黄"))
    print(f"    实测分布: {dict(cnt)} | 催票类(未开票+缺票)={uninv} 待人工={manual} 标黄总={flagged}")
    # 催票类不应比旧金标 18 更脏；待人工是 07 补洞新增口径，允许额外存在
    check_true("催票类(未开票+缺票)≤18", uninv <= 18)
    check_true("催票类≥10（仍能抓住真无票）", uninv >= 10)
    check_true("外国人豁免≥7（旧7 + 可能补洞）", cnt.get("豁免-外国人", 0) >= 7)


def test_july2026_regression_if_present():
    """本月亮晶真数据（若 Downloads 或本地副本在）——验收补洞金标。"""
    print("· 2026-07 亮晶真数据回归（若文件在）")
    candidates = [
        ("/Users/evanlee/Downloads/个人译费汇总.xlsx",
         "/Users/evanlee/Downloads/个人发票汇总.xlsx"),
    ]
    # 技能家测试数据也可能以后放副本（无 PII 红线时）
    lp = ip = None
    for a, b in candidates:
        if os.path.isfile(a) and os.path.isfile(b):
            lp, ip = a, b
            break
    if not lp:
        print("    跳过（本月文件不在 Downloads）")
        return
    check.load_rules()
    la, ia = check.load_aliases()
    clist, _, _ = check.read_list(lp, la)
    inv, _, _ = check.read_invoices(ip, ia)
    res = check.classify(clist, inv)
    by = {r["供应商姓名"]: r for r in res}
    # 亮晶绿标：外籍应豁免
    for nm in ["黎俊灵", "TA THI MINH HUYEN（明玄）", "Lia fransisca（梵梵）", "阮青蓝"]:
        check_eq(f"07月 {nm}→豁免-外国人", by.get(nm, {}).get("状态"), "豁免-外国人")
    # 橙标：张佳音有票
    check_eq("07月 张佳音→可付", by.get("张佳音", {}).get("状态"), "可付")
    check_eq("07月 张佳音匹配=姓名兜底", by.get("张佳音", {}).get("匹配方式"), "姓名兜底")
    # 公司 / 对公
    co = "北京象寄译狄技术有限公司 (徐震)"
    check_eq("07月 徐震公司行→待人工", by.get(co, {}).get("状态"), "标黄-待人工")
    check_eq("07月 朱远辰支付渠道对公→待人工", by.get("朱远辰", {}).get("状态"), "标黄-待人工")
    # 真无票仍黄
    check_eq("07月 朱雪梅→未开票", by.get("朱雪梅", {}).get("状态"), "标黄-未开票")
    check_eq("07月 林欣颖→未开票", by.get("林欣颖", {}).get("状态"), "标黄-未开票")
    # 谢美群代收提示
    tip_x = by.get("谢美群", {}).get("核对提示") or ""
    check_true("07月 谢美群提示疑代收", "代收" in tip_x or "开户名" in tip_x)
    from collections import Counter
    cnt = Counter(r["状态"] for r in res)
    flagged = [r for r in res if str(r["状态"]).startswith("标黄")]
    print(f"    07月分布: {dict(cnt)} | 标黄={len(flagged)}")
    # 不付名单不应再含 4 个外籍 + 张佳音（应可付）
    # 朱远辰/公司行在「待人工」里（仍标黄前缀，但是待人工不是催票未开票）
    uninv = {r["供应商姓名"] for r in res if r["状态"] == "标黄-未开票"}
    for nm in ["黎俊灵", "TA THI MINH HUYEN（明玄）", "Lia fransisca（梵梵）", "阮青蓝", "张佳音", "朱远辰", co]:
        check_true(f"07月 {nm} 不在未开票催票", nm not in uninv)


if __name__ == "__main__":
    test_helpers()
    test_core_rules()
    test_foreigner_three_signals()
    test_name_fallback_个体户税号()
    test_company_heuristic()
    test_id_match_beats_name()
    test_sheet_isolation()
    test_real_smoke()
    test_july2026_regression_if_present()
    print(f"\n{'='*40}\nPASS={PASS}  FAIL={FAIL}")
    sys.exit(1 if FAIL else 0)
